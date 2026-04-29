"""
transformar_profundidad.py — Convierte Base Profundidad Ripley a Tab 1, 3, 4 de Capi
=====================================================================================
Input:  Base Profundidad (301 cols, formato ancho con tiendas en columnas)
Output: Plantilla Capi v2 con Tab 1 (Maestro), Tab 3 (Ventas 4 sem), Tab 4 (Stock)
        Tab 2 queda vacía (se llena con el transformador de LY aparte)
"""

import pandas as pd
import numpy as np
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Estilos ───
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT  = Font(name='Arial', bold=True, size=12, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', size=10, color='666666')
DATA_FONT   = Font(name='Arial', size=10)
CALC_FILL   = PatternFill('solid', fgColor='E2EFDA')
CALC_FONT   = Font(name='Arial', size=10, color='008000')
GREEN_HEADER = PatternFill('solid', fgColor='548235')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data(ws, row, ncols, calc_cols=None):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if calc_cols and c in calc_cols:
            cell.fill = CALC_FILL
            cell.font = CALC_FONT

# ─── Store code → full name mapping ───
STORE_NAMES = {
    'JP': 'Jockey Plaza', 'SM': 'San Miguel', 'MIRAF': 'Miraflores',
    'SI': 'San Isidro', 'PRIM': 'Primavera', 'LO': 'Mega Plaza',
    'AQP': 'Arequipa', 'CHIC': 'Chiclayo', 'CHO': 'Chorrillos',
    'Breña': 'Breña', 'PLN': 'Plaza Norte', 'SLVR': 'Salaverry',
    'SB': 'San Borja', 'STA ANITA': 'Santa Anita', 'TRUJ': 'Trujillo',
    'CAJ': 'Cajamarca', 'CBT': 'Chimbote', 'HYO': 'Huancayo',
    'ICA': 'Ica', 'JULIACA': 'Juliaca', 'PIU2': 'Piura',
    'MAC LARCO': 'Mac Larco', 'CALLAO': 'Callao',
    'TV PI': 'Tienda Virtual PI', 'FSF MAC LO': 'FSF Mac LO',
    'TV': 'Tienda Virtual', 'ASIA': 'Asia',
    'VTAS CORP': 'Ventas Corporativas', 'ATO': 'Atocongo',
    'CAY': 'Cayma', 'PUCALPA I': 'Pucallpa',
    'MAC PLN': 'Mac Plaza Norte', 'BTQ BBB': 'Boutique BBB',
    'EST CHI': 'Estacional Chiclayo', 'PURU': 'Puruchuco',
    'EST TRU': 'Estacional Trujillo', 'EST VES': 'Estacional VES',
    'CHII': 'Chiclayo 2', 'CO': 'Comas', 'IQT': 'Iquitos',
    'OPLN': 'Outlet Plaza Norte', 'OSI': 'Outlet San Isidro', 'SJL': 'San Juan de Lurigancho',
}

# Detect store codes from columns
def detect_stores(columns):
    """Find all store prefixes that have Stk/Vta/On Order columns."""
    stk_cols = [c for c in columns if c.endswith(' Stk')]
    stores = []
    for c in stk_cols:
        prefix = c[:-4]  # Remove ' Stk'
        # Verify Vta column exists
        if f'{prefix} Vta' in columns:
            stores.append(prefix)
    return stores


def transform(input_path, output_path=None, fecha_corte='08/04/2026'):
    if output_path is None:
        output_path = os.path.join(os.path.dirname(input_path), 'Plantilla_Capi_Profundidad.xlsx')

    print(f'Leyendo {input_path}...')
    df = pd.read_excel(input_path)
    print(f'  → {len(df):,} filas, {df["Cód. Prod."].nunique()} modelos')

    # Detect stores
    stores = detect_stores(df.columns)
    print(f'  → {len(stores)} tiendas detectadas: {stores[:10]}...')

    # Clean data types
    df['Cód. Prod.'] = df['Cód. Prod.'].astype(str).str.strip()
    df['Descripción'] = df['Descripción'].astype(str).str.strip()

    for col in ['Antigüedad semanal', 'P.Blanco', 'P.Vigente PMM', 'Costo S/.',
                'Unid. Vendidas Sem. 1ant', 'Unid. Vendidas Sem. 2ant',
                'Unid. Vendidas Sem. 3ant', 'Unid. Vendidas Sem. 4ant',
                'STOCK ACTUAL (UNIDADES)', 'Total Unidades',
                'Total Tiendas Unidades']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Detectar columna de Stock CD (varios nombres posibles en Base Profundidad)
    _stock_cd_col = None
    for _candidate in ['Total  CD+Bodega Unidades',
                       'Total CD+Bodega Unid. (On-hand disponible)',
                       'Total CD+Bodega Unidades',
                       'Total  CD+Bodega Unid. (On-hand disponible)']:
        if _candidate in df.columns:
            _stock_cd_col = _candidate
            break
    # Fallback: buscar por substring
    if _stock_cd_col is None:
        for c in df.columns:
            c_lower = str(c).lower()
            if 'cd' in c_lower and 'bodega' in c_lower and ('unid' in c_lower or 'und' in c_lower):
                _stock_cd_col = c
                break
    if _stock_cd_col:
        df[_stock_cd_col] = pd.to_numeric(df[_stock_cd_col], errors='coerce').fillna(0)
        df['_stock_cd_resolved'] = df[_stock_cd_col]
        print(f'  → Stock CD columna detectada: "{_stock_cd_col}"')
    else:
        df['_stock_cd_resolved'] = 0
        print('  ⚠️ No se encontró columna de Stock CD en Base Profundidad')

    # Rango de antigüedad (categórico del reporte micro)
    if 'Antiguedad' not in df.columns:
        df['Antiguedad'] = 'Sin Rango'
    else:
        df['Antiguedad'] = df['Antiguedad'].fillna('Sin Rango').astype(str).str.strip()

    # ── Embarque / Ventana de compra ──
    _embarque_col = None
    for _cand in ['Embarque', 'embarque', 'EMBARQUE', 'Ventana', 'ventana']:
        if _cand in df.columns:
            _embarque_col = _cand
            break
    if _embarque_col is None:
        for c in df.columns:
            if 'embarque' in str(c).lower():
                _embarque_col = c
                break
    if _embarque_col:
        # Parsear: "A-A" → "A", "B-F" → "B", vacío/NaN → "NOOS"
        # NaN.astype(str) → "nan" → upper → "NAN" → slice(0,1) → "N"
        df['_ventana_compra'] = (
            df[_embarque_col]
            .astype(str).str.strip()
            .str.upper()
            .str.slice(0, 1)        # Primera letra antes del guión
        )
        # Limpiar: vacíos, NaN convertidos ("N" de "NAN"), guiones sueltos → NOOS
        _invalid = df['_ventana_compra'].isin(['', 'N', '-']) | df[_embarque_col].isna()
        df.loc[_invalid, '_ventana_compra'] = 'NOOS'
        print(f'  → Embarque columna detectada: "{_embarque_col}"')
        print(f'    Ventanas: {df["_ventana_compra"].value_counts().to_dict()}')
    else:
        df['_ventana_compra'] = 'NOOS'
        print('  ⚠️ No se encontró columna de Embarque en Base Profundidad')

    # ── Venta en Soles y Contribución (nivel SKU, no tienda) ──
    _vta_s_cols = []
    _contrib_cols = []
    for _sem in ['4ant', '3ant', '2ant', '1ant']:
        _vc = f'Vta. S/. Semana {_sem}'
        _cc = f'Contrib. S/. Sem. {_sem}'
        # Buscar variantes de nombre
        for c in df.columns:
            cn = str(c).strip()
            if cn == _vc or (f'vta' in cn.lower() and f's/' in cn.lower() and _sem in cn.lower()):
                df[cn] = pd.to_numeric(df[cn], errors='coerce').fillna(0)
                _vta_s_cols.append(cn)
                break
        for c in df.columns:
            cn = str(c).strip()
            if cn == _cc or (f'contrib' in cn.lower() and f's/' in cn.lower() and _sem in cn.lower()):
                df[cn] = pd.to_numeric(df[cn], errors='coerce').fillna(0)
                _contrib_cols.append(cn)
                break

    if _vta_s_cols:
        df['_vta_soles_4sem'] = df[_vta_s_cols].sum(axis=1)
        print(f'  → Vta S/ detectada: {len(_vta_s_cols)} semanas, total = S/{df["_vta_soles_4sem"].sum():,.0f}')
    else:
        df['_vta_soles_4sem'] = 0
        print('  ⚠️ No se encontraron columnas de Vta S/ semanal')

    if _contrib_cols:
        df['_contrib_soles_4sem'] = df[_contrib_cols].sum(axis=1)
        print(f'  → Contrib S/ detectada: {len(_contrib_cols)} semanas, total = S/{df["_contrib_soles_4sem"].sum():,.0f}')
    else:
        df['_contrib_soles_4sem'] = 0
        print('  ⚠️ No se encontraron columnas de Contrib S/ semanal')

    # Margen efectivo a nivel SKU
    df['_margen_efectivo'] = 0.0
    _vta_mask = df['_vta_soles_4sem'] > 0
    df.loc[_vta_mask, '_margen_efectivo'] = df.loc[_vta_mask, '_contrib_soles_4sem'] / df.loc[_vta_mask, '_vta_soles_4sem']
    if _vta_mask.sum() > 0:
        _margen_pond = df.loc[_vta_mask, '_contrib_soles_4sem'].sum() / df.loc[_vta_mask, '_vta_soles_4sem'].sum()
        print(f'  → Margen efectivo ponderado (SKUs con venta): {_margen_pond:.1%} ({_vta_mask.sum()} SKUs)')

    for s in stores:
        for suffix in ['Stk', 'Vta', 'On Order']:
            col = f'{s} {suffix}'
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # ════════════════════════════════════════════════════════════
    #  FILTRO DE MERCADERÍA ACTIVA
    # ════════════════════════════════════════════════════════════
    #  Reglas:
    #  1. Sin stock total = fuera
    #  2. OI con edad > 52 semanas = fuera (invierno pasado)
    #  3. PV con edad >= 40 semanas = fuera (PV viejo)
    n_before = len(df)

    # Stock total por SKU (sum across all stores)
    df['_total_stk'] = sum(df.get(f'{s} Stk', 0) for s in stores)

    excl_no_stock = df['_total_stk'] == 0
    excl_oi_old = (df['Temp.'] == 'OI') & (df['Antigüedad semanal'] > 52)
    excl_pv_old = (df['Temp.'] == 'PV') & (df['Antigüedad semanal'] >= 40)
    excluded = excl_no_stock | excl_oi_old | excl_pv_old

    n_excl_stock = excl_no_stock.sum()
    n_excl_oi = (excl_oi_old & ~excl_no_stock).sum()
    n_excl_pv = (excl_pv_old & ~excl_no_stock & ~excl_oi_old).sum()

    df = df[~excluded].copy()
    print(f'  Filtro mercadería activa: {n_before} → {len(df)} SKUs')
    print(f'    Excluidos sin stock: {n_excl_stock}')
    print(f'    Excluidos OI >52 sem: {n_excl_oi}')
    print(f'    Excluidos PV >=40 sem: {n_excl_pv}')

    df = df.drop(columns=['_total_stk'])

    # ════════════════════════════════════════════════════════════
    #  UNPIVOT: Wide → Long (SKU × Tienda)
    # ════════════════════════════════════════════════════════════
    print('Despivoteando tiendas...')

    rows_ventas = []
    rows_stock = []

    for _, r in df.iterrows():
        sku = r['Cód. Prod.']
        nombre = r['Descripción']
        categoria = r['Dpto']
        edad = r['Antigüedad semanal']

        # Total weekly sales (across all stores)
        sem1_total = r['Unid. Vendidas Sem. 1ant']
        sem2_total = r['Unid. Vendidas Sem. 2ant']
        sem3_total = r['Unid. Vendidas Sem. 3ant']
        sem4_total = r['Unid. Vendidas Sem. 4ant']

        # Per-store: Vta = last week (≈ Sem 1)
        # Calculate each store's share of Sem 1 to distribute Sem 2-4
        store_vtas = {}
        total_store_vta = 0
        for s in stores:
            v = r.get(f'{s} Vta', 0) or 0
            store_vtas[s] = v
            total_store_vta += v

        for s in stores:
            stk = r.get(f'{s} Stk', 0) or 0
            oo = r.get(f'{s} On Order', 0) or 0
            vta_sem1 = store_vtas[s]

            # Skip stores with no stock AND no sales
            if stk == 0 and oo == 0 and vta_sem1 == 0:
                continue

            # Distribute Sem 2-4 proportionally based on Sem 1 share
            if total_store_vta > 0:
                share = vta_sem1 / total_store_vta
            else:
                # No sales anywhere — if store has stock, give it equal share
                active_stores_stk = sum(1 for s2 in stores if (r.get(f'{s2} Stk', 0) or 0) > 0)
                share = (1 / active_stores_stk) if active_stores_stk > 0 and stk > 0 else 0

            vta_sem2 = round(sem2_total * share)
            vta_sem3 = round(sem3_total * share)
            vta_sem4 = round(sem4_total * share)

            tienda_name = STORE_NAMES.get(s, s)

            rows_ventas.append({
                'sku': sku,
                'nombre': nombre,
                'tienda': tienda_name,
                'vta_sem1': int(vta_sem1),
                'vta_sem2': int(vta_sem2),
                'vta_sem3': int(vta_sem3),
                'vta_sem4': int(vta_sem4),
            })

            rows_stock.append({
                'sku': sku,
                'nombre': nombre,
                'tienda': tienda_name,
                'fecha_corte': fecha_corte,
                'stock_uds': int(stk),
                'stock_transito': int(oo),
            })

    df_ventas = pd.DataFrame(rows_ventas)
    df_stock = pd.DataFrame(rows_stock)
    print(f'  → Tab 3 Ventas: {len(df_ventas):,} filas (SKU×Tienda con actividad)')
    print(f'  → Tab 4 Stock:  {len(df_stock):,} filas')

    # ════════════════════════════════════════════════════════════
    #  TAB 1 — MAESTRO (from profundidad, much richer than LY)
    # ════════════════════════════════════════════════════════════
    print('Construyendo Maestro...')

    # Filter: only SKUs that have stock or recent sales
    active_skus = set(df_ventas['sku'].unique()) | set(df_stock[df_stock['stock_uds'] > 0]['sku'].unique())

    df_maestro = df[df['Cód. Prod.'].isin(active_skus)].copy()
    df_maestro = df_maestro.drop_duplicates(subset='Cód. Prod.', keep='first')
    df_maestro = df_maestro.sort_values('Cód. Prod.').reset_index(drop=True)
    print(f'  → {len(df_maestro)} SKUs activos (con stock o venta)')

    # ════════════════════════════════════════════════════════════
    #  ESCRIBIR EXCEL
    # ════════════════════════════════════════════════════════════
    print('Escribiendo Excel...')
    wb = Workbook()

    # ── INSTRUCCIONES ──
    ws0 = wb.active
    ws0.title = 'INSTRUCCIONES'
    ws0.sheet_properties.tabColor = '1F4E79'
    instructions = [
        ('CAPI — Plantilla de Input v2', TITLE_FONT),
        ('Generada desde Base Profundidad Ripley al ' + fecha_corte, SUBTITLE_FONT),
        ('', None),
        ('TAB 1 — MAESTRO: Completo desde Profundidad. Edad, precios y costo incluidos.', DATA_FONT),
        ('TAB 2 — HISTORIAL LY: Vacía. Llenar con transformar_ripley.py si tienes data LY.', DATA_FONT),
        ('TAB 3 — VENTAS RECIENTES: Completo. Sem 1 real por tienda, Sem 2-4 distribuidas proporcionalmente.', DATA_FONT),
        ('TAB 4 — STOCK ACTUAL: Completo. Stock y On Order por tienda.', DATA_FONT),
    ]
    for i, (text, font) in enumerate(instructions, 1):
        cell = ws0.cell(row=i, column=1, value=text)
        if font: cell.font = font
    ws0.column_dimensions['A'].width = 100

    # ── TAB 1: MAESTRO ──
    ws1 = wb.create_sheet('1. Maestro Productos')
    ws1.sheet_properties.tabColor = '2E75B6'
    ws1.cell(row=1, column=1, value='MAESTRO DE PRODUCTOS — Una fila por SKU').font = TITLE_FONT

    headers1 = [
        'SKU\nCódigo Producto', 'Nombre\nProducto', 'Marca',
        'Categoría /\nLínea', 'Departamento', 'Edad\n(Semanas)',
        'Precio\nBlanco (S/)', 'Precio\nVigente (S/)', 'Costo\nUnitario (S/)',
        'Temporada', 'Rango\nAntigüedad', 'Stock CD\n(Unidades)',
        'Vta Sem1\n(uds total)', 'Vta Sem2\n(uds total)',
        'Vta Sem3\n(uds total)', 'Vta Sem4\n(uds total)',
        'Notas',
        '% Descuento\nActual', 'IMU%\n(Margen Inicial)', 'Margen\nVigente %',
        'Ventana\nCompra',
        'Vta S/\n4 sem', 'Contrib S/\n4 sem', 'Margen\nEfectivo %',
    ]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=2, column=c, value=h)
    style_header(ws1, 2, len(headers1))
    for c in [18, 19, 20]:
        ws1.cell(row=2, column=c).fill = GREEN_HEADER

    ws1.merge_cells('A1:X1')

    for i, (_, r) in enumerate(df_maestro.iterrows(), 3):
        ws1.cell(row=i, column=1, value=str(r['Cód. Prod.']))
        ws1.cell(row=i, column=2, value=r['Descripción'])
        ws1.cell(row=i, column=3, value=r.get('Marca', ''))
        ws1.cell(row=i, column=4, value=r.get('Línea', ''))
        ws1.cell(row=i, column=5, value=r.get('Dpto', ''))
        ws1.cell(row=i, column=6, value=int(r['Antigüedad semanal']) if pd.notna(r['Antigüedad semanal']) else '')
        ws1.cell(row=i, column=7, value=round(r['P.Blanco'], 2) if r['P.Blanco'] > 0.01 else '')
        ws1.cell(row=i, column=8, value=round(r['P.Vigente PMM'], 2) if r['P.Vigente PMM'] > 0.01 else '')
        ws1.cell(row=i, column=9, value=round(r['Costo S/.'], 2) if r['Costo S/.'] > 0 else '')
        ws1.cell(row=i, column=10, value=r.get('Temp.', ''))
        ws1.cell(row=i, column=11, value=r.get('Antiguedad', 'Sin Rango'))
        _stock_cd_val = r.get('_stock_cd_resolved', 0)
        ws1.cell(row=i, column=12, value=int(_stock_cd_val) if pd.notna(_stock_cd_val) else 0)
        # Venta semanal total (todas las tiendas) — dato agregado, no por tienda
        ws1.cell(row=i, column=13, value=int(r.get('Unid. Vendidas Sem. 1ant', 0)))
        ws1.cell(row=i, column=14, value=int(r.get('Unid. Vendidas Sem. 2ant', 0)))
        ws1.cell(row=i, column=15, value=int(r.get('Unid. Vendidas Sem. 3ant', 0)))
        ws1.cell(row=i, column=16, value=int(r.get('Unid. Vendidas Sem. 4ant', 0)))
        ws1.cell(row=i, column=17, value='')  # Notas
        ws1.cell(row=i, column=18, value=f'=IF(G{i}>0,1-H{i}/G{i},"")')
        ws1.cell(row=i, column=19, value=f'=IF(G{i}>0,(G{i}/1.18-I{i})/(G{i}/1.18),"")')
        ws1.cell(row=i, column=20, value=f'=IF(H{i}>0,(H{i}/1.18-I{i})/(H{i}/1.18),"")')
        ws1.cell(row=i, column=21, value=r.get('_ventana_compra', 'NOOS'))
        # Venta S/, Contribución S/ y Margen Efectivo (nivel SKU, 4 semanas)
        _vta_s = r.get('_vta_soles_4sem', 0)
        _contrib_s = r.get('_contrib_soles_4sem', 0)
        _margen_e = r.get('_margen_efectivo', 0)
        ws1.cell(row=i, column=22, value=round(_vta_s, 2) if _vta_s > 0 else 0)
        ws1.cell(row=i, column=23, value=round(_contrib_s, 2) if _contrib_s != 0 else 0)
        ws1.cell(row=i, column=24, value=round(_margen_e, 4) if _vta_s > 0 else '')
        style_data(ws1, i, 24, calc_cols={18, 19, 20})
        for c in [7, 8, 9]:
            ws1.cell(row=i, column=c).number_format = '#,##0.00'
        for c in [12, 13, 14, 15, 16]:
            ws1.cell(row=i, column=c).number_format = '#,##0'
        for c in [18, 19, 20, 24]:
            ws1.cell(row=i, column=c).number_format = '0.0%'
        for c in [22, 23]:
            ws1.cell(row=i, column=c).number_format = '#,##0.00'

    widths1 = [14, 45, 20, 18, 22, 10, 12, 12, 12, 8, 14, 12, 10, 10, 10, 10, 15, 12, 12, 12, 10, 14, 14, 12]
    for c, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 2: HISTORIAL LY (vacía — estructura) ──
    ws2 = wb.create_sheet('2. Historial LY (año pasado)')
    ws2.sheet_properties.tabColor = 'BF8F00'
    ws2.cell(row=1, column=1, value='HISTORIAL LY — Llenar con script transformar_ripley.py o dejar vacía').font = TITLE_FONT
    ws2.merge_cells('A1:F1')
    headers2 = ['SKU\nCódigo Producto', 'Tienda /\nCanal', 'Semana\n(nro 1-52)', 'Fecha\nLunes', 'Vta Uds\nLY', 'Vta Soles\nLY (S/)']
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=c, value=h)
    style_header(ws2, 2, len(headers2))

    # ── TAB 3: VENTAS RECIENTES ──
    ws3 = wb.create_sheet('3. Ventas Recientes (4 sem)')
    ws3.sheet_properties.tabColor = '548235'
    ws3.cell(row=1, column=1,
        value='VENTAS RECIENTES — Últimas 4 semanas por SKU × Tienda').font = TITLE_FONT
    ws3.merge_cells('A1:H1')

    headers3 = [
        'SKU\nCódigo Producto', 'Nombre\nProducto', 'Tienda /\nCanal',
        'Vta Uds\nSem 1', 'Vta Uds\nSem 2', 'Vta Uds\nSem 3', 'Vta Uds\nSem 4',
        'Prom. Vta\nUds/Semana',
    ]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=c, value=h)
    style_header(ws3, 2, len(headers3))
    ws3.cell(row=2, column=8).fill = GREEN_HEADER

    print(f'  Escribiendo {len(df_ventas):,} filas de ventas...')
    for i, (_, r) in enumerate(df_ventas.iterrows(), 3):
        ws3.cell(row=i, column=1, value=str(r['sku']))
        ws3.cell(row=i, column=2, value=r['nombre'])
        ws3.cell(row=i, column=3, value=r['tienda'])
        ws3.cell(row=i, column=4, value=r['vta_sem1'])
        ws3.cell(row=i, column=5, value=r['vta_sem2'])
        ws3.cell(row=i, column=6, value=r['vta_sem3'])
        ws3.cell(row=i, column=7, value=r['vta_sem4'])
        ws3.cell(row=i, column=8, value=f'=AVERAGE(D{i}:G{i})')
        ws3.cell(row=i, column=8).number_format = '0.0'
        for c in range(1, 9):
            cell = ws3.cell(row=i, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 8:
                cell.fill = CALC_FILL
                cell.font = CALC_FONT

    widths3 = [14, 45, 22, 10, 10, 10, 10, 12]
    for c, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 4: STOCK ACTUAL ──
    ws4 = wb.create_sheet('4. Stock Actual')
    ws4.sheet_properties.tabColor = 'C00000'
    ws4.cell(row=1, column=1,
        value=f'STOCK ACTUAL POR TIENDA — Corte al {fecha_corte}').font = TITLE_FONT

    headers4 = [
        'SKU\nCódigo Producto', 'Nombre\nProducto', 'Tienda /\nCanal', 'Fecha\nCorte',
        'Stock\nUnidades', 'Stock\nTránsito Uds', 'Stock\nTotal Uds', 'Notas',
    ]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=2, column=c, value=h)
    style_header(ws4, 2, len(headers4))
    ws4.cell(row=2, column=7).fill = GREEN_HEADER

    ws4.merge_cells('A1:H1')

    print(f'  Escribiendo {len(df_stock):,} filas de stock...')
    for i, (_, r) in enumerate(df_stock.iterrows(), 3):
        ws4.cell(row=i, column=1, value=str(r['sku']))
        ws4.cell(row=i, column=2, value=r['nombre'])
        ws4.cell(row=i, column=3, value=r['tienda'])
        ws4.cell(row=i, column=4, value=r['fecha_corte'])
        ws4.cell(row=i, column=5, value=r['stock_uds'])
        ws4.cell(row=i, column=6, value=r['stock_transito'])
        ws4.cell(row=i, column=7, value=f'=E{i}+F{i}')
        ws4.cell(row=i, column=8, value='')
        for c in range(1, 9):
            cell = ws4.cell(row=i, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 7:
                cell.fill = CALC_FILL
                cell.font = CALC_FONT

    widths4 = [14, 45, 22, 12, 10, 12, 10, 20]
    for c, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    # ── GUARDAR ──
    wb.save(output_path)

    # Stats
    n_con_venta = len(df_ventas[df_ventas[['vta_sem1','vta_sem2','vta_sem3','vta_sem4']].sum(axis=1) > 0])
    n_sin_venta = len(df_ventas) - n_con_venta
    n_con_stock = len(df_stock[df_stock['stock_uds'] > 0])

    print(f'\n✅ Plantilla generada: {output_path}')
    print(f'   Tab 1 Maestro:  {len(df_maestro):,} SKUs')
    print(f'   Tab 2 LY:       VACÍA (llenar aparte)')
    print(f'   Tab 3 Ventas:   {len(df_ventas):,} combos ({n_con_venta:,} con venta, {n_sin_venta:,} sin venta)')
    print(f'   Tab 4 Stock:    {len(df_stock):,} combos ({n_con_stock:,} con stock)')

    return output_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Uso: python3 transformar_profundidad.py "base_profundidad.xlsx" [--output "salida.xlsx"] [--fecha "DD/MM/YYYY"]')
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = None
    fecha = '08/04/2026'
    if '--output' in sys.argv:
        output_file = sys.argv[sys.argv.index('--output') + 1]
    if '--fecha' in sys.argv:
        fecha = sys.argv[sys.argv.index('--fecha') + 1]

    transform(input_file, output_file, fecha)
