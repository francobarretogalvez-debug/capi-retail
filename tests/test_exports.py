"""Suite de regresión de exports (vistas_excel + reportes_marcas) contra la
data real del corte 04.08. Origen: auditoría de formato 2026-08-05 + Fase 0
2026-08-23. Requiere el archivo de referencia local; si no está, skip limpio."""
import io
import os
import re
import time
import zipfile

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from conftest import REPO

import reportes_marcas
import vistas_excel

BASE = "/Users/francobarreto/Claude_Context/Ripley.md/Reportes Capi"
REF = f"{BASE}/Capi_Analisis al 04.08.xlsx"


@pytest.fixture(scope="module")
def ref_04_08():
    if not os.path.exists(REF):
        pytest.skip(f"sin archivo de referencia {REF}")
    # _add_pricing_cols real extraído del app (sin importar streamlit)
    src = open(os.path.join(REPO, "app_streamlit.py")).read()
    m = re.search(r"(def _add_pricing_cols.*?\n    return df_out\n)", src, re.S)
    assert m, "no se encontró _add_pricing_cols"
    ns = {"np": np, "pd": pd, "get_column_letter": get_column_letter, "vistas_excel": vistas_excel}
    exec(m.group(1), ns)
    xl = pd.ExcelFile(REF)
    df = pd.read_excel(xl, "Cobertura").rename(columns={
        "vta_sem_ult": "prom_vta_uds", "Precio Vigente": "precio_vigente",
        "Precio Blanco": "precio_blanco", "Costo": "costo"})
    df = df.drop(columns=[c for c in ("Nuevo Precio", "Nuevo Margen") if c in df.columns])
    # el archivo de referencia (04.08) es anterior al renombre ALTO → PRE-SOBRESTOCK (2026-09-06)
    if "estado" in df.columns:
        df["estado"] = df["estado"].replace({"ALTO": "PRE-SOBRESTOCK", "MUERTO": "OBSOLETO", "LIQUIDAR": "PRE-OBSOLETO"})
    return dict(add_pricing=ns["_add_pricing_cols"], df=df,
                df_rep=pd.read_excel(xl, "Reposiciones Detalle"),
                df_trans=pd.read_excel(xl, "Transferencias"),
                df_prec=pd.read_excel(xl, "Acciones Precio"),
                df_alertas=pd.read_excel(xl, "Alertas IA"))


def test_vistas_excel_todos_estados(ref_04_08):
    df, _add_pricing_cols = ref_04_08["df"], ref_04_08["add_pricing"]
    _dl_cols = ["sku", "nombre", "marca", "temporada", "rango_antiguedad", "tienda",
                "stock_total", "prom_vta_uds", "cobertura_sem", "stock_valor_costo",
                "edad_semanas", "estado", "pct_descuento"]
    t0 = time.time()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        vistas_excel.hoja_resumen_ejecutivo(w, df)
        vistas_excel.hoja_liquidacion_sku(w, df, _add_pricing_cols)
        vistas_excel.hojas_terceras_por_marca(w, df)
        vistas_excel.hoja_cascada_dscto(w, df)
        vistas_excel.hoja_estados_temporada(w, df)
        _add_pricing_cols(df[_dl_cols].sort_values(["estado", "stock_valor_costo"],
                                                   ascending=[True, False]),
                          df, "Todos los estados", w)
    t_gen = time.time() - t0
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    assert wb.sheetnames[0] == "Resumen Ejecutivo" and wb.sheetnames[-1] == "Todos los estados"
    ws = wb["Resumen Ejecutivo"]
    tot = ws.cell(row=ws.max_row, column=ws.max_column).value
    assert abs(tot - 20268342) < 5, f"Resumen Ejecutivo total {tot:,.0f} ≠ 20,268,342"
    assert sum(1 for s in wb.sheetnames if s.startswith("T. ")) == 9
    hdrs = [c.value for c in wb["Liquidación x SKU"][1]]
    assert all(h in hdrs for h in ["Nuevo Precio", "Nuevo Margen", "Nuevo Dscto", "temporada"])
    assert t_gen <= 8, f"generación {t_gen:.1f}s > 8s"


def test_reportes_marcas_zip_9_marcas(ref_04_08):
    r = ref_04_08
    zbytes = reportes_marcas.generar_zip_reportes(r["df"], r["df_rep"], r["df_trans"],
                                                  r["df_prec"], r["df_alertas"], corte="04.08.2026")
    z = zipfile.ZipFile(io.BytesIO(zbytes))
    assert len(z.namelist()) == 9, z.namelist()
    n_sug = n_malos = n_sin_piso = 0
    for n in z.namelist():
        wbx = load_workbook(io.BytesIO(z.read(n)), read_only=True)
        assert "Resumen" in wbx.sheetnames and "Leyenda" in wbx.sheetnames, n
        for hoja in ("1. Liquidar", "2. Activar"):
            if hoja not in wbx.sheetnames:
                continue
            rows = wbx[hoja].iter_rows(min_row=2, values_only=True)
            hd = list(next(rows))
            if "P. Sugerido" not in hd:
                continue
            i_s, i_v = hd.index("P. Sugerido"), hd.index("P. Vigente")
            i_p = hd.index("P. Mínimo (piso)") if "P. Mínimo (piso)" in hd else None
            i_a = hd.index("Acción") if "Acción" in hd else None
            for row in rows:
                if (i_p is not None and row[i_p] is None
                        and not (i_a is not None and "Sin costo" in str(row[i_a] or ""))):
                    n_sin_piso += 1
                if row[i_s] is not None:
                    n_sug += 1
                    if row[i_v] is not None and row[i_s] > row[i_v] + 0.01:
                        n_malos += 1
                    if i_p is not None and row[i_p] is not None and row[i_s] < row[i_p] - 0.01:
                        n_malos += 1
    assert n_malos == 0, f"{n_sug} sugerencias, {n_malos} violan vigente/piso"
    assert n_sin_piso == 0, f"{n_sin_piso} filas sin P. Mínimo (fix B7)"
