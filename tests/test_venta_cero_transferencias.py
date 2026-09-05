"""S7/S8 (2026-09-05): venta cero con Pareto 80% reutilizable y consolidado de transferencias."""
import io

import pandas as pd
from openpyxl import load_workbook

import vistas_excel


def _cob():
    return pd.DataFrame([
        # tienda A: capital 80/15/5 → el primero ya cubre 80% pero SIEMPRE es TOP; el 2º no
        dict(tienda='A', marca='M', sku='1', nombre='p1', categoria='L', stock_total=8, stock_valor_costo=80, prom_vta_uds=0, precio_vigente=10, pct_descuento=0.2),
        dict(tienda='A', marca='M', sku='2', nombre='p2', categoria='L', stock_total=2, stock_valor_costo=15, prom_vta_uds=0, precio_vigente=10, pct_descuento=0),
        dict(tienda='A', marca='M', sku='3', nombre='p3', categoria='L', stock_total=1, stock_valor_costo=5, prom_vta_uds=0, precio_vigente=10, pct_descuento=0),
        # con venta → no es venta cero
        dict(tienda='A', marca='M', sku='4', nombre='p4', categoria='L', stock_total=5, stock_valor_costo=50, prom_vta_uds=2, precio_vigente=10, pct_descuento=0),
        # tienda B: 50/50 → ambos TOP (antes del 2º el acumulado era 50% < 80%)
        dict(tienda='B', marca='M', sku='5', nombre='p5', categoria='L', stock_total=1, stock_valor_costo=50, prom_vta_uds=0, precio_vigente=10, pct_descuento=0),
        dict(tienda='B', marca='M', sku='6', nombre='p6', categoria='L', stock_total=1, stock_valor_costo=50, prom_vta_uds=0, precio_vigente=10, pct_descuento=0),
    ])


def test_venta_cero_pareto_y_accion():
    vc = vistas_excel.venta_cero(_cob(), min_capital=0, tipo_evento_map={'1': 'MD1'})
    assert set(vc['sku']) == {'1', '2', '3', '5', '6'}
    top = dict(zip(vc['sku'], vc['top_80']))
    assert top['1'].startswith('⭐') and top['2'] == '' and top['3'] == ''
    assert top['5'].startswith('⭐') and top['6'].startswith('⭐')
    acc = dict(zip(vc['sku'], vc['accion']))
    assert acc['1'].startswith('🏷️') and acc['2'].startswith('👁️')
    assert vistas_excel.venta_cero(_cob(), min_capital=20)['sku'].tolist() == ['1', '5', '6']


def test_hoja_venta_cero_formato():
    vc = vistas_excel.venta_cero(_cob())
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        vistas_excel.hoja_venta_cero(w, vc, hoja='5. Venta Cero')
    ws = load_workbook(io.BytesIO(buf.getvalue()))['5. Venta Cero']
    assert [c.value for c in ws[2]][:3] == ['Tienda', 'Marca', 'SKU']
    assert ws.freeze_panes == 'A3' and ws.max_row == 2 + len(vc)


def test_consolidado_transferencias():
    tp = pd.DataFrame([
        dict(sku='1', _marca='M', tienda_origen='A', tienda_destino='B', uds_transferir=10, ganancia_esperada=100),
        dict(sku='2', _marca='N', tienda_origen='A', tienda_destino='B', uds_transferir=5, ganancia_esperada=50),
        dict(sku='1', _marca='M', tienda_origen='C', tienda_destino='D', uds_transferir=3, ganancia_esperada=-10),
    ])
    dest, orig = vistas_excel.consolidar_transferencias_por_tienda(tp)
    b = dest[dest['Tienda'] == 'B'].iloc[0]
    assert b['Movimientos'] == 2 and b['Uds a recibir'] == 15 and b['Ganancia esperada S/'] == 150 and 'M, N' == b['Marcas']
    assert dest['Ganancia esperada S/'].sum() == tp['ganancia_esperada'].sum()
    a = orig[orig['Tienda'] == 'A'].iloc[0]
    assert a['Uds a despachar'] == 15
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        vistas_excel.hoja_transferencias(w, tp.rename(columns={'ganancia_esperada': 'Ganancia S/'}), dest, orig, universo='Todas')
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    assert wb.sheetnames == ['Transferencias Todas', 'Por tienda destino', 'Por tienda origen']
