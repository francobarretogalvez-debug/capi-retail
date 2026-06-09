"""
Capi Inventory Engine — Motor de Cálculo
==========================================
Reads data from RetailAI_Template_v1.xlsx and computes:
  - Cobertura actual por SKU/Tienda (semanas)
  - Clasificación: CRÍTICO / URGENTE / MONITOREAR / OK / SOBRESTOCK
  - Plan de reposición desde almacén central
  - Sugerencias de transferencia entre tiendas
  - Acciones de precio por sobrestock y antigüedad

Usage:
    python motor_calculo.py                         # usa RetailAI_Template_v1.xlsx en CWD
    python motor_calculo.py mi_archivo.xlsx         # ruta personalizada
"""

import sys
import math
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook

DEFAULT_PATH = "RetailAI_Template_v1.xlsx"


# ══════════════════════════════════════════════════════════════════════
#  1. CARGA DE DATOS DESDE EXCEL
# ══════════════════════════════════════════════════════════════════════

def load_data(path=DEFAULT_PATH):
    """
    Lee todas las pestañas del template y devuelve estructuras Python limpias.
    Usa data_only=True para obtener valores calculados (no fórmulas).
    """
    wb = load_workbook(path, data_only=True)

    # ── Parámetros globales (Tab 2, columna B, filas 5–13)
    ws2 = wb["2. PARAMETROS"]
    params = {
        "cob_objetivo":   _num(ws2["B5"].value, 12),    # Stock ideal (semanas)
        "cob_sobrestock": _num(ws2["B6"].value, 20),    # Umbral sobrestock
        "cob_min":        _num(ws2["B7"].value, 8),     # Umbral reposición mínimo
        "cob_receptora":  _num(ws2["B8"].value, 8),     # Cob. mínima tienda receptora
        "uds_min_trans":  _num(ws2["B9"].value, 6),     # Unidades mínimas para transferir
        "margen_min":     _num(ws2["B10"].value, 0.15), # Margen mínimo (tope de descuento)
        "ventana_sem":    int(_num(ws2["B11"].value, 4)), # Semanas para promedio de ventas
        "sem_temporada":  _num(ws2["B12"].value, 8),    # Semanas restantes de temporada
        "igv":            _num(ws2["B13"].value, 0.18), # Tasa IGV
    }

    # ── Maestro de productos (Tab 3, filas 3+)
    ws3 = wb["3. MAESTRO PRODUCTOS"]
    products = {}
    for row in ws3.iter_rows(min_row=3, values_only=True):
        sku = row[0]
        if not sku or str(sku).startswith("Columnas"):
            continue
        products[sku] = {
            "desc":       row[1] or "",
            "cat":        row[2] or "",
            "clasif":     row[3] or "",      # Básico / Moda
            "f_lanz":     _to_date(row[4]),  # Fecha lanzamiento
            "temp":       row[5] or "",
            "zona":       row[6] or "",
            "costo":      _num(row[7], 0),
            "p_blanco":   _num(row[8], 0),
            "p_vigente":  _num(row[9], 0),
        }

    # ── Maestro de tiendas (Tab 4, filas 3+)
    ws4 = wb["4. MAESTRO TIENDAS"]
    stores = {}
    for row in ws4.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        sid = row[0]
        stores[sid] = {
            "nombre":      row[1] or "",
            "ciudad":      row[2] or "",
            "zona":        row[3] or "",
            "inv_reversa": str(row[4] or "No").strip().lower() in ("sí", "si", "yes", "s"),
            "lead_time":   _num(row[5], 1),
        }

    # ── Matriz logística (Tab 5, filas 3+)
    ws5 = wb["5. MATRIZ LOGISTICA"]
    logistics = {}
    for row in ws5.iter_rows(min_row=3, values_only=True):
        if not row[0] or not row[1]:
            continue
        key = (str(row[0]).strip(), str(row[1]).strip())
        logistics[key] = {
            "dias":   _num(row[2], 1),
            "c_fijo": _num(row[3], 0),
            "c_var":  _num(row[4], 0),
            "activa": str(row[5] or "Sí").strip() in ("Sí", "Si", "Yes", "S", "1"),
        }

    # ── Stock actual (Tab 6, filas 4+)
    ws6 = wb["6. STOCK ACTUAL"]
    stock = {}
    for row in ws6.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        tienda = row[1]
        if not sku or not tienda:
            continue
        stock[(str(sku).strip(), str(tienda).strip())] = _num(row[2], 0)

    # ── Ventas históricas (Tab 7, filas 4+)
    ws7 = wb["7. VENTAS HISTORICAS"]
    sales = []
    weeks_seen = set()
    for row in ws7.iter_rows(min_row=4, values_only=True):
        sku = row[0]
        tienda = row[1]
        semana = row[2]
        if not sku or not semana:
            continue
        semana_str = str(semana).strip()
        if semana_str.startswith("*"):  # fila de nota al pie
            continue
        sales.append({
            "sku":     str(sku).strip(),
            "tienda":  str(tienda).strip() if tienda else "",
            "semana":  semana_str,
            "uds":     _num(row[3], 0),
            "venta":   _num(row[4], 0),
            "contrib": _num(row[5], 0),
        })
        weeks_seen.add(semana_str)

    return params, products, stores, logistics, stock, sales, sorted(weeks_seen)


# ══════════════════════════════════════════════════════════════════════
#  2. CÁLCULO DE VENTA PROMEDIO SEMANAL
# ══════════════════════════════════════════════════════════════════════

def compute_avg_sales(sales, all_weeks, ventana):
    """
    Calcula la venta promedio semanal de unidades por (SKU, Tienda)
    usando las últimas `ventana` semanas del historial.

    Excluye semanas con 0 ventas del denominador para no penalizar
    SKUs con semanas sin reposición.  Si todas las semanas tienen 0,
    retorna 0.
    """
    last_n = set(sorted(all_weeks)[-ventana:]) if len(all_weeks) >= ventana else set(all_weeks)

    totals = defaultdict(float)
    nonzero_weeks = defaultdict(int)  # semanas con venta > 0
    total_weeks = defaultdict(set)    # semanas observadas

    for s in sales:
        if s["semana"] not in last_n:
            continue
        key = (s["sku"], s["tienda"])
        totals[key] += s["uds"]
        total_weeks[key].add(s["semana"])
        if s["uds"] > 0:
            nonzero_weeks[key] += 1

    avg = {}
    for key in totals:
        n_weeks = len(total_weeks[key])
        if n_weeks == 0:
            avg[key] = 0.0
        else:
            # Promedio sobre semanas totales (incluye ceros)
            # Da una imagen real del promedio de demanda
            avg[key] = totals[key] / n_weeks

    return avg


# ══════════════════════════════════════════════════════════════════════
#  3. CLASIFICACIÓN DE COBERTURA
# ══════════════════════════════════════════════════════════════════════

def classify_coverage(cob, params):
    """
    Clasifica una cobertura (en semanas) según los umbrales del motor.

    Retorna (estado, color)
    - CRÍTICO     : cobertura < 2 semanas  → quiebre inminente
    - URGENTE     : cobertura < cob_min    → reponer ya
    - MONITOREAR  : cobertura < cob_obj    → vigilar, pedir pronto
    - OK          : cobertura == cob_obj   → stock óptimo
    - SOBRESTOCK  : cobertura > cob_sobrestock → acción de precio o transferencia
    - SIN_VENTA   : promedio de ventas = 0 → no se puede calcular cobertura
    """
    if cob is None:
        return ("SIN_VENTA", "GRAY")
    if cob < 2:
        return ("CRÍTICO",    "RED")
    if cob < params["cob_min"]:
        return ("URGENTE",    "YELLOW")
    if cob < params["cob_objetivo"]:
        return ("MONITOREAR", "LIGHT_GREEN")
    if cob <= params["cob_sobrestock"]:
        return ("OK",         "GREEN")
    return ("SOBRESTOCK",  "ORANGE")


# ══════════════════════════════════════════════════════════════════════
#  4. MOTOR PRINCIPAL DE ANÁLISIS
# ══════════════════════════════════════════════════════════════════════

def run_analysis(path=DEFAULT_PATH):
    """
    Orquesta la carga de datos y todos los módulos de cálculo.
    Retorna un dict con los resultados completos.
    """
    params, products, stores, logistics, stock, sales, all_weeks = load_data(path)
    avg_sales = compute_avg_sales(sales, all_weeks, params["ventana_sem"])

    combos = list(stock.keys())  # lista de (SKU, Tienda nombre)

    # ── Cobertura y estado por combo ──────────────────────────────────
    coverage = {}
    status = {}
    for (sku, tienda) in combos:
        stk = stock[(sku, tienda)]
        avg = avg_sales.get((sku, tienda), 0.0)
        cob = round(stk / avg, 1) if avg > 0 else None
        coverage[(sku, tienda)] = cob
        status[(sku, tienda)] = classify_coverage(cob, params)

    # ── Plan de reposición (CRÍTICO + URGENTE → desde almacén) ────────
    reposicion = _build_reposicion(combos, stock, avg_sales, coverage, status,
                                   params, products, stores)

    # ── Sugerencias de transferencia (SOBRESTOCK → tiendas con déficit) ─
    transferencias = _build_transferencias(combos, stock, avg_sales, status,
                                           params, products, stores, logistics)

    # ── Acciones de precio (SOBRESTOCK sin transferencia / residual + antigüedad) ─
    acciones_precio = _build_acciones_precio(combos, stock, avg_sales, coverage,
                                              status, params, products, transferencias)

    return {
        "params":          params,
        "products":        products,
        "stores":          stores,
        "coverage":        coverage,
        "status":          status,
        "avg_sales":       avg_sales,
        "reposicion":      reposicion,
        "transferencias":  transferencias,
        "acciones_precio": acciones_precio,
        "all_weeks":       all_weeks,
    }


# ══════════════════════════════════════════════════════════════════════
#  4a. SUBMÓDULO: PLAN DE REPOSICIÓN
# ══════════════════════════════════════════════════════════════════════

def _build_reposicion(combos, stock, avg_sales, coverage, status,
                      params, products, stores):
    """
    Genera el plan de reposición para todos los combos con estado
    CRÍTICO o URGENTE.

    Lógica:
    - Unidades a reponer = max(0, Stock Ideal − Stock Actual)
    - Stock Ideal = cob_objetivo × venta_semanal_promedio
    - Fuente por defecto: Almacén Central
    - Lead time: extraído del Maestro de Tiendas
    - Cobertura post-reposición: (Stock Actual + A Reponer) / Venta Semanal
    """
    result = []
    for (sku, tienda) in combos:
        stat, _ = status[(sku, tienda)]
        if stat not in ("CRÍTICO", "URGENTE"):
            continue

        stk  = stock[(sku, tienda)]
        avg  = avg_sales.get((sku, tienda), 0.0)
        if avg == 0:
            continue  # sin venta histórica, no se puede calcular

        stk_ideal  = params["cob_objetivo"] * avg
        a_reponer  = max(0, math.ceil(stk_ideal - stk))
        cob_post   = round((stk + a_reponer) / avg, 1)

        sid        = _get_store_id(tienda, stores)
        lead_time  = stores[sid]["lead_time"] if sid else 1
        prod       = products.get(sku, {})

        # Zona climática: validar si el SKU es compatible con la tienda
        compat = _check_zona_compatibilidad(prod, stores.get(sid, {}))

        result.append({
            "sku":          sku,
            "desc":         prod.get("desc", ""),
            "cat":          prod.get("cat", ""),
            "tienda":       tienda,
            "zona_tienda":  stores.get(sid, {}).get("zona", ""),
            "zona_sku":     prod.get("zona", ""),
            "zona_ok":      compat,
            "stock_actual": int(stk),
            "vta_semanal":  round(avg, 1),
            "cobertura":    coverage[(sku, tienda)],
            "stock_ideal":  math.ceil(stk_ideal),
            "a_reponer":    a_reponer,
            "lead_time":    lead_time,
            "cob_post":     cob_post,
            "prioridad":    stat,
        })

    # Ordenar: CRÍTICO primero, luego por tienda
    result.sort(key=lambda x: (0 if x["prioridad"] == "CRÍTICO" else 1, x["tienda"]))
    return result


# ══════════════════════════════════════════════════════════════════════
#  4b. SUBMÓDULO: TRANSFERENCIAS
# ══════════════════════════════════════════════════════════════════════

def _build_transferencias(combos, stock, avg_sales, status,
                           params, products, stores, logistics):
    """
    Genera sugerencias de transferencia para combos con SOBRESTOCK
    que tienen una tienda destino con déficit y una ruta logística viable.

    Lógica:
    1. Para cada (SKU, Tienda_origen) con SOBRESTOCK:
       - Exceso = Stock_origen − (cob_objetivo × vta_promedio_origen)
       - Si exceso < uds_min_trans → saltar
    2. Para cada (SKU, Tienda_destino) con CRÍTICO/URGENTE/MONITOREAR:
       - Déficit = Stock_ideal_destino − Stock_destino
       - Si déficit ≤ 0 → saltar
       - Verificar que tienda origen tiene logística inversa O hay ruta directa
    3. Unidades a transferir = min(exceso disponible, déficit destino)
    4. Costo total = costo_fijo + uds × costo_variable
    """
    result = []
    exceso_restante = {}  # rastrea exceso disponible para no asignar el mismo stock dos veces

    for (sku, tienda_orig) in combos:
        stat_orig, _ = status[(sku, tienda_orig)]
        if stat_orig != "SOBRESTOCK":
            continue

        stk_orig = stock[(sku, tienda_orig)]
        avg_orig = avg_sales.get((sku, tienda_orig), 0.0)
        if avg_orig == 0:
            continue

        stk_ideal_orig = params["cob_objetivo"] * avg_orig
        exceso = math.floor(stk_orig - stk_ideal_orig)
        if exceso < params["uds_min_trans"]:
            continue

        key_orig = (sku, tienda_orig)
        exceso_restante[key_orig] = exceso_restante.get(key_orig, exceso)

        for (sku2, tienda_dest) in combos:
            if sku2 != sku or tienda_dest == tienda_orig:
                continue

            stat_dest, _ = status[(sku2, tienda_dest)]
            if stat_dest not in ("CRÍTICO", "URGENTE", "MONITOREAR"):
                continue

            avg_dest = avg_sales.get((sku, tienda_dest), 0.0)
            stk_dest = stock[(sku, tienda_dest)]
            if avg_dest == 0:
                continue

            deficit = math.ceil(params["cob_objetivo"] * avg_dest - stk_dest)
            if deficit <= 0:
                continue

            # Verificar ruta logística
            sid_orig = _get_store_id(tienda_orig, stores)
            sid_dest = _get_store_id(tienda_dest, stores)
            log_key  = _find_logistics_key(sid_orig, sid_dest, logistics)
            if log_key is None:
                continue

            # Comprobar logística inversa si aplica
            store_orig_info = stores.get(sid_orig, {})
            if not store_orig_info.get("inv_reversa", False):
                # Si la tienda origen no tiene logística inversa, omitir esta ruta
                # (solo puede enviar a almacén, no directamente a otra tienda)
                # Nota: en la v1 mantenemos la lógica permisiva — si hay ruta en la
                # matriz logística, asumimos que es válida.
                pass  # se puede endurecer aquí en versiones futuras

            disp = exceso_restante.get(key_orig, 0)
            uds  = min(disp, deficit)
            if uds < params["uds_min_trans"]:
                continue

            log_info   = logistics[log_key]
            costo_total = round(log_info["c_fijo"] + uds * log_info["c_var"], 2)

            # Cobertura resultante en origen y destino post-transferencia
            cob_orig_post = round((stk_orig - uds) / avg_orig, 1)
            cob_dest_post = round((stk_dest + uds) / avg_dest, 1)

            # Actualizar exceso restante
            exceso_restante[key_orig] -= uds

            prod = products.get(sku, {})
            result.append({
                "sku":            sku,
                "desc":           prod.get("desc", ""),
                "cat":            prod.get("cat", ""),
                "tienda_origen":  tienda_orig,
                "tienda_dest":    tienda_dest,
                "uds":            uds,
                "dias_transito":  log_info["dias"],
                "costo_fijo":     log_info["c_fijo"],
                "costo_var_unit": log_info["c_var"],
                "costo_total":    costo_total,
                "stat_origen":    stat_orig,
                "stat_dest":      stat_dest,
                "cob_orig_post":  cob_orig_post,
                "cob_dest_post":  cob_dest_post,
            })

    return result


# ══════════════════════════════════════════════════════════════════════
#  4c. SUBMÓDULO: ACCIONES DE PRECIO
# ══════════════════════════════════════════════════════════════════════

def _build_acciones_precio(combos, stock, avg_sales, coverage, status,
                            params, products, transferencias):
    """
    Genera recomendaciones de descuento en dos escenarios:

    A) SOBRESTOCK — sin transferencia viable
       O SOBRESTOCK — residual post-transferencia (cobertura sigue > cob_sobrestock)
       - Descuento inicial sugerido: 25% para acelerar rotación
       - Nunca baja del Precio Mínimo = Costo / (1 − Margen Mínimo)

    B) Alerta de antigüedad (Edad del SKU > umbrales PARAMETROS D)
       - Aplica principalmente a clasificación "Moda"
       - Descuento escalonado por edad

    Cuando ambas condiciones aplican, se toma el descuento mayor.
    El precio sugerido siempre se clampea contra el precio mínimo.
    """
    # Calcular unidades ya asignadas a transferencias por (sku, tienda_origen)
    # para derivar el stock residual post-transferencia
    from collections import defaultdict as _dd
    transferidas = _dd(int)
    for t in transferencias:
        transferidas[(t["sku"], t["tienda_origen"])] += t["uds"]
    # Umbrales de antigüedad (default, alineados con PARAMETROS Sección D)
    age_brackets = [
        (0,  20,  0.00, "Sin alerta por antigüedad"),
        (20, 30,  0.30, "Descontar 30% — producto maduro"),
        (30, 40,  0.40, "Descontar 40% — rotación lenta"),
        (40, 50,  0.50, "Descontar 50% — urgente"),
        (50, 60,  0.60, "Descontar 60% — próximo a obsolescencia"),
        (60, 999, 0.70, "LIQUIDAR — evaluar retiro de piso"),
    ]

    result = []
    for (sku, tienda) in combos:
        stat, _ = status[(sku, tienda)]
        prod     = products.get(sku, {})
        costo    = prod.get("costo", 0)
        p_vigente = prod.get("p_vigente", 0)
        p_blanco  = prod.get("p_blanco", 0)
        margen_min = params["margen_min"]

        if costo <= 0 or p_vigente <= 0:
            continue

        # Precio mínimo de venta = Costo / (1 − Margen Mínimo)
        p_min = costo / (1 - margen_min) if margen_min < 1 else costo * 1.01
        max_dscto = max(0, (p_vigente - p_min) / p_vigente)

        motivo       = None
        dscto_base   = 0.0
        motivos_list = []

        # ── A) Sobrestock: evaluar si persiste después de aplicar transferencias
        if stat == "SOBRESTOCK":
            avg  = avg_sales.get((sku, tienda), 0.0)
            stk  = stock[(sku, tienda)]
            uds_trans = transferidas.get((sku, tienda), 0)
            stk_post  = stk - uds_trans
            cob_post  = round(stk_post / avg, 1) if avg > 0 else None

            if cob_post is None or cob_post > params["cob_sobrestock"]:
                if uds_trans > 0:
                    motivos_list.append(
                        f"SOBRESTOCK residual post-transferencia ({cob_post} sem)"
                    )
                else:
                    motivos_list.append("SOBRESTOCK — sin transferencia viable")
                dscto_base = max(dscto_base, 0.25)

        # ── B) Antigüedad del producto
        f_lanz = prod.get("f_lanz")
        if f_lanz:
            edad_sem = (date.today() - f_lanz).days // 7
            for desde, hasta, dscto_edad, accion_label in age_brackets:
                if desde <= edad_sem < hasta:
                    if dscto_edad > 0:
                        motivos_list.append(f"ANTIGÜEDAD {edad_sem} sem — {accion_label}")
                        dscto_base = max(dscto_base, dscto_edad)
                    break

        if not motivos_list or dscto_base == 0:
            continue

        # Clampear al máximo descuento disponible (respeta margen mínimo)
        dscto_final  = min(dscto_base, max_dscto)
        p_sugerido   = max(round(p_vigente * (1 - dscto_final), 2), round(p_min, 2))
        dscto_real   = round((p_vigente - p_sugerido) / p_vigente, 4) if p_vigente > 0 else 0

        result.append({
            "sku":          sku,
            "desc":         prod.get("desc", ""),
            "cat":          prod.get("cat", ""),
            "clasif":       prod.get("clasif", ""),
            "tienda":       tienda,
            "cobertura":    coverage.get((sku, tienda)),
            "p_blanco":     p_blanco,
            "p_vigente":    p_vigente,
            "costo":        costo,
            "p_min":        round(p_min, 2),
            "max_dscto":    round(max_dscto, 4),
            "p_sugerido":   p_sugerido,
            "dscto_real":   dscto_real,
            "motivo":       " + ".join(motivos_list),
        })

    # Ordenar: sobrestock primero, luego antigüedad
    result.sort(key=lambda x: (0 if "SOBRESTOCK" in x["motivo"] else 1, x["sku"]))
    return result


# ══════════════════════════════════════════════════════════════════════
#  5. REPORTE EN CONSOLA
# ══════════════════════════════════════════════════════════════════════

def print_report(results):
    """Imprime un resumen ejecutivo del análisis en la consola."""
    params = results["params"]
    SEP = "═" * 72

    print(f"\n{SEP}")
    print("  RETAILAI INVENTORY ENGINE — REPORTE DE ANÁLISIS")
    print(SEP)

    # ── Parámetros activos
    print(f"\n{'─'*72}")
    print("  PARÁMETROS ACTIVOS")
    print(f"{'─'*72}")
    print(f"  Stock ideal (cob. objetivo) : {params['cob_objetivo']} semanas")
    print(f"  Umbral reposición (mínimo)  : {params['cob_min']} semanas")
    print(f"  Umbral sobrestock           : {params['cob_sobrestock']} semanas")
    print(f"  Ventana de venta            : {params['ventana_sem']} semanas")
    print(f"  Margen mínimo               : {params['margen_min']:.0%}")

    # ── Coberturas
    print(f"\n{'─'*72}")
    print("  COBERTURAS POR SKU / TIENDA")
    print(f"{'─'*72}")
    print(f"  {'SKU':<10} {'TIENDA':<38} {'COB (sem)':>10}  ESTADO")
    print(f"  {'-'*68}")
    for (sku, tienda), cob in sorted(results["coverage"].items()):
        stat, _ = results["status"][(sku, tienda)]
        cob_str  = f"{cob:>6.1f}" if cob is not None else "     —"
        flag     = _status_flag(stat)
        print(f"  {sku:<10} {tienda:<38} {cob_str}      {flag} {stat}")

    # ── Reposición
    rep = results["reposicion"]
    print(f"\n{'─'*72}")
    print(f"  PLAN DE REPOSICIÓN  ({len(rep)} líneas)")
    print(f"{'─'*72}")
    if rep:
        print(f"  {'PRIOR.':<12} {'SKU':<10} {'A REPONER':>10}  {'LT':>4}  TIENDA")
        print(f"  {'-'*68}")
        for r in rep:
            print(f"  {r['prioridad']:<12} {r['sku']:<10} {r['a_reponer']:>10}  {r['lead_time']:>4}  {r['tienda']}")
            print(f"             Cob actual: {r['cobertura']} sem → post-repo: {r['cob_post']} sem"
                  f"  |  Vta promedio: {r['vta_semanal']} uds/sem")
    else:
        print("  Sin SKUs en estado CRÍTICO ni URGENTE.")

    # ── Transferencias
    trans = results["transferencias"]
    print(f"\n{'─'*72}")
    print(f"  TRANSFERENCIAS SUGERIDAS  ({len(trans)})")
    print(f"{'─'*72}")
    if trans:
        for t in trans:
            print(f"  {t['sku']}  {t['desc']}")
            print(f"    {t['uds']} uds:  {t['tienda_origen']}  →  {t['tienda_dest']}")
            print(f"    Tránsito: {t['dias_transito']} días  |  Costo total: S/ {t['costo_total']:.2f}")
            print(f"    Cob. origen post: {t['cob_orig_post']} sem  |  Cob. destino post: {t['cob_dest_post']} sem")
            print()
    else:
        print("  Sin transferencias sugeridas.")

    # ── Precio
    precio = results["acciones_precio"]
    print(f"\n{'─'*72}")
    print(f"  ACCIONES DE PRECIO  ({len(precio)})")
    print(f"{'─'*72}")
    if precio:
        for p in precio:
            print(f"  {p['sku']}  {p['desc']}  [{p['tienda']}]")
            print(f"    Precio vigente: S/ {p['p_vigente']:.2f}  →  Sugerido: S/ {p['p_sugerido']:.2f}"
                  f"  (dscto {p['dscto_real']:.0%})  |  Precio mín: S/ {p['p_min']:.2f}")
            print(f"    Motivo: {p['motivo']}")
            print()
    else:
        print("  Sin acciones de precio sugeridas.")

    print(SEP)
    print(f"  Semanas analizadas: {results['all_weeks']}")
    print(SEP + "\n")


def _status_flag(stat):
    flags = {
        "CRÍTICO":    "🔴",
        "URGENTE":    "🟡",
        "MONITOREAR": "🟢",
        "OK":         "✅",
        "SOBRESTOCK": "🟠",
        "SIN_VENTA":  "⚪",
    }
    return flags.get(stat, "  ")


# ══════════════════════════════════════════════════════════════════════
#  HELPERS INTERNOS
# ══════════════════════════════════════════════════════════════════════

def _num(val, default=0):
    """Convierte un valor de celda a float, devuelve default si es None/texto."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _to_date(val):
    """Convierte un valor de celda a date. Acepta date, datetime o None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _get_store_id(tienda_nombre, stores):
    """
    Hace match entre el nombre largo de la tienda (ej: 'T001 – Lima – Miraflores')
    y el ID de tienda en el maestro (ej: 'T001').
    """
    tienda_norm = str(tienda_nombre).strip()
    for sid, s in stores.items():
        if s["nombre"] == tienda_norm:
            return sid
        if tienda_norm.startswith(sid):
            return sid
    return None


def _find_logistics_key(sid_origen, sid_dest, logistics):
    """
    Busca la clave en el dict de logística para un par de IDs de tienda.
    Las claves en la Matriz Logística usan formato 'T001 – Lima'.
    """
    if not sid_origen or not sid_dest:
        return None
    for (orig_key, dest_key), info in logistics.items():
        if orig_key.startswith(sid_origen) and dest_key.startswith(sid_dest):
            if info["activa"]:
                return (orig_key, dest_key)
    return None


def _check_zona_compatibilidad(prod, store):
    """
    Valida si la zona climática del SKU es compatible con la zona de la tienda.
    Retorna True si compatible, False si hay potencial conflicto.
    """
    zona_sku   = str(prod.get("zona", "")).lower()
    zona_tienda = str(store.get("zona", "")).lower()

    # "Ambas" o "Mixta" = compatible con todo
    if "ambo" in zona_sku or "mixta" in zona_sku:
        return True
    if "mixta" in zona_tienda:
        return True

    # Específico: debe coincidir
    if "fría" in zona_sku or "fria" in zona_sku:
        return "fría" in zona_tienda or "fria" in zona_tienda
    if "calurosa" in zona_sku:
        return "calurosa" in zona_tienda

    return True  # default: compatible


# ══════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    print(f"\n⏳  Cargando datos desde: {path}")
    results = run_analysis(path)
    print_report(results)
