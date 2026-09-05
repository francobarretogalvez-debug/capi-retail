"""
vistas_excel.py — Hojas agregadas de decisión para los exports de Capi.

Origen: Auditoría de formato de reportes 2026-08-05 (feedback Franco).
Todas las vistas colapsan el grano SKU×tienda a nivel SKU-cadena; la data
cruda se mantiene intacta en sus hojas originales.

Convención de cobertura (decisión 2026-08-05): cobertura cadena =
stock cadena ÷ promedio de venta de 4 semanas (vta_sem1..4_total, que en el
motor son totales cadena por SKU). Si no hay columnas semanales, cae a
prom_vta_uds sumado (última semana).
"""

from datetime import date

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import agente_terceras

# Orden de trabajo de Franco: primero lo muerto (liquidar), luego activar.
ORDEN_CRITICIDAD = [
    "MUERTO", "LIQUIDAR", "DORMIDO", "ESTANCADO", "SOBRESTOCK",
    "ALTO", "QUIEBRE", "PRE-QUIEBRE", "NUEVO SIN VENTA", "ÓPTIMO",
]

ESTADOS_ACTIVACION_TERCERAS = ["DORMIDO", "ESTANCADO", "SOBRESTOCK"]

# Marcas que técnicamente son licencias propias (MARCAS_PROPIAS_SET) pero que
# Franco también negocia como terceras → tienen su hoja de activación
# (pedido Franco 2026-08-05).
MARCAS_EXTRA_ACTIVACION = {"OSCAR DE LA RENTA", "US POLO"}

_SEM_COLS = ["vta_sem1_total", "vta_sem2_total", "vta_sem3_total", "vta_sem4_total"]

_FMT_MONEDA = "#,##0"
_FMT_UDS = "#,##0"
_FMT_COB = "0.0"
_FMT_PCT = "0.0%"

# ──────────────────────────────────────────────────────────────
#  Sistema de estilos (paleta del mockup, auditoría 2026-08-05)
# ──────────────────────────────────────────────────────────────

_INK = "1C2430"       # texto principal
_INK2 = "5B6472"      # texto secundario
_LINE = "DCDFE5"      # bordes
_ACCENT = "6D3B8E"    # morado Capi
_ACCENT_SOFT = "F1EAF6"
_HEAD_BG = "EEF0F3"   # gris de headers
_ZEBRA_BG = "F6F7F9"
_INPUT_BG = "FBF7EE"  # ámbar de columnas de input
_GOOD = "2E7D5B"

_FUENTE = "Avenir Next"  # en Windows Excel sustituye solo; en Mac es nativa

F_TITULO = Font(name=_FUENTE, bold=True, size=12, color=_ACCENT)
F_HEADER = Font(name=_FUENTE, bold=True, size=10, color=_INK2)
F_SUB = Font(name=_FUENTE, bold=True, size=10, color=_INK)
F_CHIP = Font(name=_FUENTE, bold=True, size=9, color="FFFFFF")
F_FORMULA = Font(name=_FUENTE, size=10, color=_GOOD, italic=True)
FILL_HEAD = PatternFill("solid", fgColor=_HEAD_BG)
FILL_SUB = PatternFill("solid", fgColor=_HEAD_BG)
FILL_ACCENT_SOFT = PatternFill("solid", fgColor=_ACCENT_SOFT)
FILL_INPUT = PatternFill("solid", fgColor=_INPUT_BG)
FILL_ZEBRA = PatternFill("solid", fgColor=_ZEBRA_BG)
BORDE_INF = Border(bottom=Side(style="thin", color=_LINE))
CENTRADO = Alignment(horizontal="center")

# Colores por estado (idénticos a los chips del mockup)
COLORES_ESTADO = {
    "MUERTO": "A03028", "LIQUIDAR": "C25B2A", "DORMIDO": "B98514",
    "ESTANCADO": "8F6B1E", "SOBRESTOCK": "6E7D27", "ALTO": "4A7D9E",
    "QUIEBRE": "5B4A9E", "PRE-QUIEBRE": "8577C2",
    "NUEVO SIN VENTA": "6B7280", "ÓPTIMO": "2E7D5B",
}
FILLS_ESTADO = {e: PatternFill("solid", fgColor=c) for e, c in COLORES_ESTADO.items()}
# Tintes morados para mapas de calor (3 escalones)
_FILLS_HEAT = [PatternFill("solid", fgColor=c) for c in ("F1EAF6", "DCC8E8", "BFA0D6")]


def _estilo_header(ws, fila=1):
    for cell in ws[fila]:
        if cell.value is not None:
            cell.font = F_HEADER
            cell.fill = FILL_HEAD
            cell.border = BORDE_INF


def _chip_estado(cell):
    fill = FILLS_ESTADO.get(str(cell.value).strip() if cell.value else "")
    if fill:
        cell.fill = fill
        cell.font = F_CHIP
        cell.alignment = CENTRADO


def _heat_fill(cell, valor, maximo):
    if not valor or not maximo or valor <= 0:
        return
    k = valor / maximo
    cell.fill = _FILLS_HEAT[2] if k > 0.45 else _FILLS_HEAT[1] if k > 0.18 else _FILLS_HEAT[0]


def estilizar_hoja_pricing(ws, fila_header=1):
    """Estilo estándar para hojas con bloque de repricing (la llama
    _add_pricing_cols del app): header gris, chips de estado, columna
    Nuevo Precio en ámbar, fórmulas en verde."""
    _estilo_header(ws, fila_header)
    headers = [c.value for c in ws[fila_header]]

    def _col(nombre):
        return headers.index(nombre) + 1 if nombre in headers else None

    c_est = _col("estado")
    c_np = _col("Nuevo Precio") or _col("nuevo_precio")
    cols_fx = [c for c in (_col("Nuevo Margen") or _col("nuevo_margen"),
                           _col("Nuevo Dscto") or _col("nuevo_dscto")) if c]
    for row in ws.iter_rows(min_row=fila_header + 1):
        if c_est:
            _chip_estado(row[c_est - 1])
        if c_np:
            row[c_np - 1].fill = FILL_INPUT
        for cf in cols_fx:
            row[cf - 1].font = F_FORMULA


# ──────────────────────────────────────────────────────────────
#  Helpers de agregación
# ──────────────────────────────────────────────────────────────

def _vta_4sem_por_sku(df: pd.DataFrame) -> pd.Series:
    """Venta semanal promedio (4 sem) a nivel cadena, indexada por sku.

    vta_semX_total ya es total cadena por SKU (se repite en cada fila
    SKU×tienda), así que se toma first por sku y se promedia.
    """
    if all(c in df.columns for c in _SEM_COLS):
        base = df.drop_duplicates("sku").set_index("sku")[_SEM_COLS]
        return base.mean(axis=1)
    # Fallback: suma de la última semana por tienda
    return df.groupby("sku")["prom_vta_uds"].sum()


def agregar_por_sku(df: pd.DataFrame) -> pd.DataFrame:
    """Colapsa SKU×tienda → estado×SKU cadena con capital, venta y cobertura."""
    agg = {
        "nombre": ("nombre", "first"),
        "marca": ("marca", "first"),
        "stock_cadena": ("stock_total", "sum"),
        "n_tiendas": ("tienda", "nunique"),
        "capital_costo": ("stock_valor_costo", "sum"),
        "edad_semanas": ("edad_semanas", "max"),
    }
    if "categoria" in df.columns:
        agg["categoria"] = ("categoria", "first")
    if "temporada" in df.columns:
        agg["temporada"] = ("temporada", "first")
    if "pct_descuento" in df.columns:
        agg["pct_descuento"] = ("pct_descuento", "max")
    for pc in ("precio_blanco", "precio_vigente", "costo"):
        if pc in df.columns:
            agg[pc] = (pc, "first")

    g = df.groupby(["estado", "sku"], as_index=False).agg(**agg)

    vta4 = _vta_4sem_por_sku(df)
    g["vta_sem_prom4"] = g["sku"].map(vta4).round(1)
    g["cobertura_cadena"] = np.where(
        g["vta_sem_prom4"] > 0,
        (g["stock_cadena"] / g["vta_sem_prom4"]).round(1),
        np.nan,
    )
    orden = {e: i for i, e in enumerate(ORDEN_CRITICIDAD)}
    g["_orden"] = g["estado"].map(orden).fillna(99)
    g = g.sort_values(["_orden", "capital_costo"], ascending=[True, False])
    return g.drop(columns="_orden").reset_index(drop=True)


def _autoajustar(ws, widths: dict = None, freeze: str = "A2", autofilter: bool = True):
    if autofilter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    if freeze:
        ws.freeze_panes = freeze
    for col_idx, width in (widths or {}).items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _formatear_columnas(ws, formatos: dict):
    """formatos: {nombre_header: number_format} aplicado a toda la columna."""
    headers = [c.value for c in ws[1]]
    for nombre, fmt in formatos.items():
        if nombre in headers:
            ltr = get_column_letter(headers.index(nombre) + 1)
            for row in range(2, ws.max_row + 1):
                ws[f"{ltr}{row}"].number_format = fmt


# ──────────────────────────────────────────────────────────────
#  V3 — Resumen Ejecutivo (estado × marca)
# ──────────────────────────────────────────────────────────────

def hoja_resumen_ejecutivo(writer, df: pd.DataFrame, nombre_hoja: str = "Resumen Ejecutivo"):
    """Primera hoja de cada export: capital a costo por marca × estado + fecha."""
    piv = df.pivot_table(index="marca", columns="estado",
                         values="stock_valor_costo", aggfunc="sum", fill_value=0)
    piv = piv[[e for e in ORDEN_CRITICIDAD if e in piv.columns]]
    piv["TOTAL"] = piv.sum(axis=1)
    piv = piv.sort_values("TOTAL", ascending=False)
    total = piv.sum(axis=0).to_frame().T
    total.index = ["TOTAL"]
    piv = pd.concat([piv, total])
    piv = piv.round(0).reset_index().rename(columns={"index": "marca"})

    # Fila 1 reservada para la fecha de corte; la tabla arranca en fila 2
    piv.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)
    ws = writer.sheets[nombre_hoja]
    ws["A1"] = f"Capital a costo (S/) por marca × estado — generado {date.today():%d.%m.%Y}"
    ws["A1"].font = F_TITULO

    # Header: cada estado con su color de chip; marca/TOTAL en gris
    for cell in ws[2]:
        if cell.value in FILLS_ESTADO:
            cell.fill = FILLS_ESTADO[cell.value]
            cell.font = F_CHIP
            cell.alignment = CENTRADO
        elif cell.value is not None:
            cell.font = F_HEADER
            cell.fill = FILL_HEAD

    # Heatmap morado sobre las celdas de capital (sin TOTALes)
    _max_val = float(piv.iloc[:-1, 1:-1].max().max()) if len(piv) > 1 else 0
    for row in range(3, ws.max_row + 1):
        es_total = ws.cell(row=row, column=1).value == "TOTAL"
        ws.cell(row=row, column=1).font = F_SUB if es_total else Font(name=_FUENTE, bold=True, size=10, color=_INK)
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = _FMT_MONEDA
            es_col_total = ws.cell(row=2, column=col).value == "TOTAL"
            if es_total or es_col_total:
                cell.font = F_SUB
                cell.fill = FILL_SUB
            else:
                _heat_fill(cell, cell.value, _max_val)
    _autoajustar(ws, widths={1: 20}, freeze="A3", autofilter=False)


# ──────────────────────────────────────────────────────────────
#  V1 — Plan de Liquidación x SKU (agregado cadena)
# ──────────────────────────────────────────────────────────────

def hoja_liquidacion_sku(writer, df: pd.DataFrame, add_pricing_cols,
                         nombre_hoja: str = "Liquidación x SKU"):
    """Hoja estado×SKU cadena, orden criticidad → capital. Reemplaza los
    pivots manuales. `add_pricing_cols` es el helper del app (agrega Nuevo
    Precio / Nuevo Margen / Nuevo Dscto con fórmulas vivas)."""
    g = agregar_por_sku(df)
    cols = [c for c in ["estado", "sku", "nombre", "marca", "categoria",
                        "temporada", "edad_semanas", "stock_cadena", "n_tiendas",
                        "capital_costo", "vta_sem_prom4", "cobertura_cadena",
                        "pct_descuento", "precio_blanco", "precio_vigente",
                        "costo"] if c in g.columns]
    out = add_pricing_cols(g[cols], df, nombre_hoja, writer)
    ws = writer.sheets[nombre_hoja]
    _formatear_columnas(ws, {
        "capital_costo": _FMT_MONEDA, "stock_cadena": _FMT_UDS,
        "cobertura_cadena": _FMT_COB, "vta_sem_prom4": _FMT_COB,
        "pct_descuento": _FMT_PCT,
    })
    # Header explícito: la cobertura de esta hoja usa promedio 4 semanas
    for cell in ws[1]:
        if cell.value == "vta_sem_prom4":
            cell.value = "vta sem (prom 4)"
        elif cell.value == "cobertura_cadena":
            cell.value = "cob cadena (prom 4 sem)"
    _autoajustar(ws, widths={3: 34})
    return out


# ──────────────────────────────────────────────────────────────
#  V6 — Terceras por modelo (una hoja por marca) ⭐ prioridad Franco
# ──────────────────────────────────────────────────────────────

def hojas_terceras_por_marca(writer, df: pd.DataFrame):
    """Para cada marca tercera, hoja lista para enviar al proveedor:
    SKUs DORMIDO/ESTANCADO/SOBRESTOCK a nivel modelo (sin tiendas) con
    capital, venta, stock y cobertura."""
    propias = {m.upper() for m in agente_terceras.MARCAS_PROPIAS_SET}
    propias -= MARCAS_EXTRA_ACTIVACION
    df_t = df[~df["marca"].str.upper().str.strip().isin(propias)]
    df_t = df_t[df_t["estado"].isin(ESTADOS_ACTIVACION_TERCERAS)]
    if df_t.empty:
        return

    g = agregar_por_sku(df_t)
    orden_estado = {e: i for i, e in enumerate(ESTADOS_ACTIVACION_TERCERAS)}

    for marca in sorted(g["marca"].unique()):
        gm = g[g["marca"] == marca].copy()
        gm["_o"] = gm["estado"].map(orden_estado)
        gm = gm.sort_values(["_o", "capital_costo"], ascending=[True, False]).drop(columns="_o")
        cols = [c for c in ["estado", "sku", "nombre", "categoria", "temporada",
                            "stock_cadena", "vta_sem_prom4", "cobertura_cadena",
                            "capital_costo"] if c in gm.columns]
        out = gm[cols].rename(columns={
            "estado": "Estado", "sku": "SKU", "nombre": "Modelo",
            "categoria": "Línea", "temporada": "Temporada", "stock_cadena": "Stock (uds)",
            "vta_sem_prom4": "Vta sem (prom 4)",
            "cobertura_cadena": "Cobertura (sem)",
            "capital_costo": "Capital S/ (costo)",
        })
        hoja = f"T. {marca}"[:31]
        # Fila 1: título con total, para reenviar la pestaña tal cual
        out.to_excel(writer, sheet_name=hoja, index=False, startrow=1)
        ws = writer.sheets[hoja]
        ws["A1"] = (f"{marca} — stock por activar (DORMIDO/ESTANCADO/SOBRESTOCK): "
                    f"{len(out)} modelos · S/ {gm['capital_costo'].sum():,.0f} a costo · "
                    f"corte {date.today():%d.%m.%Y}")
        ws["A1"].font = F_TITULO
        ws["A1"].fill = FILL_ACCENT_SOFT
        _estilo_header(ws, fila=2)
        for i, row in enumerate(ws.iter_rows(min_row=3)):
            _chip_estado(row[0])
            if i % 2 == 1:
                for cell in row[1:]:
                    cell.fill = FILL_ZEBRA
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.freeze_panes = "A3"
        headers = [c.value for c in ws[2]]
        fmts = {"Capital S/ (costo)": _FMT_MONEDA, "Stock (uds)": _FMT_UDS,
                "Vta sem (prom 4)": _FMT_COB, "Cobertura (sem)": _FMT_COB}
        for nombre, fmt in fmts.items():
            if nombre in headers:
                ltr = get_column_letter(headers.index(nombre) + 1)
                for row in range(3, ws.max_row + 1):
                    ws[f"{ltr}{row}"].number_format = fmt
        ws.column_dimensions["C"].width = 36


# ──────────────────────────────────────────────────────────────
#  V5a — Cascada de liquidación (estado crítico × banda de descuento)
# ──────────────────────────────────────────────────────────────

def hoja_cascada_dscto(writer, df: pd.DataFrame, nombre_hoja: str = "Cascada Dscto"):
    """Capital y # SKUs por estado crítico × banda de descuento actual."""
    if "pct_descuento" not in df.columns:
        return
    criticos = df[df["estado"].isin(["MUERTO", "LIQUIDAR", "DORMIDO",
                                     "ESTANCADO", "SOBRESTOCK"])].copy()
    if criticos.empty:
        return
    bandas = ["Sin dscto (0%)", "1-30%", "31-50%", ">50%"]
    criticos["banda"] = pd.cut(criticos["pct_descuento"].fillna(0),
                               bins=[-0.01, 0.0, 0.30, 0.50, 1.01], labels=bandas)
    cap = criticos.pivot_table(index="estado", columns="banda",
                               values="stock_valor_costo", aggfunc="sum",
                               fill_value=0, observed=False)
    sk = criticos.pivot_table(index="estado", columns="banda", values="sku",
                              aggfunc=pd.Series.nunique, fill_value=0, observed=False)
    orden = [e for e in ORDEN_CRITICIDAD if e in cap.index]
    cap, sk = cap.reindex(orden), sk.reindex(orden)
    cap["TOTAL"] = cap.sum(axis=1)

    cap.round(0).reset_index().to_excel(writer, sheet_name=nombre_hoja,
                                        index=False, startrow=1)
    fila_skus = len(cap) + 4
    sk.reset_index().rename(columns={"estado": "estado (# SKUs)"}).to_excel(
        writer, sheet_name=nombre_hoja, index=False, startrow=fila_skus)
    ws = writer.sheets[nombre_hoja]
    ws["A1"] = ("Capital a costo (S/) por estado × banda de descuento vigente — "
                "lo sin descuento es el primer candidato a activar precio")
    ws["A1"].font = F_TITULO
    _estilo_header(ws, fila=2)
    _estilo_header(ws, fila=fila_skus + 1)
    _max_cas = float(cap.iloc[:, :-1].max().max()) if not cap.empty else 0
    for row in range(3, len(cap) + 3):
        _chip_estado(ws.cell(row=row, column=1))
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = _FMT_MONEDA
            if ws.cell(row=2, column=col).value == "TOTAL":
                cell.font = F_SUB
                cell.fill = FILL_SUB
            else:
                _heat_fill(cell, cell.value, _max_cas)
    for row in range(fila_skus + 2, ws.max_row + 1):
        _chip_estado(ws.cell(row=row, column=1))
    _autoajustar(ws, widths={1: 18}, freeze=None, autofilter=False)


# ──────────────────────────────────────────────────────────────
#  V5d — Estados × temporada con filtro por marca (pivot simulada)
# ──────────────────────────────────────────────────────────────

def hoja_estados_temporada(writer, df: pd.DataFrame,
                           nombre_hoja: str = "Estados x Temporada"):
    """Matriz estado × temporada de capital, con dropdown de marca (B1) y
    SUMIFS sobre una hoja de datos oculta — se comporta como pivot con
    filtro sin necesitar 'Actualizar'."""
    if "temporada" not in df.columns:
        return
    base = (df.groupby(["marca", "estado", "temporada"])["stock_valor_costo"]
              .sum().round(0).reset_index())
    if base.empty:
        return
    hoja_datos = "_datos_estados_temp"
    base.to_excel(writer, sheet_name=hoja_datos, index=False)
    writer.sheets[hoja_datos].sheet_state = "hidden"

    ws = writer.book.create_sheet(nombre_hoja)

    marcas = sorted(base["marca"].unique().tolist())
    temporadas = [t for t in ["OI", "PV", "TT"] if t in set(base["temporada"])]
    temporadas += sorted(set(base["temporada"]) - set(temporadas))
    estados = [e for e in ORDEN_CRITICIDAD if e in set(base["estado"])]

    ws["A1"] = "Marca (elige en B1):"
    ws["A1"].font = F_SUB
    ws["B1"] = "TODAS"
    ws["B1"].fill = FILL_INPUT
    ws["B1"].font = F_SUB
    ws["B1"].border = Border(bottom=Side(style="thin", color=_INK2),
                             top=Side(style="thin", color=_INK2),
                             left=Side(style="thin", color=_INK2),
                             right=Side(style="thin", color=_INK2))
    _lista = ",".join(["TODAS"] + marcas)
    if len(_lista) <= 255:  # límite de Excel para listas inline
        dv = DataValidation(type="list", formula1=f'"{_lista}"',
                            allow_blank=False, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(ws["B1"])

    ws["A2"] = "estado"
    for j, t in enumerate(temporadas):
        ws.cell(row=2, column=2 + j, value=t)
    ws.cell(row=2, column=2 + len(temporadas), value="TOTAL")

    crit = 'IF($B$1="TODAS","*",$B$1)'
    for i, e in enumerate(estados):
        r = 3 + i
        ws.cell(row=r, column=1, value=e)
        for j, _t in enumerate(temporadas):
            c = 2 + j
            tl = get_column_letter(c)
            ws.cell(row=r, column=c).value = (
                f"=SUMIFS('{hoja_datos}'!$D:$D,'{hoja_datos}'!$A:$A,{crit},"
                f"'{hoja_datos}'!$B:$B,$A{r},'{hoja_datos}'!$C:$C,{tl}$2)")
            ws.cell(row=r, column=c).number_format = _FMT_MONEDA
        ctot = 2 + len(temporadas)
        ws.cell(row=r, column=ctot).value = (
            f"=SUM(B{r}:{get_column_letter(ctot - 1)}{r})")
        ws.cell(row=r, column=ctot).number_format = _FMT_MONEDA

    rt = 3 + len(estados)
    ws.cell(row=rt, column=1, value="TOTAL")
    for c in range(2, 3 + len(temporadas)):
        ltr = get_column_letter(c)
        ws.cell(row=rt, column=c).value = f"=SUM({ltr}3:{ltr}{rt - 1})"
        ws.cell(row=rt, column=c).number_format = _FMT_MONEDA

    # Estilo: header gris, chips de estado, fila TOTAL resaltada
    _estilo_header(ws, fila=2)
    for i in range(len(estados)):
        _chip_estado(ws.cell(row=3 + i, column=1))
    for c in range(1, 3 + len(temporadas)):
        ws.cell(row=rt, column=c).font = F_SUB
        ws.cell(row=rt, column=c).fill = FILL_SUB
    ws.column_dimensions["A"].width = 20


# ──────────────────────────────────────────────────────────────
#  V2 — Resumen Obsoletos (rango × marca)
# ──────────────────────────────────────────────────────────────

def hoja_resumen_obsoletos(writer, df_obs: pd.DataFrame,
                           nombre_hoja: str = "Resumen Obsoletos"):
    """Rango × marca con subtotales y cobertura cadena recalculada (no la
    suma de ratios que invalidaba el pivot manual)."""
    vta4 = _vta_4sem_por_sku(df_obs)
    tmp = df_obs.copy()
    tmp["_vta4_sku"] = tmp["sku"].map(vta4)

    def _grupo(df_g, keys):
        g = df_g.groupby(keys, as_index=False).agg(
            skus=("sku", "nunique"), uds=("stock_total", "sum"),
            capital=("stock_valor_costo", "sum"))
        vta = (df_g.drop_duplicates("sku").groupby(keys)["_vta4_sku"]
               .sum().reset_index(name="vta_sem_prom4"))
        g = g.merge(vta, on=keys, how="left")
        g["cobertura_cadena"] = np.where(
            g["vta_sem_prom4"] > 0, (g["uds"] / g["vta_sem_prom4"]).round(1), np.nan)
        g["vta_sem_prom4"] = g["vta_sem_prom4"].round(1)
        return g

    detalle = _grupo(tmp, ["rango_antiguedad", "marca"])
    orden_rango = {"RANGO 12_99": 0, "RANGO 9_12": 1, "RANGO 6_9": 2}
    detalle["_o"] = detalle["rango_antiguedad"].map(orden_rango).fillna(9)
    detalle = detalle.sort_values(["_o", "capital"], ascending=[True, False])

    filas = []
    for rango, grupo in detalle.groupby("_o", sort=True):
        nombre_rango = grupo["rango_antiguedad"].iloc[0]
        sub = _grupo(tmp[tmp["rango_antiguedad"] == nombre_rango], ["rango_antiguedad"])
        filas.append({
            "rango": nombre_rango, "marca": "— SUBTOTAL —",
            "skus": int(sub["skus"].iloc[0]), "uds": int(sub["uds"].iloc[0]),
            "capital": float(sub["capital"].iloc[0]),
            "vta_sem_prom4": float(sub["vta_sem_prom4"].iloc[0]),
            "cobertura_cadena": sub["cobertura_cadena"].iloc[0],
        })
        for _, r in grupo.iterrows():
            filas.append({
                "rango": r["rango_antiguedad"], "marca": r["marca"],
                "skus": int(r["skus"]), "uds": int(r["uds"]),
                "capital": float(r["capital"]),
                "vta_sem_prom4": r["vta_sem_prom4"],
                "cobertura_cadena": r["cobertura_cadena"],
            })
    tot = _grupo(tmp.assign(_all="TOTAL"), ["_all"])
    filas.append({
        "rango": "TOTAL GENERAL", "marca": "",
        "skus": int(tmp["sku"].nunique()), "uds": int(tot["uds"].iloc[0]),
        "capital": float(tot["capital"].iloc[0]),
        "vta_sem_prom4": float(tot["vta_sem_prom4"].iloc[0]),
        "cobertura_cadena": tot["cobertura_cadena"].iloc[0],
    })

    out = pd.DataFrame(filas)
    out.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=1)
    ws = writer.sheets[nombre_hoja]
    ws["A1"] = (f"Inventario >6 meses por rango × marca — cobertura = uds ÷ venta "
                f"prom 4 sem (cadena) — corte {date.today():%d.%m.%Y}")
    ws["A1"].font = F_TITULO
    _estilo_header(ws, fila=2)
    for row in ws.iter_rows(min_row=3):
        marca_val = str(row[1].value or "")
        rango_val = str(row[0].value or "")
        if marca_val.startswith("—") or rango_val == "TOTAL GENERAL":
            fill = FILL_ACCENT_SOFT if rango_val == "TOTAL GENERAL" else FILL_SUB
            for cell in row:
                cell.font = F_SUB
                cell.fill = fill
    headers = [c.value for c in ws[2]]
    for nombre, fmt in {"capital": _FMT_MONEDA, "uds": _FMT_UDS,
                        "vta_sem_prom4": _FMT_COB, "cobertura_cadena": _FMT_COB}.items():
        if nombre in headers:
            ltr = get_column_letter(headers.index(nombre) + 1)
            for row in range(3, ws.max_row + 1):
                ws[f"{ltr}{row}"].number_format = fmt
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20


# ══════════════════════════════════════════════════════════════
#  TRANSFERENCIAS — S8 (2026-09-05, pedido Franco)
#  Descargable con Marca/Departamento/Línea + consolidado por tienda destino
#  (qué recibe y cuánto gana) y por tienda origen (qué despacha).
# ══════════════════════════════════════════════════════════════

def consolidar_transferencias_por_tienda(df_tp: pd.DataFrame):
    """Devuelve (por_destino, por_origen) a partir del plan de transferencias.

    df_tp: filas SKU×origen×destino con tienda_origen, tienda_destino,
    uds_transferir, ganancia_esperada y (opcional) _marca.
    La ganancia se atribuye a la tienda DESTINO: es la que vende la unidad.
    """
    if df_tp is None or df_tp.empty:
        vacio = pd.DataFrame(columns=['Tienda', 'Movimientos', 'SKUs', 'Uds a recibir', 'Ganancia esperada S/', 'Marcas'])
        return vacio, vacio.rename(columns={'Uds a recibir': 'Uds a despachar'})
    d = df_tp.copy()
    d['ganancia_esperada'] = pd.to_numeric(d.get('ganancia_esperada'), errors='coerce').fillna(0)
    d['uds_transferir'] = pd.to_numeric(d.get('uds_transferir'), errors='coerce').fillna(0)
    marca_col = '_marca' if '_marca' in d.columns else None

    def _agg(col, uds_label):
        g = d.groupby(col).agg(
            Movimientos=('sku', 'size'), SKUs=('sku', 'nunique'),
            Uds=('uds_transferir', 'sum'), Ganancia=('ganancia_esperada', 'sum'),
            **({'Marcas': (marca_col, lambda x: ', '.join(sorted(set(str(v) for v in x if pd.notna(v)))[:6]))}
               if marca_col else {}))
        g = g.rename(columns={'Uds': uds_label, 'Ganancia': 'Ganancia esperada S/'})
        g = g.reset_index().rename(columns={col: 'Tienda'})
        return g.sort_values('Ganancia esperada S/', ascending=False).reset_index(drop=True)

    return _agg('tienda_destino', 'Uds a recibir'), _agg('tienda_origen', 'Uds a despachar')


def hoja_transferencias(writer, df_detalle: pd.DataFrame, por_destino: pd.DataFrame,
                        por_origen: pd.DataFrame, universo: str = "Todas"):
    """Excel de transferencias: 3 hojas (detalle, por destino, por origen) con formato."""
    hojas = [
        (f"Transferencias {universo}"[:31], df_detalle,
         f"PLAN DE TRANSFERENCIAS ({universo}) — cada fila es un movimiento SKU × origen → destino",
         {'Ganancia S/': '#,##0', 'Uds a Transferir': '#,##0', 'Cob Origen (pre)': '0.0',
          'Cob Destino (pre)': '0.0', 'Cob Origen (post)': '0.0', 'Cob Destino (post)': '0.0'}),
        ("Por tienda destino", por_destino,
         "QUÉ RECIBE CADA TIENDA — ganancia esperada = lo que la tienda vende de más al recibir la mercadería",
         {'Ganancia esperada S/': '#,##0', 'Uds a recibir': '#,##0', 'Movimientos': '#,##0', 'SKUs': '#,##0'}),
        ("Por tienda origen", por_origen,
         "QUÉ DESPACHA CADA TIENDA — para que el inventory manager arme los envíos",
         {'Ganancia esperada S/': '#,##0', 'Uds a despachar': '#,##0', 'Movimientos': '#,##0', 'SKUs': '#,##0'}),
    ]
    for nombre, df, titulo, fmts in hojas:
        _tabla_con_titulo(writer, nombre, titulo, df if df is not None else pd.DataFrame(), fmts)


def _tabla_con_titulo(writer, hoja: str, titulo: str, df: pd.DataFrame, formatos: dict,
                      anchos: dict = None):
    """Hoja estándar Capi: título en A1, header en fila 2, formatos por columna,
    autofiltro y panel congelado bajo el header. Devuelve el worksheet."""
    df.to_excel(writer, sheet_name=hoja, index=False, startrow=1)
    ws = writer.sheets[hoja]
    ws["A1"] = titulo
    ws["A1"].font = F_TITULO
    ws["A1"].fill = FILL_ACCENT_SOFT
    _estilo_header(ws, fila=2)
    headers = [c.value for c in ws[2]]
    for nombre, fmt in (formatos or {}).items():
        if nombre in headers:
            ltr = get_column_letter(headers.index(nombre) + 1)
            for r in range(3, ws.max_row + 1):
                ws[f"{ltr}{r}"].number_format = fmt
    if ws.max_row > 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A3"
    for nombre, w in (anchos or {}).items():
        if nombre in headers:
            ws.column_dimensions[get_column_letter(headers.index(nombre) + 1)].width = w
    return ws


# ══════════════════════════════════════════════════════════════
#  VENTA CERO — S7 (2026-09-05). Antes vivía inline en la vista
#  📲 Productos Venta Cero; ahora es reutilizable (vista + reporte por marca).
# ══════════════════════════════════════════════════════════════

COLS_VENTA_CERO = ["tienda", "marca", "sku", "nombre", "categoria", "stock_total", "stock_valor_costo",
                   "pct_acum_tienda", "top_80", "precio_vigente", "pct_descuento", "tipo_evento",
                   "edad_semanas", "accion"]
REN_VENTA_CERO = {
    "tienda": "Tienda", "marca": "Marca", "sku": "SKU", "nombre": "Producto", "categoria": "Línea",
    "stock_total": "Stock (uds)", "stock_valor_costo": "Capital S/", "pct_acum_tienda": "% acum. en tienda",
    "top_80": "Prioridad", "precio_vigente": "Precio", "pct_descuento": "Dscto", "tipo_evento": "Tipo evento",
    "edad_semanas": "Edad (sem)", "accion": "Acción"}
FMT_VENTA_CERO = {"Capital S/": "#,##0", "Stock (uds)": "#,##0", "% acum. en tienda": "0%",
                  "Dscto": "0%", "Precio": "#,##0.00", "Edad (sem)": "0"}


def _accion_venta_cero(r) -> str:
    _tp = str(r.get("tipo_evento", "") or "")
    _dsc = float(r.get("pct_descuento", 0) or 0)
    if _dsc > 0 and _tp == "MD1":
        return "🏷️ Etiquetar mercadería (precio ya impreso)"
    if _dsc > 0 and _tp == "PTR":
        return "📋 Colocar cartel de precio"
    return "👁️ Revisar exhibición"


def venta_cero(df_cob: pd.DataFrame, min_capital: float = 0, tipo_evento_map: dict = None) -> pd.DataFrame:
    """SKU×tienda con stock y sin venta la última semana, con acción de piso y Pareto 80%
    del capital por tienda. Regla del TOP 80% (auditoría C1 2026-08-26): el acumulado
    ANTES del SKU no llegaba a 80%, así el primer SKU de cada tienda siempre es TOP."""
    if df_cob is None or df_cob.empty or "prom_vta_uds" not in df_cob.columns:
        return pd.DataFrame(columns=COLS_VENTA_CERO)
    d = df_cob[(df_cob["prom_vta_uds"].fillna(0) == 0) & (df_cob["stock_total"].fillna(0) > 0)].copy()
    if "stock_valor_costo" not in d.columns:
        d["stock_valor_costo"] = 0.0
    d = d[d["stock_valor_costo"].fillna(0) >= float(min_capital or 0)]
    if d.empty:
        return pd.DataFrame(columns=COLS_VENTA_CERO)
    d["tipo_evento"] = d["sku"].map(tipo_evento_map or {}).fillna("") if "sku" in d.columns else ""
    d["accion"] = d.apply(_accion_venta_cero, axis=1)
    d = d.sort_values(["tienda", "stock_valor_costo"], ascending=[True, False], kind="mergesort")  # estable: empates conservan orden
    tot = d.groupby("tienda")["stock_valor_costo"].transform("sum").replace(0, np.nan)
    share = d["stock_valor_costo"] / tot
    acum = d.groupby("tienda")["stock_valor_costo"].cumsum() / tot
    d["pct_acum_tienda"] = acum.fillna(0)
    # round(6): sin esto 0.95−0.15 da 0.7999… y un SKU exactamente en el borde entra al TOP
    d["top_80"] = np.where((acum - share).fillna(0).round(6) < 0.80, "⭐ TOP 80%", "")
    for c in COLS_VENTA_CERO:
        if c not in d.columns:
            d[c] = np.nan
    return d[COLS_VENTA_CERO].reset_index(drop=True)


def hoja_venta_cero(writer, df_vc: pd.DataFrame, hoja: str = "Venta Cero x Tienda",
                    titulo: str = None):
    """Hoja para tiendas: cada jefe filtra SU tienda y ataca los ⭐ TOP 80%."""
    titulo = titulo or ("REVISIÓN DE PISO — filtra TU tienda y ataca primero los ⭐ TOP 80% "
                        "(concentran el 80% del capital sin venta de tu tienda)")
    out = df_vc[[c for c in COLS_VENTA_CERO if c in df_vc.columns]].rename(columns=REN_VENTA_CERO)
    return _tabla_con_titulo(writer, hoja, titulo, out, FMT_VENTA_CERO, anchos={"Producto": 34})
