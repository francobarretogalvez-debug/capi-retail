"""
reportes_marcas.py — Reporte por Marca Tercera (enviable al proveedor).

Origen: decisión 2026-08-05 "Reporte Semanal por Marca Tercera" + calibraciones
de Franco: leyenda de estados incluida, costos/márgenes visibles, capital a
costo, transferencias agregadas por modelo con umbral ≥12 uds y ≥S/1.000,
descuento sugerido por pirámide oficial + precio y margen resultantes.

Un Excel por marca con pestañas: Resumen · 1. Liquidar · 2. Activar ·
3. Reponer (guía OTB) · 4. Transferir · 5. Venta Cero · Tendencias · Leyenda.
Todo a nivel modelo (sin tiendas ni rutas — eso es ejecución interna Ripley).
"""

import io
import zipfile
from datetime import date

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

import agente_terceras
import vistas_excel
from vistas_excel import (
    CENTRADO, F_CHIP, F_HEADER, F_SUB, F_TITULO, FILL_ACCENT_SOFT, FILL_HEAD,
    FILL_SUB, FILL_ZEBRA, FILLS_ESTADO, _estilo_header, _chip_estado,
)

# Umbral transferencias (calibrado con data 04.08: captura 98% del valor)
TRANSF_MIN_UDS = 12
TRANSF_MIN_VALOR = 1000.0

_RANGO_LABEL = {
    "RANGO 0": "0-3 meses", "RANGO 0_3": "0-3 meses", "RANGO 3_6": "3-6 meses",
    "RANGO 6_9": "6-9 meses", "RANGO 9_12": "9-12 meses", "RANGO 12_99": "+12 meses",
}

_FMT_S = "#,##0"
_FMT_P = "#,##0.00"
_FMT_PCT = "0.0%"
_FMT_C = "0.0"

_SUPUESTOS = ("Supuestos: descuentos compartidos 50/50 según acuerdo vigente · "
              "la reposición es guía para priorizar OTB, NO compromiso de compra · "
              "capital valorizado a costo.")

_LEYENDA_ESTADOS = [
    ("QUIEBRE", "Cobertura menor a 4 semanas — riesgo de perder venta por falta de stock."),
    ("PRE-QUIEBRE", "Cobertura entre 4 y 8 semanas — próxima a quebrar, planificar reposición."),
    ("ÓPTIMO", "Cobertura entre 8 y 16 semanas — nivel saludable (target ~12)."),
    ("PRE-SOBRESTOCK", "Cobertura entre 16 y 26 semanas — antesala del sobrestock (antes ALTO)."),
    ("SOBRESTOCK", "Cobertura entre 26 y 52 semanas — exceso claro de stock vs. su venta."),
    ("ESTANCADO", "Más de 52 semanas de cobertura con producto aún joven (≤26 sem) — rota muy lento."),
    ("PRE-OBSOLETO", "6–9 meses en tienda sin rotación (o >52 sem de cobertura) — descuento fuerte antes de los 9 meses."),
    ("NUEVO SIN VENTA", "Sin ventas y menos de 8 semanas en tienda — aún en ventana de lanzamiento."),
    ("DORMIDO", "Sin ventas entre 8 y 26 semanas — necesita acción de precio o exhibición."),
    ("OBSOLETO", "9 meses a más sin rotación — liquidación directa o negociación con la marca."),
]

_PIRAMIDE = [
    ("8-11 sem", "20%", "Eventual"), ("12-18 sem", "30%", "Eventual"),
    ("19-29 sem", "30%", "Fijo"), ("30-34 sem", "40%", "Fijo"),
    ("35-38 sem", "50%", "Fijo"), ("39-43 sem", "60%", "Fijo"),
    ("44-47 sem", "70%", "Fijo"), ("48+ sem", "80%", "Fijo"),
]


def marcas_reporte(df_cob: pd.DataFrame) -> list:
    """Marcas que reciben reporte: terceras + licencias que se negocian como
    terceras (ODLR, US Polo). Derivado siempre del set canónico en runtime."""
    propias = {m.upper() for m in agente_terceras.MARCAS_PROPIAS_SET}
    propias -= vistas_excel.MARCAS_EXTRA_ACTIVACION
    todas = {str(m).upper().strip() for m in df_cob["marca"].dropna().unique()}
    return sorted(todas - propias)


def _con_sugerencias(g: pd.DataFrame, precio_min_map: dict) -> pd.DataFrame:
    """Agrega dscto sugerido (pirámide), precio sugerido, margen resultante y
    precio mínimo a un agregado por SKU."""
    g = g.copy()
    sug = g["edad_semanas"].apply(lambda e: agente_terceras.descuento_sugerido(e))
    g["dscto_sugerido"] = sug.map(lambda t: t[0])
    g["tipo_dscto"] = sug.map(lambda t: t[1])
    # Piso de margen UNIVERSAL (fix B7 auditoría 2026-08-23): df_prec solo trae
    # precio_minimo para estados críticos; para el resto se calcula igual que
    # el motor: costo/(1-margen_min) llevado a base con IGV.
    g["precio_minimo"] = g["sku"].map(precio_min_map) if precio_min_map else np.nan
    if "costo" in g.columns:
        try:
            from motor_v2 import DEFAULT_PARAMS as _DP
            _margen_min = float(_DP.get("margen_min", 0.15))
        except Exception:
            _margen_min = 0.15
        _piso_calc = (g["costo"] / (1 - _margen_min) * 1.18).round(2)
        g["precio_minimo"] = g["precio_minimo"].fillna(_piso_calc)
    # Guards: una base sin columnas de precio no debe tumbar el reporte
    # (auditoría 2026-08-16) — las columnas ausentes simplemente no salen.
    #
    # Regla de negocio (fix 2026-08-17): NUNCA sugerir un precio mayor al
    # vigente. El P. Sugerido solo aparece cuando implica BAJAR; si el dscto
    # actual ya supera la pirámide, o el piso de margen no deja bajar más,
    # la acción es Mantener.
    if "precio_blanco" in g.columns and "precio_vigente" in g.columns:
        p_obj = (g["precio_blanco"] * (1 - g["dscto_sugerido"])).round(2)
        limitado_piso = pd.Series(False, index=g.index)
        if "precio_minimo" in g.columns:
            piso = g["precio_minimo"]
            limitado_piso = piso.notna() & (p_obj < piso)
            p_obj = np.where(limitado_piso, piso, p_obj)
        # Sin piso conocido (costo faltante en la base) no se sugiere precio
        # al proveedor — se pide revisar el costo primero.
        _sin_piso = g["precio_minimo"].isna() if "precio_minimo" in g.columns else pd.Series(False, index=g.index)
        bajar = (p_obj < (g["precio_vigente"] - 0.01)) & ~_sin_piso
        g["precio_sugerido"] = np.where(bajar, p_obj, np.nan)
        g["accion"] = np.where(
            _sin_piso,
            "⚠️ Sin costo en base — revisar antes de sugerir",
            np.where(
                bajar,
                np.where(limitado_piso, "⬇️ Bajar (limitado por piso)", "⬇️ Bajar al sugerido"),
                np.where(limitado_piso | (g["precio_vigente"] <= g.get("precio_minimo", pd.Series(np.nan, index=g.index)).fillna(-1)),
                         "✋ Mantener — ya en/bajo piso de margen",
                         "✓ Mantener — dscto actual ≥ sugerido"),
            ),
        )
        if "costo" in g.columns:
            _base = g["precio_sugerido"] / 1.18
            g["margen_resultante"] = np.where(_base > 0, (_base - g["costo"]) / _base, np.nan)
    return g


_COLS_PRECIO = [
    ("estado", "Estado"), ("sku", "SKU"), ("nombre", "Modelo"),
    ("categoria", "Línea"), ("temporada", "Temporada"), ("edad_semanas", "Edad (sem)"),
    ("rango_antiguedad", "Antigüedad"), ("stock_cadena", "Stock (uds)"),
    ("vta_sem_prom4", "Vta sem (prom 4)"), ("cobertura_cadena", "Cobertura (sem)"),
    ("capital_costo", "Capital S/ (costo)"), ("costo", "Costo unit."),
    ("pct_descuento", "Dscto actual"), ("dscto_sugerido", "Dscto sugerido"),
    ("tipo_dscto", "Tipo"), ("accion", "Acción"), ("precio_blanco", "P. Blanco"),
    ("precio_vigente", "P. Vigente"), ("precio_sugerido", "P. Sugerido"),
    ("margen_resultante", "Margen result."), ("precio_minimo", "P. Mínimo (piso)"),
]

_FMTS_PRECIO = {
    "Edad (sem)": "0", "Stock (uds)": _FMT_S, "Vta sem (prom 4)": _FMT_C,
    "Cobertura (sem)": _FMT_C, "Capital S/ (costo)": _FMT_S, "Costo unit.": _FMT_P,
    "Dscto actual": _FMT_PCT, "Dscto sugerido": _FMT_PCT, "P. Blanco": _FMT_P,
    "P. Vigente": _FMT_P, "P. Sugerido": _FMT_P, "Margen result.": _FMT_PCT,
    "P. Mínimo (piso)": _FMT_P,
}


def _escribir_tabla(writer, hoja, titulo, df, formatos, zebra=True, chips_col=None):
    df.to_excel(writer, sheet_name=hoja, index=False, startrow=1)
    ws = writer.sheets[hoja]
    ws["A1"] = titulo
    ws["A1"].font = F_TITULO
    ws["A1"].fill = FILL_ACCENT_SOFT
    _estilo_header(ws, fila=2)
    headers = [c.value for c in ws[2]]
    for i, row in enumerate(ws.iter_rows(min_row=3)):
        if chips_col and chips_col in headers:
            _chip_estado(row[headers.index(chips_col)])
        if zebra and i % 2 == 1:
            for cell in row:
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = FILL_ZEBRA
    for nombre, fmt in formatos.items():
        if nombre in headers:
            ltr = get_column_letter(headers.index(nombre) + 1)
            for r in range(3, ws.max_row + 1):
                ws[f"{ltr}{r}"].number_format = fmt
    if ws.max_row > 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A3"
    if "Modelo" in headers:
        ws.column_dimensions[get_column_letter(headers.index("Modelo") + 1)].width = 34
    return ws


def _hoja_precio(writer, hoja, titulo, g):
    if g.empty:
        return
    cols = [c for c, _ in _COLS_PRECIO if c in g.columns]
    out = g[cols].rename(columns=dict(_COLS_PRECIO))
    if "Antigüedad" in out.columns:
        out["Antigüedad"] = out["Antigüedad"].map(_RANGO_LABEL).fillna(out["Antigüedad"])
    _escribir_tabla(writer, hoja, titulo, out, _FMTS_PRECIO, chips_col="Estado")


def generar_reporte_marca(marca, df_cob, df_rep=None, df_trans=None,
                          df_prec=None, df_alertas=None, corte: str = None) -> bytes:
    """Genera el Excel de una marca. df_cob a nivel SKU×tienda con columnas
    del motor (prom_vta_uds, precio_blanco, etc.). `corte` es la fecha de la
    BASE (no la de generación) — pasar siempre que se conozca."""
    m_up = str(marca).upper().strip()
    dfm = df_cob[df_cob["marca"].str.upper().str.strip() == m_up].copy()
    corte = corte or f"{date.today():%d.%m.%Y}"

    precio_min_map = {}
    if df_prec is not None and not df_prec.empty and "precio_minimo" in df_prec.columns:
        _pm = df_prec[df_prec["sku"].isin(dfm["sku"])]
        precio_min_map = _pm.drop_duplicates("sku").set_index("sku")["precio_minimo"].to_dict()

    g_all = vistas_excel.agregar_por_sku(dfm)
    # rango por sku (para columna Antigüedad)
    _rango = dfm.drop_duplicates("sku").set_index("sku").get("rango_antiguedad")
    if _rango is not None:
        g_all["rango_antiguedad"] = g_all["sku"].map(_rango)
    g_all = _con_sugerencias(g_all, precio_min_map)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        # ── Resumen ──
        est = g_all.groupby("estado").agg(
            modelos=("sku", "count"), uds=("stock_cadena", "sum"),
            capital=("capital_costo", "sum")).reindex(
            [e for e in vistas_excel.ORDEN_CRITICIDAD
             if e in set(g_all["estado"])]).reset_index()
        est.columns = ["Estado", "Modelos", "Stock (uds)", "Capital S/ (costo)"]
        ws = _escribir_tabla(
            w, "Resumen",
            f"{marca} — Resumen de mercadería en Ripley · corte {corte}",
            est, {"Stock (uds)": _FMT_S, "Capital S/ (costo)": _FMT_S},
            zebra=False, chips_col="Estado")
        fila = ws.max_row + 2
        ws.cell(row=fila, column=1, value="TOTAL").font = F_SUB
        ws.cell(row=fila, column=2, value=int(est["Modelos"].sum())).font = F_SUB
        ws.cell(row=fila, column=3, value=int(est["Stock (uds)"].sum())).font = F_SUB
        c4 = ws.cell(row=fila, column=4, value=float(est["Capital S/ (costo)"].sum()))
        c4.font = F_SUB
        c4.number_format = _FMT_S
        for c in range(1, 5):
            ws.cell(row=fila, column=c).fill = FILL_SUB
        # Obsolescencia por rango
        if "rango_antiguedad" in dfm.columns:
            obs = dfm[dfm["rango_antiguedad"].isin({"RANGO 6_9", "RANGO 9_12", "RANGO 12_99"})]
            if not obs.empty:
                fila += 2
                ws.cell(row=fila, column=1,
                        value="Mercadería obsoleta (más de 6 meses en tienda):").font = F_TITULO
                og = obs.groupby("rango_antiguedad").agg(
                    modelos=("sku", "nunique"), uds=("stock_total", "sum"),
                    capital=("stock_valor_costo", "sum"))
                for rango in ("RANGO 6_9", "RANGO 9_12", "RANGO 12_99"):
                    if rango in og.index:
                        fila += 1
                        r = og.loc[rango]
                        ws.cell(row=fila, column=1, value=_RANGO_LABEL[rango]).font = F_HEADER
                        ws.cell(row=fila, column=2, value=int(r["modelos"]))
                        ws.cell(row=fila, column=3, value=int(r["uds"])).number_format = _FMT_S
                        cc = ws.cell(row=fila, column=4, value=float(r["capital"]))
                        cc.number_format = _FMT_S
        fila += 2
        ws.cell(row=fila, column=1, value=_SUPUESTOS).font = F_HEADER
        ws.column_dimensions["A"].width = 24
        ws.auto_filter.ref = None

        # ── 1. Liquidar / 2. Activar ──
        _hoja_precio(w, "1. Liquidar",
                     f"{marca} — LIQUIDAR: pre-obsoleto y obsoleto (más de 6 meses sin rotación) · corte {corte}",
                     g_all[g_all["estado"].isin(["OBSOLETO", "PRE-OBSOLETO"])])
        _hoja_precio(w, "2. Activar",
                     f"{marca} — ACTIVAR precio/exhibición (dormido, estancado, sobrestock) · corte {corte}",
                     g_all[g_all["estado"].isin(["DORMIDO", "ESTANCADO", "SOBRESTOCK"])])

        # ── 3. Reponer (guía OTB) ──
        if df_rep is not None and not df_rep.empty and "marca" in df_rep.columns:
            rep_m = df_rep[df_rep["marca"].str.upper().str.strip() == m_up]
            if not rep_m.empty:
                _agg_rep = dict(
                    urgencia=("urgencia", "first"), vta_sem=("prom_vta_sem", "sum"),
                    stock_cadena=("stock_actual", "sum"), stock_cd=("stock_cd", "first"),
                    a_reponer=("a_reponer", "sum"))
                _cols_rep = ["SKU", "Modelo", "Línea", "Urgencia", "Vta sem (uds)",
                             "Stock cadena", "Stock CD", "Uds sugeridas"]
                if "temporada" in rep_m.columns:
                    _agg_rep = {"temporada": ("temporada", "first"), **_agg_rep}
                    _cols_rep.insert(3, "Temporada")
                rg = rep_m.groupby(["sku", "nombre", "categoria"], as_index=False).agg(**_agg_rep)
                rg = rg.sort_values("a_reponer", ascending=False)
                rg.columns = _cols_rep
                _escribir_tabla(
                    w, "3. Reponer",
                    f"{marca} — Guía de reposición para priorizar OTB (no es orden de compra) · corte {corte}",
                    rg, {"Vta sem (uds)": _FMT_C, "Stock cadena": _FMT_S,
                         "Stock CD": _FMT_S, "Uds sugeridas": _FMT_S})

        # ── 4. Transferir (agregado por modelo, umbral) ──
        if df_trans is not None and not df_trans.empty:
            tr_m = df_trans[df_trans["sku"].isin(dfm["sku"])].copy()
            if not tr_m.empty:
                tr_m["valor"] = tr_m["uds_transferir"] * tr_m["precio_vigente"]
                _tiene_ganancia = ("ganancia_esperada" in tr_m.columns
                                   and tr_m["ganancia_esperada"].notna().any())
                _agg = dict(uds=("uds_transferir", "sum"), valor=("valor", "sum"),
                            tiendas=("tienda_destino", "nunique"))
                if _tiene_ganancia:
                    _agg["ganancia"] = ("ganancia_esperada", "sum")
                tg = tr_m.groupby(["sku", "nombre"], as_index=False).agg(**_agg)
                if _tiene_ganancia:
                    # Criterio económico (fórmula Ripley 2026-08-24): solo modelos
                    # cuya ganancia esperada supera el flete. Reemplaza al proxy
                    # de valor-venta; el mínimo de unidades se mantiene.
                    tg = tg[(tg["uds"] >= TRANSF_MIN_UDS) & (tg["ganancia"] > 0)]
                else:
                    tg = tg[(tg["uds"] >= TRANSF_MIN_UDS) & (tg["valor"] >= TRANSF_MIN_VALOR)]
                if not tg.empty:
                    _temp_map = dfm.drop_duplicates("sku").set_index("sku").get("temporada")
                    if _temp_map is not None:
                        tg.insert(2, "temporada", tg["sku"].map(_temp_map))
                    tg = tg.sort_values("ganancia" if _tiene_ganancia else "valor",
                                        ascending=False)
                    tg.columns = (["SKU", "Modelo"] +
                                  (["Temporada"] if _temp_map is not None else []) +
                                  ["Uds a mover", "Valor S/ (venta)", "Tiendas destino"] +
                                  (["Ganancia neta S/"] if _tiene_ganancia else []))
                    _escribir_tabla(
                        w, "4. Transferir",
                        f"{marca} — Rebalanceo entre tiendas (solo movimientos ≥{TRANSF_MIN_UDS} uds y ≥S/{TRANSF_MIN_VALOR:,.0f}) · corte {corte}",
                        tg, {"Uds a mover": _FMT_S, "Valor S/ (venta)": _FMT_S,
                             "Ganancia neta S/": _FMT_S})

        # ── 5. Venta Cero (S7 2026-09-05, pedido Franco): SKU×tienda con stock y sin
        #    venta la última semana, Pareto 80% del capital por tienda. Es lo que el
        #    proveedor necesita para atacar exhibición/precio tienda por tienda. ──
        try:
            _vc = vistas_excel.venta_cero(dfm, min_capital=0)
        except Exception:
            _vc = pd.DataFrame()
        if not _vc.empty:
            vistas_excel.hoja_venta_cero(
                w, _vc, hoja="5. Venta Cero",
                titulo=f"{marca} — SKUs con stock y SIN venta la última semana, por tienda "
                       f"(⭐ TOP 80% = concentran el 80% del capital sin venta de esa tienda) · corte {corte}")

        # ── Tendencias (alertas del motor para la marca) ──
        if df_alertas is not None and not df_alertas.empty and "marca" in df_alertas.columns:
            al = df_alertas[df_alertas["marca"].str.upper().str.strip() == m_up]
            _cols_al = [c for c in ["sku", "nombre", "categoria", "temporada",
                                    "estado_actual", "stock_total", "capital_stock",
                                    "prom_4sem", "variacion_pct", "tipo_alerta",
                                    "detalle"] if c in al.columns]
            if not al.empty and _cols_al:
                alo = al[_cols_al].rename(columns={
                    "sku": "SKU", "nombre": "Modelo", "categoria": "Línea",
                    "temporada": "Temporada",
                    "estado_actual": "Estado", "stock_total": "Stock (uds)",
                    "capital_stock": "Capital S/", "prom_4sem": "Vta prom 4 sem",
                    "variacion_pct": "Variación", "tipo_alerta": "Alerta",
                    "detalle": "Detalle"})
                _escribir_tabla(
                    w, "Tendencias",
                    f"{marca} — Modelos acelerando / frenando (últimas 4 semanas) · corte {corte}",
                    alo, {"Stock (uds)": _FMT_S, "Capital S/": _FMT_S,
                          "Vta prom 4 sem": _FMT_C, "Variación": _FMT_PCT},
                    chips_col="Estado")

        # ── Leyenda ──
        ws = w.book.create_sheet("Leyenda")
        ws["A1"] = "Cómo leer este reporte — estados de stock"
        ws["A1"].font = F_TITULO
        ws["A2"] = ("El estado se calcula por modelo×tienda combinando cobertura "
                    "(semanas de stock según su venta) y edad (semanas desde ingreso).")
        ws["A2"].font = F_HEADER
        fila = 4
        for estado, desc in _LEYENDA_ESTADOS:
            c = ws.cell(row=fila, column=1, value=estado)
            if estado in FILLS_ESTADO:
                c.fill = FILLS_ESTADO[estado]
                c.font = F_CHIP
                c.alignment = CENTRADO
            ws.cell(row=fila, column=2, value=desc)
            fila += 1
        fila += 1
        ws.cell(row=fila, column=1, value="Pirámide de descuentos por antigüedad:").font = F_TITULO
        fila += 1
        for col, h in enumerate(["Edad", "Dscto sugerido", "Tipo"], start=1):
            hc = ws.cell(row=fila, column=col, value=h)
            hc.font = F_HEADER
            hc.fill = FILL_HEAD
        for edad, d, t in _PIRAMIDE:
            fila += 1
            ws.cell(row=fila, column=1, value=edad)
            ws.cell(row=fila, column=2, value=d)
            ws.cell(row=fila, column=3, value=t)
        fila += 2
        ws.cell(row=fila, column=1,
                value="P. Mínimo (piso) = precio más bajo permitido para no vender bajo costo + margen mínimo.").font = F_HEADER
        fila += 1
        ws.cell(row=fila, column=1,
                value="P. Sugerido solo aparece cuando implica BAJAR el precio; si el dscto actual ya supera la pirámide o el piso no deja bajar más, la Acción es Mantener.").font = F_HEADER
        fila += 1
        ws.cell(row=fila, column=1,
                value="Transferir: solo movimientos con ganancia neta positiva (contribución sin IGV × uds vendibles − flete por unidad) y ≥12 uds.").font = F_HEADER
        fila += 1
        ws.cell(row=fila, column=1, value=_SUPUESTOS).font = F_HEADER
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 90

    buf.seek(0)
    return buf.read()


def generar_zip_reportes(df_cob, df_rep=None, df_trans=None,
                         df_prec=None, df_alertas=None, corte: str = None) -> bytes:
    """Zip con un Excel por marca del universo de reporte."""
    zbuf = io.BytesIO()
    corte = corte or f"{date.today():%d.%m.%Y}"
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as z:
        for marca in marcas_reporte(df_cob):
            data = generar_reporte_marca(marca, df_cob, df_rep, df_trans,
                                         df_prec, df_alertas, corte=corte)
            nombre = marca.title().replace(" ", "_")
            z.writestr(f"Reporte_{nombre}_al_{corte}.xlsx", data)
    zbuf.seek(0)
    return zbuf.read()
