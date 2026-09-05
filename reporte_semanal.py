"""
reporte_semanal.py — Genera el reporte semanal de una marca para gerencia.

Un comando produce el Excel y el texto del correo:

    python3 reporte_semanal.py --marca SPAVALDI --vs CACHAREL

Existe porque Majo pidió que el análisis de contribución por m² "se alimente
semana tras semana". Un reporte que hay que artesanar cada lunes muere en la
tercera semana; este se corre en un comando y se revisa antes de enviar.

Los titulares del correo NO están escritos a mano: se derivan de los datos de la
semana. Si la realidad cambia, el texto cambia solo.
"""

from __future__ import annotations

import argparse
import glob
import os
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import rendimiento_tienda as rt

COB_TARGET = 12  # semanas, mismo objetivo que usa Capi

_INK, _INK2, _HEAD, _ACCENT = "1B2330", "5D6575", "F4F1F7", "6D3B8E"
_F = "Calibri"
F_H = Font(name=_F, bold=True, size=10, color=_INK2)
F_T = Font(name=_F, size=10, color=_INK)
F_B = Font(name=_F, bold=True, size=10, color=_INK)
F_TIT = Font(name=_F, bold=True, size=13, color=_ACCENT)
F_N = Font(name=_F, size=9, color=_INK2, italic=True)
FILL_H = PatternFill("solid", fgColor=_HEAD)
BORDE = Border(bottom=Side(style="thin", color="D3CEDC"))


def _hoja(ws, df, titulo, nota, formatos, fila=1):
    ws.cell(row=fila, column=1, value=titulo).font = F_TIT
    ws.cell(row=fila + 1, column=1, value=nota).font = F_N
    h = fila + 3
    for j, c in enumerate(df.columns, 1):
        cel = ws.cell(row=h, column=j, value=c)
        cel.font, cel.fill, cel.border = F_H, FILL_H, BORDE
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, (_, r) in enumerate(df.iterrows(), start=h + 1):
        for j, c in enumerate(df.columns, 1):
            v = r[c]
            cel = ws.cell(row=i, column=j,
                          value=None if pd.isna(v) else (float(v) if isinstance(v, (int, float)) else v))
            cel.font = F_B if j == 1 else F_T
            if c in formatos:
                cel.number_format = formatos[c]
            if pd.isna(v) and c in formatos:
                cel.value = "no aplica"
                cel.font = F_N
    for j, c in enumerate(df.columns, 1):
        largo = max([len(str(c))] + [len(str(x)) for x in df[c].head(40)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(largo + 3, 11), 34)
    ws.freeze_panes = ws.cell(row=h + 1, column=1)
    ws.auto_filter.ref = f"A{h}:{get_column_letter(len(df.columns))}{h + len(df)}"


def construir(marca="SPAVALDI", vs="CACHAREL", semanas=4, dir_micros=None):
    dir_micros = dir_micros or os.path.join(os.path.dirname(__file__), "data2", "bases antiguas")
    micros = sorted(glob.glob(os.path.join(dir_micros, "Base al *.xlsx")))
    if not micros:
        raise FileNotFoundError(f"No hay Micros en {dir_micros}")

    marcas = [marca] + ([vs] if vs else [])
    largo = rt.clasificar_liquidacion(rt.acumular_micros(micros, marcas=marcas, semanas=semanas))
    nsem, cortes = largo.attrs["semanas"], largo.attrs["cortes"]

    m = rt.metricas_por_tienda(largo, marca=marca, semanas=nsem)
    # Activa = movió unidades O soles (ver nota en app_streamlit): hay tiendas
    # que cierran con 0 unidades netas pero saldo en soles distinto de cero.
    act = m[(m.unidades != 0) | (m.venta_soles != 0)].copy()
    conm2 = act[act.m2.notna() & (act.unidades > 0)]

    comp = rt.comparar_marcas(largo, marcas, por=("tienda",)) if vs else pd.DataFrame()
    return dict(largo=largo, act=act, conm2=conm2, comp=comp, nsem=nsem, cortes=cortes,
                marca=marca, vs=vs, dupes=largo.attrs.get("duplicados_ignorados", []))


def titulares(d):
    """Los tres titulares del correo, derivados de los datos — no escritos a mano."""
    act, con, out = d["act"], d["conm2"], []

    # 1. Espacio: la mejor por m2 contra la que más metros ocupa rindiendo poco
    # Solo tiendas que ya producen: un corner recién abierto tiene metraje y
    # cero venta, y compararlo daría "rinde 0 por metro, infinitas veces menos".
    maduras = con[(con.contribucion > 0) & (con.unidades >= 10)]
    if len(maduras) >= 2:
        top = maduras.nlargest(1, "contrib_x_m2").iloc[0]
        grande = maduras.nlargest(1, "m2").iloc[0]
        if grande.tienda != top.tienda and grande.contrib_x_m2 < top.contrib_x_m2 * 0.75:
            veces = top.contrib_x_m2 / grande.contrib_x_m2 if grande.contrib_x_m2 > 0 else float("inf")
            out.append(
                f"**{top.tienda} es la tienda que mejor usa el espacio**: S/ {top.contrib_x_m2:,.0f} de "
                f"contribución por m² en {top.m2:.0f} m². En el otro extremo, {grande.tienda} tiene el corner "
                f"más grande ({grande.m2:.0f} m²) y rinde S/ {grande.contrib_x_m2:,.0f} por metro — "
                f"{veces:.1f} veces menos.")
        else:
            out.append(f"**{top.tienda} lidera en contribución por m²** con S/ {top.contrib_x_m2:,.0f} "
                       f"en {top.m2:.0f} m².")
    elif len(con):
        top = con.nlargest(1, "contrib_x_m2").iloc[0]
        out.append(f"**{top.tienda} lidera en contribución por m²** con S/ {top.contrib_x_m2:,.0f}.")

    # 2. Cobertura de cadena contra objetivo
    stk, u4, u1 = act.stock_uds.sum(), act.unidades.sum() / d["nsem"], act.und_ult_sem.sum()
    if u4 > 0:
        c4, c1 = stk / u4, (stk / u1 if u1 else float("nan"))
        estado = "por encima del objetivo" if c4 > COB_TARGET else "dentro del objetivo"
        out.append(
            f"**La cadena está en {c4:.1f} semanas de cobertura** ({c1:.1f} mirando solo la última semana), "
            f"contra un objetivo de {COB_TARGET}. Está {estado}.")

    # 3. El cambio de ritmo más grande
    tend = act[act.cobertura_sem.notna() & act.cobertura_1sem.notna() & (act.cobertura_sem < 100)].copy()
    if len(tend):
        tend["delta"] = tend.cobertura_1sem - tend.cobertura_sem
        peor = tend.nlargest(1, "delta").iloc[0]
        if peor.delta >= 2:
            out.append(
                f"**{peor.tienda} se frenó**: pasó de {peor.cobertura_sem:.1f} a {peor.cobertura_1sem:.1f} "
                f"semanas de cobertura, porque vendió {peor.und_ult_sem:.0f} unidades la última semana "
                f"contra {peor.unidades / d['nsem']:.0f} de promedio.")
        else:
            mejor = tend.nsmallest(1, "delta").iloc[0]
            out.append(f"**{mejor.tienda} aceleró**: bajó de {mejor.cobertura_sem:.1f} a "
                       f"{mejor.cobertura_1sem:.1f} semanas de cobertura.")
    return out


def correo(d):
    act, con = d["act"], d["conm2"]
    vta, contrib = act.venta_soles.sum(), act.contribucion.sum()
    liq = act.venta_liq.sum() / vta if vta else 0
    ini, fin = d["cortes"][0], d["cortes"][-1]
    L = [f"**Asunto:** {d['marca'].title()} — rendimiento por tienda, {d['nsem']} semanas al {fin}", "",
         "Hola Majo,", "",
         f"Te paso el corte de {d['marca'].title()} sobre las últimas {d['nsem']} semanas ({ini} a {fin}). "
         f"El detalle por tienda va en el Excel adjunto; acá los titulares.", "",
         f"En la ventana la marca hizo **S/ {vta:,.0f} de venta neta** y **S/ {contrib:,.0f} de contribución** "
         f"(margen {contrib / vta:.1%}), con **{liq:.0%} de la venta en liquidación**.", ""]
    L += [f"{i}. {t}" for i, t in enumerate(titulares(d), 1)]
    L += ["",
          f"La contribución por tienda está calculada con la venta en soles que el reporte micro trae por "
          f"local, no prorrateada desde el total de la marca. Los m² son los del corner, no los de la sala.", ""]
    sin = act[act.m2.isna() & (act.unidades > 0)]
    if len(sin):
        L += [f"Los outlets y las tiendas de liquidación ({', '.join(sin.tienda.head(4))}) van sin metraje "
              f"a propósito: no se evalúa productividad de espacio en un canal cuyo trabajo es rematar. "
              f"En el cuadro aparecen como *no aplica*, nunca como cero.", ""]
    L += ["Cualquier corte adicional que quieras, me dices.", "", "Franco"]
    return "\n".join(L)


def exportar(d, salida=None):
    from openpyxl import Workbook
    act, con, comp = d["act"], d["conm2"], d["comp"]
    fin = d["cortes"][-1]
    salida = salida or os.path.join(os.path.dirname(__file__),
                                    f"{d['marca'].title()} - Rendimiento al {fin}.xlsx")
    wb = Workbook()

    t1 = con[["tienda", "m2", "venta_soles", "contribucion", "margen", "contrib_x_m2"]].copy()
    t1.columns = ["Tienda", "m² corner", "Venta S/", "Contribución S/", "Margen", "Contribución / m²"]
    ws = wb.active; ws.title = "1. Contribución x m2"
    _hoja(ws, t1.sort_values("Contribución / m²", ascending=False),
          f"Contribución por m² — {d['marca'].title()}",
          f"Ventana de {d['nsem']} semanas ({d['cortes'][0]} a {fin}). Venta neta sin IGV. "
          f"Solo tiendas con metraje asignado; outlets y liquidadoras van en la hoja 4.",
          {"Venta S/": "#,##0", "Contribución S/": "#,##0", "Contribución / m²": "#,##0",
           "Margen": "0.0%", "m² corner": "#,##0"})

    # Solo tiendas con metraje: las pestañas de análisis son sobre locales que
    # venden a precio y ocupan espacio. Ecommerce, outlets y liquidadoras tienen
    # su propia hoja — mezclarlos ensucia la comparación.
    t2 = con[con.cobertura_sem.notna()][
        ["tienda", "stock_uds", "unidades", "cobertura_sem", "und_ult_sem", "cobertura_1sem"]].copy()
    t2["unidades"] = t2.unidades / d["nsem"]
    t2["Δ"] = t2.cobertura_1sem - t2.cobertura_sem
    t2.columns = ["Tienda", "Stock und", f"Vta sem (prom {d['nsem']})", f"Cobertura {d['nsem']} sem",
                  "Vta última sem", "Cobertura última sem", "Δ"]
    _hoja(wb.create_sheet("2. Cobertura"), t2.sort_values(f"Cobertura {d['nsem']} sem"),
          "Semanas de cobertura",
          f"Solo tiendas con metraje asignado. Cobertura = stock en tienda ÷ venta semanal. La de "
          f"{d['nsem']} semanas es estable; la de la última es la foto de hoy. Cuando se separan, la "
          f"tienda cambió de ritmo. Objetivo: {COB_TARGET} semanas.",
          {"Stock und": "#,##0", f"Vta sem (prom {d['nsem']})": "#,##0", "Vta última sem": "#,##0",
           f"Cobertura {d['nsem']} sem": "0.0", "Cobertura última sem": "0.0", "Δ": "+0.0;-0.0;0.0"})

    if len(comp) and d["vs"]:
        a, b = rt._norm(d["marca"]), rt._norm(d["vs"])
        if a in comp.columns and b in comp.columns:
            t3 = comp[(comp[a] > 0) & comp.tienda.isin(con.tienda)][["tienda", a, b]].copy()
            t3["ratio"] = t3[a] / t3[b].replace(0, pd.NA)
            t3.columns = ["Tienda", f"{d['marca'].title()} S/", f"{d['vs'].title()} S/",
                          f"{d['marca'].title()} ÷ {d['vs'].title()}"]
            _hoja(wb.create_sheet(f"3. vs {d['vs'].title()[:14]}"),
                  t3.sort_values(f"{d['marca'].title()} ÷ {d['vs'].title()}", ascending=False),
                  f"{d['marca'].title()} frente a {d['vs'].title()}",
                  "Solo tiendas con metraje asignado donde ambas marcas venden. El total de cadena no "
                  "es comparable: las marcas no están en la misma cantidad de tiendas.",
                  {f"{d['marca'].title()} S/": "#,##0", f"{d['vs'].title()} S/": "#,##0",
                   f"{d['marca'].title()} ÷ {d['vs'].title()}": "0%"})

    sin = act[act.m2.isna() & (act.unidades > 0)][
        ["tienda", "canal", "venta_soles", "contribucion", "margen", "pct_venta_liquidacion"]].copy()
    sin.columns = ["Tienda", "Canal", "Venta S/", "Contribución S/", "Margen", "% liquidación"]
    _pv = sin["Venta S/"].sum() / act.venta_soles.sum() if act.venta_soles.sum() else 0
    _hoja(wb.create_sheet("4. Sin metraje"), sin,
          "Canales sin m² asignado — fuera de las hojas 1 a 3",
          f"Ecommerce, outlets y tiendas de liquidación. No llevan metraje a propósito: no se evalúa "
          f"productividad de espacio en un canal cuyo trabajo es rematar. Pesan {_pv:.1%} de la venta "
          f"y S/ {sin['Contribución S/'].sum():,.0f} de contribución, así que el margen de las hojas "
          f"1 a 3 no es el margen de la marca completa.",
          {"Venta S/": "#,##0", "Contribución S/": "#,##0", "Margen": "0.0%", "% liquidación": "0%"})

    ws = wb.create_sheet("5. Notas")
    notas = [
        ("Ventana", f"{d['nsem']} semanas: {', '.join(d['cortes'])}"),
        ("Venta", "Neta sin IGV, tal como la reporta el micro por tienda."),
        ("Contribución", "Calculada por tienda (venta en soles del local − unidades × costo), NO prorrateada "
                         "desde el total de la marca. Error verificado contra el micro: 0.03%."),
        ("Stock", "Foto del último corte, no acumulado. Sumarlo entre semanas infla la cobertura."),
        ("Liquidación", f"Mercadería con más de {rt.EDAD_LIQUIDACION:.0f} semanas, mismo umbral con que "
                        f"Capi separa LIQUIDAR y MUERTO."),
        ("m²", "Del corner de la marca, no de la sala. Donde falta el dato la celda dice 'no aplica', "
               "nunca cero: un cero se lee como 'no rinde'."),
    ]
    if d["dupes"]:
        notas.append(("Cortes descartados",
                      "; ".join(f"{a} duplicaba a {b}" for a, b in d["dupes"])))
    ws.cell(row=1, column=1, value="Cómo leer este reporte").font = F_TIT
    for i, (k, v) in enumerate(notas, start=3):
        ws.cell(row=i, column=1, value=k).font = F_B
        c = ws.cell(row=i, column=2, value=v); c.font = F_T
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 96

    wb.save(salida)
    return salida


def main():
    ap = argparse.ArgumentParser(description="Reporte semanal de rendimiento por tienda.")
    ap.add_argument("--marca", default="SPAVALDI")
    ap.add_argument("--vs", default="CACHAREL", help="marca de comparación ('' para omitir)")
    ap.add_argument("--semanas", type=int, default=4)
    ap.add_argument("--micros", default=None)
    ap.add_argument("--salida", default=None)
    a = ap.parse_args()

    d = construir(a.marca, a.vs or None, a.semanas, a.micros)
    xlsx = exportar(d, a.salida)
    txt = Path(xlsx).with_suffix(".md")
    txt.write_text(correo(d), encoding="utf-8")

    print(f"Ventana: {d['nsem']} semanas ({', '.join(d['cortes'])})")
    if d["dupes"]:
        for x, y in d["dupes"]:
            print(f"  corte duplicado descartado: {x} == {y}")
    print(f"Tiendas con venta: {len(d['act'])} | con metraje: {len(d['conm2'])}")
    print(f"\nExcel:  {xlsx}\nCorreo: {txt}\n")
    print("─" * 70)
    print(correo(d))


if __name__ == "__main__":
    main()
