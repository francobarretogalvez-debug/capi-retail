"""
Capi — Herramienta de Gestión Retail
========================================
Lee la plantilla del cliente (4 pestañas) y muestra:
  - Vista 1: Reposición (quiebres, cobertura, pareto tiendas)
  - Vista 2: Sobrestock (real vs aparente, obsoletos, transferencias)
  - Vista 3: Marcas Terceras (margen, contribución, sell-through)
  - Productos Venta Cero + exports accionables
"""

import io
import os
import glob as _glob
import sys
import tempfile

import altair as alt
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils import get_column_letter

# Fix para datasets grandes: aumentar límite de celdas del Pandas Styler
pd.set_option("styler.render.max_elements", 500_000)

# Motor v2 — forzar recarga para que cambios se reflejen sin reiniciar Streamlit
sys.path.insert(0, os.path.dirname(__file__))
import importlib
import motor_v2
importlib.reload(motor_v2)
import renderers_alertas_tienda as R_at
importlib.reload(R_at)
import transformar_profundidad as etl_profundidad
importlib.reload(etl_profundidad)
import vista_planificacion as vista_plan
import vista_talla_color
importlib.reload(vista_plan)

# Snapshots Engine — histórico semanal (Prompt B)
try:
    import snapshots_engine
    _HAS_SNAPSHOTS = True
except ImportError:
    _HAS_SNAPSHOTS = False

# Auto-cargar bases históricas como snapshots en cold start
if _HAS_SNAPSHOTS and not st.session_state.get("_snapshots_initialized"):
    try:
        _existing_weeks = snapshots_engine.list_available_weeks()
        if len(_existing_weeks) < 2:
            _hist_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data2", "bases antiguas")
            if os.path.isdir(_hist_dir):
                snapshots_engine.loader.load_all_bases_antiguas(_hist_dir, force=False)
        st.session_state["_snapshots_initialized"] = True
    except Exception:
        st.session_state["_snapshots_initialized"] = True

# Cargar .env antes de importar agente_terceras
try:
    from dotenv import load_dotenv
    _env_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    load_dotenv(_env_file)
except ImportError:
    # Si no hay python-dotenv, intentar leer .env manualmente
    _env_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(_env_file):
        with open(_env_file) as _f:
            for _line in _f:
                _line = _line.strip()
                if _line and not _line.startswith("#") and "=" in _line:
                    _k, _v = _line.split("=", 1)
                    os.environ[_k.strip()] = _v.strip()

# En Streamlit Cloud la API key vive en st.secrets, no en variables de entorno.
# La propagamos a os.environ para que agente_terceras (que usa
# os.getenv) la encuentren igual que en local con .env.
try:
    if "ANTHROPIC_API_KEY" in st.secrets and not os.getenv("ANTHROPIC_API_KEY"):
        os.environ["ANTHROPIC_API_KEY"] = str(st.secrets["ANTHROPIC_API_KEY"])
except Exception:
    pass

import agente_terceras
import vistas_excel
import reportes_marcas
import acciones_log
import analisis_estados
import comparativo_semanal
import otb_terceras
import rendimiento_tienda as rend_t
import reporte_semanal as rep_sem
import agente_reporte as ag_rep
import calendario_ripley as cal_rip

# ══════════════════════════════════════════════════════════════
#  CONFIG DE PÁGINA
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Capi — Gestión Retail",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Versión visible (S1 robustez, 2026-09-05): se muestra en el sidebar junto al corte
# de la base cargada, para que cualquier número citado sea trazable a una versión.
CAPI_VERSION = "2.1.0"

# ── Paleta de colores Capi (Clean Corporate: navy + light) ──
TEAL_600 = "#6D3B8E"     # Morado Ripley — primary accent (Franco 2026-08-26)
TEAL_700 = "#572E73"     # Morado oscuro — hover/active
TEAL_50  = "#F3EEF8"     # Morado-50 — light accent bg
SLATE_900 = "#1F2937"    # Gray-800 — headings, strong text
SLATE_800 = "#374151"    # Gray-700 — secondary strong
SLATE_700 = "#4B5563"    # Gray-600 — medium text
SLATE_500 = "#6B7280"    # Gray-500 — muted text
SLATE_400 = "#9CA3AF"    # Gray-400 — tertiary text
SLATE_200 = "#E5E7EB"    # Gray-200 — borders
SLATE_100 = "#F3F4F6"    # Gray-100 — surface
SLATE_50  = "#F9FAFB"    # Near-white — page bg
STATUS_CRITICO    = "#EF4444"
STATUS_PRECRITICO = "#F97316"
STATUS_OPTIMO     = "#10B981"
STATUS_ALTO       = "#F59E0B"
STATUS_SOBRESTOCK = "#E11D48"
STATUS_LIQUIDAR   = "#EC4899"
STATUS_NUEVO_SV   = "#94A3B8"
STATUS_DORMIDO    = "#78716C"
STATUS_MUERTO     = "#374151"
# Sprint 1 Capi — nombres nuevos de la taxonomía (Prompt A1)
STATUS_QUIEBRE     = STATUS_CRITICO       # alias semántico
STATUS_BAJA        = STATUS_PRECRITICO    # alias semántico
STATUS_ESTANCADO   = "#475569"            # gris medio (entre SOBRESTOCK y MUERTO)
STATUS_LANZAMIENTO = "#60A5FA"            # azul claro (positivo, "está en rampa")

# ── Tema (corporate light — sin toggle) ──
if "theme_mode" not in st.session_state:
    st.session_state["theme_mode"] = "light"

_IS_LIGHT = True  # Corporate: siempre light

# Variables Python para contextos donde CSS vars no funcionan (Plotly, etc.)
TH_TEXT_PY = SLATE_900
TH_TEXT2_PY = SLATE_500
TH_BG_CARD_PY = "#FFFFFF"
TH_BG_SURFACE_PY = SLATE_50
TH_BORDER_PY = SLATE_200
TH_PLOT_BG = "rgba(0,0,0,0)"
TH_PLOT_GRID = "#E5E7EB"

# CSS custom properties — corporate light
_CSS_VARS = f"""
    :root {{
        --capi-bg: #F9FAFB;
        --capi-bg-card: #FFFFFF;
        --capi-bg-surface: #F3F4F6;
        --capi-text: #1F2937;
        --capi-text2: #6B7280;
        --capi-text3: #9CA3AF;
        --capi-border: #E5E7EB;
        --capi-shadow: rgba(0,0,0,0.04);
        --capi-hover-shadow: rgba(0,0,0,0.08);
        --capi-tab-bg: #F3F4F6;
        --capi-tab-active: #FFFFFF;
        --capi-tab-text: #6B7280;
        --capi-tab-active-text: {TEAL_600};
        --capi-accent: {TEAL_600};
        --capi-accent-light: {TEAL_50};
    }}
"""

# Dark mode override no longer needed — corporate light only
_DARK_BG_OVERRIDE = ""

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');

    {_CSS_VARS}

    /* ── Global ──────────────────────────────── */
    html, body, [class*="css"] {{
        font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    }}
    .stApp, .main, [data-testid="stAppViewContainer"] {{
        background-color: #F9FAFB !important;
    }}
    [data-testid="stHeader"] {{
        background-color: #F9FAFB !important;
    }}
    .main .block-container {{
        padding-top: 1.5rem;
        max-width: 1400px;
    }}
    h1, h2, h3, h4, h5 {{
        font-family: 'Plus Jakarta Sans', sans-serif;
        color: var(--capi-text);
    }}

    /* ── Header (clean corporate) ──────────── */
    .main-header {{
        background: #FFFFFF;
        padding: 1.2rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.2rem;
        display: flex;
        align-items: center;
        justify-content: space-between;
        border: 1px solid {SLATE_200};
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }}
    .main-header h1 {{
        color: {SLATE_900}; margin: 0; font-size: 1.3rem; font-weight: 700;
        letter-spacing: -0.02em;
    }}
    .main-header h1 span {{ color: {TEAL_600}; }}
    .main-header p {{
        color: {SLATE_500}; margin: 0;
        font-size: 0.78rem; font-weight: 400;
    }}

    /* ── Chat input (clean corporate) ───────── */
    [data-testid="stTextInput"] input[aria-label="Pregunta a Capi"] {{
        background: #FFFFFF !important;
        border: 1px solid {SLATE_200} !important;
        border-radius: 10px !important;
        color: {SLATE_900} !important;
        padding: 14px 18px !important;
        font-size: 0.92rem !important;
        height: auto !important;
    }}
    [data-testid="stTextInput"] input[aria-label="Pregunta a Capi"]::placeholder {{
        color: {SLATE_400} !important;
    }}
    [data-testid="stTextInput"] input[aria-label="Pregunta a Capi"]:focus {{
        border-color: {TEAL_600} !important;
        box-shadow: 0 0 0 2px rgba(27,79,114,0.12) !important;
    }}

    /* ── Live badge ─────────────────────────── */
    .live-badge {{
        display: inline-block;
        background: {TEAL_50};
        color: {TEAL_600};
        font-size: 0.65rem;
        font-weight: 700;
        padding: 2px 8px;
        border-radius: 4px;
        letter-spacing: 0.05em;
        margin-left: 8px;
        vertical-align: middle;
    }}

    /* ── Section headers ──────────────────── */
    .section-header {{
        display: flex;
        align-items: center;
        gap: 8px;
        margin: 1.2rem 0 0.6rem 0;
    }}
    .section-header h3 {{
        margin: 0;
        font-size: 1.05rem;
        font-weight: 600;
    }}

    /* ── KPI cards ───────────────────────────── */
    .kpi-card {{
        background: #FFFFFF;
        border-radius: 8px;
        padding: 1.1rem 1.3rem;
        border-left: 3px solid {TEAL_600};
        border: 1px solid {SLATE_200};
        border-left: 3px solid {TEAL_600};
        box-shadow: 0 1px 2px var(--capi-shadow);
        margin-bottom: 0.6rem;
        transition: box-shadow 0.2s ease;
    }}
    .kpi-card:hover {{
        box-shadow: 0 2px 6px var(--capi-hover-shadow);
    }}
    .kpi-val {{
        font-size: 1.8rem; font-weight: 700; color: var(--capi-text);
        line-height: 1; letter-spacing: -0.02em;
    }}
    .kpi-lbl {{
        font-size: 0.74rem; color: var(--capi-text2); margin-top: 0.3rem;
        font-weight: 500; text-transform: uppercase; letter-spacing: 0.03em;
    }}
    .kpi-card.red    {{ border-left-color: {STATUS_CRITICO}; }}
    .kpi-card.red    .kpi-val {{ color: {STATUS_CRITICO}; }}
    .kpi-card.green  {{ border-left-color: #059669; }}
    .kpi-card.green  .kpi-val {{ color: #059669; }}
    .kpi-card.yellow {{ border-left-color: {STATUS_ALTO}; }}
    .kpi-card.yellow .kpi-val {{ color: #B45309; }}
    .kpi-card.orange {{ border-left-color: {STATUS_SOBRESTOCK}; }}
    .kpi-card.orange .kpi-val {{ color: #C2410C; }}
    .kpi-card.darkred {{ border-left-color: {STATUS_LIQUIDAR}; }}
    .kpi-card.darkred .kpi-val {{ color: {STATUS_LIQUIDAR}; }}
    .kpi-card.blue   {{ border-left-color: {TEAL_600}; }}
    .kpi-card.blue   .kpi-val {{ color: {TEAL_600}; }}

    /* ── Sidebar (clean corporate — light) ─── */
    [data-testid="stSidebar"] {{
        background: #FAFBFD;
        border-right: 1px solid {SLATE_200};
    }}
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {{
        padding-top: 0.5rem;
    }}
    [data-testid="stSidebar"] * {{
        color: {SLATE_500} !important;
    }}
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {{
        color: {SLATE_900} !important;
    }}
    [data-testid="stSidebar"] .stButton > button[kind="primary"] {{
        background: {TEAL_600};
        color: white !important;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        padding: 0.6rem 1rem;
        transition: all 0.2s;
        width: 100%;
    }}
    [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {{
        background: {TEAL_700};
        transform: translateY(-1px);
    }}
    [data-testid="stSidebar"] .stSlider label {{
        font-size: 0.82rem !important;
    }}
    .sidebar-nav-item {{
        display: flex;
        align-items: center;
        gap: 10px;
        padding: 8px 12px;
        border-radius: 6px;
        color: {SLATE_500};
        font-size: 0.85rem;
        font-weight: 500;
        transition: all 0.15s;
        cursor: default;
        margin-bottom: 2px;
    }}
    .sidebar-nav-item:hover {{
        background: {SLATE_100};
        color: {SLATE_900};
    }}
    .sidebar-nav-item.active {{
        background: {TEAL_50};
        color: {TEAL_600};
        font-weight: 600;
    }}
    .sidebar-nav-item .nav-icon {{
        font-size: 1rem;
        width: 20px;
        text-align: center;
    }}
    .sidebar-section-label {{
        font-size: 0.68rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: {SLATE_400} !important;
        padding: 16px 12px 6px 12px;
        font-weight: 600;
    }}
    /* Sidebar nav — secondary (inactive) */
    [data-testid="stSidebar"] .stButton > button[kind="secondary"],
    [data-testid="stSidebar"] .stButton > button[data-testid="baseButton-secondary"] {{
        background: transparent !important;
        color: {SLATE_500} !important;
        border: none !important;
        text-align: left !important;
        justify-content: flex-start !important;
        padding: 8px 12px !important;
        border-radius: 6px !important;
        font-size: 0.85rem !important;
        font-weight: 500 !important;
        transition: all 0.15s !important;
        box-shadow: none !important;
    }}
    [data-testid="stSidebar"] .stButton > button[kind="secondary"]:hover,
    [data-testid="stSidebar"] .stButton > button[data-testid="baseButton-secondary"]:hover {{
        background: {SLATE_100} !important;
        color: {SLATE_900} !important;
        border: none !important;
    }}
    /* Sidebar nav — primary (active page) */
    [data-testid="stSidebar"] .stButton > button[kind="primary"],
    [data-testid="stSidebar"] .stButton > button[data-testid="baseButton-primary"] {{
        background: {TEAL_50} !important;
        color: {TEAL_600} !important;
        font-weight: 600 !important;
        border: none !important;
        text-align: left !important;
        justify-content: flex-start !important;
        padding: 8px 12px !important;
        border-radius: 6px !important;
        font-size: 0.85rem !important;
        box-shadow: none !important;
    }}
    /* File uploader in sidebar */
    [data-testid="stSidebar"] [data-testid="stFileUploader"] {{
        border-color: {SLATE_200} !important;
    }}
    [data-testid="stSidebar"] [data-testid="stFileUploader"] section {{
        background: #FFFFFF !important;
        border: 1px dashed {SLATE_200} !important;
        border-radius: 8px !important;
    }}
    [data-testid="stSidebar"] .stExpander {{
        border-color: {SLATE_200} !important;
    }}

    /* ── Ocultar footer y menú hamburguesa ──── */
    footer {{ visibility: hidden; }}
    #MainMenu {{ visibility: hidden; }}

    /* ── Tabs ────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 4px;
        background: {SLATE_100};
        padding: 4px;
        border-radius: 8px;
        border: 1px solid {SLATE_200};
    }}
    .stTabs [data-baseweb="tab"] {{
        padding: 8px 18px;
        border-radius: 6px;
        font-size: 0.82rem;
        font-weight: 500;
        color: {SLATE_500};
        border: none;
        background: transparent;
    }}
    .stTabs [data-baseweb="tab"][aria-selected="true"] {{
        background: #FFFFFF;
        color: {TEAL_600};
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
        font-weight: 600;
    }}
    .stTabs [data-baseweb="tab-highlight"] {{
        display: none;
    }}
    .stTabs [data-baseweb="tab-border"] {{
        display: none;
    }}

    /* ── Dataframes / Tablas ─────────────────── */
    [data-testid="stDataFrame"] {{
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid {SLATE_200};
    }}

    /* ── Expanders ───────────────────────────── */
    .streamlit-expanderHeader {{
        font-weight: 600;
        font-size: 0.9rem;
        color: {SLATE_900};
    }}
    [data-testid="stExpander"] {{
        border-color: {SLATE_200} !important;
    }}

    /* ── Métricas ────────────────────────────── */
    [data-testid="stMetric"] {{
        background: #FFFFFF;
        padding: 1rem;
        border-radius: 8px;
        border: 1px solid {SLATE_200};
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }}
    [data-testid="stMetricLabel"] {{
        font-size: 0.78rem !important;
        color: {SLATE_500} !important;
        font-weight: 500 !important;
        text-transform: uppercase;
        letter-spacing: 0.03em;
    }}
    [data-testid="stMetricValue"] {{
        font-size: 1.6rem !important;
        font-weight: 700 !important;
        color: {SLATE_900} !important;
    }}

    /* ── Botones de descarga ─────────────────── */
    .stDownloadButton > button {{
        background: #FFFFFF;
        color: {TEAL_600};
        border: 1px solid {TEAL_600};
        border-radius: 6px;
        font-weight: 600;
        transition: all 0.2s;
    }}
    .stDownloadButton > button:hover {{
        background: {TEAL_50};
        border-color: {TEAL_700};
    }}

    /* ── Alertas / Dividers ──────────────────── */
    hr {{
        border: none;
        border-top: 1px solid {SLATE_200};
        margin: 1.8rem 0;
    }}

    /* ── Briefing cards ──────────────────────── */
    .briefing-card {{
        border-radius: 8px;
        padding: 14px 18px;
        margin-bottom: 6px;
        border-left: 3px solid;
        font-size: 0.92em;
        background: #FFFFFF;
    }}

    /* ── Card containers ──────────────────── */
    .nansen-card {{
        background: #FFFFFF;
        border: 1px solid {SLATE_200};
        border-radius: 8px;
        padding: 18px 22px;
        margin-bottom: 12px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }}
    .nansen-card:hover {{
        box-shadow: 0 2px 6px rgba(0,0,0,0.06);
    }}

    /* ── Selectboxes y inputs (corporate light) ── */
    .stSelectbox [data-baseweb="select"] {{
        background-color: #FFFFFF !important;
        border-color: {SLATE_200} !important;
    }}
    .stMultiSelect [data-baseweb="select"] {{
        background-color: #FFFFFF !important;
        border-color: {SLATE_200} !important;
    }}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  HELPERS DE COLOR
# ══════════════════════════════════════════════════════════════

BADGE_CSS = {
    "QUIEBRE":          f"background:{STATUS_QUIEBRE}; color:#FFFFFF",
    "PRE-QUIEBRE":      f"background:{STATUS_BAJA}; color:#FFFFFF",
    "ÓPTIMO":           f"background:{STATUS_OPTIMO}; color:#FFFFFF",
    "ALTO":             f"background:{STATUS_ALTO}; color:#FFFFFF",
    "SOBRESTOCK":       f"background:{STATUS_SOBRESTOCK}; color:#FFFFFF",
    "LIQUIDAR":         f"background:{STATUS_LIQUIDAR}; color:#FFFFFF",
    "NUEVO SIN VENTA":  f"background:{STATUS_NUEVO_SV}; color:#FFFFFF",
    "DORMIDO":          f"background:{STATUS_DORMIDO}; color:#FFFFFF",
    "MUERTO":           f"background:{STATUS_MUERTO}; color:#FFFFFF",
    "ESTANCADO":        f"background:{STATUS_ESTANCADO}; color:#FFFFFF",
}


def _badge(val):
    css = BADGE_CSS.get(str(val), "background:#E2E8F0; color:#334155")
    return f'<span style="display:inline-block;padding:3px 12px;border-radius:20px;font-size:0.73rem;font-weight:600;letter-spacing:0.02em;{css}">{val}</span>'


def color_estado(val):
    colors = {
        "QUIEBRE":          f"background-color:{STATUS_QUIEBRE}; color:#FFFFFF",
        "PRE-QUIEBRE":      f"background-color:{STATUS_BAJA}; color:#FFFFFF",
        "ÓPTIMO":           f"background-color:{STATUS_OPTIMO}; color:#FFFFFF",
        "ALTO":             f"background-color:{STATUS_ALTO}; color:#FFFFFF",
        "SOBRESTOCK":       f"background-color:{STATUS_SOBRESTOCK}; color:#FFFFFF",
        "LIQUIDAR":         f"background-color:{STATUS_LIQUIDAR}; color:#FFFFFF",
        "NUEVO SIN VENTA":  f"background-color:{STATUS_NUEVO_SV}; color:#FFFFFF",
        "DORMIDO":          f"background-color:{STATUS_DORMIDO}; color:#FFFFFF",
        "MUERTO":           f"background-color:{STATUS_MUERTO}; color:#FFFFFF",
        "ESTANCADO":        f"background-color:{STATUS_ESTANCADO}; color:#FFFFFF",
    }
    return colors.get(str(val), "")


def color_cobertura(val, params):
    try:
        v = float(val)
    except (TypeError, ValueError):
        return f"background-color:{STATUS_NUEVO_SV}; color:#FFFFFF"
    if v < params.get("umbral_critico", 4):
        return f"background-color:{STATUS_CRITICO}; color:#FFFFFF"
    elif v < params.get("umbral_precritico", 8):
        return f"background-color:{STATUS_PRECRITICO}; color:#FFFFFF"
    elif v < params.get("umbral_optimo", 12):
        return f"background-color:{STATUS_OPTIMO}; color:#FFFFFF"
    elif v < params.get("umbral_alto", 16):
        return f"background-color:{STATUS_ALTO}; color:#FFFFFF"
    else:
        return f"background-color:{STATUS_SOBRESTOCK}; color:#FFFFFF"


def color_dscto(val):
    try:
        v = float(val)
    except (TypeError, ValueError):
        return ""
    if v >= 0.40:
        return f"background-color:{STATUS_CRITICO}; color:#FFFFFF"
    elif v >= 0.25:
        return f"background-color:{STATUS_SOBRESTOCK}; color:#FFFFFF"
    elif v > 0:
        return f"background-color:{STATUS_ALTO}; color:#FFFFFF"
    return ""


def _kpi_html(valor, label, css_class="blue"):
    return f"""
    <div class="kpi-card {css_class}">
        <div class="kpi-val">{valor}</div>
        <div class="kpi-lbl">{label}</div>
    </div>"""


# ══════════════════════════════════════════════════════════════
#  ESTADO DE SESIÓN (inicializar ANTES del sidebar)
# ══════════════════════════════════════════════════════════════

if "results" not in st.session_state:
    st.session_state["results"] = None

# ── Modo demo (?demo=1 en la URL): nav simplificada + auto-carga de base ──
try:
    _DEMO_MODE = st.query_params.get("demo") == "1"
except Exception:
    _DEMO_MODE = False

# ══════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════

with st.sidebar:
    # ── Toggle día/noche + logo ──
    _sb_top1, _sb_top2 = st.columns([3, 1])
    with _sb_top1:
        st.markdown(f"""
        <div style="padding: 0.6rem 0 0.4rem 0; display:flex; align-items:center; gap:8px;">
            <div style="width:28px; height:28px; background:{TEAL_600}; border-radius:6px; display:flex; align-items:center; justify-content:center; color:white; font-size:12px; font-weight:700;">C</div>
            <div>
                <span style="font-size:1.2rem; font-weight:700; color:{SLATE_900}; letter-spacing:-0.03em;">Capi</span>
                <span style="font-size:0.6rem; color:{SLATE_400}; display:block; margin-top:1px; letter-spacing:0.05em;">GESTIÓN RETAIL</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with _sb_top2:
        _theme_icon = "☀️" if st.session_state["theme_mode"] == "dark" else "🌙"
        if st.button(_theme_icon, key="theme_toggle", help="Cambiar entre modo día y noche"):
            st.session_state["theme_mode"] = "light" if st.session_state["theme_mode"] == "dark" else "dark"
            st.rerun()
    st.markdown(f'<div style="border-bottom: 1px solid {SLATE_200}; margin-bottom: 0.8rem;"></div>', unsafe_allow_html=True)

    # ── Versión + corte de la base cargada (trazabilidad) ──
    _corte_txt = "sin base cargada"
    _bp = st.session_state.get("_base_profundidad_path")
    if _bp:
        try:
            _fc = etl_profundidad.fecha_corte_desde_nombre(os.path.basename(_bp))  # 'dd/mm/yyyy'
            _corte_txt = f"base al {_fc.replace('/', '.')}" if _fc else os.path.basename(_bp)
        except Exception:
            _corte_txt = os.path.basename(_bp)
    st.caption(f"Capi v{CAPI_VERSION} · {_corte_txt}")

    if _DEMO_MODE:
        st.caption("🎬 Modo demo activo")

    # ── Navegación funcional con botones ──
    _has_results = st.session_state["results"] is not None

    if "nav_page" not in st.session_state:
        st.session_state["nav_page"] = "🏠 Dashboard"

    if _has_results:
        # ── VISIÓN GENERAL ──
        st.markdown('<div class="sidebar-section-label">VISIÓN GENERAL</div>', unsafe_allow_html=True)

        _NAV_VISION = [
            ("🏠", "Dashboard"),
        ]
        for _icon, _label in _NAV_VISION:
            _full = f"{_icon} {_label}"
            _is_active = st.session_state["nav_page"] == _full
            if st.button(
                _full, key=f"nav_{_label}",
                use_container_width=True,
                type="primary" if _is_active else "secondary",
            ):
                st.session_state["nav_page"] = _full
                st.rerun()

        # ── GESTIÓN DE MARCAS PROPIAS ──
        # Mismo cluster que terceras, filtrado a las 7 marcas propias.
        # ── GESTIÓN DE MARCAS (fusión C2 2026-08-26: 8 vistas → 4 con
        #    selector de universo Propias/Terceras/Todas — decisión Franco) ──
        if not _DEMO_MODE:
            st.markdown('<div class="sidebar-section-label">GESTIÓN DE MARCAS</div>', unsafe_allow_html=True)
            _NAV_MARCAS = [
                ("📦", "Reposición"),
                ("🔄", "Transferencias"),
                ("💰", "Gestión de Precios"),
                ("🚚", "Predistribución"),
                ("🤝", "Agente Terceras"),
            ]
            for _icon, _label in _NAV_MARCAS:
                _full = f"{_icon} {_label}"
                _is_active = st.session_state["nav_page"] == _full
                if st.button(
                    _full, key=f"nav_{_label}",
                    use_container_width=True,
                    type="primary" if _is_active else "secondary",
                ):
                    st.session_state["nav_page"] = _full
                    st.rerun()

        # (Gestión de Stock y Gestión Comercial se quitaron del menú; sus cálculos
        #  siguen en el motor. La vista de cobertura por tienda se eliminó el
        #  2026-09-04: su ETL filtraba como "mercadería inactiva" todo lo PV con
        #  más de 40 semanas, así que escondía justo el stock viejo — Iquitos
        #  mostraba 6 unidades cuando tenía 806. La cobertura ahora vive en
        #  Rendimiento de Marca, que lee el stock directo del micro.)

        if not _DEMO_MODE:
            # ── ANÁLISIS PREDICTIVO ──
            st.markdown('<div class="sidebar-section-label">DISTRIBUCIÓN INTELIGENTE</div>', unsafe_allow_html=True)

            _NAV_PREDICTIVO = [
                ("🎯", "Match Producto-Plaza"),
                ("🧵", "Talla y Color"),
                ("📊", "Planificación"),
            ]

            for _icon, _label in _NAV_PREDICTIVO:
                _full = f"{_icon} {_label}"
                _is_active = st.session_state["nav_page"] == _full
                if st.button(
                    _full, key=f"nav_{_label}",
                    use_container_width=True,
                    type="primary" if _is_active else "secondary",
                ):
                    st.session_state["nav_page"] = _full
                    st.rerun()

        # ── ESTADO DEL STOCK ──
        # Vistas de diagnóstico transversal (todas las marcas): salud, cobertura,
        # venta cero y antigüedad. Separadas de Visión General (solo Dashboard).
        _NAV_ESTADO = [
            ("🩺", "Salud del Stock"),
            ("📲", "Productos Venta Cero"),
            ("📊", "Gestión por Antigüedad"),
            ("🏆", "Caso de Éxito"),
        ]
        if _DEMO_MODE:
            # Demo: solo las vistas protagonistas del guion de 3 minutos
            _NAV_ESTADO = [
                ("🩺", "Salud del Stock"),
                ("📲", "Productos Venta Cero"),
            ]
        st.markdown('<div class="sidebar-section-label">ESTADO DEL STOCK</div>', unsafe_allow_html=True)
        for _icon, _label in _NAV_ESTADO:
            _full = f"{_icon} {_label}"
            _is_active = st.session_state["nav_page"] == _full
            if st.button(
                _full, key=f"nav_{_label}",
                use_container_width=True,
                type="primary" if _is_active else "secondary",
            ):
                st.session_state["nav_page"] = _full
                st.rerun()

        st.markdown(f'<div style="border-bottom:1px solid {SLATE_200}; margin:4px 0 12px 0;"></div>', unsafe_allow_html=True)

        # ── RENDIMIENTO POR TIENDA ──
        # Sección propia: cruza marcas propias y terceras, así que no encaja en
        # ninguno de los dos clusters de gestión. Responde contribución/m², que
        # junto con EBITDA es lo que mira el dueño para área comercial.
        _NAV_RENDIMIENTO = [("📐", "Rendimiento de Marca")]
        st.markdown('<div class="sidebar-section-label">RENDIMIENTO DE MARCA</div>', unsafe_allow_html=True)
        for _icon, _label in _NAV_RENDIMIENTO:
            _full = f"{_icon} {_label}"
            _is_active = st.session_state["nav_page"] == _full
            if st.button(
                _full, key=f"nav_{_label}",
                use_container_width=True,
                type="primary" if _is_active else "secondary",
            ):
                st.session_state["nav_page"] = _full
                st.rerun()

        st.markdown(f'<div style="border-bottom:1px solid {SLATE_200}; margin:4px 0 12px 0;"></div>', unsafe_allow_html=True)

    nav_page = st.session_state["nav_page"] if _has_results else None

    # ── Upload ──
    uploaded = st.file_uploader(
        "Sube tu archivo Excel",
        type=["xlsx"],
        help="Sube directamente la Base Profundidad de Ripley o una Plantilla Capi. Se detecta y transforma automáticamente.",
    )
    formato_input = 'auto'

    # ── Ejecutar ──
    run_btn = st.button("🚀 Ejecutar análisis", use_container_width=True, type="primary")

    # ── Paso 3: Parámetros avanzados (colapsados por defecto) ──
    with st.expander("⚙️ Configuración avanzada", expanded=False):
        st.markdown("**Umbrales de cobertura**")
        umbral_critico    = st.slider("QUIEBRE — menor a (sem)",       min_value=1,  max_value=8,  value=4,  step=1)
        umbral_precritico = st.slider("PRE-QUIEBRE — hasta (sem)",   min_value=4,  max_value=12, value=8,  step=1)
        umbral_optimo     = st.slider("ÓPTIMO — hasta (sem)",         min_value=6,  max_value=16, value=12, step=1)
        umbral_alto       = st.slider("ALTO — hasta (sem)",           min_value=8,  max_value=32, value=16, step=1)
        umbral_edad       = st.slider("LIQUIDAR — antigüedad (sem)",  min_value=12, max_value=52, value=26, step=2)

        st.markdown("**Parámetros de cálculo**")
        cob_target    = st.slider("Cobertura objetivo (sem)",       min_value=4,  max_value=16, value=12,  step=1)
        uds_min_trans = st.slider("Uds. mínimas por transferencia", min_value=1,  max_value=20, value=3,  step=1)
        costo_transf_unit = st.number_input(
            "Costo transferencia (S/ por unidad)", min_value=0.0, max_value=20.0,
            value=3.5, step=0.5,
            help="Flete unitario del cálculo de rentabilidad (hoja Ripley 'Evaluar "
                 "posible transferencia'). Ganancia = uds × contribución sin IGV − flete × uds.")
        margen_min    = st.slider("Margen mínimo para precio (%)",  min_value=5,  max_value=40, value=15, step=5) / 100

        st.markdown("**Alertas Sobrestock (Tiendas)**")
        alertas_tienda_cob_min = st.slider(
            "Cobertura mín. para alertar (sem)",
            min_value=8, max_value=32, value=16, step=2,
            help="SKUs con cobertura ≥ este valor disparan alerta de revisión a tienda (exhibición + precio)."
        )
        alertas_tienda_edad_min = st.slider(
            "Edad mín. para alertar (sem)",
            min_value=1, max_value=8, value=2, step=1,
            help="Leadtime de llegada/exhibición. Productos más nuevos no se alertan."
        )
        alertas_tienda_top_n = st.slider(
            "Máx. ítems por tienda (sobrestock)",
            min_value=10, max_value=50, value=30, step=5,
            help="Tope de SKUs por tienda, ordenados por capital parado. Menos = WhatsApp más corto."
        )

        st.markdown("**Regla de descuento (Reposición)**")
        excluir_descuento_alto = st.checkbox(
            "Excluir SKUs con descuento ≥40% del plan de reposición",
            value=True, key="cfg_excl_dscto",
            help="Regla Majo: no reponer productos que ya tienen descuento alto (≥40%). Se aplica en el motor de cálculo."
        )
        umbral_descuento_repo = st.slider(
            "Umbral de descuento (%)",
            min_value=20, max_value=60, value=40, step=5, key="cfg_umbral_dscto",
            help="SKUs con descuento ≥ este % se excluyen del plan de reposición."
        ) / 100

        st.markdown("**Alertas Venta Cero (Tiendas)**")
        alertas_vc_capital_min = st.slider(
            "Capital mín. para alertar (S/)",
            min_value=200, max_value=5000, value=1000, step=100,
            help="Solo alertar SKUs con stock a costo ≥ este valor."
        )
        alertas_vc_top_por_marca = st.slider(
            "Máx. SKUs por marca×tienda",
            min_value=5, max_value=30, value=15, step=5,
            help="Top N SKUs por marca dentro de cada tienda, ordenados por capital parado."
        )

        st.markdown("**Confiabilidad del Stock CD (ATP)**")
        cd_prometible_pct = st.slider(
            "Stock CD prometible (%)",
            min_value=10, max_value=100, value=60, step=5,
            help="El reporte de CD no es tiempo real y el stock varía entre cortes. "
                 "Solo este % del CD reportado se considera disponible para prometer despachos. "
                 "Calibración Franco 2026-06-12: 60%."
        )

    params_ui = {
        "umbral_critico":    umbral_critico,
        "umbral_precritico": umbral_precritico,
        "umbral_optimo":     umbral_optimo,
        "umbral_alto":       umbral_alto,
        "umbral_edad":       umbral_edad,
        "cob_target":        cob_target,
        "uds_min_trans":  uds_min_trans,
        "costo_transf_unit": costo_transf_unit,
        "margen_min":     margen_min,
        "alertas_tienda_cob_min":  alertas_tienda_cob_min,
        "alertas_tienda_edad_min": alertas_tienda_edad_min,
        "alertas_tienda_top_n":    alertas_tienda_top_n,
        "alertas_vta_cero_capital_min":   alertas_vc_capital_min,
        "alertas_vta_cero_top_por_marca": alertas_vc_top_por_marca,
        "excluir_descuento_alto":  excluir_descuento_alto,
        "umbral_descuento_repo":   umbral_descuento_repo,
    }

    st.markdown(f"""
    <div style="border-top:1px solid {SLATE_200}; margin-top:1.5rem; padding-top:0.8rem; text-align:center;">
        <div class="sidebar-nav-item" style="justify-content:center; margin-bottom:4px;">
            <span class="nav-icon">⚙️</span> Settings
        </div>
        <span style="font-size:0.65rem; color:{SLATE_400}; letter-spacing:0.05em;">
            v2.7 · Powered by AI
        </span>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  HEADER
# ══════════════════════════════════════════════════════════════

st.markdown(f"""
<div class="main-header">
    <div>
        <h1><span>Retail</span>AI</h1>
        <p>Inventory Engine · Cobertura · Reposiciones · Transferencias · Alertas</p>
    </div>
    <div style="display:flex; align-items:center; gap:12px;">
        <span style="font-size:0.75rem; color:{SLATE_400};">v2.7</span>
        <span style="background:{TEAL_600}; color:white; padding:4px 12px; border-radius:8px; font-size:0.75rem; font-weight:600;">Powered by AI</span>
    </div>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  AUTO-DETECCIÓN DE FORMATO
# ══════════════════════════════════════════════════════════════

def _is_base_profundidad(path):
    """Detecta si el archivo es una Base Profundidad (formato wide de Ripley)
    vs una Plantilla Capi (4 pestañas procesadas)."""
    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
        # La plantilla Capi tiene estas pestañas específicas
        if '1. Maestro Productos' in xl.sheet_names:
            return False
        # La Base Profundidad es una sola hoja con muchas columnas (200+)
        df_head = pd.read_excel(path, nrows=1)
        if len(df_head.columns) > 100:
            return True
        return False
    except Exception:
        return False


# ══════════════════════════════════════════════════════════════
#  EJECUCIÓN DEL ANÁLISIS
# ══════════════════════════════════════════════════════════════

if run_btn:
    if uploaded is None:
        st.warning("⚠️ Primero sube tu archivo Excel para continuar.")
    else:
        try:
            # Guardar archivo subido a temporal
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(uploaded.getvalue())
                tmp_path = tmp.name

            # Auto-detectar formato
            if _is_base_profundidad(tmp_path):
                # ── Contrato de entrada + modo seguro (S1, 2026-09-05) ──
                _problemas = etl_profundidad.validar_base(tmp_path)
                if _problemas:
                    st.error("❌ La base subida no pasa la validación y NO se analizó, para no mostrar "
                             "números mal mapeados:\n\n- " + "\n- ".join(_problemas))
                    _ult = st.session_state.get("_base_profundidad_path")
                    _bases_dir = os.path.join(os.path.dirname(__file__), "data2", "bases antiguas")
                    if not _ult and os.path.isdir(_bases_dir):
                        _cands = sorted(_glob.glob(os.path.join(_bases_dir, "Base al *.xlsx")), key=os.path.getmtime)
                        _ult = _cands[-1] if _cands else None
                    if _ult and os.path.exists(_ult):
                        st.warning(f"🛟 Modo seguro: se mantiene el último corte válido disponible "
                                   f"(**{os.path.basename(_ult)}**). Corrige el archivo y vuelve a subirlo.")
                        if st.session_state.get("results") is None:
                            st.session_state["_modo_seguro_pendiente"] = _ult
                    st.stop()
                with st.spinner("Detectada Base Profundidad de Ripley. Transformando..."):
                    # Guardar copia en data2/bases antiguas/ para Afinidad y snapshots
                    _bases_dir = os.path.join(os.path.dirname(__file__), "data2", "bases antiguas")
                    os.makedirs(_bases_dir, exist_ok=True)
                    _base_copy_name = uploaded.name if uploaded.name else "Base_subida.xlsx"
                    _base_copy_path = os.path.join(_bases_dir, _base_copy_name)
                    import shutil as _shutil_cp
                    _shutil_cp.copy2(tmp_path, _base_copy_path)
                    st.session_state["_base_profundidad_path"] = _base_copy_path

                    # Guardar snapshot ANTES de transformar (columnas originales Ripley)
                    if _HAS_SNAPSHOTS:
                        try:
                            snapshots_engine.process_micro_profundidad(_base_copy_path, force=True)
                        except Exception as _e_snap:
                            # No bloquea el análisis, pero la semana perdida se avisa (S1, 2026-09-05)
                            st.warning(f"⚠️ El snapshot semanal no se guardó: {_e_snap}. "
                                       "El análisis sigue, pero esta semana no entra al comparativo.")

                    plantilla_path = tmp_path.replace(".xlsx", "_plantilla.xlsx")
                    etl_profundidad.transform(
                        tmp_path, output_path=plantilla_path,
                        fecha_corte=etl_profundidad.fecha_corte_desde_nombre(uploaded.name))
                    os.unlink(tmp_path)
                    tmp_path = plantilla_path
                    st.toast("Base Profundidad transformada a plantilla Capi")

            with st.spinner("Ejecutando análisis..."):
                results = motor_v2.run_analysis(tmp_path, params=params_ui, formato=formato_input)

                # Snapshot para formato plantilla (non-profundidad uploads)
                if _HAS_SNAPSHOTS and not st.session_state.get("_base_profundidad_path"):
                    try:
                        snapshots_engine.process_micro_profundidad(tmp_path, force=True)
                    except Exception as _e_snap:
                        st.warning(f"⚠️ El snapshot semanal no se guardó: {_e_snap}. "
                                   "El análisis sigue, pero esta semana no entra al comparativo.")

                os.unlink(tmp_path)
                st.session_state["results"] = results
                st.rerun()  # Forzar rerun para que sidebar se re-renderice con nav

        except Exception as e:
            import logging as _logging, traceback as _tb
            _logging.getLogger("capi").exception("Error al procesar la base subida")
            st.error(f"❌ No se pudo procesar el archivo: {e}")
            with st.expander("Detalle técnico (para soporte)"):
                st.code(_tb.format_exc())

# ── Modo seguro (S1): si la base nueva falló la validación y no hay análisis en
#    memoria, ofrecer cargar el último corte válido en vez de dejar la app en blanco.
_ms = st.session_state.get("_modo_seguro_pendiente")
if _ms and st.session_state.get("results") is None and os.path.exists(_ms):
    if st.button(f"🛟 Cargar el último corte válido: {os.path.basename(_ms)}", key="btn_modo_seguro",
                 use_container_width=True, type="primary"):
        try:
            with st.spinner(f"Cargando {os.path.basename(_ms)} en modo seguro..."):
                _pl = os.path.join(tempfile.gettempdir(), "capi_modo_seguro_plantilla.xlsx")
                etl_profundidad.transform(_ms, output_path=_pl,
                                          fecha_corte=etl_profundidad.fecha_corte_desde_nombre(os.path.basename(_ms)))
                st.session_state["_base_profundidad_path"] = _ms
                st.session_state["results"] = motor_v2.run_analysis(_pl, params=params_ui, formato=formato_input)
                st.session_state["_modo_seguro_activo"] = os.path.basename(_ms)
                st.session_state.pop("_modo_seguro_pendiente", None)
            st.rerun()
        except Exception as _e_ms:
            st.error(f"No se pudo cargar el último corte válido: {_e_ms}")
if st.session_state.get("_modo_seguro_activo"):
    st.warning(f"🛟 Modo seguro: estás viendo el corte **{st.session_state['_modo_seguro_activo']}** porque la última "
               "base subida no pasó la validación. Corrige el archivo y vuelve a subirlo.")


# ══════════════════════════════════════════════════════════════
#  MODO DEMO: AUTO-CARGA DE LA ÚLTIMA BASE DISPONIBLE
#  Con ?demo=1 la app arranca con datos sin pedir upload —
#  crítico para que la demo muestre insights en segundos.
# ══════════════════════════════════════════════════════════════

if _DEMO_MODE and st.session_state["results"] is None and not st.session_state.get("_demo_autoload_done"):
    st.session_state["_demo_autoload_done"] = True
    import re as _re_demo

    _demo_dir = os.path.join(os.path.dirname(__file__), "data2", "bases antiguas")

    def _demo_fecha_archivo(nombre):
        """Extrae (año, mes, día) de nombres tipo 'Base al 01.06.26.xlsx' para elegir la más reciente."""
        _m = _re_demo.search(r"(\d{1,2})\.(\d{1,2})(?:\.(\d{2}))?", nombre)
        if not _m:
            return (0, 0, 0)
        _yy = int(_m.group(3)) if _m.group(3) else 26
        return (_yy, int(_m.group(2)), int(_m.group(1)))

    try:
        _demo_bases = [f for f in os.listdir(_demo_dir)
                       if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    except OSError:
        _demo_bases = []

    if _demo_bases:
        _demo_base = max(_demo_bases, key=_demo_fecha_archivo)
        _demo_path = os.path.join(_demo_dir, _demo_base)
        try:
            with st.spinner(f"Modo demo: cargando {_demo_base}…"):
                if _is_base_profundidad(_demo_path):
                    _demo_plantilla = os.path.join(tempfile.gettempdir(), "capi_demo_plantilla.xlsx")
                    etl_profundidad.transform(_demo_path, output_path=_demo_plantilla)
                    _demo_input = _demo_plantilla
                    # Rendimiento por Tienda relee el Micro crudo (la plantilla ya
                    # perdió las columnas por tienda), así que hay que dejar la ruta.
                    st.session_state["_base_profundidad_path"] = _demo_path
                else:
                    _demo_input = _demo_path
                st.session_state["results"] = motor_v2.run_analysis(_demo_input, params=params_ui, formato=formato_input)
            st.rerun()
        except Exception as _demo_err:
            st.warning(f"Modo demo: no se pudo auto-cargar la base ({_demo_err}). Sube un archivo manualmente.")


# ══════════════════════════════════════════════════════════════
#  PANTALLA DE BIENVENIDA (sin datos)
# ══════════════════════════════════════════════════════════════

if st.session_state["results"] is None:
    st.markdown(f"""
    <div style="text-align:center; padding:60px 20px;">
        <div style="font-size:2.2rem; font-weight:700; color:var(--capi-text); margin-bottom:8px;">
            ¿Qué está pasando con tu inventario?
        </div>
        <p style="color:var(--capi-text2); font-size:1rem; margin-bottom:30px;">
            Sube tu Base Profundidad para desbloquear el análisis completo.
        </p>
        <div style="background:var(--capi-bg-surface); border:1px solid var(--capi-border); border-radius:14px; padding:24px; max-width:500px; margin:0 auto; text-align:left;">
            <div style="font-weight:600; color:var(--capi-text); margin-bottom:12px;">Cómo empezar:</div>
            <div style="color:var(--capi-text); font-size:0.9rem; line-height:1.8;">
                <span style="color:{TEAL_600}; font-weight:600;">1.</span> Sube tu archivo Excel en el sidebar<br>
                <span style="color:{TEAL_600}; font-weight:600;">2.</span> Ajusta umbrales si lo necesitas<br>
                <span style="color:{TEAL_600}; font-weight:600;">3.</span> Haz clic en <strong>Ejecutar análisis</strong>
            </div>
            <div style="margin-top:14px; padding:10px 14px; background:var(--capi-bg-card); border-radius:8px; border:1px solid var(--capi-border);">
                <span style="font-size:0.82rem; color:var(--capi-text2);">
                    Acepta <strong>Base Profundidad</strong> de Ripley o Plantilla Capi. Se detecta y transforma automáticamente.
                </span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ══════════════════════════════════════════════════════════════
#  RESULTADOS
# ══════════════════════════════════════════════════════════════

res    = st.session_state["results"]
s      = res.get("summary", {})
params = res["params"]

df_cob       = res.get("cobertura", pd.DataFrame())
df_rep       = res.get("reposiciones", pd.DataFrame())
df_rep_pivot = res.get("reposiciones_pivot", pd.DataFrame())
df_trans     = res.get("transferencias", pd.DataFrame())
df_prec      = res.get("acciones_precio", pd.DataFrame())
df_alertas   = res.get("alertas", pd.DataFrame())
df_anomalias = res.get("anomalias_tienda", pd.DataFrame())
alertas_tienda_dict = res.get("alertas_tienda", {})
alertas_venta_cero_dict = res.get("alertas_venta_cero", {})
briefing     = res.get("briefing") or {}
briefing_items  = briefing.get('items', [])
briefing_tablas = briefing.get('tablas', {})
aging        = res.get("aging", {})
df_aging     = aging.get('df_aging', pd.DataFrame())
aging_kpis   = aging.get('kpis', {})
aging_top_ejemplos = aging.get('top_ejemplos', {})
embarque     = res.get("embarque") or {}
df_embarque  = embarque.get('df_embarque', pd.DataFrame())
por_ventana  = embarque.get('por_ventana', pd.DataFrame())
embarque_kpis = embarque.get('kpis', {})
embarque_recs = embarque.get('recomendaciones', {})
embarque_top  = embarque.get('top_problemas', {})
predist       = res.get("predistribucion") or {}
df_retenidos_cd   = predist.get('retenidos_cd', pd.DataFrame())
df_gaps_dist      = predist.get('gaps_distribucion', pd.DataFrame())
predist_kpis      = predist.get('kpis', {})

# Margen efectivo (Contribución / VtasMF)
_margen_global = s.get('margen_efectivo_global', None)
_vta_soles_total = s.get('vta_soles_4sem_total', 0)
_contrib_soles_total = s.get('contrib_soles_4sem_total', 0)

# LY Comparison + Ticket Promedio
ly_comparison = res.get('ly_comparison', {})
_margen_por_marca = s.get('margen_por_marca', [])

# ══════════════════════════════════════════════════════════════
#  FILTRO GLOBAL: Solo marcas vigentes en Ripley
# ══════════════════════════════════════════════════════════════
# Solo se analizan estas marcas. El resto se descarta antes de
# cualquier visualización para no ensuciar los datos.

_MARCAS_VIGENTES = {
    "MARQUIS", "NAVIGATA", "CACHAREL", "SPAVALDI",
    "OSCAR DE LA RENTA", "US POLO", "JOHN HOLDEN",
    "PIERRE CARDIN", "LACOSTE", "DOCKERS", "SILBON",
    "NORTON", "NAUTICA", "SELECTED",
}

def _filtrar_marcas(df, col="marca"):
    """Filtra un DataFrame para quedarse solo con marcas vigentes."""
    if df.empty or col not in df.columns:
        return df
    return df[df[col].str.upper().isin(_MARCAS_VIGENTES)].reset_index(drop=True)


# ── Rendimiento por Tienda: lectura del Micro CRUDO ──
# El motor necesita las 6 columnas por tienda del Micro (Stk, Unidades, Vta S/.,
# On Order, UME, Precio Prom). La plantilla transformada ya las perdió, así que
# se relee el archivo original que quedó guardado al correr el análisis.
@st.cache_data(show_spinner=False)
def _rt_cargar_micro(_path, marca, semanas=4):
    """Ventana de N semanas apilando los Micros del mismo directorio.

    Majo pidio la rentabilidad por m2 sobre las ultimas 4 semanas, no sobre una.
    El Micro trae venta por tienda de una sola semana, asi que hay que apilar
    varios cortes. Si no hay suficientes, cae a la semana del archivo cargado y
    lo declara.
    """
    import glob as _glob
    hermanos = sorted(_glob.glob(os.path.join(os.path.dirname(_path), "Base al *.xlsx")))
    if len(hermanos) >= 2:
        try:
            largo = rend_t.acumular_micros(hermanos, marcas=marca, semanas=semanas)
            return largo, largo.attrs.get("semanas", 1), largo.attrs.get("cortes", [])
        except (ValueError, rend_t.FormatoMicroError):
            pass  # cortes con huecos o formato viejo: se usa solo el actual
    return rend_t.desde_micro(pd.read_excel(_path), marcas=marca), 1, []


# ── Rendimiento de Marca: base transaccional (color y talla) ──
# El Micro llega hasta el estilo. Para "qué colores se venden" hace falta la
# base de ventas a nivel línea de ticket, que es la única con color y talla.
@st.cache_data(show_spinner=False)
def _rt_cargar_trans(_bytes, marca):
    import io as _io
    df = pd.read_excel(_io.BytesIO(_bytes))
    t = rend_t.cargar_transaccional(df)
    return t[t["marca"].map(rend_t._norm) == rend_t._norm(marca)] if marca else t


# ── Helper: agregar columnas de precio + fórmula Nuevo Margen a Excel ──
def _add_pricing_cols(df_in, df_ref, sheet_name, writer):
    """
    Agrega precio_blanco, precio_vigente, costo al df, escribe a Excel,
    y luego agrega columnas 'Nuevo Precio' (vacía) y 'Nuevo Margen' (fórmula Excel).

    df_in: DataFrame a escribir
    df_ref: DataFrame de referencia para obtener precios (df_cob típicamente)
    sheet_name: nombre de la hoja Excel
    writer: pd.ExcelWriter activo
    """
    df_out = df_in.copy()
    # Encabezado claro en el Excel: desde el fix B6/F3 (2026-08-23),
    # prom_vta_uds es la venta sem1 por tienda SUAVIZADA por el factor
    # cadena de 4 semanas — misma base que el estado y la cobertura.
    if 'prom_vta_uds' in df_out.columns:
        df_out = df_out.rename(columns={'prom_vta_uds': 'vta_sem_prom'})
    # Agregar columnas de precio si no las tiene
    _precio_cols = ['precio_blanco', 'precio_vigente', 'costo']
    for _pc in _precio_cols:
        if _pc not in df_out.columns and _pc in df_ref.columns and 'sku' in df_out.columns:
            _map = df_ref.drop_duplicates('sku').set_index('sku')[_pc].to_dict()
            df_out[_pc] = df_out['sku'].map(_map)

    # Agregar columna vacía para Nuevo Precio
    df_out['nuevo_precio'] = np.nan
    # Placeholders para las fórmulas (se sobreescriben después)
    df_out['nuevo_margen'] = np.nan
    df_out['nuevo_dscto'] = np.nan

    df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    # Escribir fórmulas Excel en "nuevo_margen" y "nuevo_dscto"
    ws = writer.sheets[sheet_name]
    # Encontrar índices de columnas
    headers = [cell.value for cell in ws[1]]
    _col_np = headers.index('nuevo_precio') + 1  # 1-based
    _col_nm = headers.index('nuevo_margen') + 1
    _col_nd = headers.index('nuevo_dscto') + 1
    _col_costo = headers.index('costo') + 1 if 'costo' in headers else None
    _col_pb = headers.index('precio_blanco') + 1 if 'precio_blanco' in headers else None

    if _col_costo:
        _ltr_np = get_column_letter(_col_np)
        _ltr_costo = get_column_letter(_col_costo)
        _ltr_nm = get_column_letter(_col_nm)
        for row in range(2, ws.max_row + 1):
            # Nuevo Margen = (NuevoPrecio/1.18 - Costo) / (NuevoPrecio/1.18)
            ws[f'{_ltr_nm}{row}'] = f'=IF({_ltr_np}{row}="","",({_ltr_np}{row}/1.18-{_ltr_costo}{row})/({_ltr_np}{row}/1.18))'
        # Formatear como porcentaje
        for row in range(2, ws.max_row + 1):
            ws[f'{_ltr_nm}{row}'].number_format = '0.0%'
        # Formato moneda para precio cols
        for _pc_name in ['precio_blanco', 'precio_vigente', 'costo', 'nuevo_precio']:
            if _pc_name in headers:
                _ltr = get_column_letter(headers.index(_pc_name) + 1)
                for row in range(2, ws.max_row + 1):
                    ws[f'{_ltr}{row}'].number_format = '#,##0.00'

    if _col_pb:
        # Nuevo Dscto = 1 - NuevoPrecio / PrecioBlanco (pedido Franco 2026-08-05)
        _ltr_np = get_column_letter(_col_np)
        _ltr_pb = get_column_letter(_col_pb)
        _ltr_nd = get_column_letter(_col_nd)
        for row in range(2, ws.max_row + 1):
            ws[f'{_ltr_nd}{row}'] = f'=IF(OR({_ltr_np}{row}="",{_ltr_pb}{row}=""),"",1-{_ltr_np}{row}/{_ltr_pb}{row})'
            ws[f'{_ltr_nd}{row}'].number_format = '0.0%'

    # Autofiltro + panel congelado de fábrica (auditoría formato 2026-08-05)
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    # Renombrar headers a español
    _rename = {'precio_blanco': 'Precio Blanco', 'precio_vigente': 'Precio Vigente',
               'costo': 'Costo', 'nuevo_precio': 'Nuevo Precio', 'nuevo_margen': 'Nuevo Margen',
               'nuevo_dscto': 'Nuevo Dscto'}
    for cell in ws[1]:
        if cell.value in _rename:
            cell.value = _rename[cell.value]

    # Estilo visual del mockup: header gris, chips de estado, input ámbar,
    # fórmulas en verde (auditoría formato 2026-08-05)
    vistas_excel.estilizar_hoja_pricing(ws)

    return df_out


df_cob   = _filtrar_marcas(df_cob)
df_rep   = _filtrar_marcas(df_rep)
df_trans = _filtrar_marcas(df_trans)
df_prec  = _filtrar_marcas(df_prec)
df_alertas   = _filtrar_marcas(df_alertas)
df_anomalias = _filtrar_marcas(df_anomalias, col="marca" if "marca" in df_anomalias.columns else "___skip")

# Filtrar pivot de reposiciones (solo SKUs que sobrevivieron el filtro)
if not df_rep_pivot.empty and "marca" in df_rep_pivot.columns:
    df_rep_pivot = _filtrar_marcas(df_rep_pivot)
elif not df_rep_pivot.empty:
    _skus_rep_vivos = set(df_rep["sku"].unique()) if not df_rep.empty else set()
    df_rep_pivot = df_rep_pivot[df_rep_pivot["sku"].isin(_skus_rep_vivos)].reset_index(drop=True)

# Filtrar alertas por tienda: reconstruir solo con ítems de marcas vigentes
if alertas_tienda_dict:
    _at_filtrado = {}
    for _t_name, _t_payload in alertas_tienda_dict.items():
        _items = _t_payload.get('items', [])
        _items_ok = [it for it in _items
                     if str(it.get('marca', '')).upper() in _MARCAS_VIGENTES]
        if _items_ok:
            _new_payload = dict(_t_payload)
            _new_payload['items'] = _items_ok
            # Recalcular resumen
            _new_payload['resumen'] = {
                'n_items': len(_items_ok),
                'capital_parado_sol': sum(it.get('capital_parado', 0) for it in _items_ok),
                'n_con_descuento': sum(1 for it in _items_ok if it.get('pct_descuento', 0) > 0),
            }
            _at_filtrado[_t_name] = _new_payload
    alertas_tienda_dict = _at_filtrado

# Filtrar alertas venta cero: solo marcas vigentes
if alertas_venta_cero_dict:
    _vc_filtrado = {}
    for _t_name, _t_payload in alertas_venta_cero_dict.items():
        _pm = _t_payload.get('por_marca', {})
        _pm_ok = {m: v for m, v in _pm.items() if m.upper() in _MARCAS_VIGENTES}
        if _pm_ok:
            n_skus = sum(v['n_skus'] for v in _pm_ok.values())
            capital = sum(v['capital'] for v in _pm_ok.values())
            _vc_filtrado[_t_name] = {
                'tienda': _t_payload['tienda'],
                'fecha_corte': _t_payload.get('fecha_corte', ''),
                'resumen': {'n_skus': n_skus, 'n_marcas': len(_pm_ok), 'capital_parado_total': capital},
                'por_marca': _pm_ok,
            }
    alertas_venta_cero_dict = _vc_filtrado

# Recalcular summary con datos filtrados
s["total_combos"]      = len(df_cob)
s["n_critico"]         = int((df_cob["estado"] == "QUIEBRE").sum()) if not df_cob.empty else 0
s["n_precritico"]      = int((df_cob["estado"] == "PRE-QUIEBRE").sum()) if not df_cob.empty else 0
s["n_optimo"]          = int((df_cob["estado"] == "ÓPTIMO").sum()) if not df_cob.empty else 0
s["n_alto"]            = int((df_cob["estado"] == "ALTO").sum()) if not df_cob.empty else 0
s["n_sobrestock"]      = int((df_cob["estado"] == "SOBRESTOCK").sum()) if not df_cob.empty else 0
s["n_liquidar"]        = int((df_cob["estado"] == "LIQUIDAR").sum()) if not df_cob.empty else 0
s["n_nuevo_sv"]        = int((df_cob["estado"] == "NUEVO SIN VENTA").sum()) if not df_cob.empty else 0
s["n_dormido"]         = int((df_cob["estado"] == "DORMIDO").sum()) if not df_cob.empty else 0
s["n_muerto"]          = int((df_cob["estado"] == "MUERTO").sum()) if not df_cob.empty else 0
s["n_estancado"]       = int((df_cob["estado"] == "ESTANCADO").sum()) if not df_cob.empty else 0
s["uds_reponer"]       = int(df_rep["a_reponer"].sum()) if not df_rep.empty else 0
s["uds_desde_cd"]      = int(df_rep["desde_cd"].sum()) if (not df_rep.empty and "desde_cd" in df_rep.columns) else 0
s["uds_pendiente_cd"]  = int(df_rep["pendiente"].sum()) if (not df_rep.empty and "pendiente" in df_rep.columns) else 0
s["uds_transferir"]    = int(df_trans["uds_transferir"].sum()) if not df_trans.empty else 0
s["n_acciones_precio"] = len(df_prec)

# Nota: marcas propias sin stock CD ya están excluidas por motor_v2.build_reposiciones()
# Marcas propias: MARQUIS, NAVIGATA, CACHAREL, SPAVALDI, OSCAR DE LA RENTA, US POLO


# ══════════════════════════════════════════════════════════════
#  DASHBOARD VISUAL  (solo se renderiza en vista Dashboard)
# ══════════════════════════════════════════════════════════════

# ── Configuración global de Plotly (usada en múltiples vistas) ──
_plotly_layout = dict(
    font=dict(family="Inter, -apple-system, sans-serif", size=12, color=TH_TEXT_PY),
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    margin=dict(l=20, r=20, t=40, b=20),
    hoverlabel=dict(bgcolor="white", font_size=12, font_family="Inter"),
)

_estado_color_map = {
    "QUIEBRE":      STATUS_QUIEBRE,
    "PRE-QUIEBRE":  STATUS_BAJA,
    "ÓPTIMO":       STATUS_OPTIMO,
    "ALTO":         STATUS_ALTO,
    "SOBRESTOCK":   STATUS_SOBRESTOCK,
    "ESTANCADO":    STATUS_ESTANCADO,
    "LIQUIDAR":     STATUS_LIQUIDAR,
    "NUEVO SIN VENTA":  STATUS_LANZAMIENTO,
    "DORMIDO":      STATUS_DORMIDO,
    "MUERTO":       STATUS_MUERTO,
}


def _build_excel_terceras(_dfc, _dfr, _dft):
    """Excel-paquete de Marcas Terceras: una pestaña por componente de la
    sección (Capital Parado, SKUs Críticos, Quiebres, Reposición,
    Transferencias, Precios). Sin cache: se genera fresco al descargar."""
    _M = agente_terceras.MARCAS_AGENTE
    _buf = io.BytesIO()
    with pd.ExcelWriter(_buf, engine="openpyxl") as _w:
        # 1. SKUs críticos por marca×línea (con criticidad)
        _crit = agente_terceras.top5_por_marca_linea(_dfc, top_n=5)
        (_crit if not _crit.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="SKUs Criticos", index=False)
        # 2. Reposición terceras (con margen + edad cruzados)
        _rep = _dfr[_dfr['marca'].str.upper().str.strip().isin(_M)].copy() if not _dfr.empty and 'marca' in _dfr.columns else pd.DataFrame()
        if not _rep.empty:
            _rep = _rep[_rep['a_reponer'] > 0] if 'a_reponer' in _rep.columns else _rep
            if 'sku' in _dfc.columns and 'margen_efectivo' in _dfc.columns:
                _mg = _dfc.drop_duplicates('sku').set_index('sku')['margen_efectivo']
                _rep['margen_efectivo_pct'] = (_rep['sku'].map(_mg).fillna(0) * 100).round(1)
            if 'sku' in _dfc.columns and 'edad_semanas' in _dfc.columns:
                _rep['edad_sem'] = _rep['sku'].map(_dfc.groupby('sku')['edad_semanas'].max())
        (_rep if not _rep.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Reposicion", index=False)
        # 3. Transferencias terceras (cruce por sku)
        if not _dft.empty and 'sku' in _dft.columns and 'marca' in _dfc.columns:
            _s2m = dict(zip(_dfc['sku'], _dfc['marca'].str.upper().str.strip()))
            _tr = _dft.copy()
            _tr['marca'] = _tr['sku'].map(_s2m)
            _tr = _tr[_tr['marca'].isin(_M)]
        else:
            _tr = pd.DataFrame()
        (_tr if not _tr.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Transferencias", index=False)
        # 4. Gestión de precios (pirámide)
        _pr = agente_terceras.sugerencias_precio_terceras(_dfc)
        (_pr if not _pr.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Precios", index=False)
    _buf.seek(0)
    return _buf.read()


def _build_excel_propias(_dfc, _dfr, _dft, _gaps, _ret):
    """Excel-paquete de Marcas Propias: Reposición · Transferencias · Precios ·
    Predistribución (gaps limpios + retenidos CD). Filtrado a las 7 propias."""
    _P = agente_terceras.MARCAS_PROPIAS_SET
    _buf = io.BytesIO()
    with pd.ExcelWriter(_buf, engine="openpyxl") as _w:
        # 1. Reposición propias (con margen + edad)
        _rep = _dfr[_dfr['marca'].str.upper().str.strip().isin(_P)].copy() if not _dfr.empty and 'marca' in _dfr.columns else pd.DataFrame()
        if not _rep.empty:
            _rep = _rep[_rep['a_reponer'] > 0] if 'a_reponer' in _rep.columns else _rep
            if 'sku' in _dfc.columns and 'margen_efectivo' in _dfc.columns:
                _mg = _dfc.drop_duplicates('sku').set_index('sku')['margen_efectivo']
                _rep['margen_efectivo_pct'] = (_rep['sku'].map(_mg).fillna(0) * 100).round(1)
            if 'sku' in _dfc.columns and 'edad_semanas' in _dfc.columns:
                _rep['edad_sem'] = _rep['sku'].map(_dfc.groupby('sku')['edad_semanas'].max())
        (_rep if not _rep.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Reposicion", index=False)
        # 2. Transferencias propias (cruce por sku)
        if not _dft.empty and 'sku' in _dft.columns and 'marca' in _dfc.columns:
            _s2m = dict(zip(_dfc['sku'], _dfc['marca'].str.upper().str.strip()))
            _tr = _dft.copy(); _tr['marca'] = _tr['sku'].map(_s2m); _tr = _tr[_tr['marca'].isin(_P)]
        else:
            _tr = pd.DataFrame()
        (_tr if not _tr.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Transferencias", index=False)
        # 3. Precios propias (pirámide)
        _pr = agente_terceras.sugerencias_precio_terceras(_dfc, marcas=_P)
        (_pr if not _pr.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Precios", index=False)
        # 4. Predistribución — gaps limpios (stock CD>0, edad<=8) + retenidos CD
        _g = _gaps[_gaps['marca'].str.upper().str.strip().isin(_P)].copy() if not _gaps.empty and 'marca' in _gaps.columns else pd.DataFrame()
        if not _g.empty:
            if 'stock_cd' in _g.columns:
                _g = _g[_g['stock_cd'] > 0]
            if 'edad_semanas' in _g.columns:
                _g = _g[_g['edad_semanas'].fillna(999) <= 8]
        (_g if not _g.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Predist Gaps", index=False)
        _r = _ret[_ret['marca'].str.upper().str.strip().isin(_P)].copy() if not _ret.empty and 'marca' in _ret.columns else pd.DataFrame()
        (_r if not _r.empty else pd.DataFrame({"sin datos": []})).to_excel(_w, sheet_name="Retenidos CD", index=False)
    _buf.seek(0)
    return _buf.read()


def _capi_base_path():
    """Path de la Base Profundidad SUBIDA EN ESTA SESIÓN (para leer campos
    crudos que no pasan por el ETL). Fix B8 auditoría 2026-08-23: se eliminó
    el fallback a data2/bases antiguas/ — mezclaba silenciosamente datos de
    otra base con el análisis actual (caso John Holden S/1.01M vs S/1.26M)."""
    _p = st.session_state.get("_base_profundidad_path")
    if _p and os.path.exists(_p):
        return _p
    return None


@st.cache_data(show_spinner=False)
def _tipo_evento_map(_path):
    """Mapa SKU → tipo de evento de precio (MD1=etiquetado / PTR=cartel / MTR=sin evento),
    leído directo de la Base Profundidad. Cacheado por path (el archivo no cambia)."""
    try:
        _df = pd.read_excel(_path, usecols=['Cód. Prod.', 'Tipo de Evento Vigente'])
        return dict(zip(_df['Cód. Prod.'], _df['Tipo de Evento Vigente'].astype(str).str.strip().str.upper()))
    except Exception:
        return {}


if nav_page == "🏠 Dashboard":
    st.markdown(f'<div class="section-header"><h3>Dashboard</h3><span class="live-badge">LIVE</span></div>', unsafe_allow_html=True)

    # ── Filtros del dashboard ──
    st.markdown(f"""<div style="background:var(--capi-bg-surface); border:1px solid var(--capi-border); border-radius:12px; padding:12px 16px; margin-bottom:16px;">
    <span style="font-weight:600; color:var(--capi-text); font-size:0.9rem;">Filtros del Dashboard</span>
    </div>""", unsafe_allow_html=True)

    has_temporada = "temporada" in df_cob.columns
    has_rango = "rango_antiguedad" in df_cob.columns
    fcol1, fcol2 = st.columns(2)

    with fcol1:
        if has_temporada:
            temps_disponibles = sorted([t for t in df_cob["temporada"].unique() if t and str(t).strip()])
            temps_opciones = ["Todas"] + temps_disponibles
            f_dash_temp = st.selectbox("Temporada", temps_opciones, key="dash_temporada")
        else:
            f_dash_temp = "Todas"
            st.caption("Sin columna Temporada en los datos")

    with fcol2:
        if has_rango:
            _rango_orden = {"RANGO 0": 0, "RANGO 0_3": 1, "RANGO 3_6": 2, "RANGO 6_9": 3,
                            "RANGO 9_12": 4, "RANGO 12_99": 5, "Sin Rango": 6}
            rangos_disponibles = sorted(
                [r for r in df_cob["rango_antiguedad"].unique() if r and str(r).strip()],
                key=lambda x: _rango_orden.get(x, 99)
            )
            rangos_opciones = ["Todos"] + rangos_disponibles
            f_dash_rango = st.selectbox("Rango de Antigüedad", rangos_opciones, key="dash_rango",
                                         help="Filtra por rango de antigüedad del reporte micro")
        else:
            f_dash_rango = "Todos"
            st.caption("Sin columna Rango Antigüedad en los datos")

    # Aplicar filtros al df_cob para el dashboard
    df_dash = df_cob.copy()
    if f_dash_temp != "Todas" and has_temporada:
        df_dash = df_dash[df_dash["temporada"] == f_dash_temp]
    if f_dash_rango != "Todos" and has_rango:
        df_dash = df_dash[df_dash["rango_antiguedad"] == f_dash_rango]

    st.caption(f"Mostrando {len(df_dash):,} de {len(df_cob):,} combos SKU×Tienda después de filtros")

    # ── Fila 1: Donut + Leyenda de estados ──
    _marcas_donut = ["Todas"] + sorted(df_dash["marca"].unique().tolist()) if "marca" in df_dash.columns else ["Todas"]
    _marca_donut_sel = st.selectbox("Filtrar por Marca", _marcas_donut, index=0, key="donut_marca_filter")
    _df_donut = df_dash if _marca_donut_sel == "Todas" else df_dash[df_dash["marca"] == _marca_donut_sel]

    dash_c1, dash_c2 = st.columns([1, 1])

    with dash_c1:
        estado_counts = _df_donut["estado"].value_counts().reset_index()
        estado_counts.columns = ["Estado", "Cantidad"]
        estado_order = [
            "QUIEBRE", "PRE-QUIEBRE", "ÓPTIMO", "ALTO", "SOBRESTOCK",
            "ESTANCADO", "LIQUIDAR", "NUEVO SIN VENTA", "DORMIDO", "MUERTO",
        ]
        estado_counts["Estado"] = pd.Categorical(estado_counts["Estado"], categories=estado_order, ordered=True)
        estado_counts = estado_counts.sort_values("Estado").dropna(subset=["Estado"])

        # Display labels: hoy todos los estados ya son su propio label (no se renombran).
        estado_counts["Label"] = estado_counts["Estado"].astype(str)

        # Capital por estado para hover
        _capital_por_estado = _df_donut.groupby("estado")["stock_valor_costo"].sum()
        estado_counts["Capital"] = estado_counts["Estado"].astype(str).map(_capital_por_estado).fillna(0)

        _donut_title = f"Distribución por Estado — {_marca_donut_sel}" if _marca_donut_sel != "Todas" else "Distribución por Estado"
        fig_donut = go.Figure(data=[go.Pie(
            labels=estado_counts["Label"],
            values=estado_counts["Cantidad"],
            hole=0.55,
            marker=dict(colors=[_estado_color_map.get(e, "#CBD5E1") for e in estado_counts["Estado"]]),
            textinfo="label+percent",
            textfont=dict(size=11),
            customdata=estado_counts[["Capital"]].values,
            hovertemplate="<b>%{label}</b><br>%{value:,} combos<br>%{percent}<br>Capital: S/ %{customdata[0]:,.0f}<extra></extra>",
        )])
        fig_donut.update_layout(
            **_plotly_layout,
            title=dict(text=_donut_title, font=dict(size=14, color=TH_TEXT_PY)),
            showlegend=False,
            height=380,
        )
        total_donut = len(_df_donut)
        fig_donut.add_annotation(
            text=f"<b>{total_donut:,}</b><br><span style='font-size:10px;color:var(--capi-text2)'>SKU×Tienda</span>",
            showarrow=False, font=dict(size=18, color=TH_TEXT_PY),
        )
        st.plotly_chart(fig_donut, use_container_width=True)

    with dash_c2:
        # ── Leyenda: criterio de cada estado (10 estados de la taxonomía v2) ──
        _estado_criterios = {
            "QUIEBRE":     {"color": STATUS_QUIEBRE,     "icon": "🔴", "regla": "Cobertura < 4 semanas"},
            "PRE-QUIEBRE": {"color": STATUS_BAJA,        "icon": "🟠", "regla": "Cobertura 4–8 semanas"},
            "ÓPTIMO":      {"color": STATUS_OPTIMO,     "icon": "🟢", "regla": "Cobertura 8–16 semanas"},
            "ALTO":        {"color": STATUS_ALTO,       "icon": "🟡", "regla": "Cobertura 16–26 semanas"},
            "SOBRESTOCK":  {"color": STATUS_SOBRESTOCK, "icon": "🟤", "regla": "Cobertura 26–52 semanas"},
            "ESTANCADO":   {"color": STATUS_ESTANCADO,  "icon": "⬛", "regla": "Cob > 52 sem · edad ≤ 6 meses"},
            "LIQUIDAR":    {"color": STATUS_LIQUIDAR,   "icon": "💀", "regla": "Cob > 52 sem · edad > 6 meses"},
            "NUEVO SIN VENTA": {"color": STATUS_LANZAMIENTO,"icon": "🆕", "regla": "Sin venta · edad < 2 meses"},
            "DORMIDO":     {"color": STATUS_DORMIDO,    "icon": "😴", "regla": "Sin venta · edad 2–6 meses"},
            "MUERTO":      {"color": STATUS_MUERTO,     "icon": "⚫", "regla": "Sin venta · edad > 6 meses"},
        }

        st.markdown(f"""<div style="font-weight:600; color:var(--capi-text); font-size:0.95rem; margin-bottom:10px;">
        Criterios de Clasificación
        </div>""", unsafe_allow_html=True)

        for _est_name, _est_info in _estado_criterios.items():
            # Contar combos en este estado
            _n_est = int((_df_donut["estado"] == _est_name).sum()) if not _df_donut.empty else 0
            _cap_est = _df_donut.loc[_df_donut["estado"] == _est_name, "stock_valor_costo"].sum() if _n_est > 0 else 0
            _pct_est = (_n_est / total_donut * 100) if total_donut > 0 else 0

            st.markdown(f"""<div style="display:flex; align-items:center; gap:10px; padding:5px 10px; margin-bottom:3px;
                border-left:3px solid {_est_info['color']}; border-radius:0 6px 6px 0;
                background:{'rgba(0,0,0,0.02)' if _n_est > 0 else 'transparent'};">
                <span style="font-size:0.85rem; min-width:18px;">{_est_info['icon']}</span>
                <div style="flex:1;">
                    <span style="font-weight:600; color:var(--capi-text); font-size:0.82rem;">{_est_info.get('label', _est_name)}</span>
                    <span style="color:var(--capi-text2); font-size:0.78rem;"> — {_est_info['regla']}</span>
                </div>
                <div style="text-align:right; min-width:80px;">
                    <span style="font-weight:700; color:{_est_info['color']}; font-size:0.85rem;">{_n_est:,}</span>
                    <span style="color:var(--capi-text2); font-size:0.72rem;"> ({_pct_est:.0f}%)</span>
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown(f"""<div style="margin-top:10px; padding:8px 12px; background:var(--capi-bg-surface); border-radius:8px; font-size:0.78rem; color:var(--capi-text2);">
        Capital total: <strong style="color:var(--capi-text);">S/ {_df_donut['stock_valor_costo'].sum():,.0f}</strong> &nbsp;·&nbsp;
        {total_donut:,} combos SKU×Tienda
        </div>""", unsafe_allow_html=True)

    # ── 🧠 Análisis de estados desplegable (pedido Franco 2026-08-24) ──
    if _HAS_SNAPSHOTS:
        with st.expander("🧠 Analizar movimientos de estados (vs corte anterior)", expanded=False):
            try:
                from snapshots_engine.storage import list_available_weeks as _law_dash
                _an_dw = _law_dash()
            except Exception:
                _an_dw = []
            if len(_an_dw) < 2:
                st.info("Se necesitan al menos 2 snapshots para comparar cortes.")
            else:
                _an_da, _an_db = _an_dw[-2], _an_dw[-1]
                st.caption(f"Comparando {analisis_estados.etiqueta_semana(_an_da)} → "
                           f"{analisis_estados.etiqueta_semana(_an_db)} (los dos últimos cortes). "
                           f"Análisis completo y migraciones: vista 🏆 Caso de Éxito.")
                for _an_dc in analisis_estados.conclusiones(_an_da, _an_db, acciones_log.cargar()):
                    _an_dr = {"positivo": st.success, "atencion": st.warning,
                              "critico": st.error}.get(_an_dc["nivel"], st.info)
                    _an_dr(f"**{_an_dc['titulo']}** — {_an_dc['detalle']}")

    # ── Desglose por marca del estado seleccionado + descarga ──
    st.markdown("---")

    # Display labels: hoy cada estado se muestra con su propio nombre.
    _est_opciones = [e for e in estado_order if e in _df_donut["estado"].values]
    _est_default = _est_opciones.index("QUIEBRE") if "QUIEBRE" in _est_opciones else 0

    _det_c1, _det_c2 = st.columns([1, 2])
    with _det_c1:
        _est_sel = st.selectbox("Ver detalle por estado", _est_opciones, index=_est_default, key="donut_estado_sel")

    _df_est = _df_donut[_df_donut["estado"] == _est_sel]

    if not _df_est.empty:
        # Resumen por marca
        if "marca" in _df_est.columns:
            _est_marca = _df_est.groupby("marca").agg(
                combos=("sku", "count"),
                skus=("sku", "nunique"),
                capital=("stock_valor_costo", "sum"),
                vta_sem=("prom_vta_uds", "sum"),
            ).sort_values("capital", ascending=False).reset_index()
            _est_marca["Participación"] = (_est_marca["combos"] / _est_marca["combos"].sum() * 100).round(1)
            _est_marca = _est_marca.rename(columns={
                "marca": "Marca", "combos": "Combos", "skus": "SKUs",
                "capital": "Capital S/", "vta_sem": "Vta Sem (uds)",
            })

            if _est_sel == "SOBRESTOCK" and "sobrestock_aparente" in _df_est.columns:
                _apar = _df_est[_df_est["sobrestock_aparente"] == True]
                if len(_apar):
                    st.caption(f"👀 De este sobrestock, S/ {_apar['stock_valor_costo'].sum():,.0f} "
                               f"({len(_apar)} combos) es APARENTE: >60% del stock retenido en CD — "
                               "es problema de distribución, no de compra. Empujar antes de tocar precio.")
            st.markdown(f"""<div style="background:{'#FEF2F2' if _est_sel in ('QUIEBRE','PRE-QUIEBRE') else TH_BG_SURFACE_PY};
            border-left:4px solid {_estado_color_map.get(_est_sel, SLATE_500)};
            padding:10px 14px; border-radius:10px; margin-bottom:10px;">
            <strong style="color:{_estado_color_map.get(_est_sel, TH_TEXT_PY)};">{_est_sel}</strong>
            <span style="color:var(--capi-text2); font-size:0.82em;"> &nbsp;·&nbsp; {len(_df_est):,} combos · {_df_est['sku'].nunique()} SKUs · S/ {_df_est['stock_valor_costo'].sum():,.0f}</span>
            </div>""", unsafe_allow_html=True)

            st.dataframe(
                _est_marca.style.format({
                    "Capital S/": "S/ {:,.0f}", "Participación": "{:.1f}%",
                    "Vta Sem (uds)": "{:,.0f}",
                }),
                use_container_width=True, hide_index=True, height=250,
            )

        # Detalle de SKUs expandible
        with st.expander(f"Ver {min(50, len(_df_est)):,} combos en estado {_est_sel}", expanded=False):
            _det_cols = ["sku", "nombre", "marca", "tienda", "stock_total", "prom_vta_uds",
                         "cobertura_sem", "stock_valor_costo", "edad_semanas"]
            if "pct_descuento" in _df_est.columns:
                _det_cols.append("pct_descuento")
            _det_cols = [c for c in _det_cols if c in _df_est.columns]
            _df_est_disp = _df_est[_det_cols].sort_values("stock_valor_costo", ascending=False).head(50)
            _df_est_disp = _df_est_disp.rename(columns={
                "sku": "SKU", "nombre": "Nombre", "marca": "Marca", "tienda": "Tienda",
                "stock_total": "Stock", "prom_vta_uds": "Vta Sem (uds)",
                "cobertura_sem": "Cob (sem)", "stock_valor_costo": "Capital S/",
                "edad_semanas": "Edad (sem)", "pct_descuento": "Dscto",
            })
            _det_fmt = {"Capital S/": "S/ {:,.0f}", "Cob (sem)": "{:.1f}", "Vta Sem (uds)": "{:.0f}"}
            if "Dscto" in _df_est_disp.columns:
                _det_fmt["Dscto"] = "{:.0%}"
            st.dataframe(
                _df_est_disp.style.format(_det_fmt, na_rep="—"),
                use_container_width=True, hide_index=True, height=350,
            )

        # Botón de descarga de TODOS los estados (para que el usuario filtre)
        _dl_cols = ["sku", "nombre", "marca", "temporada", "rango_antiguedad", "tienda",
                    "stock_total", "prom_vta_uds", "cobertura_sem", "stock_valor_costo",
                    "edad_semanas", "estado"]
        if "pct_descuento" in _df_donut.columns:
            _dl_cols.append("pct_descuento")
        _dl_cols = [c for c in _dl_cols if c in _df_donut.columns]
        _dl_buf = io.BytesIO()
        with pd.ExcelWriter(_dl_buf, engine="openpyxl") as _w_dl:
            # Vistas agregadas primero (auditoría formato 2026-08-05):
            # Resumen → Liquidación x SKU → Terceras por marca → Cascada → Temporadas
            vistas_excel.hoja_resumen_ejecutivo(_w_dl, _df_donut)
            vistas_excel.hoja_liquidacion_sku(_w_dl, _df_donut, _add_pricing_cols)
            vistas_excel.hojas_terceras_por_marca(_w_dl, _df_donut)
            vistas_excel.hoja_cascada_dscto(_w_dl, _df_donut)
            vistas_excel.hoja_estados_temporada(_w_dl, _df_donut)
            # Data cruda al final (mismo contenido de siempre)
            _add_pricing_cols(
                _df_donut[_dl_cols].sort_values(["estado", "stock_valor_costo"], ascending=[True, False]),
                df_cob, "Todos los estados", _w_dl
            )
        _dl_buf.seek(0)
        st.download_button(
            f"📥 Descargar {len(_df_donut):,} combos SKU×Tienda — Todos los estados (.xlsx)",
            data=_dl_buf.getvalue(),
            file_name="Capi_todos_estados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_all_estados",
        )

    # ══════════════════════════════════════════════════════════
    #  💸 VENTAS PERDIDAS POR QUIEBRE (lost sales)
    #  Banda histórica (snapshots, nivel SKU) + quiebre actual
    #  (df_cob, SKU×tienda). Feature estilo RELEX/Blue Yonder.
    # ══════════════════════════════════════════════════════════
    if _HAS_SNAPSHOTS:
        @st.cache_data(show_spinner=False)
        def _vp_calcular(_clave: str):
            return snapshots_engine.estimate_lost_sales(marcas=_MARCAS_VIGENTES)

        @st.cache_data(show_spinner=False)
        def _vp_procedencia_map(_path: str):
            """Mapa SKU → procedencia (NAC/IMP) leído de la columna 'Proced.' de la Base Profundidad.
            Define si un quiebre sin stock en cadena tiene reorder viable (NAC, semanas)
            o es estructural (IMP, ~2-3 meses — no llega para esta venta)."""
            try:
                _dfp = pd.read_excel(_path, usecols=['Cód. Prod.', 'Proced.'])
                return dict(zip(_dfp['Cód. Prod.'],
                                _dfp['Proced.'].astype(str).str.strip().str.upper()))
            except Exception:
                return {}

        def _vp_base_path():
            # Fix B8 (2026-08-23): solo la base de ESTA sesión, sin fallback
            # silencioso a archivos del repo.
            _p = st.session_state.get("_base_profundidad_path")
            if _p and os.path.exists(_p):
                return _p
            return None

        @st.cache_data(show_spinner=False)
        def _vp_cd_reliability():
            return snapshots_engine.estimate_cd_reliability()

        try:
            _vp = _vp_calcular("marcas-vigentes")
        except Exception:
            _vp = None

        if _vp and _vp.get('banda_max', 0) > 0:
            # Componente B: quiebre ACTUAL por SKU×tienda (base cargada)
            _vp_q = pd.DataFrame()
            if 'estado' in df_cob.columns and 'prom_vta_uds' in df_cob.columns and 'precio_vigente' in df_cob.columns:
                _vp_q = df_cob[(df_cob['estado'] == 'QUIEBRE') & (df_cob['prom_vta_uds'].fillna(0) > 0)].copy()
                if not _vp_q.empty:
                    _vp_q['perdida_sem_soles'] = (_vp_q['prom_vta_uds'] * _vp_q['precio_vigente'].fillna(0)).round(0)
                    # Contribución en riesgo: lo que de verdad pierde el P&L (feedback Franco:
                    # "si meto más descuento vendo más, pero no necesariamente mejor contribución")
                    _vp_q['contrib_riesgo_sem'] = (
                        (_vp_q['perdida_sem_soles'] * _vp_q['margen_efectivo'].fillna(0)).round(0)
                        if 'margen_efectivo' in _vp_q.columns else _vp_q['perdida_sem_soles']
                    )
                    _vp_q['evitable'] = _vp_q['stock_cd'].fillna(0) > 0 if 'stock_cd' in _vp_q.columns else False

                    # ── Solución por quiebre: cruzar con el plan del motor ──
                    # 1) Plan de reposición (despacho CD u orden a proveedor)
                    if not df_rep.empty and 'a_reponer' in df_rep.columns and 'tienda' in df_rep.columns:
                        _vp_rep = df_rep[df_rep['a_reponer'] > 0][
                            ['sku', 'tienda', 'a_reponer'] +
                            (['requiere_proveedor'] if 'requiere_proveedor' in df_rep.columns else [])
                        ].drop_duplicates(['sku', 'tienda'])
                        _vp_q = _vp_q.merge(_vp_rep, on=['sku', 'tienda'], how='left')
                    # 1b) ATP del CD: el reporte no es tiempo real → solo cd_prometible_pct%
                    #     del CD reportado es prometible; deriva/volatilidad desde snapshots
                    try:
                        _vp_cdrel = _vp_cd_reliability()
                    except Exception:
                        _vp_cdrel = pd.DataFrame()
                    if not _vp_cdrel.empty:
                        _vp_q = _vp_q.merge(_vp_cdrel, on='sku', how='left')
                    if 'cd_deriva_sem' not in _vp_q.columns:
                        _vp_q['cd_deriva_sem'] = 0.0
                        _vp_q['cd_volatil'] = False
                    # clip(0): el reporte trae stock CD negativo en algunos SKUs (ajustes/tránsito)
                    _vp_q['cd_atp'] = (_vp_q['stock_cd'].fillna(0).clip(lower=0) * cd_prometible_pct / 100).astype(int) if 'stock_cd' in _vp_q.columns else 0
                    # Asignación del ATP por SKU: el CD es un pool ÚNICO compartido entre
                    # todas las tiendas quebradas del mismo SKU — se reparte por prioridad
                    # de venta en riesgo (no se promete el mismo stock dos veces)
                    _vp_q['uds_despacho'] = 0
                    if 'a_reponer' in _vp_q.columns:
                        # Prioridad por CONTRIBUCIÓN en riesgo, no por venta (decisión Franco 2026-06)
                        _vp_q = _vp_q.sort_values('contrib_riesgo_sem', ascending=False)
                        for _vp_sku_g, _vp_grp in _vp_q.groupby('sku', sort=False):
                            _vp_rem = int(_vp_grp['cd_atp'].iloc[0] or 0)
                            for _vp_gi, _vp_gr in _vp_grp.iterrows():
                                _vp_rp = _vp_gr.get('a_reponer')
                                if _vp_rp is None or pd.isna(_vp_rp) or _vp_rp <= 0 or _vp_gr.get('requiere_proveedor') is True:
                                    continue
                                _vp_asig = min(int(_vp_rp), max(_vp_rem, 0))
                                _vp_q.at[_vp_gi, 'uds_despacho'] = _vp_asig
                                _vp_rem -= _vp_asig
                    # 2) Transferencias sugeridas (desde tiendas con exceso)
                    if not df_trans.empty and 'tienda_destino' in df_trans.columns:
                        _vp_tr = df_trans.groupby(['sku', 'tienda_destino']).agg(
                            _vp_uds_tr=('uds_transferir', 'sum'),
                            _vp_origen=('tienda_origen', 'first'),
                            _vp_n_orig=('tienda_origen', 'nunique'),
                        ).reset_index().rename(columns={'tienda_destino': 'tienda'})
                        _vp_q = _vp_q.merge(_vp_tr, on=['sku', 'tienda'], how='left')

                    # Regla Franco 2026-06-12: categorías producibles en Perú (reorder
                    # nacional viable) vs solo-importado (costo nacional alto / mala calidad)
                    _VP_NACIONALIZABLES = {'POLOS M/C', 'CAMISAS M/L', 'CAMISAS M/C',
                                           'PANTALONES', 'JEANS', 'POLERONES'}
                    _VP_SOLO_IMPORTADO = {'CASACAS', 'CHOMPAS'}

                    def _vp_accion(r):
                        _vol = " ⚠️CD volátil" if r.get('cd_volatil') is True else ""
                        _rep = r.get('a_reponer')
                        if _rep is not None and not pd.isna(_rep) and _rep > 0:
                            if r.get('requiere_proveedor') is True:
                                return f"📋 Orden a proveedor — {int(_rep)} uds"
                            _asig = int(r.get('uds_despacho', 0) or 0)
                            if _asig <= 0:
                                return f"⏳ ATP del CD agotado por tiendas de mayor prioridad — transferir u ordenar (plan pedía {int(_rep)}){_vol}"
                            if _asig < int(_rep):
                                return f"🚚 Despachar del CD — {_asig} uds (ATP {cd_prometible_pct}% compartido; plan pedía {int(_rep)}){_vol}"
                            return f"🚚 Despachar del CD — {_asig} uds{_vol}"
                        _tr = r.get('_vp_uds_tr')
                        if _tr is not None and not pd.isna(_tr) and _tr > 0:
                            _org = str(r.get('_vp_origen', ''))
                            _extra = f" (+{int(r['_vp_n_orig']) - 1} tiendas)" if r.get('_vp_n_orig', 1) > 1 else ""
                            return f"🔄 Transferir {int(_tr)} uds desde {_org}{_extra}"
                        if r.get('evitable'):
                            return f"🚚 Stock en CD — revisar (excluido del plan por regla de dscto){_vol}"
                        _proc = str(r.get('procedencia', '') or '')
                        if _proc.startswith('IMP'):
                            _cat = str(r.get('categoria', '') or '').strip().upper()
                            if _cat in _VP_NACIONALIZABLES:
                                return "⛔→🇵🇪 Importado agotado (reorder no llega) — VIABLE producir nacional: esta categoría se hace en Perú"
                            if _cat in _VP_SOLO_IMPORTADO:
                                return "⛔ Estructural — solo importado (esta categoría no se hace nacional): sustituto o asumir pérdida"
                            return "⛔ Importado sin stock en cadena — reorder ~2-3 meses, no llega: evaluar sustituto o asumir pérdida"
                        if _proc.startswith('NAC'):
                            return "📋 Reorder nacional — gestionar con proveedor local (semanas)"
                        return "📋 Orden de compra (sin stock en cadena)"

                    _vp_base = _vp_base_path()
                    _vp_proc = _vp_procedencia_map(_vp_base) if _vp_base else {}
                    _vp_q['procedencia'] = _vp_q['sku'].map(_vp_proc).fillna('')

                    _vp_q['accion_sugerida'] = _vp_q.apply(_vp_accion, axis=1)
                    _vp_q = _vp_q.sort_values('contrib_riesgo_sem', ascending=False)

            st.markdown("---")
            st.markdown(f'<div class="section-header"><h3>💸 Ventas Perdidas por Quiebre</h3><span class="live-badge">NUEVO</span></div>', unsafe_allow_html=True)

            _vp_sem = _vp['semanas_analizadas']
            _vp_evit_n = int(_vp_q['evitable'].sum()) if not _vp_q.empty else 0
            _vp_evit_soles = float(_vp_q.loc[_vp_q['evitable'], 'perdida_sem_soles'].sum()) if not _vp_q.empty else 0.0

            # S2 (2026-09-05): el motor ya calcula ingreso NETO (descontada sustitución) y
            # MARGEN perdido; antes la pantalla solo mostraba el bruto viejo.
            _vp_recap = int(round(_vp.get('tasa_recaptura', 0.30) * 100))
            _vp_c1, _vp_c1b, _vp_c2, _vp_c3 = st.columns([1.4, 1.4, 1, 1])
            with _vp_c1:
                st.markdown(f"""
                <div style="background:#FEF2F2; border-radius:12px; padding:16px 20px; border-left:4px solid #DC2626;">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Venta perdida NETA — semanas {_vp_sem[0]} a {_vp_sem[-1]}</div>
                    <div style="font-size:1.6rem; font-weight:700; color:#DC2626;">S/ {_vp.get('ingreso_neto_min', 0):,.0f} – S/ {_vp.get('ingreso_neto_max', 0):,.0f}</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">Ingreso que no se recuperó con otro SKU (descontada sustitución {_vp_recap}%). Bruto: S/ {_vp['banda_min']:,.0f} – {_vp['banda_max']:,.0f}</div>
                </div>""", unsafe_allow_html=True)
            with _vp_c1b:
                st.markdown(f"""
                <div style="background:#FFF7ED; border-radius:12px; padding:16px 20px; border-left:4px solid #EA580C;">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Margen perdido NETO (lo que mueve el P&L)</div>
                    <div style="font-size:1.6rem; font-weight:700; color:#EA580C;">S/ {_vp.get('margen_neto_min', 0):,.0f} – S/ {_vp.get('margen_neto_max', 0):,.0f}</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">Ingreso neto × margen contable del SKU (contribución ÷ venta), sin IGV</div>
                </div>""", unsafe_allow_html=True)
            with _vp_c2:
                st.markdown(f"""
                <div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {STATUS_CRITICO};">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">SKUs con quiebre en el período</div>
                    <div style="font-size:1.6rem; font-weight:700; color:{STATUS_CRITICO};">{_vp['n_skus_afectados']:,}</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">con venta comprobada antes del quiebre</div>
                </div>""", unsafe_allow_html=True)
            with _vp_c3:
                st.markdown(f"""
                <div style="background:#FFFBEB; border-radius:12px; padding:16px 20px; border-left:4px solid #D97706;">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Quiebres evitables HOY</div>
                    <div style="font-size:1.6rem; font-weight:700; color:#D97706;">{_vp_evit_n:,}</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">combos con stock en CD · S/ {_vp_evit_soles:,.0f}/sem en juego</div>
                </div>""", unsafe_allow_html=True)

            # Fórmula visible: de dónde sale el número
            st.markdown(f"""
            <div style="background:var(--capi-bg-surface); border:1px solid var(--capi-border); border-radius:12px; padding:14px 18px; margin-top:12px;">
                <div style="font-size:0.78rem; font-weight:600; color:var(--capi-text); margin-bottom:6px;">📐 Cómo se calcula (por cada SKU en quiebre)</div>
                <div style="font-size:0.98rem; color:var(--capi-text); margin-bottom:8px;">
                    Ingreso bruto = <strong>velocidad semanal</strong> &times; <strong>semanas en quiebre</strong> &times; <strong>precio realizado</strong>
                    &nbsp;→&nbsp; Neto = bruto &times; (1 − {_vp_recap}% sustitución) &nbsp;→&nbsp; Margen perdido = neto &times; margen contable
                </div>
                <div style="font-size:0.72rem; color:var(--capi-text2); line-height:1.5;">
                    <strong>Velocidad semanal</strong> = venta real por semana del SKU (serie reconstruida de los snapshots; banda = promedio simple vs ponderado reciente) &nbsp;·&nbsp;
                    <strong>Semanas en quiebre</strong> = cierres con stock 0 (0.5 sem) + semanas entre cierres confirmadas sin venta (0.5–1.0) &nbsp;·&nbsp;
                    <strong>Precio realizado</strong> = venta S/ ÷ unidades del SKU, <strong>sin IGV</strong> (no el precio de lista) &nbsp;·&nbsp;
                    <strong>Margen contable</strong> = contribución ÷ venta del SKU (no precio − costo: el costo falta en ~48% de los SKUs).<br>
                    El <strong>neto</strong> es lo que de verdad no se vendió; el <strong>margen perdido</strong> es lo que dejó de entrar al P&L. Solo SKUs con venta comprobada antes del quiebre (DORMIDO/MUERTO no cuentan).
                </div>
            </div>""", unsafe_allow_html=True)

            st.caption(" · ".join(_vp['supuestos']))

            with st.expander(f"Ver detalle y soluciones — top quiebres actuales por tienda ({len(_vp_q):,} combos)", expanded=False):
                if _vp_q.empty:
                    st.info("No hay quiebres activos con venta en la base cargada.")
                else:
                    # Mix de soluciones (resumen accionable)
                    _vp_n_cd = int(_vp_q['accion_sugerida'].str.startswith('🚚').sum())
                    _vp_n_tr = int(_vp_q['accion_sugerida'].str.startswith('🔄').sum())
                    _vp_n_oc = int(_vp_q['accion_sugerida'].str.startswith('📋').sum())
                    _vp_n_ag = int(_vp_q['accion_sugerida'].str.startswith('⏳').sum())
                    _vp_n_est = int(_vp_q['accion_sugerida'].str.startswith('⛔').sum())
                    st.markdown(
                        f"**Cómo se resuelven:** 🚚 **{_vp_n_cd:,}** despachos desde CD · "
                        f"🔄 **{_vp_n_tr:,}** transferencias entre tiendas · "
                        f"📋 **{_vp_n_oc:,}** órdenes de compra/proveedor"
                        + (f" · ⏳ **{_vp_n_ag:,}** en cola (ATP del CD agotado)" if _vp_n_ag else "")
                        + (f" · ⛔ **{_vp_n_est:,}** estructurales (importado agotado{', ' + str(int(_vp_q['accion_sugerida'].str.startswith('⛔→').sum())) + ' con opción de producir nacional' if _vp_q['accion_sugerida'].str.startswith('⛔→').any() else ''})" if _vp_n_est else "")
                    )
                    _vp_cols = [c for c in ['marca', 'sku', 'nombre', 'tienda', 'procedencia', 'prom_vta_uds',
                                             'precio_vigente', 'pct_descuento', 'margen_efectivo',
                                             'perdida_sem_soles', 'contrib_riesgo_sem', 'accion_sugerida'] if c in _vp_q.columns]
                    _vp_show = _vp_q[_vp_cols].head(20).rename(columns={
                        'marca': 'Marca', 'sku': 'SKU', 'nombre': 'Producto', 'tienda': 'Tienda',
                        'procedencia': 'Origen', 'prom_vta_uds': 'Vta/sem (uds)', 'precio_vigente': 'Precio',
                        'pct_descuento': 'Dscto', 'margen_efectivo': 'Margen efect.',
                        'perdida_sem_soles': 'Venta en riesgo S//sem',
                        'contrib_riesgo_sem': 'Contrib. en riesgo S//sem', 'accion_sugerida': 'Acción sugerida',
                    })
                    st.dataframe(_vp_show.style.format({
                        'Vta/sem (uds)': '{:.1f}', 'Precio': 'S/ {:,.2f}',
                        'Dscto': '{:.0%}', 'Margen efect.': '{:.0%}',
                        'Venta en riesgo S//sem': 'S/ {:,.0f}', 'Contrib. en riesgo S//sem': 'S/ {:,.0f}',
                    }, na_rep="—"), use_container_width=True, hide_index=True)
                    st.markdown(f"""
                    <div style="background:var(--capi-bg-surface); border:1px solid var(--capi-border); border-radius:10px; padding:12px 16px; margin-top:8px; font-size:0.74rem; color:var(--capi-text2); line-height:1.5;">
                        🚚 <strong style="color:var(--capi-text);">ATP = Available To Promise</strong> (stock del CD disponible para prometer). Como el reporte del CD <strong>no es en tiempo real</strong>, solo se compromete el <strong>{cd_prometible_pct}%</strong> del stock reportado (configurable en ⚙️); el resto es colchón de seguridad. Ese stock es un <strong>pool único compartido</strong> entre las tiendas que piden el SKU: se reparte por prioridad y lo que no alcanza va a <strong>cola</strong> (transferencia u orden a proveedor). Los SKUs con CD volátil entre cortes van flagueados ⚠️.
                    </div>""", unsafe_allow_html=True)
                    st.caption("La tabla y la asignación del ATP se priorizan por CONTRIBUCIÓN en riesgo (venta × margen efectivo), no por venta: un SKU con descuento alto vende más unidades pero puede aportar menos al P&L. Dscto y Margen efect. dan el contexto para decidir si la velocidad es demanda real o evento de precio.")
                    st.caption("La acción sale del plan del motor: despacho si hay stock en CD (cantidades del plan de reposición), transferencia si otra tienda tiene exceso, orden de compra si no hay stock en la cadena.")

                _vp_xl_buf = io.BytesIO()
                with pd.ExcelWriter(_vp_xl_buf, engine='openpyxl') as _vp_w:
                    if not _vp['df_detalle'].empty:
                        _vp['df_detalle'].to_excel(_vp_w, sheet_name='Perdida Historica por SKU', index=False)
                    if not _vp_q.empty:
                        _vp_q[_vp_cols].to_excel(_vp_w, sheet_name='Quiebre Actual SKUxTienda', index=False)
                _vp_xl_buf.seek(0)
                st.download_button(
                    "📥 Descargar análisis de ventas perdidas (.xlsx)",
                    data=_vp_xl_buf.getvalue(),
                    file_name="Capi_Ventas_Perdidas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_ventas_perdidas",
                )

    # ── Capital + Venta a Costo + Cobertura por marca (stacked bar) ──
    st.markdown("---")

    group_col = "marca" if "marca" in df_dash.columns else "categoria"
    group_label = "Marca" if group_col == "marca" else "Categoría"

    # Calcular capital, venta a costo y cobertura por marca
    _has_vta = "vta_soles_4sem" in df_dash.columns and "contrib_soles_4sem" in df_dash.columns

    # Capital: sumar a nivel SKU×Tienda (cada fila es un combo, correcto)
    capital_grp = (
        df_dash.groupby(group_col)["stock_valor_costo"]
        .sum()
        .sort_values(ascending=False)
        .head(10)
        .reset_index()
    )
    capital_grp.columns = [group_label, "stock_valor_costo"]

    # Venta a costo: vta_soles_4sem y contrib_soles_4sem son a nivel SKU (no SKU×Tienda)
    # → deduplicar por SKU antes de sumar para no inflar por número de tiendas
    if _has_vta:
        _df_sku_vta = df_dash.drop_duplicates("sku")[[group_col, "vta_soles_4sem", "contrib_soles_4sem"]].copy()
        _vta_marca = _df_sku_vta.groupby(group_col).agg(
            vta_soles=("vta_soles_4sem", "sum"),
            contrib_soles=("contrib_soles_4sem", "sum"),
        ).reset_index()
        _vta_marca.columns = [group_label, "_vta_soles", "_contrib_soles"]
        _vta_marca["vta_costo"] = (_vta_marca["_vta_soles"] - _vta_marca["_contrib_soles"]).clip(lower=0)
        capital_grp = capital_grp.merge(_vta_marca[[group_label, "vta_costo"]], on=group_label, how="left")
        capital_grp["vta_costo"] = capital_grp["vta_costo"].fillna(0)
        # En SEMANAS (decisión Franco C1 2026-08-26): vta_costo es de 4 semanas
        # → semanal = vta_costo/4; cobertura = capital / semanal
        capital_grp["cobertura_meses"] = capital_grp.apply(
            lambda r: round(r["stock_valor_costo"] / (r["vta_costo"] / 4), 1) if r["vta_costo"] > 0 else None, axis=1
        )
    else:
        capital_grp["vta_costo"] = 0
        capital_grp["cobertura_meses"] = None

    capital_grp = capital_grp.sort_values("stock_valor_costo", ascending=True)

    _max_total = (capital_grp["stock_valor_costo"] + capital_grp["vta_costo"]).max()

    # ── Renderizar stacked bar con HTML (capital + venta a costo + badge cobertura) ──
    st.markdown(f"""<div style="margin-bottom:12px;">
    <div style="font-size:14px; font-weight:600; color:var(--capi-text); margin-bottom:4px;">Inventario a Costo por {group_label} (Top 10)</div>
    <div style="display:flex; gap:16px; font-size:12px; color:var(--capi-text2); margin-bottom:12px;">
        <span style="display:flex; align-items:center; gap:4px;"><span style="width:10px; height:10px; border-radius:2px; background:{TEAL_700}; display:inline-block;"></span>Capital a costo</span>
        <span style="display:flex; align-items:center; gap:4px;"><span style="width:10px; height:10px; border-radius:2px; background:#5DCAA5; display:inline-block;"></span>Venta a costo (4 sem)</span>
        <span style="display:flex; align-items:center; gap:4px;"><span style="background:var(--capi-bg-surface); border:1px solid var(--capi-border); border-radius:4px; padding:0 5px; font-size:10px; color:var(--capi-text);">12.5</span>Cobertura (semanas)</span>
    </div>
    </div>""", unsafe_allow_html=True)

    _bars_html = ""
    for _, _row in capital_grp.iloc[::-1].iterrows():
        _cap = _row["stock_valor_costo"]
        _vta = _row["vta_costo"]
        _cob = _row.get("cobertura_meses", None)
        _marca_name = _row[group_label]

        _total = _cap + _vta
        _bar_w_pct = max(5, _total / _max_total * 100) if _max_total > 0 else 5
        _cap_pct = (_cap / _total * 100) if _total > 0 else 100
        _vta_pct = 100 - _cap_pct

        # Texto dentro de barras
        _cap_txt = f"S/ {_cap/1e6:.1f}M" if _cap >= 1e6 else f"S/ {_cap:,.0f}"
        _vta_txt = f"S/ {_vta/1e6:.1f}M" if _vta >= 1e6 else f"S/ {_vta:,.0f}"

        # Semáforo cobertura en SEMANAS: verde ≤12, neutro ≤20, rojo >20
        if _cob is not None:
            if _cob <= 12:
                _cob_bg, _cob_color, _cob_border = "#ECFDF5", "#059669", "#A7F3D0"
            elif _cob <= 20:
                _cob_bg, _cob_color, _cob_border = SLATE_100, SLATE_700, SLATE_200
            else:
                _cob_bg, _cob_color, _cob_border = "#FEF2F2", "#DC2626", "#FECACA"
            _cob_html = f'<span style="background:{_cob_bg}; color:{_cob_color}; border:1px solid {_cob_border}; border-radius:6px; padding:2px 8px; font-size:11px; font-weight:600; min-width:50px; text-align:center; white-space:nowrap;">{_cob:.1f}</span>'
        else:
            _cob_html = f'<span style="background:var(--capi-bg-surface); color:var(--capi-text2); border:1px solid var(--capi-border); border-radius:6px; padding:2px 8px; font-size:11px; min-width:50px; text-align:center;">—</span>'

        _bars_html += f"""<div style="display:flex; align-items:center; gap:8px; margin-bottom:5px;">
            <span style="width:130px; font-size:12px; font-weight:500; color:{SLATE_800}; text-align:right; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; flex-shrink:0;">{_marca_name}</span>
            <div style="flex:1; display:flex; align-items:center; gap:6px;">
                <div style="width:{_bar_w_pct:.1f}%; display:flex; height:22px; border-radius:4px; overflow:hidden;">
                    <div style="width:{_cap_pct:.0f}%; background:{TEAL_700}; display:flex; align-items:center; justify-content:flex-end; padding-right:5px; min-width:40px;">
                        <span style="font-size:10px; color:white; font-weight:500; white-space:nowrap;">{_cap_txt}</span>
                    </div>
                    <div style="width:{_vta_pct:.0f}%; background:#5DCAA5; display:flex; align-items:center; padding-left:4px; min-width:{'40px' if _vta > 0 else '0'};">
                        <span style="font-size:10px; color:white; font-weight:500; white-space:nowrap;">{_vta_txt if _vta > 0 else ''}</span>
                    </div>
                </div>
                {_cob_html}
            </div>
        </div>"""

    st.markdown(_bars_html, unsafe_allow_html=True)

    # ── Tabs ───────────────────────────────────────────────────────

    st.markdown("---")

    # Contar combos obsoletos (rangos 6_9, 9_12, 12_99)
    _RANGOS_OBSOLETOS_TAB = {"RANGO 6_9", "RANGO 9_12", "RANGO 12_99"}
    _n_obs = 0
    if "rango_antiguedad" in df_cob.columns:
        _n_obs = len(df_cob[df_cob["rango_antiguedad"].isin(_RANGOS_OBSOLETOS_TAB)])

# Navegación controlada por sidebar radio (nav_page)
# Si nav_page es None (antes del análisis), no se renderiza nada aquí


# ─── TAB 0: Acciones del Día ─────────────────────────────────
#  Resumen por MARCA con desplegable por TIENDA
#  Filtros de temporada y rango de antigüedad (obsoletos excluidos por defecto)

if nav_page == "🏠 Dashboard":

    # ══════════════════════════════════════════════════════════
    #  FOTO DEL INVENTARIO: HOY vs HACE 1 SEMANA vs HACE 1 MES (S4)
    #  Decisión Franco 2026-09-05: lo que importa es la foto actual y si las
    #  acciones la mueven. KPIs de stock con Δ semanal y Δ mensual; sin YoY
    #  (no hay calendario de eventos ni stock 2025); la venta es contexto, sin flechas.
    # ══════════════════════════════════════════════════════════
    if _HAS_SNAPSHOTS and not _DEMO_MODE:
        try:
            _fa = comparativo_semanal.foto_actual()
        except Exception as _e_cs:
            _fa = {}
            st.caption(f"Comparativo semanal no disponible: {_e_cs}")
        if _fa and _fa.get("semana_prev"):
            st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
            st.markdown(f'<div class="section-header"><h3>📈 Foto del inventario: ¿se mueve?</h3><span class="live-badge">SEMANAL · MENSUAL</span></div>', unsafe_allow_html=True)
            _act, _prev, _mes = _fa["actual"], _fa["semana_prev"], _fa.get("semana_mes")
            _et = lambda w: analisis_estados.etiqueta_semana(w, corta=True) if w else "—"
            st.caption(f"Hoy: **{_act}** ({_et(_act)}) · hace 1 semana: {_prev} ({_et(_prev)})"
                       + (f" · hace ~1 mes: {_mes} ({_et(_mes)})" if _mes else " · aún no hay snapshot de hace un mes")
                       + ". Verde = la foto mejoró. Fuente: snapshots semanales a nivel cadena.")
            _k = _fa["kpis"][_act]
            _cards = [
                ("Capital inmovilizado", "capital_inmovilizado", "soles"),
                ("% del capital inmovilizado", "pct_inmovilizado", "pct"),
                ("Pre-obsoleto (6–9 meses)", "capital_preobsoleto", "soles"),
                ("Obsoleto (9 meses a más)", "capital_obsoleto", "soles"),
                ("SKUs en quiebre", "skus_quiebre", "int"),
                ("% SKUs con stock sin venta", "pct_venta_cero", "pct"),
                ("Cobertura (semanas)", "cobertura_sem", "num1"),
            ]
            def _fmtv(v, f):
                if pd.isna(v): return "—"
                return f"S/ {v:,.0f}" if f == "soles" else f"{v*100:.1f}%" if f == "pct" else f"{v:.1f}" if f == "num1" else f"{v:,.0f}"
            def _fmtd(d, f):
                if not d or pd.isna(d[1]): return "—"
                return f"{d[1]:+.1f} pp" if f == "pct" else f"{d[1]:+.1f}%"
            _cols = st.columns(len(_cards))
            for _col, (_lab, _key, _f) in zip(_cols, _cards):
                _ds = _fa["delta_sem"].get(_key); _dm = _fa["delta_mes"].get(_key)
                _col.metric(_lab, _fmtv(_k.get(_key), _f), _fmtd(_ds, _f) + " sem", delta_color="inverse",
                            help=f"Δ semanal vs {_prev}" + (f" · Δ mensual vs {_mes}: {_fmtd(_dm, _f)}" if _mes else ""))
                if _mes:
                    _dmv = _dm[1] if _dm else float("nan")
                    _clr = "#6B7280" if pd.isna(_dmv) else ("#10b981" if _dmv < 0 else "#ef4444")
                    _col.markdown(f'<div style="font-size:0.72rem; color:{_clr}; margin-top:-8px;">mes: {_fmtd(_dm, _f)}</div>', unsafe_allow_html=True)

            # Segunda fila (revisión Franco 2026-09-05): ¿qué viene y dónde está la plata?
            _cards2 = [
                ("Entra a pre-obsoleto en 4 sem", "capital_por_entrar", "soles"),
                ("Pasa a obsoleto en 4 sem", "capital_por_pasar", "soles"),
                ("Lanzamientos sin venta", "capital_nuevo_sin_venta", "soles"),
                ("% del capital en CD", "pct_capital_cd", "pct"),
                ("Capital en liquidación (≥40%)", "capital_liquidacion", "soles"),
                ("On order (uds)", "on_order_uds", "int"),
            ]
            _cols2 = st.columns(len(_cards2))
            for _col, (_lab, _key, _f) in zip(_cols2, _cards2):
                _ds = _fa["delta_sem"].get(_key); _dm = _fa["delta_mes"].get(_key)
                _col.metric(_lab, _fmtv(_k.get(_key), _f), _fmtd(_ds, _f) + " sem",
                            delta_color=("off" if _key == "on_order_uds" else "inverse"),
                            help=f"Δ semanal vs {_prev}" + (f" · Δ mensual vs {_mes}: {_fmtd(_dm, _f)}" if _mes else ""))
                if _mes and _key != "on_order_uds":
                    _dmv = _dm[1] if _dm else float("nan")
                    _clr = "#6B7280" if pd.isna(_dmv) else ("#10b981" if _dmv < 0 else "#ef4444")
                    _col.markdown(f'<div style="font-size:0.72rem; color:{_clr}; margin-top:-8px;">mes: {_fmtd(_dm, _f)}</div>', unsafe_allow_html=True)
            st.caption("Fila 1: dónde está el problema hoy. Fila 2: qué viene (entra a pre-obsoleto = llega a 6 meses en 4 semanas; pasa a obsoleto = llega a 9 meses en 4 semanas; "
                       "lanzamientos sin venta = NUEVO SIN VENTA) y dónde está la plata (en CD sin bajar al piso, o ya en liquidación). "
                       "On order solo existe en cortes desde el 30-ago.")

            with st.expander("Ver la serie completa (stock y venta) de las últimas 5 semanas", expanded=False):
                _cs = comparativo_semanal.panel_4_semanas()
                if len(_cs.get("semanas", [])) >= 2:
                    _t = _cs["tabla"].copy()
                    _fmt_map = {l: f for _, l, f, _ in comparativo_semanal.KPIS}
                    def _fmt_row(row):
                        f = _fmt_map.get(row.name, "num1")
                        return pd.Series([("—" if pd.isna(v) else f"S/ {v:,.0f}" if f == "soles" else f"S/ {v:,.2f}" if f == "soles2"
                                           else f"{v*100:.1f}%" if f == "pct" else f"{v:,.0f}" if f == "int" else f"{v:.1f}") for v in row], index=row.index)
                    _t_disp = _t.apply(_fmt_row, axis=1)
                    _t_disp.columns = [f"{w}" + ("  ◀ hoy" if w == _act else "") for w in _t_disp.columns]
                    st.dataframe(_t_disp, use_container_width=True, height=440)
                    st.caption("Venta, contribución y margen van como contexto: semana a semana los mueven los eventos de precio, "
                               "no las acciones sobre el inventario. Capital inmovilizado = DORMIDO + ESTANCADO + SOBRESTOCK + LIQUIDAR + MUERTO. "
                               "Pre-obsoleto = 6–9 meses en tienda · Obsoleto = 9 meses a más (definición Franco 2026-09-05, por antigüedad). Cuando haya más cortes, el comparativo mensual pasa a periodo comercial Ripley.")
                try:
                    _rp = comparativo_semanal.resumen_pareto(_act)
                    if _rp:
                        st.markdown(f"**Pareto del capital inmovilizado ({_act}):** {_rp['n_skus_top80']:,} SKUs "
                                    f"({_rp['pct_skus_top80']*100:.0f}% de los {_rp['n_skus_exceso']:,} en exceso) concentran "
                                    f"S/ {_rp['capital_top80']:,.0f} de S/ {_rp['capital_exceso']:,.0f}.")
                        _pdf = comparativo_semanal.pareto_inmovilizado(_act)
                        _pdf = _pdf[_pdf["top_80"] != ""].head(300)
                        st.dataframe(_pdf.style.format({"capital": "S/ {:,.0f}", "share": "{:.1%}", "pct_acum": "{:.0%}",
                                                        "cobertura_sem": "{:.1f}"}, na_rep="—"),
                                     use_container_width=True, hide_index=True, height=300)
                except Exception as _e_p:
                    st.caption(f"Pareto no disponible: {_e_p}")

    # ══════════════════════════════════════════════════════════
    #  MARGEN EFECTIVO — Contribución / VtasMF (4 semanas)
    # ══════════════════════════════════════════════════════════
    if _margen_global is not None and _vta_soles_total > 0:
        st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="section-header"><h3>💰 Margen Efectivo</h3><span class="live-badge">RENTABILIDAD</span></div>', unsafe_allow_html=True)
        st.caption("Contribución / Venta MF — últimas 4 semanas, ponderado por volumen de venta")

        # ── KPIs de margen ──
        _mk1, _mk2, _mk3 = st.columns(3)
        _mg_pct = _margen_global * 100
        _mg_color = "#10b981" if _mg_pct >= 35 else ("#f59e0b" if _mg_pct >= 25 else "#ef4444")
        _mg_bg = "#F0FDF4" if _mg_pct >= 35 else ("#FFFBEB" if _mg_pct >= 25 else "#FEF2F2")
        with _mk1:
            st.markdown(f"""<div style="background:{_mg_bg}; border-radius:12px; padding:16px 20px; border-left:4px solid {_mg_color};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Margen efectivo global</div>
                <div style="font-size:1.8rem; font-weight:700; color:{_mg_color};">{_mg_pct:.1f}%</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Contribución / Venta (4 sem)</div>
            </div>""", unsafe_allow_html=True)
        with _mk2:
            st.markdown(f"""<div style="background:#EFF6FF; border-radius:12px; padding:16px 20px; border-left:4px solid #3b82f6;">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Venta total (4 sem)</div>
                <div style="font-size:1.8rem; font-weight:700; color:#3b82f6;">S/ {_vta_soles_total:,.0f}</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Sin IGV</div>
            </div>""", unsafe_allow_html=True)
        with _mk3:
            st.markdown(f"""<div style="background:#F0FDF4; border-radius:12px; padding:16px 20px; border-left:4px solid {TEAL_600};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Contribución total (4 sem)</div>
                <div style="font-size:1.8rem; font-weight:700; color:{TEAL_700};">S/ {_contrib_soles_total:,.0f}</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Venta - Costo</div>
            </div>""", unsafe_allow_html=True)

        # ── Tabla de margen por marca ──
        if _margen_por_marca:
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
            st.markdown(f"<h4 style='color:{SLATE_800}; margin:0 0 8px 0;'>Margen por marca</h4>", unsafe_allow_html=True)

            _max_vta_marca = max(r['vta_soles'] for r in _margen_por_marca) if _margen_por_marca else 1
            _rows_html = ""
            for _mr in _margen_por_marca:
                _m_pct = _mr['margen_efectivo'] * 100
                _bar_w = max(2, int(_mr['vta_soles'] / _max_vta_marca * 100))
                _m_clr = "#10b981" if _m_pct >= 35 else ("#f59e0b" if _m_pct >= 25 else "#ef4444")
                _rows_html += f"""<tr>
                    <td style="padding:8px 12px; font-weight:500; white-space:nowrap;">{_mr['marca']}</td>
                    <td style="padding:8px 12px; text-align:right;">S/ {_mr['vta_soles']:,.0f}</td>
                    <td style="padding:8px 12px; text-align:right;">S/ {_mr['contrib_soles']:,.0f}</td>
                    <td style="padding:8px 12px; text-align:right; font-weight:600; color:{_m_clr};">{_m_pct:.1f}%</td>
                    <td style="padding:8px 12px; width:120px;">
                        <div style="background:#E2E8F0; border-radius:4px; height:14px; width:100%;">
                            <div style="background:{_m_clr}; border-radius:4px; height:14px; width:{_bar_w}%;"></div>
                        </div>
                    </td>
                </tr>"""

            st.markdown(f"""<div style="overflow-x:auto;">
            <table style="width:100%; border-collapse:collapse; font-size:0.85rem;">
                <thead>
                    <tr style="background:var(--capi-bg-surface); border-bottom:2px solid var(--capi-border);">
                        <th style="padding:8px 12px; text-align:left;">Marca</th>
                        <th style="padding:8px 12px; text-align:right;">Venta S/</th>
                        <th style="padding:8px 12px; text-align:right;">Contribución S/</th>
                        <th style="padding:8px 12px; text-align:right;">Margen %</th>
                        <th style="padding:8px 12px; text-align:left;">Vol. relativo</th>
                    </tr>
                </thead>
                <tbody>{_rows_html}</tbody>
            </table></div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    #  TICKET PROMEDIO + COMPARATIVO VS AÑO PASADO (LY)
    # ══════════════════════════════════════════════════════════
    # En modo demo se oculta la sección YoY para aligerar el Dashboard (guion 3 min)
    if (not _DEMO_MODE) and ly_comparison and ly_comparison.get('ticket_actual_global', 0) > 0:
        st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
        st.markdown(f'<div class="section-header"><h3>🎫 Ticket Promedio & vs Año Pasado</h3><span class="live-badge">YoY</span></div>', unsafe_allow_html=True)

        _ly_g = ly_comparison.get('ly_global')
        _sem_act = ly_comparison.get('semana_actual', '?')
        _ticket_act = ly_comparison.get('ticket_actual_global', 0)
        # S3 (2026-09-05): si la base no trae fecha de corte, el motor cae a la semana del
        # reloj (date.today) y el YoY compara contra la semana equivocada. Se avisa, no se calla.
        if 'semana_corte_base' in s and s.get('semana_corte_base') is None:
            st.error("⚠️ El comparativo vs año pasado usa la **semana de hoy**, no la de la base: el archivo "
                     "subido no tiene fecha en el nombre (ej. 'Base al 30.08.xlsx'). Renómbralo y vuelve a subir "
                     "antes de citar estos números.")

        _tk1, _tk2, _tk3, _tk4 = st.columns(4)
        with _tk1:
            st.markdown(f"""<div style="background:#EFF6FF; border-radius:12px; padding:16px 20px; border-left:4px solid #3b82f6;">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Ticket promedio actual</div>
                <div style="font-size:1.8rem; font-weight:700; color:#3b82f6;">S/ {_ticket_act:,.0f}</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Venta S/ / Unidades (4 sem)</div>
            </div>""", unsafe_allow_html=True)

        if _ly_g:
            _ticket_ly = _ly_g.get('ticket_ly', 0)
            _d_ticket = _ly_g.get('delta_ticket_pct', 0)
            _d_vta = _ly_g.get('delta_vta_soles_pct', 0)
            _d_uds = _ly_g.get('delta_vta_uds_pct', 0)
            _year_ly = _ly_g.get('año_ly', '?')

            _clr_ticket = "#10b981" if _d_ticket >= 0 else "#ef4444"
            _clr_vta = "#10b981" if _d_vta >= 0 else "#ef4444"
            _clr_uds = "#10b981" if _d_uds >= 0 else "#ef4444"
            _arrow_ticket = "▲" if _d_ticket >= 0 else "▼"
            _arrow_vta = "▲" if _d_vta >= 0 else "▼"
            _arrow_uds = "▲" if _d_uds >= 0 else "▼"

            with _tk2:
                st.markdown(f"""<div style="background:{'#F0FDF4' if _d_ticket >= 0 else '#FEF2F2'}; border-radius:12px; padding:16px 20px; border-left:4px solid {_clr_ticket};">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Δ Ticket vs LY (sem {_sem_act})</div>
                    <div style="font-size:1.8rem; font-weight:700; color:{_clr_ticket};">{_arrow_ticket} {abs(_d_ticket):.1f}%</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">LY: S/ {_ticket_ly:,.0f} ({_year_ly})</div>
                </div>""", unsafe_allow_html=True)
            with _tk3:
                st.markdown(f"""<div style="background:{'#F0FDF4' if _d_vta >= 0 else '#FEF2F2'}; border-radius:12px; padding:16px 20px; border-left:4px solid {_clr_vta};">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Δ Venta S/ vs LY</div>
                    <div style="font-size:1.8rem; font-weight:700; color:{_clr_vta};">{_arrow_vta} {abs(_d_vta):.1f}%</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">Sem actual vs misma sem {_year_ly}</div>
                </div>""", unsafe_allow_html=True)
            with _tk4:
                st.markdown(f"""<div style="background:{'#F0FDF4' if _d_uds >= 0 else '#FEF2F2'}; border-radius:12px; padding:16px 20px; border-left:4px solid {_clr_uds};">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Δ Unidades vs LY</div>
                    <div style="font-size:1.8rem; font-weight:700; color:{_clr_uds};">{_arrow_uds} {abs(_d_uds):.1f}%</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">Sem actual vs misma sem {_year_ly}</div>
                </div>""", unsafe_allow_html=True)

            # ── Tabla comparativa por marca ──
            _ly_marca_data = ly_comparison.get('ly_marca', [])
            if _ly_marca_data:
                st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
                st.markdown(f"<h4 style='color:{SLATE_800}; margin:0 0 8px 0;'>Comparativo por Marca — Semana {_sem_act} actual vs {_year_ly}</h4>", unsafe_allow_html=True)

                _ly_rows_html = ""
                for _lm in _ly_marca_data:
                    _m_name = _lm.get('marca', '')
                    _m_ticket = _lm.get('ticket', 0)
                    _m_ticket_ly = _lm.get('ticket_ly', 0)
                    _m_dvta = _lm.get('delta_vta_pct', 0)
                    _m_dticket = _lm.get('delta_ticket_pct', 0)
                    _m_mg, _m_mg_ly, _m_dmg = _lm.get('margen_pct', 0), _lm.get('margen_ly_pct', 0), _lm.get('delta_margen_pp', 0)
                    _clr_mv = "#10b981" if _m_dvta >= 0 else "#ef4444"
                    _clr_mt = "#10b981" if _m_dticket >= 0 else "#ef4444"
                    _clr_mg = "#10b981" if _m_dmg >= 0 else "#ef4444"
                    _arr_mv = "▲" if _m_dvta >= 0 else "▼"
                    _arr_mt = "▲" if _m_dticket >= 0 else "▼"
                    _arr_mg = "▲" if _m_dmg >= 0 else "▼"
                    _ly_rows_html += f"""<tr>
                        <td style="padding:6px 10px; font-weight:500;">{_m_name}</td>
                        <td style="padding:6px 10px; text-align:right;">S/ {_m_ticket:,.0f}</td>
                        <td style="padding:6px 10px; text-align:right; color:var(--capi-text2);">S/ {_m_ticket_ly:,.0f}</td>
                        <td style="padding:6px 10px; text-align:right; font-weight:600; color:{_clr_mt};">{_arr_mt} {abs(_m_dticket):.1f}%</td>
                        <td style="padding:6px 10px; text-align:right; font-weight:600; color:{_clr_mv};">{_arr_mv} {abs(_m_dvta):.1f}%</td>
                        <td style="padding:6px 10px; text-align:right;">{_m_mg:.1f}%</td>
                        <td style="padding:6px 10px; text-align:right; color:var(--capi-text2);">{_m_mg_ly:.1f}%</td>
                        <td style="padding:6px 10px; text-align:right; font-weight:600; color:{_clr_mg};">{_arr_mg} {abs(_m_dmg):.1f} pp</td>
                    </tr>"""

                st.markdown(f"""<div style="overflow-x:auto; max-height:400px; overflow-y:auto;">
                <table style="width:100%; border-collapse:collapse; font-size:0.82rem;">
                    <thead>
                        <tr style="background:var(--capi-bg-surface); border-bottom:2px solid var(--capi-border); position:sticky; top:0;">
                            <th style="padding:8px 10px; text-align:left;">Marca</th>
                            <th style="padding:8px 10px; text-align:right;">Ticket Actual</th>
                            <th style="padding:8px 10px; text-align:right;">Ticket LY</th>
                            <th style="padding:8px 10px; text-align:right;">Δ Ticket</th>
                            <th style="padding:8px 10px; text-align:right;">Δ Venta S/</th>
                            <th style="padding:8px 10px; text-align:right;">Margen actual</th>
                            <th style="padding:8px 10px; text-align:right;">Margen LY</th>
                            <th style="padding:8px 10px; text-align:right;">Δ Margen</th>
                        </tr>
                    </thead>
                    <tbody>{_ly_rows_html}</tbody>
                </table></div>""", unsafe_allow_html=True)
        else:
            with _tk2:
                st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {SLATE_400};">
                    <div style="font-size:0.75rem; color:var(--capi-text2);">Comparativo LY</div>
                    <div style="font-size:1rem; color:var(--capi-text2);">Sin data para sem {_sem_act}</div>
                </div>""", unsafe_allow_html=True)

    # (Sección "Acciones del Día" eliminada — info disponible en vistas del sidebar)

# ═══════════════════════════════════════════════════════════════
#  SALUD DEL STOCK (Prompt C)
#  Health Score compuesto: cobertura + quiebre + eficiencia + margen
# ═══════════════════════════════════════════════════════════════

elif nav_page == "🩺 Salud del Stock":
    st.markdown(f'<div class="section-header"><h3>Salud del Stock</h3><span class="live-badge">HEALTH SCORE</span></div>', unsafe_allow_html=True)

    if not _HAS_SNAPSHOTS:
        st.warning("El modulo de snapshots no esta disponible. Se necesitan snapshots para calcular Salud del Stock.")
    else:
        _hs_weeks = snapshots_engine.list_available_weeks()
        if not _hs_weeks:
            st.info("No hay snapshots disponibles. Carga datos para generar el Health Score.")
        else:
            _hs_sem = _hs_weeks[-1]
            _hs_df = snapshots_engine.api.get_snapshot(_hs_sem)
            # Filtrar a marcas con presencia real (saca colas residuales: Arrow,
            # Brooksfield, Penguin, Psycho Bunny, etc. que ensucian el análisis)
            if _hs_df is not None and not _hs_df.empty and 'marca' in _hs_df.columns:
                _hs_df = _hs_df[_hs_df['marca'].str.upper().str.strip().isin(agente_terceras.MARCAS_CON_PRESENCIA)].copy()

            if _hs_df is None or _hs_df.empty:
                st.warning(f"Snapshot de semana {_hs_sem} esta vacio.")
            else:
                st.info(f"📸 Calculado sobre el SNAPSHOT {analisis_estados.etiqueta_semana(_hs_sem)} "
                        f"— {len(_hs_df):,} SKUs (marcas con presencia). Si acabas de subir una base "
                        f"más nueva, este score aún corresponde al corte anterior.")

                # Calcular scores
                _hs_global = motor_v2.build_health_score(_hs_df)
                _hs_marca = motor_v2.build_health_score(_hs_df, grupo_cols=['marca'])
                _hs_detail = motor_v2.build_health_detail(_hs_df)

                # ── Delta temporal: comparar con semana anterior ──
                _hs_prev_global = None
                _hs_prev_marca = None
                if len(_hs_weeks) >= 2:
                    _hs_sem_prev = _hs_weeks[-2]
                    _hs_df_prev = snapshots_engine.api.get_snapshot(_hs_sem_prev)
                    if _hs_df_prev is not None and not _hs_df_prev.empty:
                        _hs_prev_global = motor_v2.build_health_score(_hs_df_prev)
                        _hs_prev_marca = motor_v2.build_health_score(_hs_df_prev, grupo_cols=['marca'])

                _hs_tabs = st.tabs([
                    "🎯 Diagnostico Rapido",
                    "🏆 Ranking Marcas",
                    "🔍 Drill-down",
                    "📋 Detalle SKU",
                    "📊 Ranking x Componente",
                ])

                # ── Tab 1: Diagnostico Rapido ──
                with _hs_tabs[0]:
                    if _hs_global is None or _hs_global.empty:
                        # Fix C1 2026-08-26: antes un st.stop() mataba TODA la
                        # página. Placeholder neutro: el tab muestra ceros con
                        # el warning visible y el resto de la app sigue vivo.
                        st.warning("No se pudo calcular el Health Score con los datos actuales.")
                        _hs_global = pd.DataFrame([{
                            'health_score': 0.0, 'semaforo': '—', 'n_skus': 0,
                            'capital_total': 0.0, 'score_cobertura': 0.0,
                            'score_quiebre': 0.0, 'score_sobrestock': 0.0,
                            'score_eficiencia': 0.0, 'score_margen': 0.0,
                            'pct_optimo_alto': 0.0, 'pct_quiebre': 0.0,
                            'pct_exceso': 0.0, 'capital_parado': 0.0,
                            'margen_pct': 0.0, 'n_con_stock': 0,
                            'venta_en_riesgo': 0.0, 'rotacion': 0.0,
                        }])
                    _g = _hs_global.iloc[0]
                    _hs_score = float(_g['health_score'])
                    _hs_semaforo = str(_g['semaforo'])

                    # Colores de semaforo
                    _sem_colors = {
                        'SALUDABLE': '#059669',
                        'ACEPTABLE': '#0ea5e9',
                        'EN RIESGO': '#f59e0b',
                        'CRITICO': '#dc2626',
                    }
                    _sem_color = _sem_colors.get(_hs_semaforo, SLATE_500)

                    # Gauge visual con HTML + delta semanal
                    _gauge_pct = min(_hs_score / 100, 1.0)
                    _delta_html = ""
                    if _hs_prev_global is not None and not _hs_prev_global.empty:
                        _prev_score = float(_hs_prev_global.iloc[0]['health_score'])
                        _delta = _hs_score - _prev_score
                        if abs(_delta) >= 0.1:
                            _d_arrow = "▲" if _delta > 0 else "▼"
                            _d_color = "#059669" if _delta > 0 else "#dc2626"
                            _delta_html = f'<span style="font-size:1rem; color:{_d_color}; margin-left:8px;">{_d_arrow} {abs(_delta):.1f}</span>'
                    st.markdown(f"""
                    <div style="text-align:center; padding:24px 0 16px 0;">
                        <div style="font-size:4rem; font-weight:800; color:{_sem_color}; line-height:1;">{_hs_score:.0f}{_delta_html}</div>
                        <div style="font-size:0.85rem; font-weight:600; color:{_sem_color}; letter-spacing:2px; margin-top:4px;">{_hs_semaforo}</div>
                        <div style="margin:12px auto 0; width:200px; height:8px; background:{SLATE_200}; border-radius:4px; overflow:hidden;">
                            <div style="width:{_gauge_pct*100:.0f}%; height:100%; background:{_sem_color}; border-radius:4px;"></div>
                        </div>
                        <div style="font-size:0.7rem; color:var(--capi-text2); margin-top:6px;">Health Score (0-100){' | vs semana anterior' if _delta_html else ''}</div>
                    </div>
                    """, unsafe_allow_html=True)

                    # 5 componentes en cards
                    _comp_data = [
                        ("Cobertura", f"{float(_g['pct_optimo_alto'])*100:.0f}%", f"{float(_g['score_cobertura']):.0f}/100", "25%",
                         "SKUs en OPTIMO + ALTO", '#059669' if float(_g['score_cobertura']) >= 70 else '#f59e0b' if float(_g['score_cobertura']) >= 40 else '#dc2626'),
                        ("Quiebre", f"{float(_g['pct_quiebre'])*100:.0f}%", f"{float(_g['score_quiebre']):.0f}/100", "20%",
                         "En QUIEBRE + PRE-QUIEBRE", '#059669' if float(_g['score_quiebre']) >= 70 else '#f59e0b' if float(_g['score_quiebre']) >= 40 else '#dc2626'),
                        ("Sobrestock", f"{float(_g['pct_exceso'])*100:.0f}%", f"{float(_g['score_sobrestock']):.0f}/100", "15%",
                         "Capital parado (menos = mejor)", '#059669' if float(_g['score_sobrestock']) >= 70 else '#f59e0b' if float(_g['score_sobrestock']) >= 40 else '#dc2626'),
                        ("Eficiencia", f"{float(_g['rotacion'])*100:.1f}%", f"{float(_g['score_eficiencia']):.0f}/100", "20%",
                         "Rotacion a costo", '#059669' if float(_g['score_eficiencia']) >= 70 else '#f59e0b' if float(_g['score_eficiencia']) >= 40 else '#dc2626'),
                        ("Margen", f"{float(_g['margen_pct'])*100:.1f}%", f"{float(_g['score_margen']):.0f}/100", "20%",
                         "Contribucion / Venta", '#059669' if float(_g['score_margen']) >= 70 else '#f59e0b' if float(_g['score_margen']) >= 40 else '#dc2626'),
                    ]
                    _hc1, _hc2, _hc3, _hc4, _hc5 = st.columns(5)
                    for _col_hs, (_title, _val, _sc, _weight, _desc, _color) in zip([_hc1, _hc2, _hc3, _hc4, _hc5], _comp_data):
                        with _col_hs:
                            st.markdown(f"""
                            <div style="background:var(--capi-bg-surface); border:1px solid {SLATE_200}; border-radius:10px; padding:14px; text-align:center; border-top:3px solid {_color};">
                                <div style="font-size:0.7rem; color:var(--capi-text2); font-weight:600; letter-spacing:1px;">{_title.upper()}</div>
                                <div style="font-size:1.6rem; font-weight:700; color:{_color}; margin:4px 0;">{_val}</div>
                                <div style="font-size:0.75rem; color:var(--capi-text2);">Score: {_sc} | Peso: {_weight}</div>
                                <div style="font-size:0.65rem; color:{SLATE_400}; margin-top:4px;">{_desc}</div>
                            </div>
                            """, unsafe_allow_html=True)

                    # KPIs de contexto
                    st.markdown("---")
                    _kc1, _kc2, _kc3, _kc4 = st.columns(4)
                    _kc1.metric("Total SKUs", f"{int(_g['n_skus']):,}")
                    _kc2.metric("Con Stock", f"{int(_g['n_con_stock']):,}")
                    _kc3.metric("Capital Parado", f"S/{float(_g['capital_parado']):,.0f}")
                    _kc4.metric("Venta en Riesgo", f"S/{float(_g['venta_en_riesgo']):,.0f}")

                # ── Tab 2: Ranking Marcas ──
                    # ── 💊 Qué hacer por marca (Franco C1: el score debe
                    # recomendar acción, no solo señalar) ──
                    st.markdown("---")
                    st.markdown("##### 💊 Qué hacer por marca — el componente que castiga su score")
                    try:
                        _hs_rec = motor_v2.recomendaciones_salud(_hs_marca)
                    except Exception:
                        _hs_rec = None
                    if _hs_rec is not None and not _hs_rec.empty:
                        _n_inflados = int(_hs_rec["score_inflado"].sum())
                        if _n_inflados:
                            st.caption(f"⚠️ {_n_inflados} marcas con score INFLADO: no quiebran "
                                       "porque casi no venden (el caso Spavaldi) — su acción lo indica.")
                        _hs_rec_disp = _hs_rec.drop(columns=["score_inflado"]).rename(columns={
                            "marca": "Marca", "health_score": "Score", "n_skus": "SKUs",
                            "problema": "Problema principal", "peor_score": "Score del problema",
                            "accion": "Acción recomendada"})
                        st.dataframe(_hs_rec_disp.style.format(
                            {"Score": "{:.0f}", "Score del problema": "{:.0f}"}),
                            use_container_width=True, hide_index=True, height=380)

                with _hs_tabs[1]:
                    st.markdown("##### Ranking de marcas por Health Score")

                    # ── Filtro de materialidad: excluir marcas insignificantes ──
                    _MIN_SKUS_MATERIAL = 5
                    _MIN_VENTA_MATERIAL = 500
                    _hs_marca_material = _hs_marca[
                        (_hs_marca['n_skus'] >= _MIN_SKUS_MATERIAL) |
                        (_hs_marca['venta_total'] >= _MIN_VENTA_MATERIAL)
                    ].copy()
                    _n_excluidas = len(_hs_marca) - len(_hs_marca_material)
                    if _n_excluidas > 0:
                        st.caption(f"{_n_excluidas} marcas excluidas por baja materialidad (<{_MIN_SKUS_MATERIAL} SKUs activos y <S/{_MIN_VENTA_MATERIAL:,} venta)")

                    # Filtro de semaforo
                    _hs_fc1, _hs_fc2 = st.columns([2, 1])
                    with _hs_fc1:
                        _hs_sem_filter = st.multiselect(
                            "Filtrar por semaforo",
                            ['SALUDABLE', 'ACEPTABLE', 'EN RIESGO', 'CRITICO'],
                            default=['SALUDABLE', 'ACEPTABLE', 'EN RIESGO', 'CRITICO'],
                            key="hs_sem_filter"
                        )
                    with _hs_fc2:
                        _hs_sort_by = st.selectbox(
                            "Ordenar por",
                            ['Health Score', 'Capital Parado', 'Venta en Riesgo', 'Impacto Negocio'],
                            key="hs_sort_by"
                        )
                    _hs_marca_f = _hs_marca_material[_hs_marca_material['semaforo'].isin(_hs_sem_filter)].copy()

                    if _hs_marca_f.empty:
                        st.info("No hay marcas con el filtro seleccionado.")
                    else:
                        # ── Delta semanal por marca ──
                        if _hs_prev_marca is not None and not _hs_prev_marca.empty:
                            _prev_map = _hs_prev_marca.set_index('marca')['health_score'].to_dict()
                            _hs_marca_f['_prev_score'] = _hs_marca_f['marca'].map(_prev_map)
                            _hs_marca_f['delta_score'] = (_hs_marca_f['health_score'] - _hs_marca_f['_prev_score'].fillna(_hs_marca_f['health_score'])).round(1)
                            _hs_marca_f = _hs_marca_f.drop(columns=['_prev_score'])
                        else:
                            _hs_marca_f['delta_score'] = 0.0

                        # Ordenar según selección
                        _sort_col_map = {
                            'Health Score': 'health_score',
                            'Capital Parado': 'capital_parado',
                            'Venta en Riesgo': 'venta_en_riesgo',
                            'Impacto Negocio': 'impacto',
                        }
                        _sort_col = _sort_col_map.get(_hs_sort_by, 'health_score')
                        # Impacto ordena desc (mayor impacto primero); el resto asc (peor score primero)
                        _sort_asc = False if _sort_col == 'impacto' else True
                        _hs_marca_sorted = _hs_marca_f.sort_values(_sort_col, ascending=_sort_asc)

                        # Horizontal bar chart con Plotly
                        _bar_colors = []
                        for _s in _hs_marca_sorted['semaforo']:
                            _bar_colors.append(_sem_colors.get(str(_s), SLATE_400))

                        _fig_rank = go.Figure(go.Bar(
                            x=_hs_marca_sorted['health_score'],
                            y=_hs_marca_sorted['marca'],
                            orientation='h',
                            marker_color=_bar_colors,
                            text=[f"{v:.0f}" for v in _hs_marca_sorted['health_score']],
                            textposition='outside',
                            textfont=dict(size=11),
                        ))
                        _fig_rank.update_layout(
                            height=max(300, len(_hs_marca_sorted) * 28),
                            margin=dict(l=0, r=40, t=10, b=10),
                            xaxis=dict(range=[min(_hs_marca_sorted['health_score'].min() - 10, 0), 105], title="Health Score"),
                            yaxis=dict(automargin=True),
                            plot_bgcolor='rgba(0,0,0,0)',
                            paper_bgcolor='rgba(0,0,0,0)',
                        )
                        # Lineas de referencia para semaforo
                        for _thresh, _lbl in [(40, 'CRITICO'), (60, 'EN RIESGO'), (75, 'ACEPTABLE')]:
                            _fig_rank.add_vline(x=_thresh, line_dash="dot", line_color=SLATE_400, line_width=1,
                                                annotation_text=_lbl, annotation_position="top")
                        st.plotly_chart(_fig_rank, use_container_width=True)

                        # ── Scatter Plot: Score vs Capital Parado ──
                        st.markdown("##### Mapa de Impacto: Score vs Capital")
                        _fig_scatter = go.Figure()
                        _scatter_colors = [_sem_colors.get(str(s), SLATE_400) for s in _hs_marca_f['semaforo']]
                        # Bubble size = impacto (si existe) o n_skus
                        if 'impacto' in _hs_marca_f.columns and _hs_marca_f['impacto'].max() > 0:
                            _bubble_raw = _hs_marca_f['impacto'].fillna(0)
                            _bubble_size = (_bubble_raw / _bubble_raw.max() * 35 + 8).clip(upper=50)
                            _hover_extra = "Impacto: %{customdata:.1f}<br>"
                            _customdata = _hs_marca_f['impacto']
                        else:
                            _bubble_size = (_hs_marca_f['n_skus'] / _hs_marca_f['n_skus'].max() * 30 + 8).clip(upper=45)
                            _hover_extra = ""
                            _customdata = None
                        _fig_scatter.add_trace(go.Scatter(
                            x=_hs_marca_f['health_score'],
                            y=_hs_marca_f['capital_parado'],
                            mode='markers+text',
                            marker=dict(
                                size=_bubble_size,
                                color=_scatter_colors,
                                opacity=0.8,
                                line=dict(width=1, color='white'),
                            ),
                            text=_hs_marca_f['marca'],
                            textposition='top center',
                            textfont=dict(size=9),
                            customdata=_customdata,
                            hovertemplate="<b>%{text}</b><br>Score: %{x:.0f}<br>Capital Parado: S/%{y:,.0f}<br>" + _hover_extra + "<extra></extra>",
                        ))
                        # Cuadrante peligroso: bajo score + alto capital
                        _fig_scatter.add_shape(
                            type="rect", x0=-100, x1=60, y0=_hs_marca_f['capital_parado'].median(), y1=_hs_marca_f['capital_parado'].max() * 1.1,
                            fillcolor="rgba(220,38,38,0.05)", line=dict(width=0),
                        )
                        _fig_scatter.add_annotation(
                            x=30, y=_hs_marca_f['capital_parado'].max() * 0.95,
                            text="ZONA CRITICA", showarrow=False,
                            font=dict(size=10, color="#dc2626"), opacity=0.6,
                        )
                        _fig_scatter.update_layout(
                            height=400,
                            margin=dict(l=0, r=0, t=10, b=0),
                            xaxis=dict(title="Health Score", zeroline=True),
                            yaxis=dict(title="Capital Parado (S/)", tickformat=","),
                            plot_bgcolor='rgba(0,0,0,0)',
                            paper_bgcolor='rgba(0,0,0,0)',
                        )
                        st.plotly_chart(_fig_scatter, use_container_width=True)

                        # Tabla resumen
                        st.markdown("##### Detalle por marca")
                        _tbl_cols = ['marca', 'health_score', 'semaforo', 'n_skus', 'n_con_stock',
                                     'pct_optimo_alto', 'pct_quiebre', 'pct_exceso', 'rotacion', 'margen_pct',
                                     'capital_parado', 'venta_en_riesgo']
                        _tbl_headers = ['Marca', 'Score', 'Semaforo', 'SKUs', 'Con Stock',
                                        '% Opt+Alto', '% Quiebre', '% Exceso', 'Rotación %', 'Margen %',
                                        'Capital Parado', 'Vta en Riesgo']
                        # Agregar impacto y prioridad si existen (motor los genera a nivel marca)
                        if 'impacto' in _hs_marca_f.columns:
                            _tbl_cols.extend(['impacto', 'prioridad'])
                            _tbl_headers.extend(['Impacto', 'Prioridad'])
                        # Agregar delta si existe
                        if 'delta_score' in _hs_marca_f.columns and _hs_marca_f['delta_score'].abs().sum() > 0:
                            _tbl_cols.insert(2, 'delta_score')
                            _tbl_headers.insert(2, 'Delta')
                        _tbl_marca = _hs_marca_f[_tbl_cols].copy()
                        _tbl_marca.columns = _tbl_headers
                        _tbl_marca['% Opt+Alto'] = (_tbl_marca['% Opt+Alto'] * 100).round(1)
                        _tbl_marca['% Quiebre'] = (_tbl_marca['% Quiebre'] * 100).round(1)
                        _tbl_marca['% Exceso'] = (_tbl_marca['% Exceso'] * 100).round(1)
                        _tbl_marca['Rotación %'] = (_tbl_marca['Rotación %'] * 100).round(1)
                        _tbl_marca['Margen %'] = (_tbl_marca['Margen %'] * 100).round(1)
                        _tbl_marca['Capital Parado'] = _tbl_marca['Capital Parado'].apply(lambda x: f"S/{x:,.0f}")
                        _tbl_marca['Vta en Riesgo'] = _tbl_marca['Vta en Riesgo'].apply(lambda x: f"S/{x:,.0f}")
                        _tbl_marca = _tbl_marca.sort_values('Score', ascending=False)
                        st.dataframe(_tbl_marca, use_container_width=True, hide_index=True, height=500)

                        # Tags de prioridad con color (debajo de la tabla)
                        if 'impacto' in _hs_marca_f.columns:
                            _foco = _hs_marca_f[_hs_marca_f['prioridad'] == 'FOCO URGENTE']['marca'].tolist()
                            _monit = _hs_marca_f[_hs_marca_f['prioridad'] == 'MONITOREAR']['marca'].tolist()
                            if _foco:
                                _foco_tags = " ".join([f"<span style='background:#dc2626;color:white;padding:2px 8px;border-radius:4px;font-size:0.82em;font-weight:600;'>{m}</span>" for m in _foco])
                                st.markdown(f"🔴 **FOCO URGENTE**: {_foco_tags}", unsafe_allow_html=True)
                            if _monit:
                                _monit_tags = " ".join([f"<span style='background:#f59e0b;color:white;padding:2px 8px;border-radius:4px;font-size:0.82em;font-weight:600;'>{m}</span>" for m in _monit])
                                st.markdown(f"🟡 **MONITOREAR**: {_monit_tags}", unsafe_allow_html=True)

                        # Excel download
                        _buf_marca = io.BytesIO()
                        _tbl_marca.to_excel(_buf_marca, index=False, sheet_name='Ranking Marcas')
                        st.download_button(
                            "Descargar ranking Excel",
                            data=_buf_marca.getvalue(),
                            file_name=f"health_ranking_marcas_{_hs_sem}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="hs_dl_ranking"
                        )

                # ── Tab 3: Drill-down ──
                with _hs_tabs[2]:
                    st.markdown("##### Drill-down: Marca > Linea > Temporada")

                    _dd_marcas_list = sorted(_hs_marca['marca'].unique())
                    # Arrancar en la marca de PEOR health score (la que más necesita diagnóstico)
                    _dd_idx = 0
                    if not _hs_marca.empty:
                        _dd_peor = _hs_marca.sort_values('health_score').iloc[0]['marca']
                        if _dd_peor in _dd_marcas_list:
                            _dd_idx = _dd_marcas_list.index(_dd_peor)
                    _dd_marca_sel = st.selectbox(
                        "Seleccionar marca (arranca en la de menor Health Score)",
                        _dd_marcas_list,
                        index=_dd_idx,
                        key="hs_dd_marca"
                    )

                    if _dd_marca_sel and not _hs_marca[_hs_marca['marca'] == _dd_marca_sel].empty:
                        # Score de la marca
                        _dd_marca_row = _hs_marca[_hs_marca['marca'] == _dd_marca_sel].iloc[0]
                        _dd_color = _sem_colors.get(str(_dd_marca_row['semaforo']), SLATE_500)
                        st.markdown(f"""
                        <div style="background:var(--capi-bg-surface); border:1px solid {SLATE_200}; border-radius:10px; padding:12px 16px; margin-bottom:12px; display:flex; align-items:center; gap:16px;">
                            <div style="font-size:2rem; font-weight:800; color:{_dd_color};">{float(_dd_marca_row['health_score']):.0f}</div>
                            <div>
                                <div style="font-size:1rem; font-weight:600; color:var(--capi-text);">{_dd_marca_sel}</div>
                                <div style="font-size:0.75rem; color:{_dd_color}; font-weight:600;">{_dd_marca_row['semaforo']}</div>
                            </div>
                            <div style="margin-left:auto; display:flex; gap:20px;">
                                <div style="text-align:center;"><div style="font-size:0.65rem; color:var(--capi-text2);">SKUs</div><div style="font-weight:600;">{int(_dd_marca_row['n_skus'])}</div></div>
                                <div style="text-align:center;"><div style="font-size:0.65rem; color:var(--capi-text2);">Rotación %</div><div style="font-weight:600;">{float(_dd_marca_row['rotacion'])*100:.1f}%</div></div>
                                <div style="text-align:center;"><div style="font-size:0.65rem; color:var(--capi-text2);">Margen</div><div style="font-weight:600;">{float(_dd_marca_row['margen_pct'])*100:.1f}%</div></div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        # Drill-down por linea
                        _dd_df_marca = _hs_df[_hs_df['marca'] == _dd_marca_sel]
                        _dd_linea = motor_v2.build_health_score(_dd_df_marca, grupo_cols=['linea'])

                        if not _dd_linea.empty:
                            st.markdown(f"###### Por Linea ({len(_dd_linea)} lineas)")
                            _dd_linea_sorted = _dd_linea.sort_values('health_score', ascending=False)
                            _dd_bar_colors = [_sem_colors.get(str(s), SLATE_400) for s in _dd_linea_sorted['semaforo']]

                            _fig_dd = go.Figure(go.Bar(
                                x=_dd_linea_sorted['linea'],
                                y=_dd_linea_sorted['health_score'],
                                marker_color=_dd_bar_colors,
                                text=[f"{v:.0f}" for v in _dd_linea_sorted['health_score']],
                                textposition='outside',
                            ))
                            _fig_dd.update_layout(
                                height=350,
                                margin=dict(l=0, r=0, t=10, b=0),
                                yaxis=dict(range=[0, 105], title="Health Score"),
                                xaxis=dict(automargin=True),
                                plot_bgcolor='rgba(0,0,0,0)',
                                paper_bgcolor='rgba(0,0,0,0)',
                            )
                            st.plotly_chart(_fig_dd, use_container_width=True)

                        # Drill-down por temporada dentro de la marca
                        _dd_temp = motor_v2.build_health_score(_dd_df_marca, grupo_cols=['temporada'])
                        if not _dd_temp.empty:
                            st.markdown("###### Por Temporada")
                            _tc1, _tc2, _tc3 = st.columns(3)
                            for _col_t, _temp in zip([_tc1, _tc2, _tc3], ['OI', 'PV', 'TT']):
                                with _col_t:
                                    _tr = _dd_temp[_dd_temp['temporada'] == _temp]
                                    if not _tr.empty:
                                        _tr = _tr.iloc[0]
                                        _tc = _sem_colors.get(str(_tr['semaforo']), SLATE_400)
                                        st.markdown(f"""
                                        <div style="background:var(--capi-bg-surface); border:1px solid {SLATE_200}; border-radius:8px; padding:12px; text-align:center;">
                                            <div style="font-size:0.7rem; color:var(--capi-text2); font-weight:600;">{_temp}</div>
                                            <div style="font-size:1.8rem; font-weight:700; color:{_tc};">{float(_tr['health_score']):.0f}</div>
                                            <div style="font-size:0.65rem; color:{_tc};">{_tr['semaforo']}</div>
                                            <div style="font-size:0.6rem; color:var(--capi-text2); margin-top:4px;">{int(_tr['n_skus'])} SKUs | {float(_tr['pct_quiebre'])*100:.0f}% quiebre</div>
                                        </div>
                                        """, unsafe_allow_html=True)
                                    else:
                                        st.markdown(f"""
                                        <div style="background:var(--capi-bg-surface); border:1px solid {SLATE_200}; border-radius:8px; padding:12px; text-align:center;">
                                            <div style="font-size:0.7rem; color:var(--capi-text2);">{_temp}</div>
                                            <div style="font-size:1rem; color:{SLATE_400};">Sin data</div>
                                        </div>
                                        """, unsafe_allow_html=True)

                        # Heatmap linea × temporada
                        _dd_lt = motor_v2.build_health_score(_dd_df_marca, grupo_cols=['linea', 'temporada'])
                        if not _dd_lt.empty and len(_dd_lt) > 1:
                            st.markdown("###### Heatmap: Linea x Temporada")
                            _pivot_hs = _dd_lt.pivot_table(index='linea', columns='temporada', values='health_score', aggfunc='first')
                            # Reorder temporada columns
                            _temp_order = [t for t in ['OI', 'PV', 'TT'] if t in _pivot_hs.columns]
                            _pivot_hs = _pivot_hs[_temp_order]

                            _fig_hm = go.Figure(go.Heatmap(
                                z=_pivot_hs.values,
                                x=_pivot_hs.columns.tolist(),
                                y=_pivot_hs.index.tolist(),
                                colorscale=[[0, '#dc2626'], [0.4, '#f59e0b'], [0.6, '#0ea5e9'], [0.75, '#059669'], [1, '#059669']],
                                zmin=0, zmax=100,
                                text=[[f"{v:.0f}" if pd.notna(v) else "" for v in row] for row in _pivot_hs.values],
                                texttemplate="%{text}",
                                textfont=dict(size=14, color="white"),
                                hovertemplate="Linea: %{y}<br>Temporada: %{x}<br>Health Score: %{z:.0f}<extra></extra>",
                                colorbar=dict(title="Score", tickvals=[0, 25, 50, 75, 100]),
                            ))
                            _fig_hm.update_layout(
                                height=max(200, len(_pivot_hs) * 35 + 60),
                                margin=dict(l=0, r=0, t=10, b=0),
                                yaxis=dict(automargin=True),
                                plot_bgcolor='rgba(0,0,0,0)',
                                paper_bgcolor='rgba(0,0,0,0)',
                            )
                            st.plotly_chart(_fig_hm, use_container_width=True)

                # ── Tab 4: Detalle SKU ──
                with _hs_tabs[3]:
                    st.markdown("##### Detalle a nivel SKU")

                    # Filtros
                    _dt_c1, _dt_c2, _dt_c3 = st.columns(3)
                    with _dt_c1:
                        _dt_marca = st.selectbox("Marca", ['Todas'] + sorted(_hs_detail['marca'].unique()), key="hs_dt_marca")
                    with _dt_c2:
                        _lineas_avail = sorted(_hs_detail['linea'].dropna().unique()) if _dt_marca == 'Todas' else sorted(_hs_detail[_hs_detail['marca'] == _dt_marca]['linea'].dropna().unique())
                        _dt_linea = st.selectbox("Linea", ['Todas'] + list(_lineas_avail), key="hs_dt_linea")
                    with _dt_c3:
                        _dt_estado = st.multiselect("Estado", sorted(_hs_detail['estado'].unique()), default=[], key="hs_dt_estado")

                    _dt_df = _hs_detail.copy()
                    if _dt_marca != 'Todas':
                        _dt_df = _dt_df[_dt_df['marca'] == _dt_marca]
                    if _dt_linea != 'Todas':
                        _dt_df = _dt_df[_dt_df['linea'] == _dt_linea]
                    if _dt_estado:
                        _dt_df = _dt_df[_dt_df['estado'].isin(_dt_estado)]

                    st.caption(f"{len(_dt_df):,} SKUs filtrados")

                    # Tabla
                    _dt_show = _dt_df[['sku', 'descripcion', 'marca', 'linea', 'temporada', 'estado',
                                       'cobertura_sem', 'stock_total', 'stock_cd', 'venta_soles',
                                       'contribucion_soles', 'stock_valor_costo', 'pct_descuento',
                                       'edad_semanas']].copy()
                    _dt_show.columns = ['SKU', 'Descripcion', 'Marca', 'Linea', 'Temp', 'Estado',
                                        'Cob (sem)', 'Stock Total', 'Stock CD', 'Venta S/',
                                        'Contrib S/', 'Stock Costo', '% Dscto', 'Edad (sem)']
                    _dt_show['Cob (sem)'] = _dt_show['Cob (sem)'].round(1)
                    _dt_show['Venta S/'] = _dt_show['Venta S/'].round(0)
                    _dt_show['Contrib S/'] = _dt_show['Contrib S/'].round(0)
                    _dt_show['Stock Costo'] = _dt_show['Stock Costo'].round(0)
                    _dt_show['% Dscto'] = (_dt_show['% Dscto'] * 100).round(0)

                    st.dataframe(_dt_show, use_container_width=True, hide_index=True, height=500)

                    # Excel download
                    _buf_detail = io.BytesIO()
                    _dt_show.to_excel(_buf_detail, index=False, sheet_name='Detalle SKU')
                    st.download_button(
                        "Descargar detalle Excel",
                        data=_buf_detail.getvalue(),
                        file_name=f"health_detalle_sku_{_hs_sem}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="hs_dl_detail"
                    )

                # ── Tab 5: Ranking x Componente ──
                with _hs_tabs[4]:
                    st.markdown("##### Ranking de marcas por componente del Health Score")
                    st.caption("Cada componente en escala 0-100 (verde = fortaleza, rojo = a mejorar). "
                               "Clic en el encabezado de una columna para ordenar por ese componente.")
                    _hc = _hs_marca.copy()
                    _hc_num = ['score_cobertura', 'score_quiebre', 'score_sobrestock', 'score_eficiencia', 'score_margen', 'health_score']
                    _hc_cols = ['marca'] + [c for c in _hc_num if c in _hc.columns]
                    _hc_disp = _hc[_hc_cols].rename(columns={
                        'marca': 'Marca', 'score_cobertura': 'Cobertura', 'score_quiebre': 'Quiebre',
                        'score_sobrestock': 'Sobrestock', 'score_eficiencia': 'Eficiencia',
                        'score_margen': 'Margen', 'health_score': 'Health Score',
                    }).sort_values('Health Score', ascending=False)
                    _hc_color = [c for c in ['Cobertura', 'Quiebre', 'Sobrestock', 'Eficiencia', 'Margen', 'Health Score'] if c in _hc_disp.columns]

                    def _hc_celda(v):
                        # Verde→rojo por tramos (sin matplotlib): mayor score = más verde
                        if pd.isna(v):
                            return ''
                        if v >= 75:   return 'background-color: #C8E6C9'
                        if v >= 60:   return 'background-color: #DCEDC8'
                        if v >= 40:   return 'background-color: #FFF9C4'
                        if v >= 20:   return 'background-color: #FFE0B2'
                        return 'background-color: #FFCDD2'

                    _hc_sty = _hc_disp.style.format({c: '{:.0f}' for c in _hc_color}, na_rep="—") \
                        .map(_hc_celda, subset=_hc_color)
                    st.dataframe(_hc_sty, use_container_width=True, hide_index=True, height=520)
                    st.caption("Pesos: Cobertura 25% · Quiebre 20% · Sobrestock 15% · Eficiencia 20% · Margen 20%. "
                               "El Health Score es la suma ponderada de los 5.")
                    _hc_buf = io.BytesIO()
                    with pd.ExcelWriter(_hc_buf, engine='openpyxl') as _hc_w:
                        _hc_disp.to_excel(_hc_w, sheet_name='Ranking Componentes', index=False)
                    _hc_buf.seek(0)
                    st.download_button("📥 Descargar ranking por componente (.xlsx)", data=_hc_buf.getvalue(),
                                       file_name="Capi_Ranking_Componentes.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="hs_dl_componentes")


# ═══════════════════════════════════════════════════════════════
#  DETALLE — Vistas de análisis granular
# ═══════════════════════════════════════════════════════════════

elif nav_page == "📦 Reposición":
    st.markdown(f'<div class="section-header"><h3>📦 Reposición</h3><span class="live-badge">POR MARCA</span></div>', unsafe_allow_html=True)
    _uni_rp = st.radio("Universo", ["Propias", "Terceras", "Todas"], horizontal=True, key="uni_repo")
    _SET_P = {m.upper() for m in agente_terceras.MARCAS_PROPIAS_SET}
    _SET_T = {m.upper() for m in agente_terceras.MARCAS_AGENTE}
    _set_rp = _SET_P if _uni_rp == "Propias" else _SET_T if _uni_rp == "Terceras" else (_SET_P | _SET_T)
    if _uni_rp == "Propias":
        st.download_button(
            "📦 Descargar TODO el análisis de Marcas Propias (.xlsx)",
            data=_build_excel_propias(df_cob, df_rep, df_trans, df_gaps_dist, df_retenidos_cd),
            file_name="Capi_Analisis_Marcas_Propias.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key="dl_pack_propias",
            help="5 pestañas: Reposición · Transferencias · Precios · Predist Gaps · Retenidos CD")
    _rp = df_rep[df_rep['marca'].str.upper().str.strip().isin(_set_rp)].copy() if not df_rep.empty and 'marca' in df_rep.columns else pd.DataFrame()
    if _rp.empty:
        st.info(f"No hay reposiciones sugeridas para marcas {_uni_rp.lower()} con la base actual.")
    else:
        _rp = _rp[_rp['a_reponer'] > 0] if 'a_reponer' in _rp.columns else _rp
        if 'sku' in df_cob.columns:
            if 'margen_efectivo' in df_cob.columns:
                _mgp = df_cob.drop_duplicates('sku').set_index('sku')['margen_efectivo']
                _rp['margen_efectivo'] = (_rp['sku'].map(_mgp).fillna(0) * 100).round(1)
            if 'edad_semanas' in df_cob.columns:
                _rp['edad'] = _rp['sku'].map(df_cob.groupby('sku')['edad_semanas'].max())
        st.caption(f"{len(_rp):,} líneas en {_rp['marca'].nunique()} marcas · {int(_rp['a_reponer'].sum()):,} uds a reponer")
        _rp_sel = st.selectbox("Marca", ["Todas"] + sorted(_rp['marca'].unique().tolist()), key="rp_marca")
        _rp_v = _rp if _rp_sel == "Todas" else _rp[_rp['marca'] == _rp_sel]
        _rp_cols = [c for c in ['marca', 'sku', 'nombre', 'categoria', 'tienda', 'edad', 'stock_actual',
                                'prom_vta_sem', 'cobertura_actual', 'a_reponer', 'cob_post_rep', 'stock_cd',
                                'pct_descuento', 'margen_efectivo', 'urgencia'] if c in _rp_v.columns]
        _rp_disp = _rp_v[_rp_cols].rename(columns={
            'marca': 'Marca', 'sku': 'SKU', 'nombre': 'Producto', 'categoria': 'Línea', 'edad': 'Edad (sem)',
            'tienda': 'Tienda', 'stock_actual': 'Stock', 'prom_vta_sem': 'Vta/sem', 'cobertura_actual': 'Cob (sem)',
            'a_reponer': 'A reponer (uds)', 'cob_post_rep': 'Cob post', 'stock_cd': 'Stock CD',
            'pct_descuento': 'Dscto', 'margen_efectivo': 'Margen efect. %', 'urgencia': 'Urgencia'})
        st.dataframe(_rp_disp.style.format({'Vta/sem': '{:.1f}', 'Cob (sem)': '{:.1f}', 'Cob post': '{:.1f}', 'Dscto': '{:.0%}'}, na_rep="—"),
                     use_container_width=True, hide_index=True, height=440)
        # ── Matriz SKU×Tienda (rescatada de la Reposición clásica, C3) ──
        with st.expander("🗺️ Matriz SKU × Tienda (unidades a reponer)", expanded=False):
            if not df_rep_pivot.empty and 'sku' in df_rep_pivot.columns:
                _mx = df_rep_pivot[df_rep_pivot['sku'].isin(_rp_v['sku'])]
                if not _mx.empty:
                    _mx_tcols = [c for c in _mx.columns if c not in
                                 ('sku', 'nombre', 'categoria', 'marca', 'CD', 'Total Repo')]
                    _rk = _mx[_mx_tcols].sum().sort_values(ascending=False).head(10)
                    st.caption("Top tiendas por unidades a reponer: " +
                               " · ".join(f"{t} ({int(u):,})" for t, u in _rk.items()))
                    st.dataframe(_mx.head(200), use_container_width=True,
                                 hide_index=True, height=320)
                else:
                    st.info("Sin filas en la matriz para este universo.")
            else:
                st.info("La matriz de reposición no está disponible con la base actual.")

        _rp_buf = io.BytesIO()
        with pd.ExcelWriter(_rp_buf, engine='openpyxl') as _w:
            _rp_v[_rp_cols].to_excel(_w, sheet_name=f'Reposicion {_uni_rp}', index=False)
        _rp_buf.seek(0)
        st.download_button(f"📥 Descargar reposición {_uni_rp.lower()} (.xlsx)", data=_rp_buf.getvalue(),
                           file_name=f"Capi_Reposicion_{_uni_rp}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_repo_uni")


elif nav_page == "🔄 Transferencias":
    st.markdown(f'<div class="section-header"><h3>🔄 Transferencias</h3><span class="live-badge">POR MARCA</span></div>', unsafe_allow_html=True)
    _uni_tr = st.radio("Universo", ["Propias", "Terceras", "Todas"], horizontal=True, key="uni_trans")
    _SET_P = {m.upper() for m in agente_terceras.MARCAS_PROPIAS_SET}
    _SET_T = {m.upper() for m in agente_terceras.MARCAS_AGENTE}
    _set_tr = _SET_P if _uni_tr == "Propias" else _SET_T if _uni_tr == "Terceras" else (_SET_P | _SET_T)
    if df_trans.empty or 'sku' not in df_trans.columns:
        st.info("No hay transferencias sugeridas con la base actual.")
    else:
        _s2m = dict(zip(df_cob['sku'], df_cob['marca'].str.upper().str.strip())) if 'sku' in df_cob.columns and 'marca' in df_cob.columns else {}
        _tp = df_trans.copy()
        _tp['_marca'] = _tp['sku'].map(_s2m)
        _tp = _tp[_tp['_marca'].isin(_set_tr)]
        if _tp.empty:
            st.info(f"No hay transferencias sugeridas para marcas {_uni_tr.lower()} con la base actual.")
        else:
            st.caption(f"{len(_tp):,} movimientos · {int(_tp['uds_transferir'].sum()):,} unidades")
            if 'ganancia_esperada' in _tp.columns and _tp['ganancia_esperada'].notna().any():
                _rent = _tp[_tp['ganancia_esperada'] > 0]
                _k1, _k2, _k3 = st.columns(3)
                _k1.metric("Ganancia neta del plan (rentables)", f"S/ {_rent['ganancia_esperada'].sum():,.0f}")
                _k2.metric("Movimientos rentables", f"{len(_rent):,} de {len(_tp):,}")
                _k3.metric("Pérdida evitada (no rentables)",
                           f"S/ {abs(_tp.loc[_tp['ganancia_esperada'] <= 0, 'ganancia_esperada'].sum()):,.0f}")
                if st.toggle("Mostrar solo rentables", value=True, key="solo_rent_uni"):
                    _tp = _rent
            # S8 (2026-09-05, pedido Franco): Marca y Departamento en el descargable +
            # consolidado por tienda con la ganancia que recibe cada una.
            _s2d = dict(zip(df_cob['sku'], df_cob['departamento'])) if 'departamento' in df_cob.columns else {}
            _tp['_departamento'] = _tp['sku'].map(_s2d).fillna('')
            _tp_cols = [c for c in ['_marca', '_departamento', 'categoria', 'sku', 'nombre', 'tienda_origen', 'tienda_destino',
                                    'uds_transferir', 'ganancia_esperada', 'veredicto', 'fuente_velocidad',
                                    'cob_origen_pre', 'cob_destino_pre', 'cob_origen_post', 'cob_destino_post', 'motivo'] if c in _tp.columns]
            _tp_ren = {
                '_marca': 'Marca', '_departamento': 'Departamento', 'categoria': 'Línea', 'sku': 'SKU', 'nombre': 'Producto',
                'tienda_origen': 'Tienda Origen', 'tienda_destino': 'Tienda Destino', 'uds_transferir': 'Uds a Transferir',
                'cob_origen_pre': 'Cob Origen (pre)', 'cob_destino_pre': 'Cob Destino (pre)',
                'cob_origen_post': 'Cob Origen (post)', 'cob_destino_post': 'Cob Destino (post)', 'motivo': 'Motivo',
                'ganancia_esperada': 'Ganancia S/', 'veredicto': 'Veredicto', 'fuente_velocidad': 'Velocidad'}
            _tp_disp = _tp[_tp_cols].rename(columns=_tp_ren)
            st.dataframe(_tp_disp.style.format({'Cob Origen (pre)': '{:.1f}', 'Cob Destino (pre)': '{:.1f}', 'Cob Origen (post)': '{:.1f}', 'Cob Destino (post)': '{:.1f}', 'Ganancia S/': 'S/ {:,.0f}'}, na_rep="—"),
                         use_container_width=True, hide_index=True, height=440)

            # Consolidado por tienda: qué recibe cada tienda y cuánto gana con eso
            _cons_dest, _cons_orig = vistas_excel.consolidar_transferencias_por_tienda(_tp)
            with st.expander(f"🏬 Consolidado por tienda destino — qué recibe cada una ({len(_cons_dest)} tiendas)", expanded=True):
                st.dataframe(_cons_dest.style.format({'Ganancia esperada S/': 'S/ {:,.0f}', 'Uds a recibir': '{:,.0f}',
                                                      'Movimientos': '{:,.0f}', 'SKUs': '{:,.0f}'}, na_rep="—"),
                             use_container_width=True, hide_index=True, height=min(60 + 35 * len(_cons_dest), 420))
            _tp_buf = io.BytesIO()
            with pd.ExcelWriter(_tp_buf, engine='openpyxl') as _w:
                vistas_excel.hoja_transferencias(_w, _tp_disp, _cons_dest, _cons_orig, universo=_uni_tr)
            _tp_buf.seek(0)
            st.download_button(f"📥 Descargar transferencias {_uni_tr.lower()} (.xlsx)", data=_tp_buf.getvalue(),
                               file_name=f"Capi_Transferencias_{_uni_tr}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_trans_uni")


elif nav_page == "💰 Gestión de Precios":
    st.markdown(f'<div class="section-header"><h3>💰 Gestión de Precios</h3><span class="live-badge">POR MARCA</span></div>', unsafe_allow_html=True)
    st.caption("Descuento sugerido por antigüedad (pirámide oficial), con gap vs el descuento vigente.")
    _uni_gp = st.radio("Universo", ["Propias", "Terceras", "Todas"], horizontal=True, key="uni_precios")
    _SET_P = {m.upper() for m in agente_terceras.MARCAS_PROPIAS_SET}
    _SET_T = {m.upper() for m in agente_terceras.MARCAS_AGENTE}
    _set_gp = _SET_P if _uni_gp == "Propias" else _SET_T if _uni_gp == "Terceras" else (_SET_P | _SET_T)
    _gpp = agente_terceras.sugerencias_precio_terceras(df_cob, marcas=_set_gp)
    if _gpp.empty:
        st.info(f"No hay SKUs de marcas {_uni_gp.lower()} para analizar con la base actual.")
    else:
        _gpp_subir = int((_gpp['gap'] >= 0.05).sum())
        _gpp_sobre = int((_gpp['gap'] <= -0.05).sum())
        _gpp_ok = int(len(_gpp) - _gpp_subir - _gpp_sobre)
        _gppc1, _gppc2, _gppc3 = st.columns(3)
        _gppc1.markdown(f"""<div style="background:#FEF2F2; border-radius:12px; padding:14px 18px; border-left:4px solid #DC2626;">
            <div style="font-size:0.75rem; color:var(--capi-text2);">SKUs a subir descuento</div>
            <div style="font-size:1.5rem; font-weight:700; color:#DC2626;">{_gpp_subir:,}</div></div>""", unsafe_allow_html=True)
        _gppc2.markdown(f"""<div style="background:#FFFBEB; border-radius:12px; padding:14px 18px; border-left:4px solid #D97706;">
            <div style="font-size:0.75rem; color:var(--capi-text2);">Sobre-descontados</div>
            <div style="font-size:1.5rem; font-weight:700; color:#D97706;">{_gpp_sobre:,}</div></div>""", unsafe_allow_html=True)
        _gppc3.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:14px 18px; border-left:4px solid #059669;">
            <div style="font-size:0.75rem; color:var(--capi-text2);">Alineados</div>
            <div style="font-size:1.5rem; font-weight:700; color:#059669;">{_gpp_ok:,}</div></div>""", unsafe_allow_html=True)
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        _gpp_filtro = st.radio("Ver", ["Solo a subir descuento", "Todos"], horizontal=True, key="gpp_filtro_uni")
        _gpp_v = _gpp[_gpp['gap'] >= 0.05] if _gpp_filtro.startswith("Solo") else _gpp
        _gpp_sel = st.selectbox("Marca", ["Todas"] + sorted(_gpp_v['marca'].unique().tolist()), key="gpp_marca_uni")
        if _gpp_sel != "Todas":
            _gpp_v = _gpp_v[_gpp_v['marca'] == _gpp_sel]
        _gpp_cols = [c for c in ['marca', 'sku', 'nombre', 'categoria', 'edad', 'dscto_actual', 'dscto_sugerido', 'tipo', 'accion', 'capital'] if c in _gpp_v.columns]
        _gpp_disp = _gpp_v[_gpp_cols].rename(columns={
            'marca': 'Marca', 'sku': 'SKU', 'nombre': 'Producto', 'categoria': 'Línea', 'edad': 'Edad (sem)',
            'dscto_actual': 'Dscto actual', 'dscto_sugerido': 'Dscto sugerido', 'tipo': 'Tipo', 'accion': 'Acción', 'capital': 'Capital S/'})
        st.dataframe(_gpp_disp.head(300).style.format({'Dscto actual': '{:.0%}', 'Dscto sugerido': '{:.0%}', 'Capital S/': 'S/ {:,.0f}'}, na_rep="—"),
                     use_container_width=True, hide_index=True, height=440)
        _gpp_buf = io.BytesIO()
        with pd.ExcelWriter(_gpp_buf, engine='openpyxl') as _w:
            _gpp_v[_gpp_cols].to_excel(_w, sheet_name=f'Precios {_uni_gp}', index=False)
        _gpp_buf.seek(0)
        st.download_button(f"📥 Descargar sugerencias de precio {_uni_gp.lower()} (.xlsx)", data=_gpp_buf.getvalue(),
                           file_name=f"Capi_Precios_{_uni_gp}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_precios_uni")


elif nav_page == "🤝 Agente Terceras":
    st.markdown(f'<div class="section-header"><h3>🤝 Agente Terceras</h3><span class="live-badge">AGENTE</span></div>', unsafe_allow_html=True)
    st.caption("Detecta oportunidades con marcas terceras y redacta el correo al proveedor. "
               "El agente genera un BORRADOR — tú lo revisas y lo envías. Nunca manda nada solo.")
    if not os.environ.get("ANTHROPIC_API_KEY"):
        st.warning("🔑 Falta configurar la API key de Anthropic (secrets del deploy o archivo "
                   ".env). Las tablas de oportunidades funcionan igual; solo la redacción "
                   "del correo con IA está deshabilitada.")

    # Paquete de trabajo: todo el análisis de terceras en un Excel
    st.download_button(
        "📦 Descargar TODO el análisis de Marcas Terceras (.xlsx)",
        data=_build_excel_terceras(df_cob, df_rep, df_trans),
        file_name="Capi_Analisis_Marcas_Terceras.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key="dl_pack_terceras",
        help="4 pestañas: SKUs Críticos · Reposición · Transferencias · Precios",
    )
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    #  REPARTO DEL OTB MENSUAL ENTRE TERCERAS — S10 / NF1 (2026-09-05)
    #  Frame Franco: maximizar venta, bajar cobertura, bajar obsoleto, negociar.
    #  Cobertura alta castiga salvo destallado (curva rota).
    # ══════════════════════════════════════════════════════════
    with st.expander("💼 Reparto del OTB mensual entre marcas terceras", expanded=False):
        st.caption("Hoy repartes por venta a costo y cobertura. Capi parte de la misma base (venta a costo 4 semanas) y la ajusta "
                   "por cobertura vs objetivo, % de stock con más de 6 meses, margen contable y quiebres. Si el reporte de stock "
                   "por variación está cargado en 🧵 Talla y Color, una marca destallada no pierde OTB por cobertura alta.")
        _o1, _o2, _o3 = st.columns([1.2, 1, 1])
        _otb_total = _o1.number_input("OTB total del mes (S/ a costo)", min_value=0.0, value=500000.0, step=10000.0, key="otb_total")
        _otb_cob = _o2.number_input("Cobertura objetivo (semanas)", 4.0, 30.0, 13.0, 1.0, key="otb_cob")
        _otb_cap = _o2.slider("Tope de cambio por marca (±%)", 0, 100, 40, 5, key="otb_cap",
                              help="Evita que una sola marca absorba lo que pierden las demás. 0 = sin tope.")
        _otb_uni = _o3.multiselect("Marcas", sorted(agente_terceras.MARCAS_AGENTE),
                                   default=sorted(agente_terceras.MARCAS_AGENTE - agente_terceras.MARCAS_PROPIAS_SET), key="otb_marcas")
        with st.expander("Pesos de los factores (1 = peso completo, 0 = ignorar)", expanded=False):
            _w1, _w2, _w3, _w4 = st.columns(4)
            _pesos = {"cobertura": _w1.slider("Cobertura", 0.0, 1.5, 1.0, 0.1, key="otb_w_cob"),
                      "obsoleto": _w2.slider("Obsoleto (>6 m)", 0.0, 1.5, 1.0, 0.1, key="otb_w_obs"),
                      "margen": _w3.slider("Margen", 0.0, 1.5, 0.5, 0.1, key="otb_w_mg"),
                      "quiebre": _w4.slider("Quiebres", 0.0, 1.5, 0.5, 0.1, key="otb_w_q")}
        # destallado desde el reporte de variación cargado en la sesión (si existe y trae marca)
        _dest = None
        for _k, _v in list(st.session_state.items()):
            if isinstance(_k, str) and _k.startswith("_tc_") and isinstance(_v, pd.DataFrame) and "marca" in _v.columns:
                try:
                    import stock_variacion as _svo
                    _cr = _svo.curva_rota_por_tienda(_v)
                    _cr = _cr.merge(_v.drop_duplicates("cod_modelo")[["cod_modelo", "marca"]], on="cod_modelo", how="left")
                    _dest = _cr.groupby(_cr["marca"].astype(str).str.upper())["curva_rota"].mean().to_dict()
                except Exception:
                    _dest = None
                break
        if _otb_uni and _otb_total > 0:
            _met = otb_terceras.metricas_marca(df_cob, set(_otb_uni), destallado=_dest)
            _rep = otb_terceras.repartir_otb(_met, _otb_total, pesos=_pesos, cob_objetivo=_otb_cob, max_delta_pct=(_otb_cap or None))
            if _rep.empty:
                st.info("Sin datos de esas marcas en la base cargada.")
            else:
                st.caption(("Destallado calculado desde el reporte de variación cargado." if _dest else
                            "Sin reporte de variación cargado: el chequeo de destallado no aplica (cobertura alta castiga siempre)."))
                _rd = _rep[["marca", "venta_costo_4sem", "cobertura_sem", "margen_pct", "pct_obsoleto", "pct_skus_quiebre", "pct_curva_rota",
                            "reparto_venta", "reparto_capi", "reparto_capi_min", "reparto_capi_max", "delta_pct", "por_que"]].rename(columns={
                    "marca": "Marca", "venta_costo_4sem": "Venta a costo 4 sem", "cobertura_sem": "Cob (sem)", "margen_pct": "Margen",
                    "pct_obsoleto": "% >6 meses", "pct_skus_quiebre": "% quiebre", "pct_curva_rota": "% curva rota",
                    "reparto_venta": "Reparto por venta (hoy)", "reparto_capi": "Reparto Capi", "reparto_capi_min": "Capi mín",
                    "reparto_capi_max": "Capi máx", "delta_pct": "Δ %", "por_que": "Por qué"})
                st.dataframe(_rd.style.format({"Venta a costo 4 sem": "S/ {:,.0f}", "Cob (sem)": "{:.0f}", "Margen": "{:.0%}",
                                               "% >6 meses": "{:.0%}", "% quiebre": "{:.0%}", "% curva rota": "{:.0%}",
                                               "Reparto por venta (hoy)": "S/ {:,.0f}", "Reparto Capi": "S/ {:,.0f}",
                                               "Capi mín": "S/ {:,.0f}", "Capi máx": "S/ {:,.0f}", "Δ %": "{:+.0f}%"}, na_rep="—"),
                             use_container_width=True, hide_index=True, height=min(60 + 35 * len(_rd), 460))
                st.caption("Reparto Capi en banda ±10%: la venta de 4 semanas y el margen contable son estimaciones del mes. "
                           "Los factores están acotados (cobertura 0.4–1.3×, obsoleto 0.5–1×, margen 0.8–1.2×, quiebre 1–1.3×).")
                st.markdown("**Argumento para cada marca**")
                for _r in _rep.itertuples(index=False):
                    st.markdown(f"- **{_r.marca}** → S/ {_r.reparto_capi:,.0f} ({_r.delta_pct:+.0f}% vs por venta): {_r.argumento_negociacion}")
                _ob = io.BytesIO()
                with pd.ExcelWriter(_ob, engine="openpyxl") as _wo:
                    vistas_excel._tabla_con_titulo(_wo, "Reparto OTB", f"Reparto del OTB S/ {_otb_total:,.0f} entre terceras — base venta a costo 4 sem × factores",
                                                   _rd, {"Venta a costo 4 sem": "#,##0", "Reparto por venta (hoy)": "#,##0", "Reparto Capi": "#,##0",
                                                         "Capi mín": "#,##0", "Capi máx": "#,##0", "Margen": "0%", "% >6 meses": "0%", "% quiebre": "0%",
                                                         "% curva rota": "0%", "Cob (sem)": "0", "Δ %": "0"})
                    _rep.to_excel(_wo, sheet_name="Detalle factores", index=False)
                _ob.seek(0)
                st.download_button("📥 Excel — reparto OTB con factores y argumentos", _ob.getvalue(), file_name="Capi_Reparto_OTB_Terceras.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="otb_dl")

    _at_prov = agente_terceras.cargar_proveedores()

    _at_tipo = st.radio(
        "Tipo de oportunidad",
        ["💰 Capital parado (rebate / markdown support)", "📦 Quiebre (reorder urgente)"],
        horizontal=True, key="at_tipo",
    )

    if _at_tipo.startswith("💰"):
        _at_op = agente_terceras.detectar_capital_parado(df_cob)
        if _at_op.empty:
            st.info("No hay marcas terceras con capital parado sobre el umbral.")
        else:
            st.markdown(f"**{len(_at_op)} marcas terceras con capital inmovilizado y baja rotación:**")
            _at_show = _at_op[['marca', 'capital', 'n_skus', 'cob_prom', 'sell_through', 'margen_efectivo']].rename(columns={
                'marca': 'Marca', 'capital': 'Capital S/', 'n_skus': 'SKUs',
                'cob_prom': 'Cob (sem)', 'sell_through': 'Sell-through %', 'margen_efectivo': 'Margen %',
            })
            st.dataframe(_at_show.style.format({
                'Capital S/': 'S/ {:,.0f}', 'Cob (sem)': '{:.0f}', 'Sell-through %': '{:.0f}%', 'Margen %': '{:.0f}%',
            }, na_rep="—"), use_container_width=True, hide_index=True)

            _at_marcas_op = _at_op['marca'].tolist()
            _at_sel = st.selectbox("Marca para generar el correo", _at_marcas_op, key="at_sel_cap")
            _at_row = _at_op[_at_op['marca'] == _at_sel].iloc[0]
            _at_prov_marca = _at_prov.get(_at_sel.upper())
            if _at_prov_marca and _at_prov_marca.get('contacto'):
                st.caption(f"Destinatario sugerido: {_at_prov_marca.get('contacto','')} ({_at_prov_marca.get('empresa','')})")
            if st.button("✍️ Generar texto del correo", key="at_gen_cap", type="primary"):
                with st.spinner("Redactando con IA..."):
                    try:
                        _at_skus = agente_terceras.top_skus_marca(df_cob, _at_sel)
                        _at_det = agente_terceras.top5_por_marca_linea(df_cob)
                        if not _at_det.empty:
                            _at_det = _at_det[_at_det['marca'].str.upper() == _at_sel.upper()]
                        _at_correo = agente_terceras.generar_correo_capital_parado(
                            _at_row, _at_prov_marca, _at_skus, detalle_lineas=_at_det)
                        st.session_state["at_borrador"] = _at_correo
                    except Exception as _at_e:
                        st.error(f"No se pudo generar el texto: {_at_e}")

    else:
        _at_op = agente_terceras.detectar_quiebre_tercera(df_cob)
        if _at_op.empty:
            st.info("No hay marcas terceras en quiebre con venta relevante.")
        else:
            st.markdown(f"**{len(_at_op)} marcas terceras con quiebres que vendían bien:**")
            _at_show = _at_op.rename(columns={
                'marca': 'Marca', 'n_skus_quiebre': 'SKUs en quiebre',
                'venta_riesgo_sem': 'Venta en riesgo S//sem', 'vta_sem_uds': 'Vta/sem (uds)',
            })
            st.dataframe(_at_show.style.format({
                'Venta en riesgo S//sem': 'S/ {:,.0f}', 'Vta/sem (uds)': '{:.0f}',
            }), use_container_width=True, hide_index=True)

            _at_sel = st.selectbox("Marca para generar el correo", _at_op['marca'].tolist(), key="at_sel_q")
            _at_row = _at_op[_at_op['marca'] == _at_sel].iloc[0]
            _at_prov_marca = _at_prov.get(_at_sel.upper())
            if _at_prov_marca and _at_prov_marca.get('contacto'):
                st.caption(f"Destinatario sugerido: {_at_prov_marca.get('contacto','')} ({_at_prov_marca.get('empresa','')})")
            if st.button("✍️ Generar texto del correo", key="at_gen_q", type="primary"):
                with st.spinner("Redactando con IA..."):
                    try:
                        _at_correo = agente_terceras.generar_correo_reorder(_at_row, _at_prov_marca)
                        st.session_state["at_borrador"] = _at_correo
                    except Exception as _at_e:
                        st.error(f"No se pudo generar el texto: {_at_e}")

    # ── Borrador generado: revisar / editar / copiar ──
    _at_bor = st.session_state.get("at_borrador")
    if _at_bor:
        st.markdown("---")
        st.markdown(f"##### ✉️ Texto para {_at_bor.get('marca','')} — revisar y copiar")
        st.text_input("Asunto", value=_at_bor.get("asunto", ""), key="at_asunto")
        st.text_area("Cuerpo", value=_at_bor.get("cuerpo", ""), height=320, key="at_cuerpo")
        st.caption("Revisa y ajusta el texto, luego cópialo a tu correo de Ripley para enviarlo al proveedor. "
                   "El agente solo redacta — el envío lo haces tú.")
        _atb1, _atb2 = st.columns(2)
        if _atb1.button("📨 Ya lo envié → registrar acción", key="at_enviado", type="primary"):
            acciones_log.agregar(
                "", "Negociación Terceras", str(_at_bor.get("marca", "")).upper(),
                f"Correo al proveedor enviado: {_at_bor.get('asunto', '')[:120]}",
                origen="Sugerida por Capi", estado="Ejecutada")
            del st.session_state["at_borrador"]
            st.success("Acción registrada en el log ✅ — cuenta para el Caso de Éxito.")
        if _atb2.button("🗑️ Descartar", key="at_descartar"):
            del st.session_state["at_borrador"]
            st.rerun()


# ─── TAB 2: Gestión por Antigüedad ─────────────────────────────────
#  Ventana de Mercadería completa (4 capas) + Obsolescencia detallada
#  Sección dedicada con tabs

elif nav_page == "📊 Gestión por Antigüedad":
    st.markdown(f'<div class="section-header"><h3>📊 Gestión por Antigüedad</h3><span class="live-badge">MERCADERÍA</span></div>', unsafe_allow_html=True)
    st.caption("Análisis completo del envejecimiento del inventario — Ventana de Mercadería + Obsoletos detallados")

    # ── Resumen de 1 pantalla ANTES del detalle (Franco C1 2026-08-26:
    # "avisar rápido el problema" — el drill queda en los tabs) ──
    _ag_obs = df_cob[df_cob["rango_antiguedad"].isin({"RANGO 6_9", "RANGO 9_12", "RANGO 12_99"})] if "rango_antiguedad" in df_cob.columns else pd.DataFrame()
    if not _ag_obs.empty:
        _ag_cap = _ag_obs["stock_valor_costo"].sum()
        _ag_tot = df_cob["stock_valor_costo"].sum()
        _ag_top = _ag_obs.groupby("marca")["stock_valor_costo"].sum().sort_values(ascending=False)
        _agk1, _agk2, _agk3 = st.columns(3)
        _agk1.metric("Capital >6 meses", f"S/ {_ag_cap/1e6:,.2f}M",
                     delta=f"{_ag_cap/_ag_tot*100:.1f}% del total", delta_color="off")
        _agk2.metric("Marca más cargada", _ag_top.index[0],
                     delta=f"S/ {_ag_top.iloc[0]/1e6:,.2f}M", delta_color="off")
        _agk3.metric("SKUs >6 meses", f"{_ag_obs['sku'].nunique():,}")
        st.caption("👇 El detalle (ventana de mercadería y obsolescencia por marca/modelo) vive en los tabs.")

    _aging_tab1, _aging_tab2 = st.tabs(["🪟 Ventana de Mercadería", "⏳ Obsolescencia Detallada"])

    # ══════════════════════════════════════════════════════════
    #  TAB 1 — VENTANA DE MERCADERÍA (4 capas)
    # ══════════════════════════════════════════════════════════
    with _aging_tab1:
        if df_aging.empty:
            st.info("ℹ️ No hay datos de aging disponibles. Sube archivos con información de antigüedad.")
        else:
            import math as _math_mod  # necesario: cada elif es scope independiente

            # ────────────────────────────────────────────────────
            #  CAPA 1 — Resumen ejecutivo (KPIs + barra horizontal)
            # ────────────────────────────────────────────────────
            _ak = aging_kpis
            _cap_viejo = _ak.get('capital_viejo', 0)
            _pct_viejo = _ak.get('pct_viejo', 0)
            _edad_prom = _ak.get('edad_prom_pond', 0)
            if _edad_prom is None or (isinstance(_edad_prom, float) and _math_mod.isnan(_edad_prom)):
                _edad_prom = 0
            _n_riesgo = _ak.get('n_zona_riesgo', 0)
            _cap_total = _ak.get('capital_total', 1)

            _kc1, _kc2, _kc3 = st.columns(3)
            with _kc1:
                st.markdown(f"""<div style="background:#FEF2F2; border-radius:12px; padding:16px 20px; border-left:4px solid #ef4444;">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Capital en mercadería vieja (>16 sem)</div>
                    <div style="font-size:1.8rem; font-weight:700; color:#ef4444;">S/ {_cap_viejo:,.0f}</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">{_pct_viejo:.0f}% del inventario total</div>
                </div>""", unsafe_allow_html=True)
            with _kc2:
                st.markdown(f"""<div style="background:#FFFBEB; border-radius:12px; padding:16px 20px; border-left:4px solid #f59e0b;">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Edad promedio ponderada</div>
                    <div style="font-size:1.8rem; font-weight:700; color:#f59e0b;">{_edad_prom:.1f} sem</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">Ponderada por capital invertido</div>
                </div>""", unsafe_allow_html=True)
            with _kc3:
                st.markdown(f"""<div style="background:#F0FDF4; border-radius:12px; padding:16px 20px; border-left:4px solid {TEAL_600};">
                    <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">SKUs en zona de riesgo (8-16 sem)</div>
                    <div style="font-size:1.8rem; font-weight:700; color:{TEAL_700};">{_n_riesgo:,}</div>
                    <div style="font-size:0.7rem; color:var(--capi-text2);">Aún rescatables con empuje a piso</div>
                </div>""", unsafe_allow_html=True)

            # Barra horizontal de distribución de capital por edad
            _dist_edad = _ak.get('dist_edad', {})
            if _dist_edad and _cap_total > 0:
                _edad_colors = {'0-4 sem': '#10b981', '4-8 sem': '#84cc16', '8-16 sem': '#f59e0b', '16-26 sem': '#f97316', '26+ sem': '#ef4444'}
                _bar_parts = ""
                _legend_parts = ""
                for _lbl in ['0-4 sem', '4-8 sem', '8-16 sem', '16-26 sem', '26+ sem']:
                    _val = _dist_edad.get(_lbl, 0)
                    _pct = _val / _cap_total * 100 if _cap_total > 0 else 0
                    _clr = _edad_colors.get(_lbl, '#94A3B8')
                    if _pct > 1:
                        _bar_parts += f'<div style="width:{_pct:.1f}%; background:{_clr}; height:100%; display:inline-block;" title="{_lbl}: S/{_val:,.0f} ({_pct:.0f}%)"></div>'
                    _legend_parts += f'<span style="display:inline-flex; align-items:center; gap:4px; margin-right:12px;"><span style="width:10px; height:10px; border-radius:2px; background:{_clr}; display:inline-block;"></span><span style="font-size:0.7rem; color:var(--capi-text2);">{_lbl}</span></span>'
                st.markdown(f"""<div style="margin-top:16px; background:var(--capi-bg-card); border-radius:8px; padding:12px 16px; border:1px solid var(--capi-border);">
                    <div style="font-size:0.8rem; font-weight:600; color:var(--capi-text); margin-bottom:8px;">Distribución de capital por edad</div>
                    <div style="width:100%; height:20px; border-radius:6px; overflow:hidden; background:var(--capi-border); display:flex;">{_bar_parts}</div>
                    <div style="margin-top:6px;">{_legend_parts}</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

            # ────────────────────────────────────────────────────
            #  CAPA 2 — Stacked bar por categoría (click → drill a marca)
            # ────────────────────────────────────────────────────
            st.markdown(f"<h4 style='color:{SLATE_800}; margin:0 0 4px 0;'>Capital por categoría y antigüedad</h4>", unsafe_allow_html=True)
            st.caption("Cada barra = una categoría. Colores = rango de edad. Mientras más rojo, más vieja la mercadería.")

            # Filtros
            _aging_fcols = st.columns(4)
            with _aging_fcols[0]:
                _temps_aging = ["Todas"] + sorted(df_aging['temporada'].dropna().unique().tolist()) if 'temporada' in df_aging.columns else ["Todas"]
                _f_temp_aging = st.selectbox("Temporada", _temps_aging, key="aging_inv_temp")
            with _aging_fcols[1]:
                _f_tipo_marca = st.selectbox("Tipo marca", ["Todas", "Marca Tercera", "Marca Propia"], key="aging_inv_tipo_marca")
            with _aging_fcols[2]:
                _marcas_aging = ["Todas"] + sorted(df_aging['marca'].dropna().unique().tolist()) if 'marca' in df_aging.columns else ["Todas"]
                _f_marca_aging = st.selectbox("Marca", _marcas_aging, key="aging_inv_marca")
            with _aging_fcols[3]:
                _f_rango_dscto = st.selectbox("Descuento", ["Todos", "Sin descuento", "1-20%", "20-40%", "40%+"], key="aging_inv_dscto")

            # Aplicar filtros
            _df_ag = df_aging.copy()
            if _f_temp_aging != "Todas" and 'temporada' in _df_ag.columns:
                _df_ag = _df_ag[_df_ag['temporada'] == _f_temp_aging]
            if _f_tipo_marca != "Todas" and 'tipo_marca' in _df_ag.columns:
                _df_ag = _df_ag[_df_ag['tipo_marca'] == _f_tipo_marca]
            if _f_marca_aging != "Todas" and 'marca' in _df_ag.columns:
                _df_ag = _df_ag[_df_ag['marca'] == _f_marca_aging]
            if _f_rango_dscto != "Todos" and 'rango_descuento' in _df_ag.columns:
                _df_ag = _df_ag[_df_ag['rango_descuento'] == _f_rango_dscto]

            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

            if 'categoria' in _df_ag.columns and 'rango_edad_aging' in _df_ag.columns:
                _chart_data = _df_ag.groupby(
                    ['categoria', 'rango_edad_aging'], observed=True
                ).agg(capital=('stock_valor_costo', 'sum')).reset_index()
                _chart_data['rango_edad_aging'] = _chart_data['rango_edad_aging'].astype(str)

                _pivot = _chart_data.pivot_table(index='categoria', columns='rango_edad_aging', values='capital', fill_value=0)
                _orden_rangos = [r for r in ['0-4 sem', '4-8 sem', '8-16 sem', '16-26 sem', '26+ sem'] if r in _pivot.columns]
                _pivot = _pivot[_orden_rangos]
                _pivot['_total'] = _pivot.sum(axis=1)
                _pivot = _pivot.sort_values('_total', ascending=False).drop(columns='_total').head(8)

                _colores_rango = {'0-4 sem': '#10b981', '4-8 sem': '#84cc16', '8-16 sem': '#f59e0b', '16-26 sem': '#f97316', '26+ sem': '#ef4444'}

                # Plotly stacked bar con tooltips formateados
                _totales_cat = _pivot.sum(axis=1)
                _fig_aging = go.Figure()
                for _rango in _pivot.columns:
                    _vals = _pivot[_rango]
                    _pcts = (_vals / _totales_cat * 100).round(1)
                    _custom = np.column_stack([_pcts.values])
                    _fig_aging.add_trace(go.Bar(
                        name=_rango,
                        x=_pivot.index,
                        y=_vals,
                        marker_color=_colores_rango.get(_rango, '#94A3B8'),
                        customdata=_custom,
                        hovertemplate=(
                            '<b>%{x}</b><br>'
                            f'<b>{_rango}</b><br>'
                            'Capital: S/ %{y:,.0f}<br>'
                            'Peso: %{customdata[0]:.1f}%'
                            '<extra></extra>'
                        ),
                    ))
                _fig_aging.update_layout(
                    barmode='stack',
                    height=380,
                    margin=dict(l=0, r=0, t=10, b=0),
                    legend=dict(orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5,
                                font=dict(size=12)),
                    xaxis=dict(tickangle=-45, tickfont=dict(size=11)),
                    yaxis=dict(tickformat=',', tickfont=dict(size=11)),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    hoverlabel=dict(bgcolor='white', font_size=13, font_color='#1e293b', bordercolor='#e2e8f0'),
                )
                st.plotly_chart(_fig_aging, use_container_width=True, config={'displayModeBar': False})

                # Drill-down: seleccionar categoría para ver marcas
                _cats_list = _pivot.index.tolist()
                _sel_cat = st.selectbox("Drill-down: seleccionar categoría", ["—"] + _cats_list, key="aging_inv_cat_drill")
                if _sel_cat != "—":
                    _df_cat = _df_ag[_df_ag['categoria'] == _sel_cat]
                    _drill_marca = _df_cat.groupby('marca').agg(
                        capital=('stock_valor_costo', 'sum'),
                        n_skus=('sku', 'nunique'),
                        edad_prom=('edad_semanas', 'mean'),
                    ).reset_index().sort_values('capital', ascending=False).head(8)

                    st.markdown(f"<div style='background:var(--capi-bg-card); border:1px solid var(--capi-border); border-radius:8px; padding:16px; margin-top:8px;'>", unsafe_allow_html=True)
                    st.markdown(f"<div style='font-weight:600; color:{SLATE_800}; margin-bottom:8px;'>Drill-down: {_sel_cat}</div>", unsafe_allow_html=True)
                    st.caption("Top marcas por capital viejo (>16 sem)")

                    _max_cap_drill = _drill_marca['capital'].max() if not _drill_marca.empty else 1
                    for _, _dr in _drill_marca.iterrows():
                        _bar_w_d = max(5, int(_dr['capital'] / _max_cap_drill * 100))
                        _dr_color = '#ef4444' if _dr['edad_prom'] > 26 else ('#f97316' if _dr['edad_prom'] > 16 else ('#f59e0b' if _dr['edad_prom'] > 8 else '#10b981'))
                        st.markdown(f"""<div style="display:flex; align-items:center; gap:8px; margin-bottom:4px;">
                            <span style="width:120px; font-size:0.8rem; font-weight:500; color:var(--capi-text); white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">{_dr['marca']}</span>
                            <div style="flex:1; background:var(--capi-border); border-radius:4px; height:16px;">
                                <div style="background:{_dr_color}; border-radius:4px; height:16px; width:{_bar_w_d}%;"></div>
                            </div>
                            <span style="font-size:0.75rem; color:var(--capi-text2); white-space:nowrap;">S/{_dr['capital']:,.0f}</span>
                        </div>""", unsafe_allow_html=True)

                    # Nivel 3: click marca → ver SKUs
                    _marcas_drill = ["—"] + _drill_marca['marca'].tolist()
                    _sel_marca_drill = st.selectbox("Click marca → ver SKUs", _marcas_drill, key="aging_inv_marca_drill")
                    if _sel_marca_drill != "—":
                        _df_sku_drill = _df_cat[_df_cat['marca'] == _sel_marca_drill].sort_values('stock_valor_costo', ascending=False).head(10)
                        _sku_rows = ""
                        for _, _sr in _df_sku_drill.iterrows():
                            _sku_rows += f"""<tr>
                                <td style="padding:6px 8px; font-size:0.75rem;">{str(_sr.get('nombre',''))[:35]}</td>
                                <td style="padding:6px 8px; font-size:0.75rem;">{_sr.get('tienda','')}</td>
                                <td style="padding:6px 8px; font-size:0.75rem; text-align:right;">{int(_sr.get('edad_semanas',0))} sem</td>
                                <td style="padding:6px 8px; font-size:0.75rem; text-align:right;">S/{_sr.get('stock_valor_costo',0):,.0f}</td>
                                <td style="padding:6px 8px; font-size:0.75rem; text-align:center;">{_sr.get('accion_aging','')}</td>
                            </tr>"""
                        st.markdown(f"""<table style="width:100%; border-collapse:collapse; margin-top:8px;">
                            <thead><tr style="background:var(--capi-bg-surface); border-bottom:1px solid var(--capi-border);">
                                <th style="padding:6px 8px; text-align:left; font-size:0.7rem;">Modelo</th>
                                <th style="padding:6px 8px; text-align:left; font-size:0.7rem;">Tienda</th>
                                <th style="padding:6px 8px; text-align:right; font-size:0.7rem;">Edad</th>
                                <th style="padding:6px 8px; text-align:right; font-size:0.7rem;">Capital</th>
                                <th style="padding:6px 8px; text-align:center; font-size:0.7rem;">Acción</th>
                            </tr></thead>
                            <tbody>{_sku_rows}</tbody>
                        </table>""", unsafe_allow_html=True)
                    st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

            # ────────────────────────────────────────────────────
            #  CAPA 3 — Alertas inteligentes por tipo de acción
            # ────────────────────────────────────────────────────
            st.markdown(f"<h4 style='color:{SLATE_800}; margin:0 0 4px 0;'>Alertas inteligentes por tipo de acción</h4>", unsafe_allow_html=True)
            st.caption("El motor clasifica cada SKU según edad + sell-through + descuento → sugiere la acción correcta.")

            _acc_configs = [
                {'key': 'EMPUJE', 'title': 'EMPUJE A PISO', 'icon': '🟢', 'color': '#84cc16', 'bg': '#F7FEE7', 'border': '#84cc16',
                 'criteria': 'Edad 8-16 sem · ST >5% · Sin dscto',
                 'n_label': f"{_ak.get('n_skus_empuje', 0)} SKUs · S/{_ak.get('capital_empuje', 0):,.0f}",
                 'accion': 'Acción: instruir tiendas'},
                {'key': 'MARKDOWN', 'title': 'MARKDOWN PROGRESIVO', 'icon': '🟡', 'color': '#f59e0b', 'bg': '#FFFBEB', 'border': '#f59e0b',
                 'criteria': 'Edad 16-26 sem · ST 2-5% · Dscto <40%',
                 'n_label': f"{_ak.get('n_skus_markdown', 0)} SKUs · S/{_ak.get('capital_markdown', 0):,.0f}",
                 'accion': 'Acción: ajustar precio'},
                {'key': 'NEGOCIAR', 'title': 'NEGOCIAR PROVEEDOR', 'icon': '🟠', 'color': '#f97316', 'bg': '#FFF7ED', 'border': '#f97316',
                 'criteria': 'Marca tercera · >16 sem · Capital >S/50K',
                 'n_label': f"{_ak.get('n_marcas_negociar', 0)} marcas · S/{_ak.get('capital_negociar', 0):,.0f}",
                 'accion': 'Acción: reunión proveedor'},
                {'key': 'LIQUIDAR', 'title': 'LIQUIDAR', 'icon': '🔴', 'color': '#ef4444', 'bg': '#FEF2F2', 'border': '#ef4444',
                 'criteria': 'Edad >26 sem · ST <2% · Ya con dscto',
                 'n_label': f"{_ak.get('n_skus_liquidar', 0)} SKUs · S/{_ak.get('capital_liquidar', 0):,.0f}",
                 'accion': 'Acción: sacar del sistema'},
            ]

            _acc_c1, _acc_c2, _acc_c3, _acc_c4 = st.columns(4)
            for _col_ui, _cfg in zip([_acc_c1, _acc_c2, _acc_c3, _acc_c4], _acc_configs):
                with _col_ui:
                    _ejemplos_acc = aging_top_ejemplos.get(_cfg['key'], [])
                    _ej_html = ""
                    for _ej in _ejemplos_acc[:2]:
                        _ej_html += f"""<div style="margin-bottom:8px; padding:6px 8px; background:var(--capi-bg-card); border-radius:6px; border:1px solid var(--capi-border);">
                            <div style="font-size:0.72rem; font-weight:600; color:{SLATE_800};">• {_ej['nombre'][:45]}</div>
                            <div style="font-size:0.68rem; color:var(--capi-text2); margin-top:2px;">{_ej['detalle']}</div>
                            <div style="font-size:0.68rem; color:{_cfg['color']}; margin-top:2px;">→ {_ej['sugerencia'][:60]}</div>
                        </div>"""
                    if not _ej_html:
                        _ej_html = f'<div style="font-size:0.72rem; color:var(--capi-text3); padding:8px;">Sin casos detectados</div>'

                    st.markdown(f"""<div style="background:{_cfg['bg']}; border:2px solid {_cfg['border']}; border-radius:12px; padding:14px; height:100%;">
                        <div style="font-weight:700; font-size:0.85rem; color:{_cfg['color']}; margin-bottom:2px;">{_cfg['icon']} {_cfg['title']}</div>
                        <div style="font-size:0.68rem; color:var(--capi-text2); margin-bottom:10px;">{_cfg['criteria']}</div>
                        {_ej_html}
                        <div style="background:{_cfg['color']}; color:white; padding:6px 10px; border-radius:6px; font-size:0.72rem; font-weight:600; text-align:center; margin-top:8px;">{_cfg['n_label']}</div>
                        <div style="font-size:0.68rem; color:var(--capi-text2); text-align:center; margin-top:4px;">{_cfg['accion']}</div>
                    </div>""", unsafe_allow_html=True)

            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

            # ────────────────────────────────────────────────────
            #  CAPA 4 — Reglas del motor de clasificación
            # ────────────────────────────────────────────────────
            st.markdown(f"<h4 style='color:{SLATE_800}; margin:0 0 4px 0;'>Reglas del motor de clasificación</h4>", unsafe_allow_html=True)
            st.caption("Cada SKU se clasifica automáticamente según esta matriz de decisión:")

            _reglas_html = f"""<div style="overflow-x:auto;">
            <table style="width:100%; border-collapse:collapse; font-size:0.8rem;">
                <thead>
                    <tr style="background:var(--capi-bg-surface); border-bottom:2px solid var(--capi-border);">
                        <th style="padding:8px 12px; text-align:left; color:var(--capi-text);">Criterio</th>
                        <th style="padding:8px 12px; text-align:center; color:#84cc16;">🟢 Empuje</th>
                        <th style="padding:8px 12px; text-align:center; color:#f59e0b;">🟡 Markdown</th>
                        <th style="padding:8px 12px; text-align:center; color:#f97316;">🟠 Negociar</th>
                        <th style="padding:8px 12px; text-align:center; color:#ef4444;">🔴 Liquidar</th>
                    </tr>
                </thead>
                <tbody>
                    <tr style="border-bottom:1px solid var(--capi-border);">
                        <td style="padding:8px 12px; font-weight:500;">Edad (sem)</td>
                        <td style="padding:8px 12px; text-align:center;">8 – 16</td>
                        <td style="padding:8px 12px; text-align:center;">16 – 26</td>
                        <td style="padding:8px 12px; text-align:center;">16+ (terceras)</td>
                        <td style="padding:8px 12px; text-align:center;">>26</td>
                    </tr>
                    <tr style="border-bottom:1px solid var(--capi-border);">
                        <td style="padding:8px 12px; font-weight:500;">Sell-through</td>
                        <td style="padding:8px 12px; text-align:center;">>5%</td>
                        <td style="padding:8px 12px; text-align:center;">2% – 5%</td>
                        <td style="padding:8px 12px; text-align:center;">cualquiera</td>
                        <td style="padding:8px 12px; text-align:center;"><2%</td>
                    </tr>
                    <tr style="border-bottom:1px solid var(--capi-border);">
                        <td style="padding:8px 12px; font-weight:500;">Descuento actual</td>
                        <td style="padding:8px 12px; text-align:center;"><10% o ninguno</td>
                        <td style="padding:8px 12px; text-align:center;"><40%</td>
                        <td style="padding:8px 12px; text-align:center;">cualquiera</td>
                        <td style="padding:8px 12px; text-align:center;">ya con 30%+</td>
                    </tr>
                    <tr style="border-bottom:1px solid var(--capi-border);">
                        <td style="padding:8px 12px; font-weight:500;">Tipo marca</td>
                        <td style="padding:8px 12px; text-align:center;">cualquiera</td>
                        <td style="padding:8px 12px; text-align:center;">cualquiera</td>
                        <td style="padding:8px 12px; text-align:center;">solo terceras</td>
                        <td style="padding:8px 12px; text-align:center;">cualquiera</td>
                    </tr>
                    <tr style="border-bottom:1px solid var(--capi-border);">
                        <td style="padding:8px 12px; font-weight:500;">Capital mín.</td>
                        <td style="padding:8px 12px; text-align:center;">S/10K por SKU</td>
                        <td style="padding:8px 12px; text-align:center;">S/5K por SKU</td>
                        <td style="padding:8px 12px; text-align:center;">S/50K por marca</td>
                        <td style="padding:8px 12px; text-align:center;">cualquiera</td>
                    </tr>
                    <tr>
                        <td style="padding:8px 12px; font-weight:700;">Output</td>
                        <td style="padding:8px 12px; text-align:center; font-weight:600;">WhatsApp a tienda</td>
                        <td style="padding:8px 12px; text-align:center; font-weight:600;">Sugerencia de precio</td>
                        <td style="padding:8px 12px; text-align:center; font-weight:600;">Brief para reunión</td>
                        <td style="padding:8px 12px; text-align:center; font-weight:600;">Lista de corte</td>
                    </tr>
                </tbody>
            </table></div>"""
            st.markdown(_reglas_html, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    #  TAB 2 — OBSOLESCENCIA DETALLADA (por modelo + drill tienda)
    # ══════════════════════════════════════════════════════════
    with _aging_tab2:
        st.markdown(f"<h4 style='color:{SLATE_800}; margin:0 0 4px 0;'>Mercadería Obsoleta y Preobsoleta</h4>", unsafe_allow_html=True)
        st.caption("Productos con más de 6 meses de antigüedad. Analiza por modelo para tomar acciones de liquidación.")

        st.markdown("---")

        _RANGOS_OBS_DASH = {"RANGO 6_9", "RANGO 9_12", "RANGO 12_99"}

        if "rango_antiguedad" in df_cob.columns:
            df_obs = df_cob[df_cob["rango_antiguedad"].isin(_RANGOS_OBS_DASH)].copy()

            if not df_obs.empty:
                st.markdown(f"""<div style="background:#FDF4FF; border-left:4px solid {STATUS_LIQUIDAR}; padding:10px 14px; border-radius:10px; margin-bottom:10px;">
                <strong style="color:{STATUS_LIQUIDAR};">Inventario Obsoleto (>6 meses)</strong>
                <span style="color:var(--capi-text2); font-size:0.82em;"> &nbsp;·&nbsp; {len(df_obs):,} combos · S/ {df_obs['stock_valor_costo'].sum():,.0f} en capital</span>
                </div>""", unsafe_allow_html=True)

                _obs_marcas = ["Todas"] + sorted(df_obs["marca"].unique().tolist()) if "marca" in df_obs.columns else ["Todas"]
                _obs_marca_sel = st.selectbox("Filtrar por Marca", _obs_marcas, index=0, key="aging_obs_marca_filter")
                if _obs_marca_sel != "Todas":
                    df_obs = df_obs[df_obs["marca"] == _obs_marca_sel]

                # ══════════════════════════════════════════════════════
                #  S5 (2026-09-05, pedido Franco FR8): ranking por tienda,
                #  alerta "por entrar a obsoleto" y delta vs semana anterior
                # ══════════════════════════════════════════════════════
                import obsoletos as _obsm
                _df_cob_obs = df_cob if _obs_marca_sel == "Todas" else df_cob[df_cob["marca"] == _obs_marca_sel]
                _terc_set = {m.upper() for m in agente_terceras.MARCAS_AGENTE}
                # Definición Franco 2026-09-05: pre-obsoleto = 6–9 meses, obsoleto = 9 meses a más (por antigüedad)
                _def_key = "rango"
                _rk = _obsm.ranking_por_tienda(_df_cob_obs, definicion=_def_key, marcas_terceras=_terc_set)
                _tot_pre = float(_rk["capital_preobsoleto_6_9m"].sum()) if not _rk.empty else 0.0
                _tot_obs = float(_rk["capital_obsoleto_9m_mas"].sum()) if not _rk.empty else 0.0
                st.markdown(f"<h5 style='margin:8px 0 2px 0;'>🏬 Pre-obsoleto y obsoleto por tienda</h5>", unsafe_allow_html=True)
                st.caption(f"**Pre-obsoleto (6–9 meses): S/ {_tot_pre:,.0f}** · **Obsoleto (9 meses a más): S/ {_tot_obs:,.0f}** · "
                           f"total más de 6 meses: S/ {_tot_pre + _tot_obs:,.0f}. Ordenado por el total. "
                           "Para terceras se muestra también el capital a costo implícito (precio sin IGV × (1 − margen contable)), "
                           "porque el campo Costo de Ripley subestima su margen ~11.7 pp.")
                if not _rk.empty:
                    _rk_disp = _rk[[c for c in ["tienda", "capital_obsoleto", "capital_preobsoleto_6_9m", "capital_obsoleto_9m_mas",
                                                "uds_obsoletas", "skus_obsoletos", "capital_tienda", "pct_stock_tienda",
                                                "capital_implicito", "marca_top"] if c in _rk.columns]].rename(columns={
                        "tienda": "Tienda", "capital_obsoleto": "Total >6 m S/", "capital_preobsoleto_6_9m": "Pre-obsoleto 6–9 m S/",
                        "capital_obsoleto_9m_mas": "Obsoleto ≥9 m S/", "uds_obsoletas": "Uds", "skus_obsoletos": "SKUs",
                        "capital_tienda": "Capital tienda S/", "pct_stock_tienda": "% del stock de la tienda",
                        "capital_implicito": "Capital implícito terceras S/", "marca_top": "Marca que más pesa"})
                    st.dataframe(_rk_disp.style.format({"Total >6 m S/": "S/ {:,.0f}", "Pre-obsoleto 6–9 m S/": "S/ {:,.0f}",
                                                        "Obsoleto ≥9 m S/": "S/ {:,.0f}", "Uds": "{:,.0f}", "SKUs": "{:,.0f}",
                                                        "Capital tienda S/": "S/ {:,.0f}", "% del stock de la tienda": "{:.1%}",
                                                        "Capital implícito terceras S/": "S/ {:,.0f}"}, na_rep="—")
                                 .apply(lambda col: ["background-color:#FECACA" if (pd.notna(v) and v >= 0.30)
                                                     else ("background-color:#FEF3C7" if (pd.notna(v) and v >= 0.15) else "") for v in col],
                                        subset=["% del stock de la tienda"]),
                                 use_container_width=True, hide_index=True, height=min(60 + 35 * len(_rk_disp), 460))

                # ── Alerta: lo que cruza a obsoleto en N semanas ──
                st.markdown(f"<h5 style='margin:14px 0 2px 0;'>⏳ Por cruzar de nivel</h5>", unsafe_allow_html=True)
                _pe_c0, _pe_c00 = st.columns([1, 2])
                with _pe_c0:
                    _hacia_sel = st.radio("Hacia", ["Pre-obsoleto (llega a 6 meses)", "Obsoleto (llega a 9 meses)"],
                                          horizontal=True, key="obs_hacia")
                with _pe_c00:
                    _n_sem = st.select_slider("Horizonte", options=[2, 3, 4, 6, 8], value=2, key="obs_horizonte",
                                              format_func=lambda x: f"{x} semanas")
                _hacia = "obsoleto" if _hacia_sel.startswith("Obsoleto") else "preobsoleto"
                _pe = _obsm.por_entrar(_df_cob_obs, semanas=_n_sem, definicion=_def_key, hacia=_hacia)
                if _pe.empty:
                    st.success(f"Nada cruza a {_hacia_sel.split(' (')[0].lower()} en las próximas {_n_sem} semanas.")
                else:
                    _pe_res = _obsm.resumen_por_entrar(_pe)
                    _pe_tot = float(_pe["capital"].sum()); _pe_sin = float(_pe_res["capital_sin_dscto"].sum())
                    st.markdown(f"""<div style="background:#FFF7ED; border-left:4px solid #EA580C; padding:10px 14px; border-radius:10px; margin-bottom:8px;">
                    <strong style="color:#EA580C;">S/ {_pe_tot:,.0f}</strong> <span style="color:var(--capi-text2); font-size:0.85em;">a costo cruzan a {_hacia_sel.split(' (')[0].lower()} en {_n_sem} semanas
                    ({len(_pe):,} combos · {int(_pe['stock_total'].sum()):,} uds). De eso, <strong>S/ {_pe_sin:,.0f}</strong> todavía no tiene el descuento de la pirámide: atacarlo ahora, no cuando ya esté congelado.</span></div>""",
                                unsafe_allow_html=True)
                    _pe_c1, _pe_c2 = st.columns([1, 2])
                    with _pe_c1:
                        st.dataframe(_pe_res.rename(columns={"marca": "Marca", "capital": "Capital S/", "uds": "Uds", "skus": "SKUs",
                                                             "capital_sin_dscto": "Sin dscto S/"})
                                     .style.format({"Capital S/": "S/ {:,.0f}", "Sin dscto S/": "S/ {:,.0f}", "Uds": "{:,.0f}", "SKUs": "{:,.0f}"}),
                                     use_container_width=True, hide_index=True, height=min(60 + 35 * len(_pe_res), 320))
                    with _pe_c2:
                        _pe_disp = _pe.rename(columns={"tienda": "Tienda", "marca": "Marca", "sku": "SKU", "nombre": "Producto",
                                                       "edad_semanas": "Edad", "semanas_para_obsoleto": "Cruza en (sem)",
                                                       "stock_total": "Stock", "capital": "Capital S/", "prom_vta_uds": "Vta sem",
                                                       "precio_vigente": "Precio", "pct_descuento": "Dscto actual",
                                                       "dscto_sugerido": "Dscto sugerido", "precio_sugerido": "Precio sugerido", "accion": "Acción"})
                        _pe_cols = [c for c in ["Tienda", "Marca", "SKU", "Producto", "Edad", "Cruza en (sem)", "Stock", "Capital S/",
                                                "Vta sem", "Precio", "Dscto actual", "Dscto sugerido", "Precio sugerido", "Acción"] if c in _pe_disp.columns]
                        st.dataframe(_pe_disp[_pe_cols].head(300).style.format({"Capital S/": "S/ {:,.0f}", "Precio": "S/ {:,.2f}",
                                                                                 "Precio sugerido": "S/ {:,.2f}", "Dscto actual": "{:.0%}",
                                                                                 "Dscto sugerido": "{:.0%}", "Vta sem": "{:.1f}", "Edad": "{:.0f}",
                                                                                 "Cruza en (sem)": "{:.0f}"}, na_rep="—"),
                                     use_container_width=True, hide_index=True, height=320)
                    _pe_buf = io.BytesIO()
                    with pd.ExcelWriter(_pe_buf, engine="openpyxl") as _wpe:
                        vistas_excel._tabla_con_titulo(_wpe, "Por entrar a obsoleto",
                                                       f"Mercadería que cruza a {_hacia_sel.split(' (')[0].lower()} en {_n_sem} semanas — atacar con descuento ANTES del cruce",
                                                       _pe_disp[_pe_cols], {"Capital S/": "#,##0", "Precio": "#,##0.00", "Precio sugerido": "#,##0.00",
                                                                            "Dscto actual": "0%", "Dscto sugerido": "0%"})
                        if not _rk.empty:
                            vistas_excel._tabla_con_titulo(_wpe, "Obsoleto por tienda", "Pre-obsoleto (6–9 m) y obsoleto (≥9 m) por tienda",
                                                           _rk_disp, {"Total >6 m S/": "#,##0", "Pre-obsoleto 6–9 m S/": "#,##0", "Obsoleto ≥9 m S/": "#,##0",
                                                                      "Capital tienda S/": "#,##0", "% del stock de la tienda": "0.0%", "Capital implícito terceras S/": "#,##0"})
                    _pe_buf.seek(0)
                    st.download_button("📥 Excel — por entrar a obsoleto + ranking por tienda", _pe_buf.getvalue(),
                                       file_name=f"Capi_Obsoletos_por_entrar_{_n_sem}sem.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_obs_pe")

                # ── Delta vs semana anterior (snapshots, nivel cadena, MUERTO por marca) ──
                if _HAS_SNAPSHOTS:
                    try:
                        _wk = snapshots_engine.list_available_weeks()
                        if len(_wk) >= 2:
                            _dm = _obsm.delta_marca(_wk[-2], _wk[-1])
                            if not _dm.empty:
                                with st.expander(f"📉 Capital MUERTO (sin venta >26 sem, taxonomía) por marca: {_wk[-1]} vs {_wk[-2]}", expanded=False):
                                    st.dataframe(_dm.rename(columns={"marca": "Marca", "delta": "Δ S/", "delta_pct": "Δ %"})
                                                 .style.format({_wk[-2]: "S/ {:,.0f}", _wk[-1]: "S/ {:,.0f}", "Δ S/": "S/ {:+,.0f}", "Δ %": "{:+.1f}%"}, na_rep="—"),
                                                 use_container_width=True, hide_index=True, height=min(60 + 35 * len(_dm), 400))
                                    st.caption("Nivel cadena (el snapshot semanal aún no tiene tienda). El delta por tienda llega con el snapshot liviano SKU×tienda.")
                    except Exception as _e_dm:
                        st.caption(f"Delta semanal no disponible: {_e_dm}")

                st.markdown("---")
                obs_c1, obs_c2 = st.columns(2)

                # ── DONUT: % de inventario obsoleto por rango ──
                with obs_c1:
                    obs_rango = df_obs.groupby("rango_antiguedad").agg(
                        capital=("stock_valor_costo", "sum"),
                        n_combos=("sku", "count"),
                    ).reset_index()
                    obs_rango_order = {"RANGO 6_9": 0, "RANGO 9_12": 1, "RANGO 12_99": 2}
                    obs_rango = obs_rango.sort_values("rango_antiguedad", key=lambda x: x.map(obs_rango_order))

                    _obs_colors = {"RANGO 6_9": "#F59E0B", "RANGO 9_12": "#F97316", "RANGO 12_99": "#EF4444"}
                    _obs_labels = {"RANGO 6_9": "6-9 meses", "RANGO 9_12": "9-12 meses", "RANGO 12_99": ">12 meses"}
                    obs_rango["label"] = obs_rango["rango_antiguedad"].map(_obs_labels)

                    fig_obs = go.Figure(data=[go.Pie(
                        labels=obs_rango["label"],
                        values=obs_rango["capital"],
                        hole=0.55,
                        marker=dict(colors=[_obs_colors.get(r, "#CBD5E1") for r in obs_rango["rango_antiguedad"]]),
                        textinfo="label+percent",
                        textfont=dict(size=11),
                        hovertemplate="<b>%{label}</b><br>S/ %{value:,.0f}<br>%{percent}<extra></extra>",
                    )])
                    fig_obs.update_layout(
                        **_plotly_layout,
                        title=dict(text="Capital Obsoleto por Antigüedad", font=dict(size=14, color=TH_TEXT_PY)),
                        showlegend=False,
                        height=340,
                    )
                    _obs_total = obs_rango["capital"].sum()
                    _obs_pct_total = (_obs_total / df_cob["stock_valor_costo"].sum() * 100) if df_cob["stock_valor_costo"].sum() > 0 else 0
                    fig_obs.add_annotation(
                        text=f"<b>{_obs_pct_total:.1f}%</b><br><span style='font-size:10px;color:var(--capi-text2)'>del inventario</span>",
                        showarrow=False, font=dict(size=18, color=TH_TEXT_PY),
                    )
                    st.plotly_chart(fig_obs, use_container_width=True)

                # ── Tabla: stock, venta, cobertura, dscto por marca ──
                with obs_c2:
                    st.markdown(f"**Métricas por Marca** — mercadería >6 meses")
                    _obs_marca = df_obs.groupby("marca" if "marca" in df_obs.columns else "categoria").agg(
                        stock_uds=("stock_total", "sum"),
                        capital=("stock_valor_costo", "sum"),
                        vta_sem=("prom_vta_uds", "sum"),
                        n_skus=("sku", "nunique"),
                    ).reset_index()
                    _grp_label = "marca" if "marca" in df_obs.columns else "categoria"

                    # Calcular venta a costo
                    if "marca" in df_obs.columns and "costo" in df_obs.columns:
                        _obs_vta_costo = df_obs.groupby("marca").apply(
                            lambda g: (g["prom_vta_uds"] * g["costo"]).sum()
                        ).reset_index(name="vta_sem_costo")
                        _obs_marca = _obs_marca.merge(_obs_vta_costo, on="marca", how="left")
                    else:
                        _obs_marca["vta_sem_costo"] = 0

                    _obs_marca["cobertura"] = _obs_marca.apply(
                        lambda r: round(r["stock_uds"] / r["vta_sem"], 1) if r["vta_sem"] > 0 else None, axis=1
                    )

                    # Descuento promedio por marca
                    if "pct_descuento" in df_obs.columns:
                        _obs_dscto = df_obs.groupby(_grp_label)["pct_descuento"].mean().reset_index(name="dscto_prom")
                        _obs_marca = _obs_marca.merge(_obs_dscto, on=_grp_label, how="left")
                    else:
                        _obs_marca["dscto_prom"] = None

                    _obs_marca = _obs_marca.sort_values("capital", ascending=False)
                    _obs_disp = _obs_marca.rename(columns={
                        _grp_label: "Marca", "capital": "Capital S/", "stock_uds": "Stock Uds",
                        "vta_sem_costo": "Vta Sem S/", "cobertura": "Cob (sem)", "dscto_prom": "Dscto Prom",
                        "n_skus": "SKUs",
                    })
                    st.dataframe(
                        _obs_disp[["Marca", "Capital S/", "Stock Uds", "Vta Sem S/", "Cob (sem)", "Dscto Prom", "SKUs"]].style.format({
                            "Capital S/": "S/ {:,.0f}", "Stock Uds": "{:,.0f}",
                            "Vta Sem S/": "S/ {:,.0f}", "Cob (sem)": "{:.1f}", "Dscto Prom": "{:.0%}",
                        }, na_rep="—"),
                        use_container_width=True, hide_index=True, height=300,
                    )

                    # Expander: detalle por marca
                    with st.expander("Ver detalle por marca", expanded=False):
                        _sel_marca_obs = st.selectbox(
                            "Marca", sorted(df_obs[_grp_label].unique().tolist()), key="aging_obs_marca_sel"
                        )
                        _obs_det = df_obs[df_obs[_grp_label] == _sel_marca_obs].copy()
                        _obs_det_grp = _obs_det.groupby("rango_antiguedad").agg(
                            stock_uds=("stock_total", "sum"),
                            capital=("stock_valor_costo", "sum"),
                            n_skus=("sku", "nunique"),
                        ).reset_index()
                        if "pct_descuento" in _obs_det.columns:
                            _obs_det_dscto = _obs_det.groupby("rango_antiguedad")["pct_descuento"].mean().reset_index(name="dscto_prom")
                            _obs_det_grp = _obs_det_grp.merge(_obs_det_dscto, on="rango_antiguedad", how="left")
                        else:
                            _obs_det_grp["dscto_prom"] = None
                        _obs_det_grp["label"] = _obs_det_grp["rango_antiguedad"].map(_obs_labels).fillna(_obs_det_grp["rango_antiguedad"])
                        _obs_det_grp = _obs_det_grp.rename(columns={
                            "label": "Rango", "capital": "Capital S/", "stock_uds": "Stock Uds",
                            "dscto_prom": "Dscto Prom", "n_skus": "SKUs",
                        })
                        st.dataframe(
                            _obs_det_grp[["Rango", "Capital S/", "Stock Uds", "Dscto Prom", "SKUs"]].style.format({
                                "Capital S/": "S/ {:,.0f}", "Stock Uds": "{:,.0f}", "Dscto Prom": "{:.0%}",
                            }, na_rep="—"),
                            use_container_width=True, hide_index=True,
                        )

                    # ── Botón de descarga: detalle completo de obsoletos ──
                    _obs_dl_cols = ["sku", "nombre", "marca", "categoria", "temporada",
                                    "tienda", "rango_antiguedad", "stock_total",
                                    "stock_valor_costo", "prom_vta_uds", "cobertura_sem",
                                    "edad_semanas"]
                    if "pct_descuento" in df_obs.columns:
                        _obs_dl_cols.append("pct_descuento")
                    _obs_dl_cols = [c for c in _obs_dl_cols if c in df_obs.columns]
                    _obs_dl = df_obs[_obs_dl_cols].copy()
                    _obs_dl = _obs_dl.sort_values(["marca", "stock_valor_costo"], ascending=[True, False])
                    _obs_buf = io.BytesIO()
                    with pd.ExcelWriter(_obs_buf, engine="openpyxl") as _w_obs:
                        # Resumen rango × marca primero (auditoría formato 2026-08-05)
                        vistas_excel.hoja_resumen_obsoletos(_w_obs, df_obs)
                        _add_pricing_cols(_obs_dl, df_cob, 'Obsoletos', _w_obs)
                    _obs_buf.seek(0)
                    st.download_button(
                        "📥 Descargar detalle de obsoletos (.xlsx)",
                        data=_obs_buf.getvalue(),
                        file_name="detalle_obsoletos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )


        st.markdown("---")
        st.markdown("##### Detalle por Modelo")


        _OBS_RANGOS = {"RANGO 6_9", "RANGO 9_12", "RANGO 12_99"}
        _OBS_LABELS = {"RANGO 6_9": "Pre-obsoleto (6-9m)", "RANGO 9_12": "Obsoleto (9-12m)", "RANGO 12_99": "Crítico (>12m)"}

        if "rango_antiguedad" not in df_cob.columns:
            st.warning("No hay columna de rango de antigüedad en los datos. Sube un archivo con esta información.")
        else:
            df_obs_tab = df_cob[df_cob["rango_antiguedad"].isin(_OBS_RANGOS)].copy()

            if df_obs_tab.empty:
                st.success("✅ No hay mercadería obsoleta en el inventario actual.")
            else:
                # KPIs rápidos
                _obs_k1, _obs_k2, _obs_k3, _obs_k4 = st.columns(4)
                _obs_k1.metric("SKUs Obsoletos", f"{df_obs_tab['sku'].nunique():,}")
                _obs_k2.metric("Capital Parado", f"S/ {df_obs_tab['stock_valor_costo'].sum():,.0f}")
                _obs_k3.metric("Stock (uds)", f"{int(df_obs_tab['stock_total'].sum()):,}")
                _obs_pct_inv = (df_obs_tab['stock_valor_costo'].sum() / df_cob['stock_valor_costo'].sum() * 100) if df_cob['stock_valor_costo'].sum() > 0 else 0
                _obs_k4.metric("% del Inventario", f"{_obs_pct_inv:.1f}%")

                st.markdown("---")

                # Filtros
                _obs_f1, _obs_f2, _obs_f3 = st.columns(3)
                with _obs_f1:
                    _obs_rango_opts = sorted(df_obs_tab["rango_antiguedad"].unique().tolist(),
                                              key=lambda x: {"RANGO 6_9": 0, "RANGO 9_12": 1, "RANGO 12_99": 2}.get(x, 99))
                    _f_obs_rango = st.multiselect("Rango", _obs_rango_opts, default=_obs_rango_opts, key="obs_tab_rango",
                                                   format_func=lambda x: _OBS_LABELS.get(x, x))
                with _obs_f2:
                    if "marca" in df_obs_tab.columns:
                        _obs_marcas = ["Todas"] + sorted(df_obs_tab["marca"].dropna().unique().tolist())
                        _f_obs_marca = st.selectbox("Marca", _obs_marcas, key="obs_tab_marca")
                    else:
                        _f_obs_marca = "Todas"
                with _obs_f3:
                    _obs_sort = st.selectbox("Ordenar por", ["Capital S/", "Stock Uds", "Cobertura (sem)", "Dscto Prom"], key="obs_tab_sort")

                # Aplicar filtros
                df_obs_f = df_obs_tab.copy()
                if _f_obs_rango:
                    df_obs_f = df_obs_f[df_obs_f["rango_antiguedad"].isin(_f_obs_rango)]
                if _f_obs_marca != "Todas" and "marca" in df_obs_f.columns:
                    df_obs_f = df_obs_f[df_obs_f["marca"] == _f_obs_marca]

                if df_obs_f.empty:
                    st.info("Sin resultados con los filtros seleccionados.")
                else:
                    # Agrupar por modelo (sku)
                    _grp_col = "marca" if "marca" in df_obs_f.columns else "categoria"

                    _obs_modelo = df_obs_f.groupby(["sku", "nombre"]).agg(
                        marca=(_grp_col, "first"),
                        rango=("rango_antiguedad", "first"),
                        stock_uds=("stock_total", "sum"),
                        capital=("stock_valor_costo", "sum"),
                        vta_sem=("prom_vta_uds", "sum"),
                        n_tiendas=("tienda", "nunique"),
                    ).reset_index()

                    # Venta a costo
                    if "costo" in df_obs_f.columns:
                        _obs_vta_costo = df_obs_f.groupby("sku").apply(
                            lambda g: (g["prom_vta_uds"] * g["costo"]).sum()
                        ).reset_index(name="vta_costo")
                        _obs_modelo = _obs_modelo.merge(_obs_vta_costo, on="sku", how="left")
                    else:
                        _obs_modelo["vta_costo"] = 0

                    _obs_modelo["cobertura"] = _obs_modelo.apply(
                        lambda r: round(r["stock_uds"] / r["vta_sem"], 1) if r["vta_sem"] > 0 else None, axis=1
                    )

                    # Descuento promedio
                    if "pct_descuento" in df_obs_f.columns:
                        _obs_dscto = df_obs_f.groupby("sku")["pct_descuento"].mean().reset_index(name="dscto_prom")
                        _obs_modelo = _obs_modelo.merge(_obs_dscto, on="sku", how="left")
                    else:
                        _obs_modelo["dscto_prom"] = None

                    _obs_modelo["rango_label"] = _obs_modelo["rango"].map(_OBS_LABELS).fillna(_obs_modelo["rango"])

                    # Ordenar
                    _sort_map = {"Capital S/": "capital", "Stock Uds": "stock_uds",
                                 "Cobertura (sem)": "cobertura", "Dscto Prom": "dscto_prom"}
                    _sort_col = _sort_map.get(_obs_sort, "capital")
                    _obs_modelo = _obs_modelo.sort_values(_sort_col, ascending=False, na_position="last")

                    # Tabla resumen
                    _obs_disp = _obs_modelo.rename(columns={
                        "sku": "SKU", "nombre": "Nombre", "marca": "Marca",
                        "rango_label": "Antigüedad", "stock_uds": "Stock Uds",
                        "capital": "Capital S/", "vta_costo": "Vta Sem S/",
                        "cobertura": "Cob (sem)", "dscto_prom": "Dscto Prom",
                        "n_tiendas": "Tiendas",
                    })

                    st.dataframe(
                        _obs_disp[["SKU", "Nombre", "Marca", "Antigüedad", "Stock Uds", "Capital S/",
                                   "Vta Sem S/", "Cob (sem)", "Dscto Prom", "Tiendas"]].style.format({
                            "Stock Uds": "{:,.0f}", "Capital S/": "S/ {:,.0f}",
                            "Vta Sem S/": "S/ {:,.0f}", "Cob (sem)": "{:.1f}",
                            "Dscto Prom": "{:.0%}",
                        }, na_rep="—"),
                        use_container_width=True, hide_index=True, height=400,
                    )

                    st.caption(f"{len(_obs_modelo)} modelos obsoletos mostrados")

                    # Expander por modelo → detalle tienda
                    st.markdown("---")
                    st.caption("Selecciona un modelo para ver el desglose por tienda:")

                    _obs_sku_sel = st.selectbox(
                        "Modelo",
                        _obs_modelo["sku"].tolist(),
                        format_func=lambda x: f"{x} — {_obs_modelo[_obs_modelo['sku']==x]['nombre'].iloc[0][:40]}",
                        key="obs_sku_sel",
                    )

                    if _obs_sku_sel:
                        _obs_det = df_obs_f[df_obs_f["sku"] == _obs_sku_sel].copy()
                        _obs_det_cols = ["tienda", "stock_total", "stock_valor_costo", "prom_vta_uds",
                                         "cobertura_sem", "estado"]
                        if "pct_descuento" in _obs_det.columns:
                            _obs_det_cols.append("pct_descuento")
                        _obs_det_disp = _obs_det[_obs_det_cols].rename(columns={
                            "tienda": "Tienda", "stock_total": "Stock", "stock_valor_costo": "Capital S/",
                            "prom_vta_uds": "Vta Sem (uds)", "cobertura_sem": "Cob (sem)",
                            "estado": "Estado", "pct_descuento": "Dscto",
                        }).sort_values("Capital S/", ascending=False)

                        _det_fmt = {"Capital S/": "S/ {:,.0f}", "Vta Sem (uds)": "{:.1f}", "Cob (sem)": "{:.1f}"}
                        if "Dscto" in _obs_det_disp.columns:
                            _det_fmt["Dscto"] = "{:.0%}"

                        st.dataframe(
                            _obs_det_disp.style.format(_det_fmt, na_rep="—"),
                            use_container_width=True, hide_index=True,
                        )
# ─── Afinidad Producto × Plaza ────────────────────────────────

elif nav_page == "🎯 Match Producto-Plaza":
    import glob as _glob_mod
    import io as _io_af
    from afinidad_engine import build_afinidad
    from transformar_profundidad import STORE_NAMES as _STORE_NAMES_AF

    st.markdown(f'<h2 style="color:{SLATE_900};margin-bottom:4px;">🎯 Match Producto-Plaza</h2>', unsafe_allow_html=True)
    st.caption("¿Qué producto pongo en qué tienda? Empujes blindados con 4 filtros + destino final del stock mal ubicado. Análisis avanzado al fondo.")

    # Detectar base más reciente — priorizar la última subida por el usuario
    # Fix B8 (2026-08-23): solo la base subida en esta sesión — el fallback a
    # data2/bases antiguas/ analizaba silenciosamente una base distinta.
    _base_path_af = st.session_state.get("_base_profundidad_path")
    if _base_path_af and os.path.exists(_base_path_af):
        _base_name_af = os.path.basename(_base_path_af)
    else:
        _base_path_af = None
    if _base_path_af is None:
        st.warning("Esta vista necesita la Base Profundidad ORIGINAL de esta sesión. "
                   "Sube tu base (no la plantilla) y ejecuta el análisis primero.")
    else:
        st.info(f"📂 Base: **{_base_name_af}**")

        with st.spinner("Analizando afinidad producto × plaza..."):
            try:
                _af_result = build_afinidad(_base_path_af)
            except Exception as _e_af:
                st.error(f"Error en análisis de afinidad: {_e_af}")
                _af_result = None

        if _af_result is not None:
            _rm = _af_result['rotation_matrix']
            _cl = _af_result['clusters_df']
            _an = _af_result['anomalias_df']
            _emp = _af_result['empujes_df']
            _red = _af_result['redistribucion_df']
            _prod = _af_result['produccion_df']
            _tiendas_af = _af_result['tiendas_activas']

            # KPI cards
            _n_empujes = len(_emp)
            _u_empujes = int(_emp['unidades_sugeridas'].sum()) if _n_empujes > 0 else 0
            _n_redist = len(_red)
            _n_anomalias = len(_an)
            _n_prod = len(_prod)

            _kpi_html_af = f"""<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px;">
              <div style="background:{SLATE_100};border-radius:8px;padding:14px 16px;">
                <div style="font-size:12px;color:{SLATE_500};">Empujes CD→Tiendas</div>
                <div style="font-size:22px;font-weight:600;color:{TEAL_600};">{_n_empujes:,}</div>
                <div style="font-size:11px;color:{SLATE_400};">{_u_empujes:,} unidades</div>
              </div>
              <div style="background:{SLATE_100};border-radius:8px;padding:14px 16px;">
                <div style="font-size:12px;color:{SLATE_500};">Redistribuciones</div>
                <div style="font-size:22px;font-weight:600;color:{TEAL_600};">{_n_redist:,}</div>
                <div style="font-size:11px;color:{SLATE_400};">entre tiendas</div>
              </div>
              <div style="background:{SLATE_100};border-radius:8px;padding:14px 16px;">
                <div style="font-size:12px;color:{SLATE_500};">Anomalías</div>
                <div style="font-size:22px;font-weight:600;color:#B45309;">{_n_anomalias:,}</div>
                <div style="font-size:11px;color:{SLATE_400};">producto malo + mal match</div>
              </div>
              <div style="background:{SLATE_100};border-radius:8px;padding:14px 16px;">
                <div style="font-size:12px;color:{SLATE_500};">Señales Producción</div>
                <div style="font-size:22px;font-weight:600;color:#059669;">{_n_prod}</div>
                <div style="font-size:11px;color:{SLATE_400};">líneas con demanda insatisfecha</div>
              </div>
            </div>"""
            st.markdown(_kpi_html_af, unsafe_allow_html=True)

            # Tabs
            _tab_emp, _tab_mm, _tab_adv = st.tabs([
                "🚀 Empujes CD → Tienda", "🏷️ Mal match → destino final",
                "📚 Análisis avanzado"
            ])
            with _tab_adv:
                st.caption("Material de análisis estratégico — no requiere acción semanal.")
                _tab_hm, _tab_cl, _tab_red, _tab_an, _tab_prod = st.tabs([
                    "🗺️ Heatmap Rotación", "🔗 Clusters", "🔄 Redistribución",
                    "⚠️ Anomalías", "🏭 Producción"
                ])

            # ── TAB: Mal match → destino final (Fase B, plan Demo Chile) ──
            with _tab_mm:
                _mm_res = None
                try:
                    from afinidad_engine import mal_match_destino as _mmd
                    _mm_res = _mmd(_an, df_trans, df_cob)
                except Exception as _e_mm:
                    st.error(f"No se pudo cruzar mal match con transferencias: {_e_mm}")
                if _mm_res:
                    _mm_cu, _mm_hu = _mm_res["cubiertos"], _mm_res["huerfanos"]
                    _mk1, _mk2, _mk3 = st.columns(3)
                    _mk1.metric("SKUs mal ubicados", f"{len(_mm_cu) + len(_mm_hu):,}")
                    _mk2.metric("✅ Cubiertos por transferencias",
                                f"{len(_mm_cu):,}",
                                help=f"Ganancia esperada del plan: S/ {_mm_cu['ganancia'].sum():,.0f}"
                                if not _mm_cu.empty else None)
                    _mk3.metric("🏷️ Sin destino rentable",
                                f"S/ {_mm_hu['capital_parado'].sum():,.0f}"
                                if not _mm_hu.empty else "S/ 0",
                                help="Capital parado en tiendas donde el SKU no rota y "
                                     "ninguna transferencia paga el flete → acción localizada")
                    if not _mm_hu.empty:
                        st.markdown("**El stock que nadie va a rescatar** — decisión EN la tienda, "
                                    "no en la cadena:")
                        _mm_disp = _mm_hu.rename(columns={
                            "sku": "SKU", "nombre": "Producto", "marca": "Marca",
                            "n_tiendas_muertas": "Tiendas muertas",
                            "tiendas_muertas": "¿Dónde?", "stock_parado": "Stock (uds)",
                            "capital_parado": "Capital S/", "edad_semanas": "Edad (sem)",
                            "dscto_actual": "Dscto actual", "dscto_sugerido": "Dscto sugerido",
                            "accion": "Acción"})
                        _mm_fmt = {"Capital S/": "S/ {:,.0f}", "Stock (uds)": "{:,.0f}",
                                   "Edad (sem)": "{:.0f}", "Dscto sugerido": "{:.0%}"}
                        if "Dscto actual" in _mm_disp.columns:
                            _mm_fmt["Dscto actual"] = "{:.0%}"
                        st.dataframe(_mm_disp.style.format(_mm_fmt, na_rep="—"),
                                     use_container_width=True, hide_index=True, height=380)
                        _mm_buf = _io_af.BytesIO()
                        with pd.ExcelWriter(_mm_buf, engine="openpyxl") as _wmm:
                            _mm_disp.to_excel(_wmm, sheet_name="Liquidacion Localizada", index=False)
                            if not _mm_cu.empty:
                                _mm_cu.to_excel(_wmm, sheet_name="Cubiertos x Transferencia", index=False)
                        _mm_buf.seek(0)
                        st.download_button("📥 Descargar plan de destino final (.xlsx)",
                                           _mm_buf.getvalue(),
                                           file_name="Capi_Mal_Match_Destino_Final.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="dl_mal_match")
                    if not _mm_cu.empty:
                        with st.expander(f"Ver los {len(_mm_cu)} SKUs cubiertos por el plan de "
                                         f"transferencias (S/ {_mm_cu['ganancia'].sum():,.0f})"):
                            st.dataframe(_mm_cu.rename(columns={
                                "sku": "SKU", "nombre": "Producto", "movimientos": "Movs",
                                "uds": "Uds", "ganancia": "Ganancia S/"}).style.format(
                                {"Ganancia S/": "S/ {:,.0f}"}),
                                use_container_width=True, hide_index=True, height=250)

            # ── TAB 1: Heatmap ──
            with _tab_hm:
                import plotly.graph_objects as _go_af
                _rm_display = _rm.copy()
                _rm_display.index = [l.title() for l in _rm_display.index]
                _rm_display.columns = [_STORE_NAMES_AF.get(c, c) for c in _rm_display.columns]

                _fig_hm = _go_af.Figure(data=_go_af.Heatmap(
                    z=_rm_display.values,
                    x=_rm_display.columns.tolist(),
                    y=_rm_display.index.tolist(),
                    colorscale=[[0, '#FEE2E2'], [0.25, '#FECACA'], [0.5, '#FDE68A'], [0.75, '#A7F3D0'], [1, '#059669']],
                    text=[[f"{v*100:.1f}%" for v in row] for row in _rm_display.values],
                    texttemplate="%{text}",
                    textfont={"size": 10},
                    hovertemplate="Línea: %{y}<br>Tienda: %{x}<br>Rotación: %{z:.2%}<extra></extra>",
                    colorbar=dict(title=dict(text="Rotación %", side="right"), tickformat=".0%"),
                ))
                _fig_hm.update_layout(
                    height=500, margin=dict(l=120, r=20, t=30, b=80),
                    xaxis=dict(tickangle=-45, tickfont=dict(size=10)),
                    yaxis=dict(tickfont=dict(size=11)),
                    plot_bgcolor='white',
                )
                st.plotly_chart(_fig_hm, use_container_width=True)
                st.caption("Rotación = Vta / Stk. Verde = alta rotación (oportunidad). Rojo = baja rotación (capital parado).")

            # ── TAB 2: Clusters ──
            with _tab_cl:
                for _cid in sorted(_cl['cluster_id'].unique()):
                    _sub_cl = _cl[_cl['cluster_id'] == _cid]
                    _nombre_cl = _sub_cl['cluster_nombre'].iloc[0]
                    _n_tiendas_cl = len(_sub_cl)
                    _vta_cl = _sub_cl['vta_total'].sum()

                    _tiendas_list = ", ".join([
                        f"{_STORE_NAMES_AF.get(t, t)}" for t in _sub_cl['tienda'].tolist()
                    ])

                    _perfiles_af = _af_result['perfiles']
                    _perfil_cl = _perfiles_af.get(_cid, {})
                    _top3_lineas = sorted(_perfil_cl.items(), key=lambda x: -x[1])[:3]
                    _top3_str = ", ".join([f"{l.title()} ({r:.3f})" for l, r in _top3_lineas])

                    st.markdown(f"""<div style="background:{SLATE_100};border-radius:8px;padding:14px 18px;margin-bottom:10px;">
                      <div style="font-weight:600;color:{SLATE_900};font-size:14px;">{_nombre_cl}</div>
                      <div style="color:{SLATE_500};font-size:12px;margin:4px 0;">{_n_tiendas_cl} tiendas — Vta total: {_vta_cl:,}u</div>
                      <div style="color:{SLATE_700};font-size:13px;">{_tiendas_list}</div>
                      <div style="color:{TEAL_600};font-size:12px;margin-top:4px;">Top líneas: {_top3_str}</div>
                    </div>""", unsafe_allow_html=True)

            # ── TAB 3: Empujes CD→Tiendas ──
            with _tab_emp:
                if len(_emp) == 0:
                    st.info("No se encontraron oportunidades de empuje con los umbrales actuales.")
                else:
                    _col_filt1, _col_filt2 = st.columns(2)
                    with _col_filt1:
                        _marca_filt_emp = st.selectbox("Marca", ["Todas"] + sorted(_emp['marca'].unique().tolist()), key="af_emp_marca")
                    with _col_filt2:
                        _tienda_filt_emp = st.selectbox("Tienda destino", ["Todas"] + sorted(_emp['tienda'].unique().tolist()), key="af_emp_tienda")

                    _emp_show = _emp.copy()
                    if _marca_filt_emp != "Todas":
                        _emp_show = _emp_show[_emp_show['marca'] == _marca_filt_emp]
                    if _tienda_filt_emp != "Todas":
                        _emp_show = _emp_show[_emp_show['tienda'] == _tienda_filt_emp]

                    st.caption("❄️ Clima: líneas abrigadoras (GRUESO) no van a tiendas de calor. "
                               "💰 Margen: no se empuja donde la línea rota con margen realizado < 25%. "
                               "🏷️ Descuento (regla Majo): SKUs con dscto ≥40% no se empujan (es "
                               "liquidación aunque el margen aguante), ni hacia tiendas donde la línea "
                               "vendió con dscto promedio ≥40%. Umbrales editables en "
                               "config_afinidad.json / config_clima_tiendas.json.")
                    _emp_show['tienda_nombre'] = _emp_show['tienda'].map(lambda t: _STORE_NAMES_AF.get(t, t))
                    _emp_cols = ['marca', 'descripcion', 'tienda_nombre', 'stk_actual_tienda',
                                 'stock_cd', 'rotacion_linea_tienda']
                    # Nuevas columnas de cobertura (si existen en el output del motor)
                    _has_cob = 'vta_semanal_est' in _emp_show.columns and 'target_stock' in _emp_show.columns
                    if _has_cob:
                        _emp_cols += ['vta_semanal_est', 'target_stock']
                    _has_mg = 'margen_destino_pct' in _emp_show.columns
                    if _has_mg:
                        _emp_cols += ['margen_destino_pct']
                    _has_dd = 'dscto_destino_pct' in _emp_show.columns
                    if _has_dd:
                        _emp_cols += ['dscto_destino_pct']
                    _has_fill = 'es_llenado_inicial' in _emp_show.columns
                    _emp_cols += ['unidades_sugeridas']
                    if _has_fill:
                        _emp_cols += ['es_llenado_inicial']
                    _emp_cols += ['es_marca_propia']
                    _emp_display = _emp_show[_emp_cols].head(100).copy()
                    _emp_headers = ['Marca', 'Descripción', 'Tienda', 'Stk Tienda', 'Stk CD', 'Rot. %']
                    if _has_cob:
                        _emp_headers += ['Vta/Sem Est', 'Target 12s']
                    if _has_mg:
                        _emp_headers += ['Margen dest. %']
                    if _has_dd:
                        _emp_headers += ['Dscto dest. %']
                    _emp_headers += ['Empujar']
                    if _has_fill:
                        _emp_headers += ['Llenado inicial']
                    _emp_headers += ['Propia']
                    _emp_display.columns = _emp_headers
                    _emp_display['Rot. %'] = (_emp_display['Rot. %'] * 100).round(1)
                    if _has_fill:
                        _emp_display['Llenado inicial'] = _emp_display['Llenado inicial'].map(lambda x: '🆕 sí' if x else '')

                    _cob_sem = 12  # default
                    try:
                        import json as _json_emp
                        _cfg_path_emp = os.path.join(os.path.dirname(__file__), 'config_afinidad.json')
                        with open(_cfg_path_emp) as _f_emp:
                            _cob_sem = _json_emp.load(_f_emp).get('empujes', {}).get('semanas_cobertura_target', 12)
                    except Exception:
                        pass
                    _n_fill = int(_emp_show['es_llenado_inicial'].sum()) if _has_fill else 0
                    st.markdown(f"**{len(_emp_show):,}** empujes — **{_emp_show['unidades_sugeridas'].sum():,}** unidades"
                                f" — Cobertura target: **{_cob_sem} semanas**")
                    if _has_fill:
                        st.caption(f"🆕 {_n_fill:,} son **llenado inicial** (producto nuevo en la tienda, sin stock): "
                                   f"se manda mínimo 3 curvas en tallas completas para exhibir bien. Si el CD no alcanza, "
                                   f"baja a 2 o 1 curva entera; si no llega ni a 1, no se manda.")

                    with st.expander("📐 Cómo se calcula el empuje (CD → tienda)", expanded=False):
                        st.markdown(f"""
**Empuje** = lo que el CD alcanza a mandar para llevar la tienda a su **stock objetivo de {_cob_sem} semanas** de venta.
Se calcula por cada combo **SKU × tienda** candidato:

**1. ¿Quién es candidato?**
- El SKU tiene stock en el **CD** (≥ 5 uds).
- La **línea rota bien en esa tienda**: rotación ≥ umbral (`rotación alta`, hoy 6%).
- El SKU **no** está en liquidación (margen sobre el mínimo de la marca).

**2. Venta semanal estimada** — columna `Vta/Sem Est`
- Si el SKU **ya vende** en esa tienda → venta de las últimas 4 semanas ÷ 4.
- Si es **nuevo** ahí (sin venta) → se usa el **promedio semanal de la línea** como proxy.

**3. Stock objetivo** — columna `Target {_cob_sem}s`
- `Target = ⌈ Vta/Sem Est × {_cob_sem} semanas ⌉`
- Piso mínimo: el **UME** (nunca menor a 3).
- **Llenado inicial** (tienda sin stock del producto): `Target = máx(Target, 3 curvas)` — 3 corridas de tallas completas para exhibir bien.

**4. Necesidad bruta**
- `Necesidad = Target − Stk Tienda`  (acotada entre 0 y 200 uds).

**5. Reparto del pool del CD** → columna `Empujar`
- El stock del CD es **único por SKU**: se reparte entre las tiendas por **prioridad = rotación × 1.2 (si es marca propia)**.
- Tienda **sin stock**: se sirve en **curvas enteras** (si el CD no alcanza, baja a 2 o 1 curva; si no llega ni a 1, no se manda).
- Tienda **con stock**: se sirve suelto = `mín(Necesidad, CD disponible)`.

> **Empujar** = las unidades que el pool del CD efectivamente alcanzó a asignar a esa tienda (puede ser menor que la Necesidad si el CD no da para todas).
""")

                    st.dataframe(_emp_display, use_container_width=True, hide_index=True)

                    # Descarga Excel
                    _buf_emp = _io_af.BytesIO()
                    _emp_show.to_excel(_buf_emp, index=False, sheet_name="Empujes CD")
                    st.download_button("📥 Descargar empujes", _buf_emp.getvalue(),
                                       file_name="empujes_cd_tiendas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_empujes_af")

                    # ── Marcar como ejecutado → alimenta la atribución del
                    # Caso de Éxito (Fase B, plan Demo Chile 2026-08-25) ──
                    st.markdown("---")
                    _ej_ops = [f"{r['sku']} → {_STORE_NAMES_AF.get(r['tienda'], r['tienda'])} "
                               f"({int(r['unidades_sugeridas'])} uds · {r['marca']})"
                               for _, r in _emp.head(150).iterrows()]
                    _ej_sel = st.multiselect(
                        "✅ Marcar empujes como EJECUTADOS (se registran en el log de acciones)",
                        _ej_ops, key="ej_empujes_sel")
                    if st.button("Registrar en el log de acciones", key="ej_empujes_btn",
                                 disabled=not _ej_sel):
                        for _op in _ej_sel:
                            _sku_op = _op.split(" → ")[0]
                            _fila_op = _emp[_emp["sku"].astype(str) == _sku_op].iloc[0]
                            acciones_log.agregar(
                                "", "Reposición / Empuje", str(_fila_op["marca"]),
                                f"Empuje ejecutado: {_op}",
                                magnitud=f"{int(_fila_op['unidades_sugeridas'])} uds",
                                sku=_sku_op, origen="Sugerida por Capi", estado="Ejecutada")
                        st.success(f"{len(_ej_sel)} empujes registrados en el log ✅ "
                                   "(descarga el CSV en Caso de Éxito al cerrar la semana)")

            # ── TAB 4: Redistribución ──
            with _tab_red:
                if len(_red) == 0:
                    st.info("No se encontraron oportunidades de redistribución con los umbrales actuales.")
                else:
                    _red_show = _red.copy()
                    _red_show['origen_nombre'] = _red_show['tienda_origen'].map(lambda t: _STORE_NAMES_AF.get(t, t))
                    _red_show['destino_nombre'] = _red_show['tienda_destino'].map(lambda t: _STORE_NAMES_AF.get(t, t))

                    _red_display = _red_show[['marca', 'descripcion', 'linea', 'origen_nombre', 'stk_origen',
                                               'destino_nombre', 'stk_destino', 'rotacion_destino',
                                               'unidades_sugeridas', 'mismo_cluster']].head(100)
                    _red_display.columns = ['Marca', 'Descripción', 'Línea', 'Origen', 'Stk Origen',
                                            'Destino', 'Stk Destino', 'Rot. Destino %', 'Mover', 'Mismo Cluster']
                    _red_display['Rot. Destino %'] = (_red_display['Rot. Destino %'] * 100).round(1)

                    st.markdown(f"**{len(_red_show):,}** oportunidades — **{_red_show['unidades_sugeridas'].sum():,}** unidades a redistribuir")
                    st.dataframe(_red_display, use_container_width=True, hide_index=True)

                    _buf_red = _io_af.BytesIO()
                    _red_show.to_excel(_buf_red, index=False, sheet_name="Redistribución")
                    st.download_button("📥 Descargar redistribución", _buf_red.getvalue(),
                                       file_name="redistribucion_tiendas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_redist_af")

            # ── TAB 5: Anomalías ──
            with _tab_an:
                if len(_an) == 0:
                    st.info("No se detectaron anomalías cruzadas.")
                else:
                    _pm_af = _an[_an['tipo_anomalia'] == 'producto_malo']
                    _mm_af = _an[_an['tipo_anomalia'] == 'mal_match_plaza']

                    st.markdown(f"""<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px;">
                      <div style="background:#FEF2F2;border-radius:8px;padding:12px 16px;">
                        <div style="font-size:12px;color:#991B1B;">Producto malo (no vende en ninguna tienda)</div>
                        <div style="font-size:20px;font-weight:600;color:#B91C1C;">{len(_pm_af)}</div>
                        <div style="font-size:11px;color:#DC2626;">Stk parado: {_pm_af['stk_parado'].sum():,}u → Liquidar</div>
                      </div>
                      <div style="background:#FFF7ED;border-radius:8px;padding:12px 16px;">
                        <div style="font-size:12px;color:#92400E;">Mal match plaza (vende en unas, no en otras)</div>
                        <div style="font-size:20px;font-weight:600;color:#B45309;">{len(_mm_af)}</div>
                        <div style="font-size:11px;color:#D97706;">Stk parado: {_mm_af['stk_parado'].sum():,}u → Redistribuir</div>
                      </div>
                    </div>""", unsafe_allow_html=True)

                    _tipo_filt_an = st.radio("Tipo", ["Todos", "Producto malo", "Mal match plaza"], horizontal=True, key="af_an_tipo")
                    _an_show = _an.copy()
                    if _tipo_filt_an == "Producto malo":
                        _an_show = _pm_af
                    elif _tipo_filt_an == "Mal match plaza":
                        _an_show = _mm_af

                    _an_display = _an_show[['marca', 'descripcion', 'linea', 'tipo_anomalia',
                                            'n_tiendas_sin_venta', 'n_tiendas_con_venta', 'stk_parado', 'accion']].head(100)
                    _an_display.columns = ['Marca', 'Descripción', 'Línea', 'Tipo', 'Sin Venta', 'Con Venta', 'Stk Parado', 'Acción']

                    st.dataframe(_an_display, use_container_width=True, hide_index=True)

            # ── TAB 6: Producción ──
            with _tab_prod:
                if len(_prod) == 0:
                    st.info("No se detectaron señales de producción con los umbrales actuales.")
                else:
                    st.markdown("**Líneas con alta rotación + baja cobertura en múltiples tiendas del mismo cluster**")

                    _prod_display = _prod[['marca', 'linea', 'n_tiendas', 'tiendas', 'vta_total',
                                            'stk_total', 'rotacion_prom', 'es_marca_propia', 'accion']].copy()
                    _prod_display.columns = ['Marca', 'Línea', 'N Tiendas', 'Tiendas', 'Vta Total',
                                             'Stk Total', 'Rotación %', 'Propia', 'Acción']
                    _prod_display['Rotación %'] = (_prod_display['Rotación %'] * 100).round(1)

                    st.dataframe(_prod_display, use_container_width=True, hide_index=True)

                    _propias_prod = _prod[_prod['es_marca_propia']]
                    if len(_propias_prod) > 0:
                        st.success(f"🏭 **{len(_propias_prod)} señales de marcas propias** — tienes control de producción para responder.")
# ─── Rendimiento por Tienda ─────────────────────────────────

elif nav_page == "🧵 Talla y Color":
    # Vista aislada (S9 ingesta, 2026-09-05): stock y venta por talla × color × tienda.
    vista_talla_color.render(st)

elif nav_page == "📊 Planificación":
    # Toda la vista vive en vista_planificacion.py (módulo aislado): este
    # bloque solo delega. Ver la nota de scoping en el encabezado del módulo.
    vista_plan.render(st)

elif nav_page == "📐 Rendimiento de Marca":
    # El nav dice Spavaldi porque es el caso de uso vivo, pero el módulo es
    # genérico: el encabezado sigue a la marca que se elija en el selector.
    _rt_titulo = st.empty()
    _rt_titulo.markdown('<div class="section-header"><h3>📐 Rendimiento de Marca</h3>'
                        '<span class="live-badge">CONTRIBUCIÓN / M²</span></div>', unsafe_allow_html=True)

    _rt_path = st.session_state.get("_base_profundidad_path")
    if not _rt_path or not os.path.exists(_rt_path):
        st.info("Esta vista necesita el reporte Micro original. Vuelve a correr el análisis "
                "subiendo la base de profundidad (no la plantilla transformada).")
    else:
        _rt_marcas = sorted(df_cob["marca"].dropna().str.upper().str.strip().unique()) \
            if "marca" in df_cob.columns else []
        _rt_default = _rt_marcas.index("SPAVALDI") if "SPAVALDI" in _rt_marcas else 0
        _rt_c1, _rt_c2 = st.columns([1, 3])
        with _rt_c1:
            _rt_marca = st.selectbox("Marca", _rt_marcas, index=_rt_default, key="rend_t_marca")

        _rt_sem, _rt_cortes = 1, []
        _rt_titulo.markdown(
            f'<div class="section-header"><h3>📐 {_rt_marca.title()} — rendimiento por tienda</h3>'
            f'<span class="live-badge">CONTRIBUCIÓN / M²</span></div>', unsafe_allow_html=True)
        try:
            _rt_largo, _rt_sem, _rt_cortes = _rt_cargar_micro(_rt_path, _rt_marca)
        except rend_t.FormatoMicroError as _e:
            _rt_largo = None
            st.warning(f"⚠️ {_e}")
        except rend_t.TiendaSinMapearError as _e:
            _rt_largo = None
            st.error(f"❌ {_e}")

        if _rt_largo is None or _rt_largo.empty:
            st.info(f"Sin datos de {_rt_marca} en el Micro cargado.")
        else:
            _rt_largo = rend_t.clasificar_liquidacion(_rt_largo)
            _rt_m = rend_t.metricas_por_tienda(_rt_largo, marca=_rt_marca, semanas=_rt_sem)
            # Activa = movió unidades O soles. No basta con unidades != 0:
            # Chorrillos cerró la ventana con 0 unidades netas pero −S/58, porque
            # la venta y la devolución fueron a precios distintos. Con el filtro
            # por unidades la tienda desaparecía del listado pero su plata seguía
            # en los KPIs, y los dos números no cuadraban.
            _rt_vivas = _rt_m[(_rt_m["unidades"] != 0) | (_rt_m["venta_soles"] != 0)]
            # Las pestañas de análisis van solo sobre tiendas con metraje: son
            # locales que venden a precio y ocupan espacio. Ecommerce, outlets y
            # liquidadoras tienen su propio bloque — mezclarlos ensucia la
            # comparación y, de paso, mejora el margen sin decirlo.
            _rt_con = _rt_vivas[_rt_vivas["m2"].notna()]
            _rt_sin = _rt_vivas[_rt_vivas["m2"].isna()]

            _rt_und = _rt_vivas["unidades"].sum()
            _rt_vta = _rt_vivas["venta_soles"].sum()
            _rt_contrib = _rt_m["contribucion"].sum()
            _rt_liq = _rt_vivas["venta_liq"].sum()
            _k = st.columns(5)
            _k[0].markdown(_kpi_html(f"S/ {_rt_vta:,.0f}",
                                     f"Venta neta ({_rt_sem} sem)" if _rt_sem > 1 else "Venta neta (semana)"),
                           unsafe_allow_html=True)
            _k[1].markdown(_kpi_html(f"{_rt_und:,.0f}", "Unidades"), unsafe_allow_html=True)
            _k[2].markdown(_kpi_html(f"S/ {_rt_contrib:,.0f}", "Contribución",
                                     "green" if _rt_contrib >= 0 else "red"), unsafe_allow_html=True)
            _k[3].markdown(_kpi_html(f"{(_rt_contrib/_rt_vta if _rt_vta else 0):.1%}", "Margen"),
                           unsafe_allow_html=True)
            _k[4].markdown(_kpi_html(f"{(_rt_liq/_rt_vta if _rt_vta else 0):.0%}", "Venta en liquidación",
                                     "yellow"), unsafe_allow_html=True)
            _rt_vent = (f"ventana de {_rt_sem} semanas ({_rt_cortes[0]} a {_rt_cortes[-1]})"
                        if _rt_sem > 1 else "última semana cerrada del Micro")
            _rt_pv = (_rt_sin["venta_soles"].sum() / _rt_vta) if _rt_vta else 0
            st.caption(f"{len(_rt_vivas)} tiendas con venta · {_rt_vent} · venta neta ex-IGV · "
                       f"liquidación = mercadería con más de {rend_t.EDAD_LIQUIDACION:.0f} semanas "
                       f"(mismo umbral que la taxonomía de Capi)")
            if len(_rt_sin):
                st.caption(f"⚠️ Los indicadores de arriba son de la marca completa. **Los cuadros muestran "
                           f"solo las {len(_rt_con)} tiendas con m² asignado** — el {_rt_pv:.0%} restante "
                           f"de la venta ({', '.join(_rt_sin['tienda'].head(4))}) va al final, porque no "
                           f"tiene metraje contra el cual medirse.")

            _t1, _t2, _t3, _t7, _t4, _t5, _t6 = st.tabs(
                ["📋 Resumen", "📐 Rendimiento m²", "📦 Cobertura", "🏆 Best sellers",
                 "📅 Evolución", "⚖️ Comparar marcas", "✉️ Correo a gerencia"])

            # ── Resumen: P&L partido temporada / liquidación ──
            with _t1:
                st.caption("El margen total engaña cuando la tienda absorbe liquidación. "
                           "La columna que manda es **margen de temporada**.")
                _c = ["tienda", "canal", "unidades", "venta_soles", "contribucion", "margen",
                      "margen_temporada", "pct_venta_liquidacion", "edad_mediana", "n_skus"]
                _c = [x for x in _c if x in _rt_con.columns]
                _d = _rt_con[_c].rename(columns={
                    "tienda": "Tienda", "canal": "Canal", "unidades": "Und",
                    "venta_soles": "Venta S/", "contribucion": "Contribución S/",
                    "margen": "Margen", "margen_temporada": "Margen temporada",
                    "pct_venta_liquidacion": "% liquidación",
                    "edad_mediana": "Edad mediana (sem)", "n_skus": "SKUs"})
                st.dataframe(_d.style.format({
                    "Venta S/": "{:,.0f}", "Contribución S/": "{:,.0f}", "Und": "{:,.0f}",
                    "Margen": "{:.1%}", "Margen temporada": "{:.1%}", "% liquidación": "{:.0%}",
                    "Edad mediana (sem)": "{:.0f}"}, na_rep="—"),
                    use_container_width=True, hide_index=True, height=420)

                if len(_rt_sin):
                    st.markdown("**Sin m² asignado — ecommerce, outlets y liquidación**")
                    _cs = [x for x in ("tienda", "canal", "unidades", "venta_soles", "contribucion",
                                       "margen", "pct_venta_liquidacion") if x in _rt_sin.columns]
                    st.dataframe(_rt_sin[_cs].rename(columns={
                        "tienda": "Tienda", "canal": "Canal", "unidades": "Und",
                        "venta_soles": "Venta S/", "contribucion": "Contribución S/",
                        "margen": "Margen", "pct_venta_liquidacion": "% liquidación"}
                    ).style.format({"Venta S/": "{:,.0f}", "Contribución S/": "{:,.0f}",
                                    "Und": "{:,.0f}", "Margen": "{:.1%}", "% liquidación": "{:.0%}"},
                                   na_rep="—"), use_container_width=True, hide_index=True)

                _perd = rend_t.tiendas_en_perdida(_rt_vivas)
                if not _perd.empty:
                    st.markdown("**Tiendas con contribución negativa**")
                    st.dataframe(_perd.rename(columns={
                        "tienda": "Tienda", "contribucion": "Contribución S/", "margen": "Margen",
                        "margen_temporada": "Margen temporada",
                        "pct_venta_liquidacion": "% liquidación", "diagnostico": "Lectura"}
                    ).style.format({"Contribución S/": "{:,.0f}", "Margen": "{:.1%}",
                                    "Margen temporada": "{:.1%}", "% liquidación": "{:.0%}"},
                                   na_rep="—"),
                        use_container_width=True, hide_index=True)

            # ── Rendimiento m² ──
            with _t2:
                _sin_m2 = _rt_con.empty
                if _sin_m2:
                    st.warning("Falta cargar los m² de corner en `config_tiendas.json`. "
                               "Sin ese dato la métrica queda vacía a propósito — un cero se leería "
                               "como «no rinde». Outlets y tiendas liquidadoras van sin m² por diseño.")
                _c = [x for x in ["tienda", "canal", "m2", "contribucion", "contrib_x_m2",
                                  "contrib_temporada_x_m2", "venta_x_m2"] if x in _rt_con.columns]
                st.dataframe(_rt_con[_c].rename(columns={
                    "tienda": "Tienda", "canal": "Canal", "m2": "m² corner",
                    "contribucion": "Contribución S/", "contrib_x_m2": "Contrib/m²",
                    "contrib_temporada_x_m2": "Contrib temporada/m²", "venta_x_m2": "Venta/m²"}
                ).style.format({"Contribución S/": "{:,.0f}", "m² corner": "{:,.1f}",
                                "Contrib/m²": "{:,.0f}", "Contrib temporada/m²": "{:,.0f}",
                                "Venta/m²": "{:,.0f}"}, na_rep="no aplica"),
                    use_container_width=True, hide_index=True, height=420)

            # ── Cobertura ──
            with _t3:
                _c = [x for x in ["tienda", "canal", "stock_uds", "unidades", "cobertura_sem",
                                  "und_ult_sem", "cobertura_1sem", "n_skus"] if x in _rt_con.columns]
                _cob = _rt_con[_c].copy()
                _tot_stk = _rt_m["stock_uds"].sum() if "stock_uds" in _rt_m.columns else 0
                _cob_cadena = _tot_stk / _rt_und if _rt_und else float("nan")
                st.markdown(_kpi_html(f"{_cob_cadena:,.1f}", "Cobertura cadena (semanas)"),
                            unsafe_allow_html=True)
                st.caption("Cobertura = stock en tienda / venta semanal. No hay año anterior contra "
                           "el cual medirse (la marca arrancó en oct-25), así que la cobertura es el "
                           "sustituto del comparativo LY.")
                st.dataframe(_cob.rename(columns={
                    "tienda": "Tienda", "canal": "Canal", "stock_uds": "Stock und",
                    "unidades": "Venta ventana", "cobertura_sem": "Cobertura ventana",
                    "und_ult_sem": "Vta última sem", "cobertura_1sem": "Cobertura última sem",
                    "n_skus": "SKUs"}
                ).style.format({"Stock und": "{:,.0f}", "Venta ventana": "{:,.0f}",
                                "Vta última sem": "{:,.0f}", "Cobertura ventana": "{:,.1f}",
                                "Cobertura última sem": "{:,.1f}"}, na_rep="—"),
                    use_container_width=True, hide_index=True, height=420)

            # ── Evolución ──
            with _t4:
                st.caption("El mes se toma como las últimas 4 semanas, sin meterse con el calendario "
                           "comercial. Acá va la evolución semana a semana dentro de esa ventana.")
                if "semana_idx" not in _rt_largo.columns or _rt_sem < 2:
                    st.info("Con un solo corte no hay evolución que mostrar. Guarda el Micro cada "
                            "semana y la serie se arma sola.")
                else:
                    _ev = _rt_largo[_rt_largo["cod_tienda"].isin(_rt_con["cod_tienda"])]
                    _ev = _ev.groupby(["semana_idx", "fecha_corte"], dropna=False).agg(
                        unidades=("unidades", "sum"), venta_soles=("venta_soles", "sum"),
                        contribucion=("contribucion", "sum")).reset_index()
                    _ev["margen"] = np.where(_ev.venta_soles > 0,
                                             _ev.contribucion / _ev.venta_soles, np.nan)
                    _ev["Semana"] = _ev["fecha_corte"].dt.strftime("Cierre %d-%b")
                    st.dataframe(_ev[["Semana", "unidades", "venta_soles", "contribucion", "margen"]]
                                 .rename(columns={"unidades": "Und", "venta_soles": "Venta S/",
                                                  "contribucion": "Contribución S/", "margen": "Margen"})
                                 .style.format({"Und": "{:,.0f}", "Venta S/": "{:,.0f}",
                                                "Contribución S/": "{:,.0f}", "Margen": "{:.1%}"},
                                               na_rep="—"),
                                 use_container_width=True, hide_index=True)
                    st.line_chart(_ev.set_index("Semana")[["venta_soles", "contribucion"]],
                                  use_container_width=True)
                    _ev_d = _ev.venta_soles.iloc[-1] - _ev.venta_soles.iloc[0]
                    st.caption(f"Entre el primer y el último corte la venta semanal "
                               f"{'subió' if _ev_d >= 0 else 'bajó'} S/ {abs(_ev_d):,.0f}. "
                               f"Solo tiendas con m² asignado, para que cuadre con los otros cuadros.")

            # ── Best sellers ──
            with _t7:
                st.caption(f"Ordenado por venta de la ventana de {_rt_sem} semanas. **El grano es el "
                           f"ESTILO**: el reporte micro no baja a color ni talla. Para el detalle por "
                           f"color hace falta la base transaccional — está abajo. Stock del último corte.")
                _bs_n = st.slider("Cuántos mostrar", 5, 60, 20, 5, key="rend_t_bs_n")
                _bs_por = st.radio("Ordenar por", ["venta_soles", "contribucion", "unidades"],
                                   format_func=lambda x: {"venta_soles": "Venta S/",
                                                          "contribucion": "Contribución",
                                                          "unidades": "Unidades"}[x],
                                   horizontal=True, key="rend_t_bs_por")
                _bs = rend_t.bestsellers(_rt_largo, top=_bs_n, por=_bs_por)
                st.dataframe(_bs[["rk", "descripcion", "linea", "unidades", "venta_soles",
                                  "contribucion", "margen", "precio_real", "stock_uds",
                                  "cobertura_sem", "tiendas", "es_liquidacion"]].rename(columns={
                    "rk": "#", "descripcion": "Producto", "linea": "Línea", "unidades": "Und",
                    "venta_soles": "Venta S/", "contribucion": "Contribución S/", "margen": "Margen",
                    "precio_real": "Precio real", "stock_uds": "Stock", "cobertura_sem": "Cobertura",
                    "tiendas": "Tiendas", "es_liquidacion": "Tipo"}
                ).style.format({"Und": "{:,.0f}", "Venta S/": "{:,.0f}", "Contribución S/": "{:,.0f}",
                                "Margen": "{:.1%}", "Precio real": "{:,.0f}", "Stock": "{:,.0f}",
                                "Cobertura": "{:,.1f}"}, na_rep="—"),
                    use_container_width=True, hide_index=True, height=460)

                _bs_mal = _bs[_bs.contribucion < 0]
                if len(_bs_mal):
                    st.warning("⚠️ Entre los más vendidos hay productos con **contribución negativa**: "
                               + ", ".join(f"{r.descripcion} (S/ {r.contribucion:,.0f})"
                                           for _, r in _bs_mal.head(4).iterrows())
                               + ". Venden volumen pero restan plata.")

                st.markdown("---")
                st.markdown("**Detalle por color y talla**")
                st.caption("El reporte micro no baja de estilo. Para ver colores hace falta la base de "
                           "ventas por línea de ticket (día × tienda × SKU, con color y talla). "
                           "Súbela acá y el detalle aparece abajo.")
                _bs_f = st.file_uploader("Base de ventas con color y talla (.xlsx)", type=["xlsx"],
                                         key="rend_t_trans")
                if _bs_f is not None:
                    try:
                        _bs_t = _rt_cargar_trans(_bs_f.getvalue(), _rt_marca)
                    except Exception as _e:
                        _bs_t = None
                        st.error(f"No pude leer el archivo: {_e}")
                    if _bs_t is not None and len(_bs_t):
                        _bs_min, _bs_max = _bs_t["fecha"].min(), _bs_t["fecha"].max()
                        st.caption(f"{len(_bs_t):,} líneas de venta · {_bs_min:%d-%b-%y} a {_bs_max:%d-%b-%y} · "
                                   f"{_bs_t['color'].nunique()} colores · {_bs_t['talla'].nunique()} tallas")
                        # La ventana del transaccional casi nunca coincide con la
                        # del Micro. Declararlo evita comparar cifras de periodos
                        # distintos creyendo que son la misma ventana.
                        if _rt_cortes:
                            _bs_ini = pd.Timestamp(_rt_cortes[0]) - pd.Timedelta(days=7)
                            _bs_fin = pd.Timestamp(_rt_cortes[-1])
                            _bs_dentro = _bs_t[(_bs_t.fecha >= _bs_ini) & (_bs_t.fecha <= _bs_fin)]
                            if len(_bs_dentro) < len(_bs_t) * 0.5:
                                st.warning(
                                    f"⚠️ Este archivo cubre hasta el {_bs_max:%d-%b}, pero la ventana de "
                                    f"arriba va del {_bs_ini:%d-%b} al {_bs_fin:%d-%b}. Solo "
                                    f"{len(_bs_dentro):,} de {len(_bs_t):,} líneas caen dentro. "
                                    f"Elige qué periodo quieres mirar.")
                            _bs_amb = st.radio(
                                "Periodo", ["Ventana del micro", "Todo el archivo"],
                                horizontal=True, key="rend_t_bs_periodo")
                            _bs_use = _bs_dentro if _bs_amb == "Ventana del micro" else _bs_t
                        else:
                            _bs_use = _bs_t

                        if not len(_bs_use):
                            st.info("No hay ventas en ese periodo.")
                        else:
                            _bs_niv = st.radio("Nivel", ["color", "talla", "solo_color"],
                                               format_func=lambda x: {"color": "Estilo × color",
                                                                      "talla": "Estilo × color × talla",
                                                                      "solo_color": "Solo color"}[x],
                                               horizontal=True, key="rend_t_bs_nivel")
                            _bs_c = rend_t.bestsellers_color(_bs_use, top=_bs_n, por=_bs_por,
                                                             nivel=_bs_niv)
                            _bs_cols = [c for c in ("rk", "estilo", "color", "talla", "linea",
                                                    "unidades", "venta_soles", "contribucion",
                                                    "margen", "precio_real", "tiendas")
                                        if c in _bs_c.columns]
                            st.dataframe(_bs_c[_bs_cols].rename(columns={
                                "rk": "#", "estilo": "Estilo", "color": "Color", "talla": "Talla",
                                "linea": "Línea", "unidades": "Und", "venta_soles": "Venta S/",
                                "contribucion": "Contribución S/", "margen": "Margen",
                                "precio_real": "Precio real", "tiendas": "Tiendas"}
                            ).style.format({"Und": "{:,.0f}", "Venta S/": "{:,.0f}",
                                            "Contribución S/": "{:,.0f}", "Margen": "{:.1%}",
                                            "Precio real": "{:,.0f}"}, na_rep="—"),
                                use_container_width=True, hide_index=True, height=440)

            # ── Comparar marcas ──
            with _t5:
                _otras = [m for m in _rt_marcas if m != _rt_marca]
                _vs = st.multiselect("Comparar contra", _otras,
                                     default=["CACHAREL"] if "CACHAREL" in _otras else _otras[:1],
                                     key="rend_t_vs")
                if not _vs:
                    st.caption("Elige al menos una marca para comparar.")
                else:
                    _todas = rend_t.desde_micro(pd.read_excel(_rt_path), marcas=[_rt_marca] + _vs)
                    _cmp = rend_t.comparar_marcas(_todas, [_rt_marca] + _vs, por=("tienda",))
                    _cmp = _cmp[_cmp["tienda"].isin(_rt_con["tienda"])]
                    st.caption("Participación calculada solo sobre tiendas donde alguna de las marcas "
                               "vende. Dos marcas no están en las mismas tiendas, así que el total "
                               "global no es comparable sin esta salvedad.")
                    # Totalizado: Majo lo pidió explícitamente ("y también tener
                    # un totalizado"). Va como fila, no como métrica aparte.
                    _num = [c for c in _cmp.columns if c != "tienda"]
                    _tot = {"tienda": "TOTAL (tiendas con m²)"}
                    for _c in _num:
                        _tot[_c] = _cmp[_c].sum() if not str(_c).startswith(("pct_",)) and "_vs_" not in str(_c) else None
                    _a, _b = rend_t._norm(_rt_marca), rend_t._norm(_vs[0])
                    if _a in _cmp.columns and _b in _cmp.columns and _cmp[_b].sum():
                        _tot[f"{_a}_vs_{_b}"] = _cmp[_a].sum() / _cmp[_b].sum()
                    _cmp = pd.concat([_cmp, pd.DataFrame([_tot])], ignore_index=True)
                    st.dataframe(_cmp.style.format(
                        {c: "{:,.0f}" for c in _cmp.columns if c not in ("tienda",)
                         and not str(c).startswith("pct_") and not str(c).endswith("_vs_" + _vs[0])}
                        | {c: "{:.1%}" for c in _cmp.columns if str(c).startswith("pct_")},
                        na_rep="—"), use_container_width=True, hide_index=True, height=420)


            # ── Correo a gerencia ──
            with _t6:
                st.caption("Genera el borrador del correo semanal con las cifras de esta ventana. "
                           "**El agente redacta, no calcula**: recibe los números ya verificados por el "
                           "motor y solo los pone en prosa. Siempre es un borrador — nada se envía solo.")

                # Sin value=: con key, Streamlit ya persiste el contenido en
                # session_state, y pasar ambos lanza StreamlitAPIException.
                _ag_pre = st.text_area(
                    "¿Qué preguntó la gerencia? (opcional)", height=100, key="rend_t_preg",
                    help="Pega acá sus preguntas y el correo las responde punto por punto, en ese orden.")

                _ag_c1, _ag_c2 = st.columns([1, 2])
                with _ag_c1:
                    _ag_go = st.button("✉️ Redactar borrador", type="primary",
                                       use_container_width=True, key="rend_t_gen")
                _ag_hay_key = bool(os.environ.get("ANTHROPIC_API_KEY"))
                if not _ag_hay_key:
                    with _ag_c2:
                        st.caption("🔑 Sin API key: se usa la versión por reglas, que arma el mismo "
                                   "correo con frases fijas. Las cifras son idénticas.")

                if _ag_go:
                    _ag_d = {"act": _rt_vivas, "conm2": _rt_con, "comp": pd.DataFrame(),
                             "nsem": _rt_sem, "cortes": _rt_cortes or ["—"],
                             "marca": _rt_marca, "vs": None}
                    _ag_h = ag_rep.hechos(_ag_d)
                    try:
                        _ag_r = ag_rep.redactar(_ag_h, _ag_pre)
                        _ag_txt = (f"**Asunto:** {_ag_r['asunto']}\n\n{_ag_r['cuerpo']}"
                                   if _ag_r["asunto"] else _ag_r["cuerpo"])
                        _ag_sosp = ag_rep.verificar(_ag_txt, _ag_h)
                        _ag_via = "redactado con IA"
                    except (ValueError, ImportError) as _e:
                        _ag_txt = rep_sem.correo(_ag_d)
                        _ag_sosp, _ag_via = [], f"versión por reglas ({_e})"
                    st.session_state["rend_t_borrador"] = _ag_txt
                    st.session_state["rend_t_sosp"] = _ag_sosp
                    st.session_state["rend_t_via"] = _ag_via

                if st.session_state.get("rend_t_borrador"):
                    st.caption(f"Borrador — {st.session_state.get('rend_t_via', '')}")
                    if st.session_state.get("rend_t_sosp"):
                        st.error("⚠️ El borrador tiene cifras que NO están en el análisis: "
                                 + ", ".join(st.session_state["rend_t_sosp"])
                                 + ". Revísalas antes de enviar — el agente pudo haberlas inventado.")
                    st.text_area("Borrador (selecciona y copia)", st.session_state["rend_t_borrador"],
                                 height=420, key="rend_t_out")
                    st.download_button(
                        "📥 Descargar el borrador (.md)",
                        data=st.session_state["rend_t_borrador"].encode("utf-8"),
                        file_name=f"Correo_{_rt_marca}_{(_rt_cortes or ['x'])[-1]}.md",
                        mime="text/markdown", key="dl_rend_t_correo")
                    st.caption("Revisa el borrador contra los cuadros de las otras pestañas antes de "
                               "enviarlo. El agente no ve la base: solo las cifras que le pasamos.")

            # ── Export ──
            _rt_buf = io.BytesIO()
            with pd.ExcelWriter(_rt_buf, engine="openpyxl") as _rt_w:
                _rt_vivas.to_excel(_rt_w, sheet_name="Rendimiento x Tienda", index=False)
                _rt_largo.to_excel(_rt_w, sheet_name="Detalle SKU x Tienda", index=False)
            _rt_buf.seek(0)
            st.download_button(
                f"📥 Descargar rendimiento de {_rt_marca} (.xlsx)", data=_rt_buf.getvalue(),
                file_name=f"Capi_Rendimiento_Tienda_{_rt_marca}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_rend_t")
# ══════════════════════════════════════════════════════════════
#  DESCARGA DE EXCEL
# ══════════════════════════════════════════════════════════════


def _build_excel(cob_json, rep_pivot_json, rep_json, trans_json, prec_json, alertas_json, anomalias_json):
    """Genera el Excel de resultados en memoria con columnas de precio y nuevo margen.

    SIN @st.cache_data a propósito: la función llama a helpers externos
    (agente_terceras.*) cuyo código puede cambiar entre deploys. El cache de
    Streamlit solo hashea el código directo + argumentos, NO las funciones
    llamadas, así que cacheaba Excels viejos tras arreglar un helper. Generarlo
    fresco toma ~1-2 seg y elimina esa clase de bug."""
    buf = io.BytesIO()
    _df_cob_ref = pd.read_json(io.StringIO(cob_json))
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Resumen ejecutivo primero: la foto antes del detalle (auditoría 2026-08-05)
        vistas_excel.hoja_resumen_ejecutivo(writer, _df_cob_ref)
        _add_pricing_cols(_df_cob_ref, _df_cob_ref, "Cobertura", writer)
        _df_rep_piv_xl = pd.read_json(io.StringIO(rep_pivot_json))
        _df_rep_piv_xl.to_excel(writer, sheet_name="Reposiciones", index=False)
        # Fila TOTAL con SUBTOTAL (respeta filtros) al pie de la matriz — antes
        # Franco la agregaba a mano encima del header (auditoría 2026-08-05)
        _ws_rep = writer.sheets["Reposiciones"]
        if not _df_rep_piv_xl.empty:
            _num_cols = [i + 1 for i, c in enumerate(_df_rep_piv_xl.columns)
                         if pd.api.types.is_numeric_dtype(_df_rep_piv_xl[c]) and c != "sku"]
            _fila_tot = _ws_rep.max_row + 1
            _ws_rep.cell(row=_fila_tot, column=1, value="TOTAL (visible)")
            for _ci in _num_cols:
                _ltr = get_column_letter(_ci)
                _ws_rep.cell(row=_fila_tot, column=_ci,
                             value=f"=SUBTOTAL(9,{_ltr}2:{_ltr}{_fila_tot - 1})")
            _ws_rep.auto_filter.ref = f"A1:{get_column_letter(_ws_rep.max_column)}{_fila_tot - 1}"
            _ws_rep.freeze_panes = "A2"
        _df_rep_xl = pd.read_json(io.StringIO(rep_json))
        if not _df_rep_xl.empty:
            _add_pricing_cols(_df_rep_xl, _df_cob_ref, "Reposiciones Detalle", writer)
        else:
            _df_rep_xl.to_excel(writer, sheet_name="Reposiciones Detalle", index=False)
        pd.read_json(io.StringIO(trans_json)).to_excel(writer, sheet_name="Transferencias", index=False)
        _df_prec_xl = pd.read_json(io.StringIO(prec_json))
        if not _df_prec_xl.empty:
            _add_pricing_cols(_df_prec_xl, _df_cob_ref, "Acciones Precio", writer)
        else:
            _df_prec_xl.to_excel(writer, sheet_name="Acciones Precio", index=False)
        pd.read_json(io.StringIO(alertas_json)).to_excel(writer, sheet_name="Alertas IA", index=False)
        pd.read_json(io.StringIO(anomalias_json)).to_excel(writer, sheet_name="Anomalías Tienda", index=False)
        # Revisar Terceras: top 5 SKUs por marca×línea con sobrestock real
        # (cobertura >16 sem, edad ≥3 sem) + criticidad 🔴🟠🟡. Hoja única
        # accionable para revisar/negociar — reemplaza las dos hojas previas.
        _df_rev = agente_terceras.top5_por_marca_linea(_df_cob_ref, top_n=5)
        if not _df_rev.empty:
            _cols_r = [c for c in ['marca', 'categoria', 'criticidad', 'sku', 'nombre', 'edad',
                                   'cobertura', 'stock', 'vta_sem', 'dscto', 'capital',
                                   'sell_through', 'margen_efectivo'] if c in _df_rev.columns]
            _rev_out = _df_rev[_cols_r].rename(columns={
                'marca': 'Marca', 'categoria': 'Línea', 'criticidad': 'Criticidad',
                'sku': 'SKU', 'nombre': 'Producto', 'edad': 'Edad (sem)',
                'cobertura': 'Cobertura (sem)', 'stock': 'Stock (uds)',
                'vta_sem': 'Vta/sem (uds)', 'dscto': 'Dscto', 'capital': 'Capital S/ (costo)',
                'sell_through': 'Sell-through %', 'margen_efectivo': 'Margen efect. %',
            })
            _rev_out.to_excel(writer, sheet_name="Revisar Terceras", index=False)
    buf.seek(0)
    return buf.read()


# Combinar alertas + anomalías para Excel
_alertas_excel = df_alertas if not df_alertas.empty else pd.DataFrame()
_anomalias_excel = df_anomalias if not df_anomalias.empty else pd.DataFrame()

if nav_page == "🏠 Dashboard":
    st.markdown("---")
    st.markdown(f'<div class="section-header"><h3>📥 Descargar resultados</h3></div>', unsafe_allow_html=True)
    excel_bytes = _build_excel(
        df_cob.to_json(),
        df_rep_pivot.to_json() if not df_rep_pivot.empty else pd.DataFrame().to_json(),
        df_rep.to_json() if not df_rep.empty else pd.DataFrame().to_json(),
        df_trans.to_json() if not df_trans.empty else pd.DataFrame().to_json(),
        df_prec.to_json() if not df_prec.empty else pd.DataFrame().to_json(),
        _alertas_excel.to_json() if not _alertas_excel.empty else pd.DataFrame().to_json(),
        _anomalias_excel.to_json() if not _anomalias_excel.empty else pd.DataFrame().to_json(),
    )

    st.download_button(
        label="📥 Descargar análisis completo (.xlsx)",
        data=excel_bytes,
        file_name="Capi_Analisis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # ── Reportes por marca tercera: un Excel por marca, enviable al proveedor ──
    # (decisión 2026-08-05; generación bajo demanda para no alentar cada rerun)
    # Fingerprint de la base: si Franco sube otra base, el zip viejo deja de
    # ofrecerse (evita mandar cifras de un corte anterior a un proveedor).
    _fp_base_marcas = f"{len(df_cob)}|{df_cob['stock_valor_costo'].sum():.0f}"
    if st.button("📦 Generar reportes por marca tercera (un Excel por marca)",
                 use_container_width=True):
        with st.spinner("Generando reportes por marca..."):
            st.session_state["zip_reportes_marcas"] = reportes_marcas.generar_zip_reportes(
                df_cob, df_rep if not df_rep.empty else None,
                df_trans if not df_trans.empty else None,
                df_prec if not df_prec.empty else None,
                _alertas_excel if not _alertas_excel.empty else None,
            )
            st.session_state["zip_reportes_marcas_fp"] = _fp_base_marcas
    if (st.session_state.get("zip_reportes_marcas")
            and st.session_state.get("zip_reportes_marcas_fp") == _fp_base_marcas):
        st.download_button(
            label="📦 Descargar reportes por marca (.zip)",
            data=st.session_state["zip_reportes_marcas"],
            file_name="Reportes_Marcas_Terceras.zip",
            mime="application/zip",
            use_container_width=True,
        )

if nav_page == "📲 Productos Venta Cero":
    st.markdown(f'<div class="section-header"><h3>📲 Productos Venta Cero</h3><span class="live-badge">REVISIÓN TIENDA</span></div>', unsafe_allow_html=True)
    st.caption("SKUs con stock en tienda que NO vendieron la semana pasada. Para que cada tienda revise "
               "exhibición y comunicación de precio. El tipo de evento indica si etiquetar (MD1) o poner cartel (PTR).")

    # Reconstrucción desde df_cob (venta=0, con stock) + tipo de evento de la base
    _vc_min_cap = st.slider("Capital mínimo a costo por SKU (S/)", min_value=0, max_value=5000, value=1000, step=250,
                            help="Filtra SKUs sin venta con poco capital parado para enfocar la revisión.")
    _vc = df_cob[(df_cob['prom_vta_uds'].fillna(0) == 0) & (df_cob['stock_total'].fillna(0) > 0)].copy() if 'prom_vta_uds' in df_cob.columns else pd.DataFrame()
    if not _vc.empty and 'stock_valor_costo' in _vc.columns:
        _vc = _vc[_vc['stock_valor_costo'].fillna(0) >= _vc_min_cap]
    if _vc.empty:
        st.success("No hay productos con venta cero sobre el umbral con la base actual.")
    else:
        # Cruce tipo de evento (MD1 / PTR / MTR) desde la Base Profundidad
        _vc_base = _capi_base_path()
        _vc_tev = _tipo_evento_map(_vc_base) if _vc_base else {}
        _vc['tipo_evento'] = _vc['sku'].map(_vc_tev).fillna('') if 'sku' in _vc.columns else ''

        def _vc_accion(r):
            _tp = str(r.get('tipo_evento', '') or '')
            _dsc = float(r.get('pct_descuento', 0) or 0)
            if _dsc > 0 and _tp == 'MD1':
                return "🏷️ Etiquetar mercadería (precio ya impreso)"
            if _dsc > 0 and _tp == 'PTR':
                return "📋 Colocar cartel de precio"
            return "👁️ Revisar exhibición"
        _vc['accion'] = _vc.apply(_vc_accion, axis=1)

        _vc_n_tiendas = _vc['tienda'].nunique() if 'tienda' in _vc.columns else 0
        _vc_n_skus = len(_vc)
        _vc_cap = _vc['stock_valor_costo'].sum() if 'stock_valor_costo' in _vc.columns else 0
        _vcc1, _vcc2, _vcc3 = st.columns(3)
        _vcc1.markdown(f'<div style="background:#FEF2F2; border-radius:12px; padding:14px 18px; border-left:4px solid {STATUS_CRITICO};"><div style="font-size:0.75rem; color:var(--capi-text2);">Combos sin venta</div><div style="font-size:1.5rem; font-weight:700; color:{STATUS_CRITICO};">{_vc_n_skus:,}</div></div>', unsafe_allow_html=True)
        _vcc2.markdown(f'<div style="background:var(--capi-bg-surface); border-radius:12px; padding:14px 18px; border-left:4px solid {STATUS_SOBRESTOCK};"><div style="font-size:0.75rem; color:var(--capi-text2);">Tiendas</div><div style="font-size:1.5rem; font-weight:700; color:{STATUS_SOBRESTOCK};">{_vc_n_tiendas}</div></div>', unsafe_allow_html=True)
        _vcc3.markdown(f'<div style="background:var(--capi-bg-surface); border-radius:12px; padding:14px 18px; border-left:4px solid {STATUS_MUERTO};"><div style="font-size:0.75rem; color:var(--capi-text2);">Capital parado</div><div style="font-size:1.5rem; font-weight:700; color:{STATUS_MUERTO};">S/ {_vc_cap:,.0f}</div></div>', unsafe_allow_html=True)

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        _vc_tiendas = ["Todas"] + sorted(_vc['tienda'].dropna().unique().tolist()) if 'tienda' in _vc.columns else ["Todas"]
        _vc_sel = st.selectbox("Tienda", _vc_tiendas, key="vc_tienda")
        _vc_v = _vc if _vc_sel == "Todas" else _vc[_vc['tienda'] == _vc_sel]
        _vc_v = _vc_v.sort_values('stock_valor_costo', ascending=False) if 'stock_valor_costo' in _vc_v.columns else _vc_v

        _vc_cols = [c for c in ['tienda', 'marca', 'sku', 'nombre', 'categoria', 'stock_total',
                                'stock_valor_costo', 'precio_vigente', 'pct_descuento', 'tipo_evento',
                                'edad_semanas', 'accion'] if c in _vc_v.columns]
        _vc_disp = _vc_v[_vc_cols].rename(columns={
            'tienda': 'Tienda', 'marca': 'Marca', 'sku': 'SKU', 'nombre': 'Producto', 'categoria': 'Línea',
            'stock_total': 'Stock (uds)', 'stock_valor_costo': 'Capital S/', 'precio_vigente': 'Precio',
            'pct_descuento': 'Dscto', 'tipo_evento': 'Tipo evento', 'edad_semanas': 'Edad (sem)', 'accion': 'Acción',
        })
        st.dataframe(_vc_disp.head(500).style.format({'Capital S/': 'S/ {:,.0f}', 'Precio': 'S/ {:,.2f}', 'Dscto': '{:.0%}'}, na_rep="—"),
                     use_container_width=True, hide_index=True, height=460)
        st.caption("Acción: MD1 = mercadería ya etiquetada (verificar etiqueta) · PTR = colocar cartel de precio · sin evento = revisar exhibición.")
        if not _vc_tev:
            st.warning("⚠️ Esta sesión no tiene la Base Profundidad original cargada: sin tipo de "
                       "evento, todas las acciones caen a 'Revisar exhibición'. Sube tu base "
                       "original (no la plantilla) para el detalle MD1/PTR.")

        # ── Excel para tiendas: Pareto 80% (decisión Franco C1 2026-08-26) ──
        # S7 (2026-09-05): la lógica vive en vistas_excel.venta_cero() y se reusa en el
        # reporte por marca tercera (hoja "5. Venta Cero").
        _vp80 = vistas_excel.venta_cero(df_cob, min_capital=_vc_min_cap, tipo_evento_map=_vc_tev)
        _vp80_buf = io.BytesIO()
        with pd.ExcelWriter(_vp80_buf, engine='openpyxl') as _wvp:
            vistas_excel.hoja_venta_cero(_wvp, _vp80)
        _vp80_buf.seek(0)
        st.download_button(
            "📥 Excel para tiendas — Pareto 80% del capital sin venta",
            _vp80_buf.getvalue(), file_name="Capi_Venta_Cero_Tiendas_Pareto.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_vc_pareto", use_container_width=True)

        # Excel: una hoja por tienda (para repartir a cada una)
        _vc_buf = io.BytesIO()
        with pd.ExcelWriter(_vc_buf, engine='openpyxl') as _vc_w:
            _vc[_vc_cols].rename(columns={
                'tienda': 'Tienda', 'marca': 'Marca', 'sku': 'SKU', 'nombre': 'Producto', 'categoria': 'Linea',
                'stock_total': 'Stock', 'stock_valor_costo': 'Capital S/', 'precio_vigente': 'Precio',
                'pct_descuento': 'Dscto', 'tipo_evento': 'Tipo evento', 'edad_semanas': 'Edad sem', 'accion': 'Accion',
            }).to_excel(_vc_w, sheet_name='Venta Cero', index=False)
        _vc_buf.seek(0)
        st.download_button("📥 Descargar detalle por tienda (.xlsx)", data=_vc_buf.getvalue(),
                           file_name="Capi_Productos_Venta_Cero.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="dl_venta_cero")


elif nav_page == "📦 Ventana de Compra":
    st.markdown(f'<div class="section-header"><h3>📦 Ventana de Compra</h3><span class="live-badge">EMBARQUES</span></div>', unsafe_allow_html=True)
    st.caption("Análisis de performance por ventana de compra (embarques). Sell-through, cobertura y recomendaciones por ventana.")

    if por_ventana.empty:
        st.info("No hay datos de embarques/ventana de compra. Sube un archivo con la columna Embarque en Base Profundidad.")
    else:
        # ── KPIs ──
        _ek = embarque_kpis
        _ek_c1, _ek_c2, _ek_c3, _ek_c4 = st.columns(4)
        with _ek_c1:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {TEAL_600};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Ventanas activas</div>
                <div style="font-size:1.6rem; font-weight:700; color:{TEAL_700};">{_ek.get('n_ventanas', 0)}</div>
            </div>""", unsafe_allow_html=True)
        with _ek_c2:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {STATUS_OPTIMO};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Mejor ventana</div>
                <div style="font-size:1.6rem; font-weight:700; color:{STATUS_OPTIMO};">{_ek.get('mejor_ventana', '—')}</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Retorno/sem: {_ek.get('mejor_retorno', 0):.2%}</div>
            </div>""", unsafe_allow_html=True)
        with _ek_c3:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {STATUS_CRITICO};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Peor ventana</div>
                <div style="font-size:1.6rem; font-weight:700; color:{STATUS_CRITICO};">{_ek.get('peor_ventana', '—')}</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Retorno/sem: {_ek.get('peor_retorno', 0):.2%}</div>
            </div>""", unsafe_allow_html=True)
        with _ek_c4:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {STATUS_PRECRITICO};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Ventanas en rojo</div>
                <div style="font-size:1.6rem; font-weight:700; color:{STATUS_PRECRITICO};">{_ek.get('n_rojos', 0)}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Tabla resumen por ventana ──
        _sem_colors = {"verde": STATUS_OPTIMO, "amarillo": STATUS_ALTO, "rojo": STATUS_CRITICO}
        _sem_icons = {"verde": "🟢", "amarillo": "🟡", "rojo": "🔴"}

        for _, vrow in por_ventana.iterrows():
            _v = vrow["ventana"]
            _sem = vrow.get("semaforo", "verde")
            _icon = _sem_icons.get(_sem, "⚪")
            _border_color = _sem_colors.get(_sem, SLATE_400)

            st.markdown(f"""<div style="background:var(--capi-bg-card); border:1px solid var(--capi-border); border-left:4px solid {_border_color}; border-radius:10px; padding:12px 16px; margin-bottom:6px;">
            <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap;">
                <div>
                    <strong style="color:var(--capi-text); font-size:1rem;">{_icon} Ventana {_v}</strong>
                    <span style="color:var(--capi-text2); font-size:0.82em;"> &nbsp;·&nbsp; {vrow['label']}</span>
                </div>
                <div style="display:flex; gap:20px; font-size:0.85em;">
                    <span><strong>ST:</strong> {vrow['sell_through']:.0%}</span>
                    <span><strong>Cob:</strong> {vrow['cobertura_prom']:.1f} sem</span>
                    <span><strong>Capital:</strong> S/ {vrow['capital']:,.0f}</span>
                    <span><strong>SKUs:</strong> {vrow['n_skus']:,}</span>
                    <span><strong>Dscto:</strong> {vrow['descuento_prom']:.0%}</span>
                </div>
            </div>
            </div>""", unsafe_allow_html=True)

            # Recomendación
            _rec = embarque_recs.get(_v, "")
            if _rec:
                with st.expander(f"Ver recomendación — Ventana {_v}", expanded=False):
                    st.markdown(_rec)

            # Top 5 SKUs problemáticos
            _top_df = embarque_top.get(_v, pd.DataFrame())
            if not _top_df.empty:
                with st.expander(f"Top 5 SKUs problemáticos — Ventana {_v}", expanded=False):
                    _top_disp = _top_df.rename(columns={
                        "sku": "SKU", "nombre": "Producto", "marca": "Marca", "tienda": "Tienda",
                        "stock_total": "Stock Uds", "stock_valor_costo": "Capital S/",
                        "sell_through": "ST %", "cobertura_sem": "Cob (sem)",
                        "pct_descuento": "Dscto", "edad_semanas": "Edad (sem)",
                    })
                    _top_show = [c for c in ["SKU", "Producto", "Marca", "Tienda", "Stock Uds", "Capital S/", "ST %", "Cob (sem)", "Dscto", "Edad (sem)"] if c in _top_disp.columns]
                    _top_fmt = {}
                    if "Capital S/" in _top_show: _top_fmt["Capital S/"] = "S/ {:,.0f}"
                    if "ST %" in _top_show: _top_fmt["ST %"] = "{:.0%}"
                    if "Cob (sem)" in _top_show: _top_fmt["Cob (sem)"] = "{:.1f}"
                    if "Dscto" in _top_show: _top_fmt["Dscto"] = "{:.0%}"
                    st.dataframe(
                        _top_disp[_top_show].style.format(_top_fmt, na_rep="—"),
                        use_container_width=True, hide_index=True,
                    )

        # ── Drill-down: detalle completo por ventana ──
        st.markdown("---")
        st.markdown("##### Detalle por Ventana de Compra")

        if not df_embarque.empty:
            _vc_ventanas = sorted(df_embarque["ventana_compra"].unique().tolist())
            _vc_sel = st.selectbox("Seleccionar ventana", _vc_ventanas, key="vc_drill_ventana")
            _vc_det = df_embarque[df_embarque["ventana_compra"] == _vc_sel].copy()

            _vc_show_cols = [c for c in ["sku", "nombre", "marca", "tienda", "stock_total",
                                          "stock_valor_costo", "sell_through", "cobertura_sem",
                                          "pct_descuento", "edad_semanas"] if c in _vc_det.columns]
            _vc_det_disp = _vc_det[_vc_show_cols].sort_values("stock_valor_costo", ascending=False).head(50)
            _vc_det_disp = _vc_det_disp.rename(columns={
                "sku": "SKU", "nombre": "Producto", "marca": "Marca", "tienda": "Tienda",
                "stock_total": "Stock Uds", "stock_valor_costo": "Capital S/",
                "sell_through": "ST %", "cobertura_sem": "Cob (sem)",
                "pct_descuento": "Dscto", "edad_semanas": "Edad (sem)",
            })
            _vc_fmt = {}
            if "Capital S/" in _vc_det_disp.columns: _vc_fmt["Capital S/"] = "S/ {:,.0f}"
            if "ST %" in _vc_det_disp.columns: _vc_fmt["ST %"] = "{:.0%}"
            if "Cob (sem)" in _vc_det_disp.columns: _vc_fmt["Cob (sem)"] = "{:.1f}"
            if "Dscto" in _vc_det_disp.columns: _vc_fmt["Dscto"] = "{:.0%}"
            st.dataframe(
                _vc_det_disp.style.format(_vc_fmt, na_rep="—"),
                use_container_width=True, hide_index=True, height=500,
            )
            st.caption(f"Top 50 combos por capital invertido en ventana {_vc_sel}. Total: {len(_vc_det):,} combos.")


# ══════════════════════════════════════════════════════════════
#  PREDISTRIBUCIÓN (Marcas Propias)
# ══════════════════════════════════════════════════════════════

elif nav_page == "🚚 Predistribución":
    st.markdown(f'<div class="section-header"><h3>🚚 Monitor de Predistribución</h3><span class="live-badge">MARCAS PROPIAS</span></div>', unsafe_allow_html=True)
    st.caption("Detecta productos propios retenidos en CD sin distribuir y gaps de distribución vs la matriz configurable de tiendas por línea.")

    # ── Tab layout ──
    _pd_tab1, _pd_tab2, _pd_tab3 = st.tabs([
        "🔴 Retenidos en CD",
        "🟡 Gaps de Distribución",
        "⚙️ Configurar Matriz",
    ])

    # ══════════════ TAB 1: Retenidos en CD ══════════════
    with _pd_tab1:
        st.markdown("#### Productos 100% en CD — sin stock ni tránsito en ninguna tienda")
        st.caption("SKUs de marcas propias con todo el inventario en Centro de Distribución y nada enviado a tiendas. Requieren predistribución urgente.")

        _n_ret = predist_kpis.get('n_retenidos', 0)
        _uds_ret = predist_kpis.get('uds_retenidas_cd', 0)
        _cap_ret = predist_kpis.get('capital_retenido', 0)

        # KPIs
        _r_c1, _r_c2, _r_c3 = st.columns(3)
        with _r_c1:
            _ret_color = STATUS_CRITICO if _n_ret > 0 else STATUS_OPTIMO
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {_ret_color};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">SKUs retenidos en CD</div>
                <div style="font-size:1.6rem; font-weight:700; color:{_ret_color};">{_n_ret:,}</div>
            </div>""", unsafe_allow_html=True)
        with _r_c2:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {TEAL_600};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Unidades retenidas</div>
                <div style="font-size:1.6rem; font-weight:700; color:{TEAL_700};">{_uds_ret:,}</div>
            </div>""", unsafe_allow_html=True)
        with _r_c3:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {STATUS_ALTO};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Capital retenido</div>
                <div style="font-size:1.6rem; font-weight:700; color:{STATUS_ALTO};">S/ {_cap_ret:,.0f}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("")

        if df_retenidos_cd.empty:
            st.success("No hay productos 100% retenidos en CD. Todos los SKUs propios tienen alguna distribución a tiendas.")
        else:
            # Tabla de retenidos
            _ret_disp = df_retenidos_cd.copy()
            _ret_disp = _ret_disp.rename(columns={
                'sku': 'SKU', 'nombre': 'Producto', 'marca': 'Marca',
                'categoria': 'Línea', 'stock_cd': 'Stock CD',
                'costo': 'Costo Unit', 'precio_vigente': 'P.Vigente',
                'capital_retenido': 'Capital Retenido S/',
            })
            _ret_fmt = {
                'Costo Unit': 'S/ {:.2f}',
                'P.Vigente': 'S/ {:.2f}',
                'Capital Retenido S/': 'S/ {:,.0f}',
            }
            st.dataframe(
                _ret_disp.style.format(_ret_fmt, na_rep="—"),
                use_container_width=True, hide_index=True, height=400,
            )
            st.caption(f"{_n_ret} SKUs propios con 100% del stock en CD. Capital total: S/ {_cap_ret:,.0f}")

            # Excel download
            _ret_xl_buf = io.BytesIO()
            with pd.ExcelWriter(_ret_xl_buf, engine='openpyxl') as _w:
                df_retenidos_cd.to_excel(_w, sheet_name='Retenidos CD', index=False)
            st.download_button(
                "📥 Descargar Retenidos CD (.xlsx)",
                data=_ret_xl_buf.getvalue(),
                file_name="predist_retenidos_cd.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_retenidos_cd",
            )

    # ══════════════ TAB 2: Gaps de Distribución ══════════════
    with _pd_tab2:
        st.markdown("#### Gaps de distribución vs matriz de tiendas")
        st.caption("SKUs propios que están en algunas tiendas pero faltan en otras según la segmentación configurada por línea de producto.")

        _n_gaps = predist_kpis.get('n_gaps', 0)
        _n_gaps_cd = predist_kpis.get('n_gaps_con_cd', 0)
        _prom_cob = predist_kpis.get('prom_cobertura_dist', 1.0)

        # KPIs
        _g_c1, _g_c2, _g_c3 = st.columns(3)
        with _g_c1:
            _gap_color = STATUS_ALTO if _n_gaps > 0 else STATUS_OPTIMO
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {_gap_color};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">SKUs con gaps</div>
                <div style="font-size:1.6rem; font-weight:700; color:{_gap_color};">{_n_gaps:,}</div>
            </div>""", unsafe_allow_html=True)
        with _g_c2:
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {TEAL_600};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Gaps con stock CD</div>
                <div style="font-size:1.6rem; font-weight:700; color:{TEAL_700};">{_n_gaps_cd:,}</div>
                <div style="font-size:0.7rem; color:var(--capi-text2);">Accionables (hay stock para enviar)</div>
            </div>""", unsafe_allow_html=True)
        with _g_c3:
            _cob_color = STATUS_OPTIMO if _prom_cob >= 0.8 else (STATUS_ALTO if _prom_cob >= 0.5 else STATUS_CRITICO)
            st.markdown(f"""<div style="background:var(--capi-bg-surface); border-radius:12px; padding:16px 20px; border-left:4px solid {_cob_color};">
                <div style="font-size:0.75rem; color:var(--capi-text2); font-weight:500;">Cobertura distribución prom.</div>
                <div style="font-size:1.6rem; font-weight:700; color:{_cob_color};">{_prom_cob:.0%}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("")

        if df_gaps_dist.empty:
            st.success("No hay gaps de distribución. Todos los SKUs propios están en las tiendas esperadas según la matriz.")
        else:
            # Filtros
            _gf_c1, _gf_c2, _gf_c3, _gf_c4 = st.columns(4)
            with _gf_c1:
                _gap_lineas = ['Todas'] + sorted(df_gaps_dist['categoria'].unique().tolist())
                _gap_linea_sel = st.selectbox("Filtrar por Línea", _gap_lineas, key="gap_linea_filter")
            with _gf_c2:
                _gap_marcas = ['Todas'] + sorted(df_gaps_dist['marca'].unique().tolist())
                _gap_marca_sel = st.selectbox("Filtrar por Marca", _gap_marcas, key="gap_marca_filter")
            with _gf_c3:
                _gap_solo_cd = st.checkbox("Solo con stock CD > 0", value=True, key="gap_solo_cd")
            with _gf_c4:
                _gap_solo_nuevos = st.checkbox("Solo 0-3 meses", value=True, key="gap_solo_nuevos",
                                               help="Filtrar solo SKUs con edad ≤ 12 semanas (0-3 meses)")

            _gaps_filtered = df_gaps_dist.copy()
            if _gap_linea_sel != 'Todas':
                _gaps_filtered = _gaps_filtered[_gaps_filtered['categoria'] == _gap_linea_sel]
            if _gap_marca_sel != 'Todas':
                _gaps_filtered = _gaps_filtered[_gaps_filtered['marca'] == _gap_marca_sel]
            if _gap_solo_cd:
                _gaps_filtered = _gaps_filtered[_gaps_filtered['stock_cd'] > 0]
            if _gap_solo_nuevos and 'edad_semanas' in _gaps_filtered.columns:
                _gaps_filtered = _gaps_filtered[_gaps_filtered['edad_semanas'] <= 12]

            if _gaps_filtered.empty:
                st.info("No hay gaps con los filtros seleccionados.")
            else:
                # Ordenar por Stock CD descendente
                _gaps_filtered = _gaps_filtered.sort_values('stock_cd', ascending=False)

                _gaps_disp = _gaps_filtered.rename(columns={
                    'sku': 'SKU', 'nombre': 'Producto', 'marca': 'Marca',
                    'categoria': 'Línea', 'edad_semanas': 'Edad (sem)',
                    'stock_cd': 'Stock CD',
                    'costo': 'Costo Unit', 'precio_vigente': 'P.Vigente',
                    'n_tiendas_esperadas': 'Tiendas Esperadas',
                    'n_tiendas_presentes': 'Tiendas Presentes',
                    'n_tiendas_faltantes': 'Tiendas Faltantes',
                    'pct_cobertura_dist': '% Cobertura',
                    'tiendas_faltantes': 'Detalle Tiendas Faltantes',
                })
                _gaps_fmt = {
                    'Costo Unit': 'S/ {:.2f}',
                    'P.Vigente': 'S/ {:.2f}',
                    '% Cobertura': '{:.0%}',
                }
                st.dataframe(
                    _gaps_disp.style.format(_gaps_fmt, na_rep="—"),
                    use_container_width=True, hide_index=True, height=500,
                )
                st.caption(f"{len(_gaps_filtered):,} SKUs con gaps ({_gaps_filtered[_gaps_filtered['stock_cd'] > 0].shape[0]} con stock CD disponible)")

                # Excel download
                _gaps_xl_buf = io.BytesIO()
                with pd.ExcelWriter(_gaps_xl_buf, engine='openpyxl') as _w:
                    _gaps_filtered.to_excel(_w, sheet_name='Gaps Distribución', index=False)
                st.download_button(
                    "📥 Descargar Gaps de Distribución (.xlsx)",
                    data=_gaps_xl_buf.getvalue(),
                    file_name="predist_gaps_distribucion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_gaps_dist",
                )

    # ══════════════ TAB 3: Configurar Matriz ══════════════
    with _pd_tab3:
        st.markdown("#### Matriz de tiendas por marca")
        st.caption("Define qué tiendas deberían recibir cada marca. Generada desde el archivo de distribución marca×tienda que pasó Franco. Editable manualmente.")

        import json as _json_mod
        _matriz_marca_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config_marca_tiendas.json')

        if os.path.exists(_matriz_marca_path):
            with open(_matriz_marca_path, 'r', encoding='utf-8') as _mf:
                _matriz_marca_actual = _json_mod.load(_mf)
        else:
            _matriz_marca_actual = {}
            st.warning("No se encontró config_marca_tiendas.json. Sube el archivo de distribución marca×tienda para generar la matriz.")

        if _matriz_marca_actual:
            # Resumen de la matriz
            _mat_rows = []
            for _marca, _info in sorted(_matriz_marca_actual.items()):
                _vta = _info.get('venta_total', 0)
                _mat_rows.append({
                    'Marca': _marca,
                    'N° Tiendas': _info.get('n_tiendas', len(_info.get('tiendas', []))),
                    'Venta S/': _vta,
                    'Tiendas': ', '.join(_info.get('tiendas', [])[:5]) + ('...' if len(_info.get('tiendas', [])) > 5 else ''),
                })
            _df_mat = pd.DataFrame(_mat_rows)
            _mat_fmt = {'Venta S/': 'S/ {:,.0f}'}
            st.dataframe(
                _df_mat.style.format(_mat_fmt, na_rep="—"),
                use_container_width=True, hide_index=True,
            )

            # Detalle editable por marca
            st.markdown("---")
            st.markdown("##### Editar tiendas por marca")
            _marcas_disponibles = sorted(_matriz_marca_actual.keys())
            _marca_editar = st.selectbox("Seleccionar marca", _marcas_disponibles, key="mat_editar_marca")

            if _marca_editar:
                _tiendas_marca = _matriz_marca_actual[_marca_editar].get('tiendas', [])
                st.markdown(f"**{_marca_editar}** — {len(_tiendas_marca)} tiendas activas:")

                # Mostrar todas las tiendas posibles y marcar las activas
                from transformar_profundidad import STORE_NAMES as _ALL_STORES
                _todas_tiendas = sorted(set(_ALL_STORES.values()) - {
                    'Tienda Virtual', 'Tienda Virtual PI', 'Ventas Corporativas',
                    'FSF Mac LO', 'Boutique BBB', 'Asia',
                    'Estacional Chiclayo', 'Estacional Trujillo', 'Estacional VES',
                    'Mac Larco', 'Mac Plaza Norte',
                    'Outlet Plaza Norte', 'Outlet San Isidro',
                })

                _tiendas_set = set(_tiendas_marca)
                _nuevas_tiendas = st.multiselect(
                    f"Tiendas para {_marca_editar}",
                    options=_todas_tiendas,
                    default=[t for t in _todas_tiendas if t in _tiendas_set],
                    key=f"mat_tiendas_{_marca_editar}",
                )

                if st.button("💾 Guardar cambios en matriz", key="mat_guardar"):
                    # Reconstruir tiendas_code desde STORE_NAMES inverso
                    _inv_store = {v: k for k, v in _ALL_STORES.items()}
                    _nuevos_codes = sorted([_inv_store[t] for t in _nuevas_tiendas if t in _inv_store])
                    _matriz_marca_actual[_marca_editar] = {
                        'tiendas': sorted(_nuevas_tiendas),
                        'tiendas_code': _nuevos_codes,
                        'n_tiendas': len(_nuevas_tiendas),
                        'venta_total': _matriz_marca_actual[_marca_editar].get('venta_total', 0),
                    }
                    with open(_matriz_marca_path, 'w', encoding='utf-8') as _mf:
                        _json_mod.dump(_matriz_marca_actual, _mf, ensure_ascii=False, indent=2)
                    st.success(f"Matriz actualizada: {_marca_editar} → {len(_nuevas_tiendas)} tiendas. Re-corre el análisis para ver el efecto.")

        # ── Editor por LÍNEA (fallback) ──
        st.markdown("---")
        st.markdown("#### Matriz de tiendas por línea (fallback)")
        st.caption("Se usa cuando una marca no tiene asignación propia. Define qué tiendas deberían recibir cada línea de producto.")

        _matriz_linea_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config_matriz_tiendas.json')

        if os.path.exists(_matriz_linea_path):
            with open(_matriz_linea_path, 'r', encoding='utf-8') as _mlf:
                _matriz_linea_actual = _json_mod.load(_mlf)
        else:
            _matriz_linea_actual = {}
            st.warning("No se encontró config_matriz_tiendas.json.")

        if _matriz_linea_actual:
            # Resumen de la matriz por línea
            _mat_lin_rows = []
            for _linea_k, _info_l in sorted(_matriz_linea_actual.items()):
                _mat_lin_rows.append({
                    'Línea': _linea_k,
                    'N° Tiendas': _info_l.get('n_tiendas', len(_info_l.get('tiendas', []))),
                    'Tiendas': ', '.join(_info_l.get('tiendas', [])[:5]) + ('...' if len(_info_l.get('tiendas', [])) > 5 else ''),
                })
            _df_mat_lin = pd.DataFrame(_mat_lin_rows)
            st.dataframe(_df_mat_lin, use_container_width=True, hide_index=True)

            # Detalle editable por línea
            st.markdown("---")
            st.markdown("##### Editar tiendas por línea")
            _lineas_disponibles = sorted(_matriz_linea_actual.keys())
            _linea_editar = st.selectbox("Seleccionar línea", _lineas_disponibles, key="mat_editar_linea")

            if _linea_editar:
                _tiendas_linea = _matriz_linea_actual[_linea_editar].get('tiendas', [])
                st.markdown(f"**{_linea_editar}** — {len(_tiendas_linea)} tiendas activas:")

                from transformar_profundidad import STORE_NAMES as _ALL_STORES_L
                _todas_tiendas_l = sorted(set(_ALL_STORES_L.values()) - {
                    'Tienda Virtual', 'Tienda Virtual PI', 'Ventas Corporativas',
                    'FSF Mac LO', 'Boutique BBB', 'Asia',
                    'Estacional Chiclayo', 'Estacional Trujillo', 'Estacional VES',
                    'Mac Larco', 'Mac Plaza Norte',
                    'Outlet Plaza Norte', 'Outlet San Isidro',
                })

                _tiendas_lin_set = set(_tiendas_linea)
                _nuevas_tiendas_l = st.multiselect(
                    f"Tiendas para {_linea_editar}",
                    options=_todas_tiendas_l,
                    default=[t for t in _todas_tiendas_l if t in _tiendas_lin_set],
                    key=f"mat_tiendas_linea_{_linea_editar}",
                )

                if st.button("💾 Guardar cambios en matriz de líneas", key="mat_guardar_linea"):
                    _inv_store_l = {v: k for k, v in _ALL_STORES_L.items()}
                    _nuevos_codes_l = sorted([_inv_store_l[t] for t in _nuevas_tiendas_l if t in _inv_store_l])
                    _matriz_linea_actual[_linea_editar] = {
                        'tiendas': sorted(_nuevas_tiendas_l),
                        'tiendas_code': _nuevos_codes_l,
                        'n_tiendas': len(_nuevas_tiendas_l),
                    }
                    with open(_matriz_linea_path, 'w', encoding='utf-8') as _mlf:
                        _json_mod.dump(_matriz_linea_actual, _mlf, ensure_ascii=False, indent=2)
                    st.success(f"Matriz actualizada: {_linea_editar} → {len(_nuevas_tiendas_l)} tiendas. Re-corre el análisis para ver el efecto.")
# ─── CASO DE ÉXITO (Fase 1-2 auditoría 2026-08-23) ───────────

elif nav_page == "🏆 Caso de Éxito":
    st.markdown("#### 🏆 Caso de Éxito — capital en exceso, semana a semana")
    st.caption("La métrica titular para gerencia: capital a costo en DORMIDO + ESTANCADO + "
               "SOBRESTOCK + LIQUIDAR + MUERTO. Cada acción registrada hace el delta atribuible.")

    if not _HAS_SNAPSHOTS:
        st.warning("Esta vista necesita el módulo de snapshots.")
    else:
        _ce_serie = snapshots_engine.api.serie_capital_exceso()
        if _ce_serie.empty or len(_ce_serie) < 2:
            st.info("Aún no hay suficientes snapshots para la serie (mínimo 2 semanas).")
        else:
            _ce_ult = _ce_serie.iloc[-1]
            _ce_prev = _ce_serie.iloc[-2]
            _c1, _c2, _c3, _c4 = st.columns(4)
            _c1.metric("Capital en exceso", f"S/ {_ce_ult['capital_exceso']:,.0f}",
                       delta=f"{_ce_ult['delta_exceso_pct']:+.1f}% vs {_ce_prev['semana_iso']}",
                       delta_color="inverse")
            _c2.metric("% del capital total", f"{_ce_ult['pct_exceso']*100:.1f}%")
            _c3.metric("SKUs en exceso", f"{int(_ce_ult['skus_exceso']):,}")
            _c4.metric("Semana del corte", analisis_estados.etiqueta_semana(_ce_ult["semana_iso"], corta=True))

            st.markdown("##### Evolución del capital en exceso")
            _ce_chart = _ce_serie.copy()
            _ce_chart["semana_iso"] = _ce_chart["semana_iso"].map(
                lambda w: analisis_estados.etiqueta_semana(w, corta=True))
            _ce_chart = _ce_chart.set_index("semana_iso")[["capital_exceso", "capital_total"]]
            _ce_chart.columns = ["Capital en exceso", "Capital total"]
            st.line_chart(_ce_chart, height=260)
            st.caption("⚠️ Las semanas no son consecutivas todavía (huecos entre cortes). "
                       "La disciplina semanal arranca en la semana 0 declarada.")

            with st.expander("Ver serie completa por estado", expanded=False):
                _ce_est = snapshots_engine.api.serie_capital_estados()
                _ce_piv = _ce_est.pivot_table(index="semana_iso", columns="estado",
                                              values="capital", aggfunc="sum", fill_value=0)
                st.dataframe(_ce_piv.style.format("S/ {:,.0f}"), use_container_width=True)

            # ── 🧠 Análisis de estados: conclusiones automáticas ──
            st.markdown("---")
            st.markdown("##### 🧠 Análisis de estados — qué se movió, quién lo impulsó y qué hacer")
            st.caption("Conclusiones generadas por reglas de retail sobre los movimientos "
                       "entre dos cortes: subidas/bajadas por estado, marca que impulsa, "
                       "migraciones (¿mejora real o deterioro encubierto?) y atribución.")
            _an_weeks = list(_ce_serie["semana_iso"])
            _an_c1, _an_c2 = st.columns(2)
            _an_a = _an_c1.selectbox("Comparar desde", _an_weeks[:-1],
                                     index=len(_an_weeks) - 2, key="an_sem_a",
                                     format_func=analisis_estados.etiqueta_semana)
            _an_b_opts = [w for w in _an_weeks if w > _an_a]
            _an_b = _an_c2.selectbox("hasta", _an_b_opts,
                                     index=len(_an_b_opts) - 1, key="an_sem_b",
                                     format_func=analisis_estados.etiqueta_semana)
            _an_conc = analisis_estados.conclusiones(_an_a, _an_b, acciones_log.cargar())
            if not _an_conc:
                st.info("Sin movimientos materiales entre esas dos semanas (umbral S/ 100K).")
            for _an_c in _an_conc:
                _an_render = {"positivo": st.success, "atencion": st.warning,
                              "critico": st.error}.get(_an_c["nivel"], st.info)
                _an_render(f"**{_an_c['titulo']}** — {_an_c['detalle']}")

            # ── 🔀 Migraciones entre estados, semana a semana ──
            st.markdown("---")
            st.markdown("##### 🔀 Migraciones entre estados — dónde están las oportunidades")
            st.caption("Capital que MEJORÓ de estado (reactivaciones: dormido que vuelve a "
                       "vender, sobrestock que drena) vs capital que se DETERIORÓ. El "
                       "deterioro recurrente marca la oportunidad de gestión.")
            _mg_serie = analisis_estados.serie_migraciones()
            if not _mg_serie.empty:
                _mg_ult = _mg_serie.iloc[-1]
                _mg_k1, _mg_k2, _mg_k3 = st.columns(3)
                _mg_k1.metric("Capital que mejoró", f"S/ {_mg_ult['capital_mejora']/1e6:,.2f}M",
                              help=f"Par {_mg_ult['par']}")
                _mg_k2.metric("Capital que se deterioró", f"S/ {_mg_ult['capital_deterioro']/1e6:,.2f}M")
                _mg_k3.metric("Neto", f"S/ {_mg_ult['neto']/1e6:+,.2f}M",
                              delta=f"{_mg_ult['neto']/1e6:+,.2f}M", delta_color="normal")
                _mg_chart = _mg_serie.set_index("par")[["capital_mejora", "capital_deterioro"]]
                _mg_chart.columns = ["Mejoró de estado", "Se deterioró"]
                st.bar_chart(_mg_chart, height=240, color=["#2E7D5B", "#A03028"])

            _mg_m = analisis_estados.matriz_migraciones(_an_a, _an_b)
            if _mg_m.empty:
                st.info(f"Sin migraciones entre {_an_a} y {_an_b}.")
            else:
                st.markdown(f"**Flujos {_an_a} → {_an_b}** (ordenados por capital)")
                _mg_show = _mg_m.head(15).copy()
                _mg_show["clase"] = _mg_show["clase"].map(
                    {"mejora": "✅ mejora", "deterioro": "🔻 deterioro",
                     "lateral": "↔ lateral", "relanzamiento": "🔁 relanzamiento"})
                _mg_show.columns = ["De", "A", "SKUs", "Capital S/", "Clase"]
                st.dataframe(_mg_show.style.format({"Capital S/": "S/ {:,.0f}"}),
                             use_container_width=True, hide_index=True, height=320)

                # Drill-down: SKUs de un flujo específico
                _mg_ops = [f"{r['estado_a']} → {r['estado_b']}" for _, r in _mg_m.iterrows()]
                _mg_sel = st.selectbox("Ver los SKUs de un flujo (para accionar)", _mg_ops,
                                       key="mg_flujo_sel")
                _mg_ea, _mg_eb = [x.strip() for x in _mg_sel.split("→")]
                _mg_det = analisis_estados.detalle_migracion(_an_a, _an_b, _mg_ea, _mg_eb)
                if not _mg_det.empty:
                    st.dataframe(_mg_det.style.format(
                        {"stock_valor_costo": "S/ {:,.0f}", "cobertura_sem": "{:.1f}",
                         "edad_semanas": "{:.0f}"}, na_rep="—"),
                        use_container_width=True, hide_index=True, height=300)

            # ── Joyas rescatadas de Evolución Semanal (poda C3 2026-08-26) ──
            with st.expander("✅ Cumplimiento de reposición — ¿se ejecutó lo que Capi sugirió?", expanded=False):
                try:
                    _cum_w = snapshots_engine.api.list_available_weeks() if hasattr(snapshots_engine.api, "list_available_weeks") else None
                except Exception:
                    _cum_w = None
                if _cum_w is None:
                    from snapshots_engine.storage import list_available_weeks as _law_ce
                    _cum_w = _law_ce()
                if len(_cum_w) >= 2:
                    # S6 v1 (2026-09-05): antes se filtraba por una columna `cumplido` que no
                    # existía y la métrica daba N/N siempre. Ahora: pedido (acciones_log) × observado (snapshots).
                    _ce = analisis_estados.cumplimiento_empujes(_cum_w[-2], _cum_w[-1])
                    _ce_df = _ce["df"]
                    _c1, _c2, _c3 = st.columns(3)
                    _c1.metric("Empujes pedidos", f"{_ce['n_pedidos']}",
                               help=f"Registrados en el log como 'Reposición / Empuje' en la semana {_cum_w[-2]}")
                    _c2.metric("Cumplimiento", (f"{_ce['pct']:.0f}%" if _ce["pct"] is not None else "—"),
                               help="Pedidos cuyo stock en tiendas subió más de lo que la venta explica. "
                                    "— = no hay pedidos registrados esa semana, no se inventa un 100%")
                    _c3.metric("Recibidos sin pedir", f"{_ce['n_sin_pedir']}",
                               help="Movimientos CD→tienda que nadie registró como empuje")
                    if _ce["n_pedidos"] == 0:
                        st.caption("Para medir cumplimiento hay que marcar los empujes enviados en "
                                   "🎯 Match Producto-Plaza (\"marcar como ejecutado\") la semana que se piden.")
                    if _ce_df.empty:
                        st.info("Sin empujes pedidos ni movimientos CD→tienda entre los dos últimos cortes.")
                    else:
                        st.dataframe(_ce_df.head(200), use_container_width=True, hide_index=True, height=280)

            with st.expander("🔮 Predicción de quiebre — qué se agota antes del próximo corte", expanded=False):
                try:
                    _pq = snapshots_engine.api.predict_stockout()
                    if _pq is None or _pq.empty:
                        st.info("Sin quiebres proyectados con la velocidad actual.")
                    else:
                        st.dataframe(_pq.head(100), use_container_width=True, hide_index=True, height=260)
                except Exception as _e_pq:
                    st.info(f"Predicción no disponible: {_e_pq}")

        # ── Registro de acciones (Gap G3: atribución) ──
        st.markdown("---")
        st.markdown("##### 📋 Acciones de gestión (lo que hace el delta atribuible)")
        _tab_reg, _tab_log = st.tabs(["➕ Registrar acción", "📜 Log de acciones"])

        with _tab_reg:
            with st.form("form_accion", clear_on_submit=True):
                _fa1, _fa2, _fa3 = st.columns(3)
                _ac_sem = _fa1.text_input("Semana ISO", value="", placeholder="2026-35 (vacío = actual)")
                _ac_tipo = _fa2.selectbox("Tipo", acciones_log.TIPOS)
                _ac_marca = _fa3.text_input("Marca", placeholder="LACOSTE")
                _ac_desc = st.text_input("Descripción", placeholder="Markdown 40% aprobado en 12 modelos MUERTO")
                _fb1, _fb2, _fb3 = st.columns(3)
                _ac_mag = _fb1.text_input("Magnitud", placeholder="40% / 500 uds / S/ 80K")
                _ac_orig = _fb2.selectbox("Origen", acciones_log.ORIGENES)
                _ac_est = _fb3.selectbox("Estado", acciones_log.ESTADOS_ACCION)
                if st.form_submit_button("Registrar", use_container_width=True):
                    if _ac_desc.strip():
                        acciones_log.agregar(_ac_sem.strip(), _ac_tipo, _ac_marca.strip().upper(),
                                             _ac_desc.strip(), magnitud=_ac_mag.strip(),
                                             origen=_ac_orig, estado=_ac_est)
                        st.success("Acción registrada ✅")
                    else:
                        st.warning("La descripción es obligatoria.")

        with _tab_log:
            _log = acciones_log.cargar()
            if _log.empty:
                st.info("Sin acciones registradas todavía. La primera se registra en la pestaña de al lado.")
            else:
                st.dataframe(_log.sort_values("fecha_registro", ascending=False),
                             use_container_width=True, hide_index=True, height=300)
                st.download_button("📥 Descargar log (CSV)",
                                   data=_log.to_csv(index=False).encode("utf-8"),
                                   file_name="acciones_log.csv", mime="text/csv")
                st.caption("☁️ En la nube el registro es temporal: descarga el CSV al terminar y "
                           "pásalo al flujo semanal (se versiona en el repo junto al snapshot).")
